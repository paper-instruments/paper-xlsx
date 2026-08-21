"""PR 8: scale, determinism, and release-gate honesty for Paper-managed pivots."""
from __future__ import annotations

import os
import sys

import pytest

from openpyxl import Workbook, load_workbook
from openpyxl.errors import BoundaryViolationError
from openpyxl.worksheet.table import Table

from .support.harness import save_and_reopen
from .support.partdiff import part_payloads


_PIVOT_PARTS = (
    "xl/pivotCache/pivotCacheDefinition1.xml",
    "xl/pivotCache/pivotCacheRecords1.xml",
    "xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels",
    "xl/pivotTables/pivotTable1.xml",
    "xl/pivotTables/_rels/pivotTable1.xml.rels",
)


def _rows(count):
    regions = ("East", "West", "North", "South")
    products = ("A", "B", "C")
    rows = []
    for index in range(count):
        rows.append((
            regions[index % 4],
            products[index % 3],
            float(index % 50),
        ))
    return rows


def _source_workbook(count):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(("Region", "Product", "Amount"))
    for row in _rows(count):
        ws.append(row)
    ws.add_table(Table(displayName="Sales", ref="A1:C%s" % (count + 1)))
    wb.create_sheet("Summary")
    return wb


def _create_by_region(workbook):
    return workbook["Summary"].pivots.create(
        name="ByRegion", source="Sales", destination="A1",
        rows=["Region"], columns=["Product"], values=["Amount"])


def _output_values(worksheet):
    cells = getattr(worksheet, "_cells", {})
    return {
        (row, column): cell.value
        for (row, column), cell in sorted(cells.items())
        if cell.value is not None
    }


def test_moderate_source_is_deterministic(tmp_path):
    count = 100000 if os.environ.get("PAPER_PIVOT_SCALE") == "1" else 2000
    source = str(tmp_path / "a.xlsx")
    _source_workbook(count).save(source)

    first = load_workbook(source, preserve=True)
    _create_by_region(first)
    first = save_and_reopen(first, str(tmp_path / "b.xlsx"), preserve=True)

    second = load_workbook(source, preserve=True)
    _create_by_region(second)
    second = save_and_reopen(second, str(tmp_path / "c.xlsx"), preserve=True)

    left = part_payloads(str(tmp_path / "b.xlsx"))
    right = part_payloads(str(tmp_path / "c.xlsx"))
    for name in _PIVOT_PARTS:
        assert left[name] == right[name]
    assert _output_values(first["Summary"]) == _output_values(second["Summary"])
    assert first["Summary"].pivots["ByRegion"].spec.rows[0].field == "Region"


def test_output_cardinality_refuses_before_cells(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(("Region", "Amount"))
    ws.append(("East", 1))
    ws.append(("West", 2))
    ws.add_table(Table(displayName="Sales", ref="A1:B3"))
    wb.create_sheet("Summary")
    path = str(tmp_path / "src.xlsx")
    wb.save(path)
    wb = load_workbook(path, preserve=True)
    from openpyxl.pivot.plan import plan_pivot
    from openpyxl.pivot.source import snapshot_from_workbook, PivotLimits
    from openpyxl.pivot.api_types import (
        PivotMeasure, PivotSource, PivotSpec, PivotAxisField,
    )
    snapshot = snapshot_from_workbook(wb, "Sales")
    spec = PivotSpec(
        name="ByRegion",
        source=PivotSource.table("Sales"),
        destination="A1",
        rows=(PivotAxisField("Region"),),
        values=(PivotMeasure("Amount"),),
    )
    with pytest.raises(BoundaryViolationError) as exc:
        plan_pivot(spec, snapshot, limits=PivotLimits(output_cells=1))
    assert exc.value.kind == "pivot-output-too-large"
    assert wb["Summary"]["A1"].value is None
    assert list(wb["Summary"].pivots) == []


def test_readme_pivot_example_is_runnable(tmp_path):
    sales = str(tmp_path / "sales.xlsx")
    out = str(tmp_path / "sales-out.xlsx")

    wb = Workbook()
    data = wb.active
    data.title = "Data"
    data.append(("Region", "Amount"))
    data.append(("East", 10))
    data.append(("West", 20))
    data.add_table(Table(displayName="RegionTable", ref="A1:B3"))
    wb.create_sheet("Summary")
    wb.save(sales)

    wb = load_workbook(sales, preserve=True)
    pivot = wb["Summary"].pivots.create(
        name="ByRegion",
        source="RegionTable",
        destination="A1",
        rows=["Region"],
        values=["Amount"],
    )
    wb["Data"]["B2"] = 99
    pivot.refresh()
    wb.save(out)

    reopened = load_workbook(out, preserve=True)
    handle = reopened["Summary"].pivots["ByRegion"]
    assert handle.origin == "paper"
    assert handle.capabilities.can_edit_layout is True
    assert reopened["Summary"]["B2"].value == 99


def test_excel_runner_is_opt_in_and_not_a_runtime_dependency(monkeypatch):
    import openpyxl  # noqa: F401

    assert not any("excel_pivot" in name for name in sys.modules)
    monkeypatch.delenv("PAPER_EXCEL_PIVOT", raising=False)
    from .support.excel_pivot import excel_available, run_transcript

    assert excel_available() is False
    with pytest.raises(RuntimeError):
        run_transcript("unused.xlsx", {})
    monkeypatch.setenv("PAPER_EXCEL_PIVOT", "1")
    with pytest.raises(NotImplementedError):
        run_transcript("unused.xlsx", {})
