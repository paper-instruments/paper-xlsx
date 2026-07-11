from __future__ import annotations

import os
from pathlib import Path
import subprocess
import sys

import pytest


def test_existing_drawing_append_is_namespace_valid_without_lxml(
        fixture_copy, tmp_path):
    pytest.importorskip("et_xmlfile")
    source = fixture_copy("features/chart_image.xlsx")
    output = tmp_path / "stdlib.xlsx"
    script = r'''
import importlib.abc
import sys
from xml.etree import ElementTree as ET
from zipfile import ZipFile

class BlockLxml(importlib.abc.MetaPathFinder):
    attempts = 0

    def find_spec(self, fullname, path=None, target=None):
        if fullname == "lxml" or fullname.startswith("lxml."):
            self.attempts += 1
            raise ImportError("lxml intentionally blocked by regression test")
        return None

blocker = BlockLxml()
sys.meta_path.insert(0, blocker)

import et_xmlfile

from openpyxl import LXML, load_workbook
from openpyxl.chart import BarChart, Reference

assert LXML is False
assert blocker.attempts > 0
source, output = sys.argv[1:]
wb = load_workbook(source, preserve=True)
ws = wb["Model"]
chart = BarChart()
chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=4))
ws.add_chart(chart, "K2")
wb.save(output)
with ZipFile(output) as archive:
    drawing = next(name for name in archive.namelist()
                   if name.startswith("xl/drawings/drawing")
                   and name.endswith(".xml"))
    ET.fromstring(archive.read(drawing))
wb2 = load_workbook(output)
assert len(wb2["Model"]._charts) == 2
'''
    env = dict(os.environ)
    env["OPENPYXL_LXML"] = "False"
    repository_root = Path(__file__).resolve().parents[2]
    subprocess.run([sys.executable, "-c", script, source, str(output)],
                   check=True, env=env, cwd=repository_root)
