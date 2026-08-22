"""Public PivotTable inspection API."""
from __future__ import annotations

import json
import os

import pytest

from openpyxl import Workbook, load_workbook
from openpyxl.errors import (
    AmbiguousTargetError,
    TargetNotFoundError,
    UnsupportedStructureError,
)
from openpyxl.pivot import (
    PivotAxisField,
    PivotCapabilities,
    PivotItemFilter,
    PivotMeasure,
    PivotSource,
    PivotSpec,
    PivotTable,
)
from openpyxl.pivot.api import TO_DICT_SCHEMA, TO_DICT_VERSION, invalidate_pivot_overlay
from openpyxl.pivot.graph import load_pivot_graph
from openpyxl.pivot.table import TableDefinition

from .conftest import FIXTURES_DIR
from .support.harness import assert_part_budget
from .test_pivot_graph import _basic_package, _sidecar_binaries, _write_package


_PIVOTS = os.path.join(FIXTURES_DIR, "pivots")
_PAPER_TAG = "paper-xlsx:pivot-v1"


def _load_preserved(tmp_path, payload, name="pivot.xlsx"):
    return load_workbook(_write_package(tmp_path, name, payload), preserve=True)


def test_public_exports_are_the_semantic_types():
    import openpyxl.pivot as pivot_mod

    assert pivot_mod.__all__ == (
        "PivotAxisField",
        "PivotCapabilities",
        "PivotItemFilter",
        "PivotMeasure",
        "PivotSource",
        "PivotSpec",
        "PivotTable",
    )
    assert not hasattr(pivot_mod, "TableDefinition")
    assert not hasattr(pivot_mod, "CacheDefinition")
    assert TableDefinition.__module__ == "openpyxl.pivot.table"


def test_source_and_axis_types_are_immutable_and_deterministic():
    table = PivotSource.table("SalesData")
    ranged = PivotSource.range("Raw Data", "A1:H5000")
    assert table.to_dict() == {"kind": "table", "name": "SalesData"}
    assert ranged.to_dict() == {
        "kind": "range", "sheet": "Raw Data", "ref": "A1:H5000"}
    assert PivotSource.parse("SalesData") == table
    assert PivotSource.parse("'Raw Data'!A1:H5000") == ranged
    with pytest.raises(ValueError):
        PivotSource.parse("A1:H5000")
    with pytest.raises(ValueError):
        PivotMeasure("Revenue", aggregate="product")
    with pytest.raises(ValueError):
        PivotItemFilter("Status", include=[])
    field = PivotAxisField("Region", items=["East", "West"])
    assert field.to_dict()["items"] == ["East", "West"]
    with pytest.raises(AttributeError):
        field.field = "Product"


def test_to_dict_schema_and_deterministic_ordering(tmp_path):
    wb = _load_preserved(tmp_path, _basic_package())
    pivot = wb["Summary"].pivots["SalesByRegion"]
    payload = pivot.to_dict()
    assert list(payload)[:4] == ["schema", "version", "name", "sheet"]
    assert payload["schema"] == TO_DICT_SCHEMA
    assert payload["version"] == TO_DICT_VERSION
    assert payload["name"] == "SalesByRegion"
    assert payload["sheet"] == "Summary"
    assert payload["source"] == {
        "kind": "range", "sheet": "Data", "ref": "A1:B5"}
    assert payload["destination"] == "B3"
    assert payload["output_range"] == "B3:C8"
    assert payload["rows"] == ["Region"]
    assert payload["values"] == [{"field": "Amount", "aggregate": "sum"}]
    assert payload["origin"] == "foreign"
    assert payload["valid"] is True
    assert payload["capabilities"] == PivotCapabilities(
        can_refresh_on_open=True).to_dict()
    assert payload["refresh_on_open_scope"] == ["Summary!SalesByRegion"]
    assert isinstance(payload["qualification_reasons"], list)
    dumped = json.dumps(payload)
    assert dumped == json.dumps(pivot.to_dict())
    assert "East" not in dumped
    assert "10" not in dumped or payload["source"]["ref"] == "A1:B5"


def test_name_lookup_exact_casefold_ambiguous_and_missing(tmp_path):
    payload = _basic_package(
        extra_pivots=(
            ("marginByRegion", "xl/pivotTables/pivotTable2.xml",
             "rIdPivot2", "1"),
        ),
    )
    wb = _load_preserved(tmp_path, payload)
    collection = wb["Summary"].pivots
    assert collection["SalesByRegion"].name == "SalesByRegion"
    assert collection["salesbyregion"].name == "SalesByRegion"
    with pytest.raises(TargetNotFoundError) as missing:
        collection["Missing"]
    assert missing.value.kind == "pivot-not-found"
    assert "SalesByRegion" in missing.value.options

    clash = _basic_package(
        extra_pivots=(
            ("SalesByRegion", "xl/pivotTables/pivotTable2.xml",
             "rIdPivot2", "1"),
        ),
    )
    other = _load_preserved(tmp_path, clash, name="clash.xlsx")
    with pytest.raises(AmbiguousTargetError) as ambiguous:
        other["Summary"].pivots["SalesByRegion"]
    assert ambiguous.value.kind == "ambiguous-pivot"


def test_preserve_false_read_only_and_write_only_refuse(
        tmp_path, fixture_copy):
    path = _write_package(tmp_path, "stock.xlsx", _basic_package())
    stock = load_workbook(path, preserve=False)
    with pytest.raises(UnsupportedStructureError) as stock_exc:
        stock["Summary"].pivots
    assert stock_exc.value.kind == "invalid-pivot-graph"
    assert stock._paper_ledger is None

    readonly = load_workbook(fixture_copy("minimal/minimal_clean.xlsx"),
                             read_only=True)
    with pytest.raises(UnsupportedStructureError) as ro_exc:
        readonly.worksheets[0].pivots
    assert ro_exc.value.kind == "invalid-pivot-graph"
    readonly.close()

    write_only = Workbook(write_only=True)
    ws = write_only.create_sheet("Summary")
    with pytest.raises(UnsupportedStructureError) as wo_exc:
        ws.pivots
    assert wo_exc.value.kind == "invalid-pivot-graph"


def test_stale_handle_after_overlay_change_and_close(tmp_path):
    wb = _load_preserved(tmp_path, _basic_package())
    pivot = wb["Summary"].pivots["SalesByRegion"]
    assert pivot.valid is True
    invalidate_pivot_overlay(wb)
    with pytest.raises(TargetNotFoundError) as stale:
        pivot.to_dict()
    assert stale.value.kind == "stale-pivot-handle"

    wb = _load_preserved(tmp_path, _basic_package(), name="close.xlsx")
    pivot = wb["Summary"].pivots["SalesByRegion"]
    wb.close()
    with pytest.raises(TargetNotFoundError) as closed:
        pivot.name
    assert closed.value.kind == "stale-pivot-handle"


def test_inspection_does_not_mutate_cells_ledger_or_relationships(
        tmp_path):
    wb = _load_preserved(tmp_path, _basic_package())
    ws = wb["Summary"]
    before_cells = {
        sheet.title: set(sheet._cells) for sheet in wb.worksheets
    }
    before_dirty = {
        sheet.title: set(wb._paper_ledger.dirty_coordinates(sheet))
        for sheet in wb.worksheets
    }
    before_rels = list(ws._rels)
    before_dimensions = (ws.dimensions, ws._current_row)
    payload = ws.pivots["SalesByRegion"].to_dict()
    assert payload["schema"] == TO_DICT_SCHEMA
    assert {
        sheet.title: set(sheet._cells) for sheet in wb.worksheets
    } == before_cells
    assert {
        sheet.title: set(wb._paper_ledger.dirty_coordinates(sheet))
        for sheet in wb.worksheets
    } == before_dirty
    assert list(ws._rels) == before_rels
    assert (ws.dimensions, ws._current_row) == before_dimensions


def test_to_dict_stays_bounded_when_cache_has_many_records(tmp_path):
    records = ['<s v="East"/><n v="%s"/>' % index for index in range(400)]
    wb = _load_preserved(tmp_path, _basic_package(
        records=records, record_count=400))
    payload = wb["Summary"].pivots["SalesByRegion"].to_dict()
    dumped = json.dumps(payload)
    assert payload["values"] == [{"field": "Amount", "aggregate": "sum"}]
    assert dumped.count("East") <= 1
    assert "399" not in dumped


def test_direct_low_level_pivot_list_is_not_the_paper_collection(tmp_path):
    wb = _load_preserved(tmp_path, _basic_package())
    ws = wb["Summary"]
    paper = ws.pivots
    assert paper is not ws._pivots
    assert [pivot.name for pivot in paper] == ["SalesByRegion"]
    assert hasattr(paper, "create")
    assert not hasattr(paper["SalesByRegion"], "create")
    assert hasattr(paper["SalesByRegion"], "update")
    assert hasattr(paper["SalesByRegion"], "refresh")
    assert hasattr(paper["SalesByRegion"], "delete")
    assert not hasattr(paper, "update")
    assert not hasattr(paper, "refresh")
    assert not hasattr(paper, "delete")


def test_noop_save_after_inspection_is_byte_identical(fixture_copy, tmp_path):
    from tests.paper.test_delivery import _with_pivot_graph

    src = fixture_copy("minimal/minimal_clean.xlsx")
    wb = load_workbook(src, preserve=True)
    _with_pivot_graph(wb, [("SalesPivot", "1")])
    path = tmp_path / "inspect.xlsx"
    path.write_bytes(wb._paper_source)
    wb.close()
    wb = load_workbook(path, preserve=True)
    pivot = wb.worksheets[0].pivots["SalesPivot"]
    assert pivot.origin == "foreign"
    assert pivot.capabilities.can_refresh_on_open is True
    out = tmp_path / "inspect-out.xlsx"
    wb.save(out)
    assert_part_budget(path.read_bytes(), out.read_bytes())


@pytest.mark.parametrize("filename", _sidecar_binaries())
def test_real_fixture_to_dict_matches_sidecar(filename):
    path = os.path.join(_PIVOTS, filename)
    with open(path + ".json") as handle:
        sidecar = json.load(handle)
    expected = sidecar.get("expected_pivot_to_dict")
    if expected is None:
        pytest.skip("sidecar has no expected_pivot_to_dict")
    wb = load_workbook(path, preserve=True)
    observed = {}
    for sheet in wb.worksheets:
        try:
            collection = sheet.pivots
        except UnsupportedStructureError:
            continue
        for pivot in collection:
            observed["%s!%s" % (sheet.title, pivot.name)] = pivot.to_dict()
    assert observed == expected
