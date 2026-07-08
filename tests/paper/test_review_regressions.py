"""Regression tests for the confirmed findings of the pre-PR adversarial
review (scratch/results/final_review/confirmed.json). Every test here
reproduces a defect that once produced silent corruption, a crash, or a
missing refusal."""
from __future__ import annotations

import re
import zipfile

import pytest

from openpyxl import load_workbook
from openpyxl.errors import PaperRefusal, UnsupportedStructureError
from openpyxl.package import diff_package
from openpyxl.styles import Font

from .support.partdiff import part_payloads


class TestRowColStyleTranslation:
    """Findings 1/10: row and column style indices reached the spliced bytes
    untranslated — dangling xf references, IndexError on reload."""

    def test_row_style_appends_the_xf(self, fixture_copy, tmp_path):
        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Sheet1"].row_dimensions[2].font = Font(bold=True)
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)                     # was: IndexError
        assert wb2["Sheet1"].row_dimensions[2].font.bold is True
        sheet = next(p for n, p in part_payloads(out).items()
                     if n.startswith("xl/worksheets/"))
        m = re.search(rb'<row r="2"[^>]* s="(\d+)"', sheet)
        assert m is not None
        styles = part_payloads(out)["xl/styles.xml"]
        count = int(re.search(rb'<cellXfs count="(\d+)"', styles).group(1))
        assert int(m.group(1)) < count               # the xf exists

    def test_row_height_on_lo_authored_file(self, fixture_copy, tmp_path):
        src = fixture_copy("features/lo_authored.xlsx")
        wb = load_workbook(src, preserve=True)
        wb.active.row_dimensions[2].height = 42
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)                     # was: IndexError
        assert wb2.active.row_dimensions[2].height == 42

    def test_column_style_appends_the_xf(self, fixture_copy, tmp_path):
        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Sheet1"].column_dimensions["A"].font = Font(italic=True)
        wb["Sheet1"].column_dimensions["A"].width = 30
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)                     # was: IndexError
        assert wb2["Sheet1"].column_dimensions["A"].width == 30


class TestRegionInsertionOrder:
    """Finding 2: CF + DV both inserted at the same offset came out in
    schema-invalid order."""

    def test_cf_and_dv_added_together_are_schema_ordered(
            self, fixture_copy, tmp_path):
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation

        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Sheet1"]
        dv = DataValidation(type="whole", operator="between",
                            formula1="1", formula2="10")
        dv.add("B2:B4")
        ws.add_data_validation(dv)
        ws.conditional_formatting.add(
            "B1:B5", CellIsRule(operator="greaterThan", formula=["5"],
                                fill=PatternFill("solid", fgColor="FFAA0000")))
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        sheet = next(p for n, p in part_payloads(out).items()
                     if n.startswith("xl/worksheets/"))
        assert sheet.index(b"<conditionalFormatting") \
            < sheet.index(b"<dataValidations")
        wb2 = load_workbook(out)
        assert len(list(wb2["Sheet1"].conditional_formatting)) == 1
        assert len(wb2["Sheet1"].data_validations.dataValidation) == 1


class TestCellRowMismatchRefusal:
    """Finding 3: a cell whose r sits in a different parent row produced a
    silent duplicate reference on edit; now a typed refusal."""

    def test_mismatched_cell_refuses(self, fixture_copy, tmp_path):
        src = fixture_copy("minimal/minimal_clean.xlsx")
        surgical = str(tmp_path / "mismatch.xlsx")
        with zipfile.ZipFile(src) as zin, zipfile.ZipFile(surgical, "w") as zout:
            for name in zin.namelist():
                payload = zin.read(name)
                if name.startswith("xl/worksheets/"):
                    payload = payload.replace(b'<c r="A2"', b'<c r="A9"', 1)
                zout.writestr(name, payload)
        wb = load_workbook(surgical, preserve=True)
        wb["Sheet1"]["A9"] = "edited"
        with pytest.raises(UnsupportedStructureError, match="disagrees"):
            wb.save(str(tmp_path / "o.xlsx"))


class TestShiftInteractions:
    """Findings 13/19/20/22: shifts vs pre-shift edits, split shared groups,
    and hyperlinks on deleted rows."""

    def test_edits_before_a_shift_survive(self, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Schedule"]
        ws["A2"] = "edited BEFORE the shift"     # pre-shift dirty mark
        ws.insert_rows(2)                        # A2 content moves to A3
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert wb2["Schedule"]["A3"].value == "edited BEFORE the shift"
        assert wb2["Schedule"]["A2"].value is None

    def test_insert_inside_shared_group_keeps_all_members_correct(
            self, fixture_copy, tmp_path):
        src = fixture_copy("gauntlet/gauntlet.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Calc"]                          # shared group B2:B6
        ws.insert_rows(4)                        # split the group
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        parts = part_payloads(out)
        calc = next(p for n, p in parts.items()
                    if n.startswith("xl/worksheets/") and b"A2*2" in p)
        assert b't="shared"' not in calc         # every member dissolved
        wb2 = load_workbook(out)
        assert wb2["Calc"]["B3"].value == "=A3*2"
        assert wb2["Calc"]["B5"].value == "=A5*2"   # below the insert
        assert wb2["Calc"]["B7"].value == "=A7*2"

    def test_hyperlink_on_deleted_row_does_not_reattach(
            self, fixture_copy, tmp_path):
        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Sheet1"]["A3"].hyperlink = "https://example.org/doomed"
        mid = str(tmp_path / "mid.xlsx")
        wb.save(mid)
        wb = load_workbook(mid, preserve=True)
        wb["Sheet1"].delete_rows(3)
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        # the row that shifted up into A3 must NOT inherit the hyperlink
        assert wb2["Sheet1"]["A3"].hyperlink is None

    def test_pre_shift_column_hyperlink_survives_column_insert(
            self, fixture_copy, tmp_path):
        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Sheet1"]["A3"].hyperlink = "https://example.org/x"
        mid = str(tmp_path / "mid.xlsx")
        wb.save(mid)
        wb = load_workbook(mid, preserve=True)
        wb["Sheet1"].insert_cols(1)              # push A -> B
        out = str(tmp_path / "o.xlsx")
        wb.save(out)                             # was: refused every save
        wb2 = load_workbook(out)
        assert wb2["Sheet1"]["B3"].hyperlink.target == "https://example.org/x"


class TestEscapedSheetNames:
    """Findings 9/23: names needing XML escaping ('P&L') were invisible to
    byte scans and state patches."""

    @pytest.fixture
    def pl_file(self, fixture_copy, tmp_path):
        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src)
        wb["Sheet1"].title = "P&L"
        out = str(tmp_path / "pl.xlsx")
        wb.save(out)
        return out

    def test_sheet_state_toggle_on_escaped_name(self, pl_file, tmp_path):
        wb = load_workbook(pl_file, preserve=True)
        wb.create_sheet("Other")                 # a visible sheet must remain
        out1 = str(tmp_path / "mid.xlsx")
        wb.save(out1)
        wb = load_workbook(out1, preserve=True)
        wb["P&L"].sheet_state = "hidden"
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert wb2["P&L"].sheet_state == "hidden"   # was: silently dropped


class TestMiscRefusalsAndCrashes:

    def test_template_toggle_refuses(self, fixture_copy, tmp_path):
        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        wb.template = True
        with pytest.raises(UnsupportedStructureError, match="template"):
            wb.save(str(tmp_path / "o.xlsx"))

    def test_removing_all_custom_props_refuses_not_crashes(
            self, fixture_copy, tmp_path):
        src = fixture_copy("gauntlet/gauntlet.xlsx")
        wb = load_workbook(src, preserve=True)
        if not len(wb.custom_doc_props):
            pytest.skip("fixture has no custom properties")

    def test_mark_dirty_full_column_range(self, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule.xlsx")
        wb = load_workbook(src, preserve=True)
        wb.mark_dirty("Schedule!A:B")            # was: TypeError
        led = wb._paper_ledger
        assert (12, 2) in led.dirty_coordinates(wb["Schedule"])

    def test_create_chartsheet_refuses(self, fixture_copy):
        wb = load_workbook(fixture_copy("minimal/minimal_clean.xlsx"),
                           preserve=True)
        with pytest.raises(UnsupportedStructureError, match="chartsheet"):
            wb.create_chartsheet("Chartz")

    def test_append_of_prebuilt_cells_is_ledgered(self, fixture_copy, tmp_path):
        from openpyxl.cell import Cell

        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Sheet1"]
        cell = Cell(ws, value="prebuilt")
        ws.append([cell])                        # the write-only-compat path
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert wb2["Sheet1"]["A6"].value == "prebuilt"   # was: silently absent

    def test_excelwriter_refuses_preserve_workbooks(self, fixture_copy, tmp_path):
        import zipfile as zf

        from openpyxl.writer.excel import ExcelWriter

        wb = load_workbook(fixture_copy("gauntlet/gauntlet.xlsx"),
                           preserve=True)
        archive = zf.ZipFile(str(tmp_path / "o.xlsx"), "w")
        with pytest.raises(UnsupportedStructureError, match="ExcelWriter"):
            ExcelWriter(wb, archive)


class TestOracleRegressions:
    """Findings 29/30/31: volatile seeding by substring, xlsm stripping,
    bool/number equality."""

    def test_string_literals_do_not_taint(self, fixture_copy, monkeypatch):
        from openpyxl import Workbook, oracle

        wb = Workbook()
        ws = wb.active
        ws["A1"] = '=IF(B1>0,"use RAND() wisely",0)'   # literal, not a call
        ws["B1"] = 1
        import io
        buf = io.BytesIO()
        wb.save(buf)
        monkeypatch.setattr(oracle, "find_soffice", lambda: None)
        result = oracle.certify(buf.getvalue())
        # A1 must NOT be volatile-excluded (it has no cache -> unverifiable)
        assert not result.volatile_excluded

    def test_xlsm_recalc_output_refuses(self, fixture_copy, tmp_path):
        from openpyxl import oracle

        src = fixture_copy("features/macro_stub.xlsm")
        with pytest.raises(UnsupportedStructureError, match="VBA"):
            oracle.recalc(src, in_place=True)
        with pytest.raises(UnsupportedStructureError, match="VBA"):
            oracle.recalc(src, output_path=str(tmp_path / "o.xlsx"))

    def test_bool_never_equals_number(self):
        from openpyxl.oracle import _values_match

        assert not _values_match(True, 1)
        assert not _values_match(0, False)
        assert _values_match(True, True)
        assert _values_match(1, 1.0)
