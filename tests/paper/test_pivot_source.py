"""Typed pivot source snapshots."""
from __future__ import annotations

from datetime import date

import pytest

from openpyxl import Workbook
from openpyxl.errors import BoundaryViolationError
from openpyxl.pivot.api_types import PivotSource
from openpyxl.pivot.source import (
    DEFAULT_LIMITS,
    KIND_BLANK,
    KIND_BOOLEAN,
    KIND_NUMBER,
    KIND_TEXT,
    PivotLimits,
    snapshot_from_matrix,
    snapshot_from_workbook,
    typed_value,
)
from openpyxl.worksheet.table import Table


def test_typed_identity_distinguishes_number_text_bool_and_date():
    assert typed_value(1) != typed_value("1")
    assert typed_value(1) != typed_value(True)
    assert typed_value(1).kind == KIND_NUMBER
    assert typed_value("1").kind == KIND_TEXT
    assert typed_value(True).kind == KIND_BOOLEAN
    assert typed_value(date(2024, 1, 1)).kind == "date"
    assert typed_value(None).kind == KIND_BLANK
    assert typed_value("").kind == KIND_BLANK
    with pytest.raises(BoundaryViolationError) as nan:
        typed_value(float("nan"))
    assert nan.value.kind == "invalid-pivot-source"
    with pytest.raises(BoundaryViolationError):
        typed_value(float("inf"))
    with pytest.raises(BoundaryViolationError):
        typed_value("#DIV/0!")


def test_duplicate_and_blank_headers_refuse():
    with pytest.raises(BoundaryViolationError) as dup:
        snapshot_from_matrix(["Region", "region"], [["East", 1]])
    assert dup.value.kind == "invalid-pivot-source"
    with pytest.raises(BoundaryViolationError):
        snapshot_from_matrix(["Region", ""], [["East", 1]])


def test_partial_and_fully_blank_rows_remain_records():
    snapshot = snapshot_from_matrix(
        ["Region", "Amount"],
        [["East", 10], [None, 7], [None, None], ["West", None]],
    )
    assert len(snapshot.records) == 4
    assert snapshot.records[1].values[0].kind == KIND_BLANK
    assert snapshot.records[2].values[0].kind == KIND_BLANK
    assert snapshot.records[2].values[1].kind == KIND_BLANK


def test_table_and_range_snapshots_match_for_identical_cells():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Region"
    ws["B1"] = "Amount"
    ws["A2"] = "East"
    ws["B2"] = 10
    ws["A3"] = "West"
    ws["B3"] = 7
    ws.add_table(Table(displayName="SalesData", ref="A1:B3"))
    table = snapshot_from_workbook(wb, PivotSource.table("SalesData"))
    ranged = snapshot_from_workbook(wb, PivotSource.range("Data", "A1:B3"))
    assert [record.values for record in table.records] == [
        record.values for record in ranged.records]
    assert table.fields == ranged.fields == ("Region", "Amount")


@pytest.mark.parametrize("attribute", ["totalsRowCount", "totalsRowShown"])
def test_table_totals_row_is_excluded_but_table_identity_is_retained(
        attribute, tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    rows = [
        ("Region", "Amount"),
        ("East", 10),
        ("West", 20),
        ("Total", 30),
    ]
    for row in rows:
        ws.append(row)
    table = Table(displayName="SalesData", ref="A1:B4")
    setattr(table, attribute, 1 if attribute == "totalsRowCount" else True)
    ws.add_table(table)
    path = tmp_path / (attribute + ".xlsx")
    wb.save(path)

    from openpyxl import load_workbook
    wb = load_workbook(path)

    snapshot = snapshot_from_workbook(wb, PivotSource.table("SalesData"))

    assert snapshot.source == PivotSource.table("SalesData")
    assert snapshot.bounds == ("Data", 1, 1, 2, 4)
    assert [record.values[0].value for record in snapshot.records] == [
        "East", "West"]
    assert [record.values[1].value for record in snapshot.records] == [10, 20]


def test_table_column_metadata_must_match_visible_headers():
    from openpyxl.worksheet.table import TableColumn

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(("Region", "Amount"))
    ws.append(("East", 10))
    table = Table(displayName="SalesData", ref="A1:B2")
    table.tableColumns = [
        TableColumn(id=1, name="Region"),
        TableColumn(id=2, name="Revenue"),
    ]
    ws.add_table(table)

    with pytest.raises(BoundaryViolationError) as exc:
        snapshot_from_workbook(wb, PivotSource.table("SalesData"))
    assert exc.value.kind == "unsupported-pivot-source"


@pytest.mark.parametrize(
    ("attribute", "value"),
    [
        ("headerRowCount", 0),
        ("headerRowCount", 2),
        ("totalsRowCount", 2),
        ("tableType", "queryTable"),
        ("connectionId", 1),
    ],
)
def test_unsupported_table_source_metadata_refuses(attribute, value):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(("Region", "Amount"))
    ws.append(("East", 10))
    ws.append(("West", 20))
    table = Table(displayName="SalesData", ref="A1:B3")
    setattr(table, attribute, value)
    ws.add_table(table)

    with pytest.raises(BoundaryViolationError) as exc:
        snapshot_from_workbook(wb, PivotSource.table("SalesData"))
    assert exc.value.kind == "unsupported-pivot-source"


def test_source_intersecting_merged_cells_refuses():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(("Region", "Amount"))
    ws.append(("East", 10))
    ws.append(("West", 20))
    ws.merge_cells("A2:A3")

    with pytest.raises(BoundaryViolationError) as exc:
        snapshot_from_workbook(wb, PivotSource.range("Data", "A1:B3"))
    assert exc.value.kind == "unsupported-pivot-source"
    assert exc.value.anchor == "Data!A2:A3"


@pytest.mark.parametrize("source", [
    PivotSource.range("Data", "A:D"),
    "'Data'!A:D",
    PivotSource.range("Data", "1:4"),
])
def test_open_axis_source_refuses_with_typed_error(source):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    with pytest.raises(BoundaryViolationError) as exc:
        snapshot_from_workbook(wb, source)
    assert exc.value.kind == "unsupported-pivot-source"


def test_hidden_rows_remain_in_the_snapshot():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Region"
    ws["B1"] = "Amount"
    ws["A2"] = "East"
    ws["B2"] = 10
    ws["A3"] = "West"
    ws["B3"] = 7
    ws.row_dimensions[3].hidden = True
    snapshot = snapshot_from_workbook(wb, PivotSource.range("Data", "A1:B3"))
    assert len(snapshot.records) == 2
    assert snapshot.records[1].values[0].value == "West"


def test_formula_coordinates_are_discovered_without_evaluation():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Region"
    ws["B1"] = "Amount"
    ws["A2"] = "East"
    ws["B2"] = "=1+2"
    snapshot = snapshot_from_workbook(wb, PivotSource.range("Data", "A1:B2"))
    assert snapshot.formula_coordinates == ("Data!B2",)
    assert snapshot.records[0].values[1].kind == KIND_TEXT
    assert snapshot.records[0].values[1].value == "=1+2"


def test_source_limits_stop_before_unbounded_allocation():
    with pytest.raises(BoundaryViolationError) as rows:
        snapshot_from_matrix(
            ["Region", "Amount"],
            [["East", 1], ["West", 2]],
            limits=PivotLimits(source_rows=1),
        )
    assert rows.value.kind == "pivot-source-too-large"
    with pytest.raises(BoundaryViolationError) as items:
        snapshot_from_matrix(
            ["Region", "Amount"],
            [["A", 1], ["B", 2], ["C", 3]],
            limits=PivotLimits(distinct_items_field=1),
        )
    assert items.value.kind == "pivot-cardinality-too-large"


def test_snapshot_does_not_create_missing_cells():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Region"
    ws["B1"] = "Amount"
    ws["A2"] = "East"
    before = set(ws._cells)
    snapshot_from_workbook(wb, PivotSource.range("Data", "A1:B3"))
    assert set(ws._cells) == before


def test_source_identity_is_stable_and_changes_with_values():
    first = snapshot_from_matrix(["Region", "Amount"], [["East", 10]])
    second = snapshot_from_matrix(["Region", "Amount"], [["East", 10]])
    changed = snapshot_from_matrix(["Region", "Amount"], [["East", 11]])
    assert first.identity == second.identity
    assert first.identity != changed.identity
    assert DEFAULT_LIMITS.source_rows == 500000
