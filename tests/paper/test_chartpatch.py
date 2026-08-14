"""Chart-range rewriting inside preserved bytes."""
from __future__ import annotations

import re
import zipfile

import pytest

from openpyxl import load_workbook
from openpyxl.errors import UnsupportedStructureError

from .support.partdiff import part_payloads


class TestChartShift:

    def test_known_non_reference_extension_does_not_block_shift(
            self, fixture_copy):
        from openpyxl.preserve.chartpatch import patch_chart

        source = fixture_copy("features/lo_authored.xlsx")
        with zipfile.ZipFile(source) as archive:
            payload = archive.read("xl/charts/chart1.xml")
        assert b"c15:showLeaderLines" in payload

        patched, changed, blockers = patch_chart(
            payload, "Model", "insert_rows", 2, 1)

        assert changed is True
        assert blockers == []
        assert b"c15:showLeaderLines" in patched
        assert b"Model!$B$3" in patched

    def test_unknown_extension_content_blocks_shift(self):
        from openpyxl.preserve.chartpatch import patch_chart

        payload = (
            b'<chartSpace xmlns="http://schemas.openxmlformats.org/'
            b'drawingml/2006/chart"><chart><plotArea><barChart><ser>'
            b'<val><numRef><f>Model!$B$2</f></numRef></val></ser>'
            b'</barChart></plotArea></chart><extLst><ext uri="vendor">'
            b'<v:formula xmlns:v="urn:vendor:opaque">Model!$B$2</v:formula>'
            b'</ext></extLst></chartSpace>')

        unchanged, changed, blockers = patch_chart(
            payload, "Model", "insert_rows", 2, 1)

        assert unchanged == payload
        assert changed is False
        assert blockers == [
            "the chart carries unrecognized extension content formula "
            "in urn:vendor:opaque"]

    def test_insert_rows_patches_series_and_anchors(self, fixture_copy, tmp_path):
        src = fixture_copy("features/chart_image.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Model"]
        ws.insert_rows(2)                     # data rows shift; header stays
        ws["A2"] = "INSERTED"
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        parts = part_payloads(out)
        chart = parts["xl/charts/chart1.xml"]
        # value series moved with their data; the row-1 headers stayed
        assert b"<f>'Model'!$B$3</f>" in chart
        assert b"<f>'Model'!B1</f>" in chart
        assert b"$B$2" not in chart
        drawing = parts["xl/drawings/drawing1.xml"]
        anchors = [int(m) for m in re.findall(rb"<row>(\d+)</row>", drawing)]
        assert anchors == [2, 20]             # chart row 1->2, image 19->20
        wb2 = load_workbook(out)
        assert len(wb2["Model"]._charts) == 1
        assert wb2["Model"]["B4"].value == 40  # shifted data intact

    @pytest.mark.lo_smoke
    def test_patched_chart_loads_in_libreoffice(self, lo, fixture_copy, tmp_path):
        src = fixture_copy("features/chart_image.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Model"].insert_rows(2)
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        assert lo.lo_loads(out)

    def test_shift_below_chart_data_leaves_chart_bytes_identical(
            self, fixture_copy, tmp_path):
        src = fixture_copy("features/chart_image.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Model"].insert_rows(40)           # far below everything charted
        wb["Model"]["A40"] = "note"
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        before = part_payloads(src)
        after = part_payloads(out)
        assert after["xl/charts/chart1.xml"] == before["xl/charts/chart1.xml"]
        assert after["xl/drawings/drawing1.xml"] == \
            before["xl/drawings/drawing1.xml"]

    def test_deleting_charted_data_refuses(self, fixture_copy):
        src = fixture_copy("features/chart_image.xlsx")
        wb = load_workbook(src, preserve=True)
        with pytest.raises(UnsupportedStructureError, match="delete data"):
            wb["Model"].delete_rows(1, 3)

    def test_column_shift_patches_chart_too(self, fixture_copy, tmp_path):
        src = fixture_copy("features/chart_image.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Model"].insert_cols(2)            # push B..E to C..F
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        chart = part_payloads(out)["xl/charts/chart1.xml"]
        assert b"<f>'Model'!$C$2</f>" in chart
        assert b"<f>'Model'!C1</f>" in chart
        assert load_workbook(out)["Model"]["C2"].value == 100
