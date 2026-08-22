"""Preserve-mode PR 4 pivot creation package contract."""
from __future__ import annotations

import zipfile

import pytest

from openpyxl import load_workbook
from openpyxl.errors import UnsupportedStructureError
from openpyxl.pivot.cache import CacheDefinition
from openpyxl.pivot.fields import Number
from openpyxl.pivot.qualify import PAPER_TAG
from openpyxl.pivot.record import RecordList
from openpyxl.pivot.table import TableDefinition
from openpyxl.xml.constants import SHEET_MAIN_NS
from openpyxl.xml.functions import fromstring

from .support.harness import assert_part_budget, save_and_reopen
from .support.partdiff import part_payloads
from .test_pivot_graph import _basic_package, _write_package


_TABLE = "features/tables.xlsx"
_EXPECTED_ADDED = {
    "xl/pivotTables/pivotTable1.xml",
    "xl/pivotTables/_rels/pivotTable1.xml.rels",
    "xl/pivotCache/pivotCacheDefinition1.xml",
    "xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels",
    "xl/pivotCache/pivotCacheRecords1.xml",
}
_EXPECTED_CHANGED = {
    "xl/worksheets/sheet1.xml",
    "xl/worksheets/_rels/sheet1.xml.rels",
    "xl/workbook.xml",
    "xl/_rels/workbook.xml.rels",
    "[Content_Types].xml",
}


def _create_by_region(ws, source="RegionTable", destination="E3", name="ByRegion"):
    return ws.pivots.create(
        name=name,
        source=source,
        destination=destination,
        rows=["Region"],
        values=["Amount"],
    )


def _output_grid(ws):
    return {
        (row, column): ws.cell(row, column).value
        for row in range(3, 10)
        for column in range(5, 7)
    }


def _expected_grid():
    return {
        (3, 5): "Sum of Amount",
        (3, 6): None,
        (4, 5): "Region",
        (4, 6): "Total",
        (5, 5): "North",
        (5, 6): 20,
        (6, 5): "South",
        (6, 6): 30,
        (7, 5): "East",
        (7, 6): 40,
        (8, 5): "West",
        (8, 6): 50,
        (9, 5): "Grand Total",
        (9, 6): 140,
    }


def test_create_from_table_saves_and_reopens(fixture_copy):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    handle = _create_by_region(wb["Data"])
    assert handle.name == "ByRegion"
    assert handle.origin == "paper"
    assert handle.valid is True
    assert handle.output_range == "E3:F9"
    assert handle.capabilities.can_delete is True
    assert _output_grid(wb["Data"]) == _expected_grid()

    dest = src + ".created.xlsx"
    reopened = save_and_reopen(wb, dest, preserve=True)
    pivot = reopened["Data"].pivots["ByRegion"]
    assert pivot.origin == "paper"
    assert pivot.valid is True
    assert pivot.spec.rows[0].field == "Region"
    assert pivot.spec.values[0].aggregate == "sum"
    assert pivot.source.kind == "table"
    assert pivot.source.name == "RegionTable"
    assert _output_grid(reopened["Data"]) == _expected_grid()
    assert reopened["Data"]["A2"].value == "North"
    assert_part_budget(
        src, dest, expect_changed=_EXPECTED_CHANGED, expect_added=_EXPECTED_ADDED)


def test_create_from_equivalent_range(fixture_copy):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    _create_by_region(wb["Data"], source="'Data'!A1:B5", name="ByRange")
    dest = src + ".range.xlsx"
    reopened = save_and_reopen(wb, dest, preserve=True)
    pivot = reopened["Data"].pivots["ByRange"]
    assert pivot.source.kind == "range"
    assert pivot.source.sheet == "Data"
    assert pivot.source.ref == "A1:B5"
    assert _output_grid(reopened["Data"]) == _expected_grid()


def test_created_parts_parse_and_indexes_match(fixture_copy):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    _create_by_region(wb["Data"])
    dest = src + ".parts.xlsx"
    wb.save(dest)
    parts = part_payloads(dest)
    cache = CacheDefinition.from_tree(fromstring(
        parts["xl/pivotCache/pivotCacheDefinition1.xml"]))
    records = RecordList.from_tree(fromstring(
        parts["xl/pivotCache/pivotCacheRecords1.xml"]))
    table = TableDefinition.from_tree(fromstring(
        parts["xl/pivotTables/pivotTable1.xml"]))
    assert cache.saveData is True
    assert cache.enableRefresh is True
    assert cache.recordCount == 4
    assert cache.recordCount == records.count
    assert len(cache.cacheFields) == 2
    assert cache.cacheFields[0].name == "Region"
    assert cache.cacheFields[0].sharedItems.count == 4
    assert cache.cacheFields[1].sharedItems.count == 0
    assert len(cache.cacheFields[1].sharedItems._fields) == 0
    cache_root = fromstring(
        parts["xl/pivotCache/pivotCacheDefinition1.xml"])
    cache_fields = cache_root.find("{%s}cacheFields" % SHEET_MAIN_NS)
    region_items = cache_fields[0].find("{%s}sharedItems" % SHEET_MAIN_NS)
    amount_items = cache_fields[1].find("{%s}sharedItems" % SHEET_MAIN_NS)
    assert "containsString" not in region_items.attrib
    assert "containsNonDate" not in region_items.attrib
    assert amount_items.attrib["containsString"] == "0"
    assert amount_items.attrib["containsSemiMixedTypes"] == "0"
    assert amount_items.attrib["containsNumber"] == "1"
    assert len(records.r) == 4
    assert [record._fields[1].v for record in records.r] == [
        20, 30, 40, 50]
    assert all(isinstance(record._fields[1], Number) for record in records.r)
    assert table.name == "ByRegion"
    assert table.cacheId == 1
    assert table.id is None
    assert table.tag == PAPER_TAG
    assert table.location.ref == "E3:F9"
    assert table.location.firstHeaderRow == 2
    assert table.location.firstDataRow == 2
    assert table.location.firstDataCol == 1
    assert len(table.pivotFields) == 2
    table_root = fromstring(parts["xl/pivotTables/pivotTable1.xml"])
    pivot_fields = table_root.find("{%s}pivotFields" % SHEET_MAIN_NS)
    assert "defaultSubtotal" not in pivot_fields[0].attrib
    assert len(table.rowFields) == 1
    assert table.rowFields[0].x == 0
    assert len(table.rowItems) == 5
    row_items = table_root.find("{%s}rowItems" % SHEET_MAIN_NS)
    assert row_items[0].attrib == {}
    assert row_items[0][0].attrib == {}
    assert row_items[1][0].attrib == {"v": "1"}
    assert len(table.dataFields) == 1
    assert table.dataFields[0].baseField == 0
    assert table.dataFields[0].baseItem == 0
    assert table.dataFields[0].fld == 1
    assert table.dataFields[0].subtotal == "sum"


def test_relationships_and_content_types(fixture_copy):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    _create_by_region(wb["Data"])
    dest = src + ".rels.xlsx"
    wb.save(dest)
    with zipfile.ZipFile(dest) as archive:
        workbook = archive.read("xl/workbook.xml")
        assert b"<pivotCaches" in workbook
        assert b'cacheId="1"' in workbook
        types = archive.read("[Content_Types].xml")
        assert b"pivotTable+xml" in types
        assert b"pivotCacheDefinition+xml" in types
        assert b"pivotCacheRecords+xml" in types
        sheet_rels = archive.read("xl/worksheets/_rels/sheet1.xml.rels")
        assert b"pivotTable" in sheet_rels
        wb_rels = archive.read("xl/_rels/workbook.xml.rels")
        assert b"pivotCacheDefinition" in wb_rels
        cache_rels = archive.read(
            "xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels")
        assert b"pivotCacheRecords" in cache_rels
        pivot_rels = archive.read("xl/pivotTables/_rels/pivotTable1.xml.rels")
        assert b"pivotCacheDefinition" in pivot_rels


def test_receipt_reports_pivot_created(fixture_copy):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    _create_by_region(wb["Data"])
    dest = src + ".receipt.xlsx"
    receipt = wb.save(dest, receipt=True)
    created = [item for item in receipt.derived_effects
               if item["kind"] == "pivot_created"]
    assert len(created) == 1
    assert created[0]["name"] == "ByRegion"
    assert created[0]["sheet"] == "Data"
    assert created[0]["output_range"] == "E3:F9"
    assert set(created[0]["parts"]) == _EXPECTED_ADDED


def test_second_save_is_byte_identical(fixture_copy):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    _create_by_region(wb["Data"])
    first = src + ".first.xlsx"
    wb.save(first)
    again = src + ".again.xlsx"
    reopened = load_workbook(first, preserve=True)
    reopened.save(again)
    assert part_payloads(first) == part_payloads(again)


def test_identical_runs_are_deterministic(fixture_copy):
    first = fixture_copy(_TABLE, "first.xlsx")
    second = fixture_copy(_TABLE, "second.xlsx")
    for path in (first, second):
        wb = load_workbook(path, preserve=True)
        _create_by_region(wb["Data"])
        wb.save(path + ".out.xlsx")
    left = part_payloads(first + ".out.xlsx")
    right = part_payloads(second + ".out.xlsx")
    for name in _EXPECTED_ADDED:
        assert left[name] == right[name]


def _prepare_surgical_workbook(path):
    from openpyxl.workbook.properties import CalcProperties

    wb = load_workbook(path, preserve=True)
    if wb.calculation is None:
        wb.calculation = CalcProperties(calcMode="auto")
    return wb


def test_create_preserves_foreign_pivot_bytes(tmp_path):
    payload = _basic_package()
    src = _write_package(tmp_path, "foreign.xlsx", payload)
    before = part_payloads(src)
    wb = _prepare_surgical_workbook(src)
    _create_by_region(
        wb["Data"], source="'Data'!A1:B2", destination="E3", name="PaperPivot")
    dest = str(tmp_path / "foreign-created.xlsx")
    wb.save(dest)
    after = part_payloads(dest)
    for name, body in before.items():
        if name in (
            "xl/worksheets/sheet2.xml",
            "xl/workbook.xml",
            "xl/_rels/workbook.xml.rels",
            "[Content_Types].xml",
        ):
            continue
        assert after[name] == body
    assert "xl/pivotTables/pivotTable1.xml" in before
    assert after["xl/pivotTables/pivotTable1.xml"] == before[
        "xl/pivotTables/pivotTable1.xml"]


def test_create_skips_gapped_and_custom_part_names(tmp_path):
    payload = _basic_package(
        cache_part="xl/custom/cache-a.xml",
        records_part="xl/custom/records-a.xml",
        pivot_part="xl/custom/report.xml",
        cache_target="/xl/custom/cache-a.xml",
        pivot_target="../custom/report.xml",
        cache_id="1",
    )
    src = _write_package(tmp_path, "custom.xlsx", payload)
    wb = _prepare_surgical_workbook(src)
    _create_by_region(
        wb["Data"], source="'Data'!A1:B2", destination="E3", name="PaperPivot")
    dest = str(tmp_path / "custom-created.xlsx")
    wb.save(dest)
    names = set(part_payloads(dest))
    assert "xl/custom/report.xml" in names
    assert "xl/custom/cache-a.xml" in names
    assert "xl/pivotTables/pivotTable1.xml" in names
    assert "xl/pivotCache/pivotCacheDefinition1.xml" in names
    table = TableDefinition.from_tree(fromstring(
        part_payloads(dest)["xl/pivotTables/pivotTable1.xml"]))
    assert table.cacheId == 2
    assert table.tag == PAPER_TAG


def test_create_fills_conventional_part_number_gaps(tmp_path):
    payload = _basic_package(
        cache_part="xl/pivotCache/pivotCacheDefinition2.xml",
        records_part="xl/pivotCache/pivotCacheRecords2.xml",
        pivot_part="xl/pivotTables/pivotTable2.xml",
        cache_id="1",
    )
    src = _write_package(tmp_path, "gapped.xlsx", payload)
    wb = _prepare_surgical_workbook(src)
    _create_by_region(
        wb["Data"], source="'Data'!A1:B2", destination="E3", name="PaperPivot")
    dest = str(tmp_path / "gapped-created.xlsx")
    wb.save(dest)
    names = set(part_payloads(dest))
    assert "xl/pivotTables/pivotTable2.xml" in names
    assert "xl/pivotTables/pivotTable1.xml" in names
    assert "xl/pivotCache/pivotCacheDefinition1.xml" in names
    assert "xl/pivotCache/pivotCacheDefinition2.xml" in names


def test_create_refuses_unknown_keywords_and_styles(fixture_copy):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    ws = wb["Data"]
    with pytest.raises(TypeError, match="unexpected keyword"):
        ws.pivots.create(
            name="Wide",
            source="RegionTable",
            destination="E3",
            rows=["Region"],
            values=["Amount"],
            showDataAs="percent",
        )
    with pytest.raises(UnsupportedStructureError) as exc:
        ws.pivots.create(
            name="Wide",
            source="RegionTable",
            destination="E3",
            rows=["Region"],
            values=["Amount"],
            style="CustomPivotTheme",
        )
    assert exc.value.kind == "unsupported-pivot-feature"


def test_libreoffice_loads_created_workbook(fixture_copy, lo):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    _create_by_region(wb["Data"])
    dest = src + ".lo.xlsx"
    wb.save(dest)
    converted = lo.lo_convert(dest, fmt="xlsx")
    assert converted[:2] == b"PK"
