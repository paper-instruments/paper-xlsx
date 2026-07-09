"""Phase 6a: the structural-edit guard — reference-aware refusals under
preserve, loud warning on the stock path (PLAN Phase 6a; PR-0 §8)."""
from __future__ import annotations

import warnings

import pytest

from openpyxl import Workbook, load_workbook
from openpyxl.errors import StructuralShiftWarning, UnsupportedStructureError


class TestPreserveRefusals:
    """Since Phase 6b, fully-modeled sheets REWRITE instead of refusing
    (test_rewrite.py); these tests cover the sheets that must still refuse —
    anything carrying unmodeled range-bearing content."""

    def test_refusal_names_blockers_and_victims(self, fixture_copy):
        src = fixture_copy("gauntlet/gauntlet.xlsx")
        wb = load_workbook(src, preserve=True)
        with pytest.raises(UnsupportedStructureError) as exc:
            wb["Model"].insert_rows(1)
        msg = str(exc.value)
        assert "extLst" in msg or "extension" in msg    # the blocker
        assert "A1:F1" in msg                 # merged banner (victim)
        assert "chart" in msg.lower()         # preserved chart bytes
        assert "xl/charts/chart1.xml" in msg
        assert "Nothing was changed" in msg

    def test_column_shift_analyzed_too(self, fixture_copy):
        src = fixture_copy("gauntlet/gauntlet.xlsx")
        wb = load_workbook(src, preserve=True)
        with pytest.raises(UnsupportedStructureError) as exc:
            wb["Model"].delete_cols(2)
        assert "'Model'!B6" in str(exc.value)  # victim formula named

    def test_refusal_is_atomic(self, fixture_copy):
        src = fixture_copy("gauntlet/gauntlet.xlsx")
        wb = load_workbook(src, preserve=True)
        before_b6 = wb["Model"]["B6"].value
        before_max = wb["Model"].max_row
        with pytest.raises(UnsupportedStructureError):
            wb["Model"].insert_rows(3)
        assert wb["Model"]["B6"].value == before_b6
        assert wb["Model"].max_row == before_max
        assert wb._paper_ledger.cells == {}
        assert wb._paper_ledger.shifts == {}

    def test_added_sheets_stay_exempt(self, fixture_copy):
        wb = load_workbook(fixture_copy("gauntlet/gauntlet.xlsx"),
                           preserve=True)
        ws = wb.create_sheet("Scratch")
        ws.append([1, 2, 3])
        ws.insert_rows(1)                     # no refusal: generated whole
        assert ws.max_row == 2


class TestStockWarning:

    def test_loaded_workbook_warns_on_shift(self, fixture_copy):
        wb = load_workbook(fixture_copy("features/schedule.xlsx"))
        with pytest.warns(StructuralShiftWarning, match="updates NOTHING"):
            wb["Schedule"].insert_rows(5)
        # stock behavior unchanged: the shift still happened
        assert wb["Schedule"]["B13"].value == "=SUM(B2:B11)"

    def test_fresh_workbook_does_not_warn(self):
        wb = Workbook()
        ws = wb.active
        ws.append([1, 2, 3])
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            ws.insert_rows(1)
        assert not [w for w in caught
                    if isinstance(w.message, StructuralShiftWarning)]


class TestAddressRemap:
    """CONVENTIONS §2 (pinned, debt paid in v0.1 Batch 1): structural
    edits return an AddressRemap; pre-edit addresses must be remapped,
    never reused."""

    def test_insert_rows_returns_remap(self, fixture_copy, tmp_path):
        from openpyxl.preserve import AddressRemap

        wb = load_workbook(fixture_copy("features/schedule.xlsx"),
                           preserve=True)
        remap = wb["Schedule"].insert_rows(3)
        assert isinstance(remap, AddressRemap)
        assert remap.map("Schedule!B12") == "Schedule!B13"
        assert remap.map("B12") == "B13"
        assert remap.map("B2") == "B2"                 # above the insert
        assert remap.map("Summary!B1") == "Summary!B1" # other sheet
        assert remap.map("'Schedule'!$B$12") == "'Schedule'!$B$13"
        assert remap.map("B2:B11") == "B2:B12"         # spanning range
        # the remapped address reads the moved cell after save→reopen
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert wb2["Schedule"]["B13"].value == "=SUM(B2:B12)"

    def test_delete_rows_maps_deleted_to_none(self, fixture_copy):
        wb = load_workbook(fixture_copy("features/schedule.xlsx"),
                           preserve=True)
        remap = wb["Schedule"].delete_rows(5)
        assert remap.map("B5") is None                 # deleted
        assert remap.map("Schedule!A5") is None
        assert remap.map("B12") == "B11"               # shifted up
        assert remap.map("B2") == "B2"

    def test_stock_path_keeps_returning_none(self, fixture_copy):
        wb = load_workbook(fixture_copy("features/schedule.xlsx"))
        with pytest.warns(StructuralShiftWarning):
            assert wb["Schedule"].insert_rows(3) is None


class TestBoundaryViolation:

    def test_insert_rows_past_sheet_limit_refuses(self, fixture_copy):
        from openpyxl.errors import BoundaryViolationError

        src = fixture_copy("features/schedule.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Schedule"]
        ws.cell(row=1048576, column=1, value="sentinel")
        with pytest.raises(BoundaryViolationError, match="1048576"):
            ws.insert_rows(1)
        # atomic: the guard fired before any cell moved
        assert ws.cell(row=1048576, column=1).value == "sentinel"
        assert ws["B12"].value == "=SUM(B2:B11)"

    def test_insert_cols_past_sheet_limit_refuses(self, fixture_copy):
        from openpyxl.errors import BoundaryViolationError

        wb = load_workbook(fixture_copy("features/schedule.xlsx"),
                           preserve=True)
        ws = wb["Schedule"]
        ws.cell(row=1, column=16384, value="sentinel")
        with pytest.raises(BoundaryViolationError, match="XFD"):
            ws.insert_cols(1)


class TestMultipleShifts:
    """PLAN-v0.1 3.3: shifts compose within one session."""

    def test_two_shifts_one_session(self, fixture_copy, tmp_path):
        wb = load_workbook(fixture_copy("features/schedule.xlsx"),
                           preserve=True)
        ws = wb["Schedule"]
        remap1 = ws.insert_rows(3, 2)      # B12 -> B14
        remap2 = ws.delete_rows(5)         # (was row 3 data) B14 -> B13
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        ws2 = wb2["Schedule"]
        total_row = remap2.map(remap1.map("B12"))
        assert ws2[total_row].value.startswith("=SUM(")
        # cross-sheet reference composed through both shifts
        assert wb2["Summary"]["B1"].value == "=Schedule!" + total_row

    def test_shift_then_edit_then_shift(self, fixture_copy, tmp_path):
        wb = load_workbook(fixture_copy("features/schedule.xlsx"),
                           preserve=True)
        ws = wb["Schedule"]
        ws.insert_rows(3)
        ws["B3"] = 777                     # edit between the shifts
        ws.insert_rows(1)
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert wb2["Schedule"]["B4"].value == 777


class TestSpanningMerges:
    """PLAN-v0.1 3.3: Excel merge semantics under shifts (expansion on
    insert-inside, shrink on delete-inside, move otherwise)."""

    def test_merge_expands_shrinks_and_moves(self, fixture_copy, tmp_path):
        src = fixture_copy("features/merged.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb.worksheets[0]                    # merges: A1:D1, A3:B4
        ws.insert_rows(4)                        # inside A3:B4 -> expand
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert sorted(str(r) for r in
                      wb2.worksheets[0].merged_cells.ranges) == \
            ["A1:D1", "A3:B5"]

        wb3 = load_workbook(fixture_copy("features/merged.xlsx"),
                            preserve=True)
        wb3.worksheets[0].delete_rows(4)         # inside -> shrink
        out2 = str(tmp_path / "o2.xlsx")
        wb3.save(out2)
        wb4 = load_workbook(out2)
        assert sorted(str(r) for r in
                      wb4.worksheets[0].merged_cells.ranges) == \
            ["A1:D1", "A3:B3"]


class TestMoveRange:
    """PLAN-v0.1 3.3: move_range as tracked cell edits."""

    def test_clean_move_lands(self, fixture_copy, tmp_path):
        wb = load_workbook(fixture_copy("features/schedule.xlsx"),
                           preserve=True)
        ws = wb["Schedule"]
        # move two label cells sideways into empty columns (nothing
        # references them)
        assert ws["A2"].value is not None
        ws.move_range("A2:A3", rows=0, cols=4)
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        ws2 = wb2["Schedule"]
        assert ws2["E2"].value is not None
        assert ws2["A2"].value is None
        assert ws2["B12"].value == "=SUM(B2:B11)"    # untouched machinery

    def test_move_of_referenced_block_refuses(self, fixture_copy):
        wb = load_workbook(fixture_copy("features/schedule.xlsx"),
                           preserve=True)
        ws = wb["Schedule"]
        with pytest.raises(UnsupportedStructureError, match="references"):
            ws.move_range("B2:B4", rows=0, cols=3)   # inside =SUM(B2:B11)
        assert ws["B2"].value is not None            # atomic


class TestStructuredRefs:
    """PLAN-v0.1 3.5: Table1[@Col] is never mis-shifted — tables on the
    shifted sheet block the shift (refusal), and bracketed operands are
    never rewritten."""

    def test_shift_on_table_sheet_refuses(self, fixture_copy):
        wb = load_workbook(fixture_copy("features/tables.xlsx"),
                           preserve=True)
        ws = wb.worksheets[0]
        with pytest.raises(UnsupportedStructureError, match="table"):
            ws.insert_rows(2)

    def test_structured_ref_text_survives_other_sheet_shift(
            self, fixture_copy, tmp_path):
        wb = load_workbook(fixture_copy("features/schedule.xlsx"),
                           preserve=True)
        wb["Summary"]["C1"] = "=SUM(RegionTable[Amount])"
        wb["Schedule"].insert_rows(3)
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert wb2["Summary"]["C1"].value == "=SUM(RegionTable[Amount])"
