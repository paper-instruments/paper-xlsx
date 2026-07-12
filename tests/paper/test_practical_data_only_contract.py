import io

import pytest

from openpyxl import Workbook, load_workbook
from openpyxl.errors import UnsupportedStructureError
from openpyxl.workbook.defined_name import DefinedName


def _source_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Model"
    sheet["A1"] = "=1+1"
    sheet["B1"] = 10
    workbook.defined_names.add(DefinedName(
        "FormulaInput", attr_text="'Model'!$A$1"))
    workbook.defined_names.add(DefinedName(
        "ValueInput", attr_text="'Model'!$B$1"))
    target = io.BytesIO()
    workbook.save(target)
    return target.getvalue()


def test_preserve_data_only_set_input_refuses_source_formula_atomically():
    source = _source_bytes()
    workbook = load_workbook(
        io.BytesIO(source), preserve=True, data_only=True)
    ledger = workbook._paper_ledger
    before_cells = {
        sheet: set(coordinates) for sheet, coordinates in ledger.cells.items()
    }
    before_overwrites = {
        sheet: set(coordinates)
        for sheet, coordinates in ledger.value_overwrites.items()
    }

    with pytest.raises(UnsupportedStructureError) as refusal:
        workbook.set_input("FormulaInput", 5)

    assert refusal.value.kind == "input-is-calculation"
    assert workbook["Model"]["A1"].value is None
    assert {sheet: set(coordinates)
            for sheet, coordinates in ledger.cells.items()} == before_cells
    assert {sheet: set(coordinates)
            for sheet, coordinates in ledger.value_overwrites.items()} == \
        before_overwrites


def test_preserve_data_only_direct_assignment_refuses_source_formula():
    workbook = load_workbook(
        io.BytesIO(_source_bytes()), preserve=True, data_only=True)
    cell = workbook["Model"]["A1"]
    before = (cell.value, cell.data_type)

    with pytest.raises(UnsupportedStructureError) as refusal:
        cell.value = 5

    assert refusal.value.kind == "input-is-calculation"
    assert (cell.value, cell.data_type) == before


def test_preserve_data_only_data_type_assignment_refuses_source_formula():
    workbook = load_workbook(
        io.BytesIO(_source_bytes()), preserve=True, data_only=True)
    cell = workbook["Model"]["A1"]
    before = (cell.value, cell.data_type)

    with pytest.raises(UnsupportedStructureError) as refusal:
        cell.data_type = "n"

    assert refusal.value.kind == "input-is-calculation"
    assert (cell.value, cell.data_type) == before


def test_preserve_data_only_deletion_refuses_source_formula():
    workbook = load_workbook(
        io.BytesIO(_source_bytes()), preserve=True, data_only=True)

    with pytest.raises(UnsupportedStructureError) as refusal:
        del workbook["Model"]["A1"]

    assert refusal.value.kind == "input-is-calculation"


def test_preserve_data_only_merge_refuses_hidden_formula():
    workbook = Workbook()
    workbook.active["B1"] = "=1+1"
    source = io.BytesIO()
    workbook.save(source)
    workbook = load_workbook(
        io.BytesIO(source.getvalue()), preserve=True, data_only=True)

    with pytest.raises(UnsupportedStructureError) as refusal:
        workbook.active.merge_cells("A1:B1")

    assert refusal.value.kind == "input-is-calculation"
    assert not workbook.active.merged_cells.ranges


def test_data_only_array_follower_refusal_does_not_materialize_cell(
        fixture_copy):
    workbook = load_workbook(
        fixture_copy("features/shared_formulas.xlsx"),
        preserve=True, data_only=True)
    sheet = workbook["Calc"]
    before_cells = dict(sheet._cells)
    before_row = sheet._current_row

    with pytest.raises(UnsupportedStructureError):
        sheet["D3"] = 7

    assert sheet._cells == before_cells
    assert sheet._current_row == before_row


def test_preserve_data_only_direct_assignment_allows_source_literal():
    workbook = load_workbook(
        io.BytesIO(_source_bytes()), preserve=True, data_only=True)

    workbook["Model"]["B1"] = 20

    assert workbook["Model"]["B1"].value == 20


def test_preserve_data_only_set_input_allows_proven_nonformula():
    source = _source_bytes()
    workbook = load_workbook(
        io.BytesIO(source), preserve=True, data_only=True)

    target = workbook.set_input("ValueInput", 20)
    assert target.coordinate == "B1"
    assert target.value == 20

    output = io.BytesIO()
    workbook.save(output, allow_formula_loss=True)
    output.seek(0)
    reopened = load_workbook(output, data_only=False)
    assert reopened["Model"]["A1"].value == "=1+1"
    assert reopened["Model"]["B1"].value == 20


def test_stock_data_only_set_input_refuses_without_source_custody():
    workbook = load_workbook(io.BytesIO(_source_bytes()), data_only=True)

    with pytest.raises(UnsupportedStructureError) as refusal:
        workbook.set_input("ValueInput", 20)

    assert refusal.value.kind == "data-only-input-model-unavailable"
    assert workbook["Model"]["B1"].value == 10


def test_preserve_data_only_loaded_sheet_copy_refuses_atomically():
    workbook = load_workbook(
        io.BytesIO(_source_bytes()), preserve=True, data_only=True)
    sheets = workbook._sheets
    before = list(sheets)

    with pytest.raises(UnsupportedStructureError) as refusal:
        workbook.copy_worksheet(workbook["Model"])

    assert refusal.value.kind == "data-only-reference-model-unavailable"
    assert workbook._sheets is sheets
    assert list(sheets) == before


def test_preserve_data_only_added_sheet_copy_remains_provable():
    workbook = load_workbook(
        io.BytesIO(_source_bytes()), preserve=True, data_only=True)
    added = workbook.create_sheet("Added")
    added["A1"] = "=1+1"

    copied = workbook.copy_worksheet(added)

    assert copied["A1"].value == "=1+1"
    assert copied in workbook._paper_ledger.added_sheets


@pytest.mark.parametrize(
    "operation",
    [
        lambda sheet: sheet.insert_rows(1),
        lambda sheet: sheet.delete_rows(1),
        lambda sheet: sheet.insert_cols(1),
        lambda sheet: sheet.delete_cols(1),
        lambda sheet: sheet.move_range("A1:B1", rows=1),
        lambda sheet: setattr(sheet, "title", "Renamed"),
        lambda sheet: sheet.parent.remove(sheet),
    ],
)
def test_data_only_loaded_sheet_structural_edits_refuse_atomically(operation):
    workbook = load_workbook(
        io.BytesIO(_source_bytes()), preserve=True, data_only=True)
    sheet = workbook["Model"]
    before_sheets = list(workbook.worksheets)
    before_cells = dict(sheet._cells)

    with pytest.raises(UnsupportedStructureError) as refusal:
        operation(sheet)

    assert refusal.value.kind == "data-only-reference-model-unavailable"
    assert list(workbook.worksheets) == before_sheets
    assert sheet.title == "Model"
    assert sheet._cells == before_cells


def test_data_only_added_sheet_structural_edits_remain_supported():
    workbook = load_workbook(
        io.BytesIO(_source_bytes()), preserve=True, data_only=True)
    added = workbook.create_sheet("Added")
    added["A1"] = "=1+1"

    added.insert_rows(1)
    added.title = "Renamed Added"

    assert added["A2"].value == "=1+1"
