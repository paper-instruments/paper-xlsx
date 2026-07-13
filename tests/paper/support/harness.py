"""The five-assertion contract harness.

Every mutating API in this fork must pass, on the relevant fixtures:
  1. save → reopen → assert (never assert on the in-memory object),
  2. intended effect present in the reopened workbook,
  3. changed-part budget: exactly the expected parts changed, all other part
     payloads byte-identical,
  4. independent-loader smoke (LibreOffice) where available (lo_smoke),
  5. refusal atomicity: typed refusal raised AND output bytes == input bytes.

These helpers implement the mechanical parts; tests supply the semantics.
"""
from __future__ import annotations

import os
import shutil

from openpyxl import load_workbook

from .partdiff import diff_parts, xml_semantic_diff


def save_and_reopen(wb, path, **load_kw):
    """Assertion-1 helper: persist and reload. Never assert on ``wb`` after
    calling this — use the returned workbook."""
    wb.save(path)
    return load_workbook(path, **load_kw)


def assert_part_budget(before, after, expect_changed=(), expect_added=(), expect_removed=()):
    """Assertion-3 helper: the package diff shows exactly the expected parts
    changed/added/removed and every other part payload byte-identical.

    ``before``/``after`` are paths or bytes. Expected sets are exact — an
    unexpected changed part is a failure even if it looks harmless, and an
    expected-but-unchanged part is also a failure (the budget is literal).
    """
    d = diff_parts(before, after)
    problems = []
    if d.changed != set(expect_changed):
        problems.append(
            "changed parts {0} != expected {1}".format(sorted(d.changed), sorted(expect_changed))
        )
    if d.added != set(expect_added):
        problems.append(
            "added parts {0} != expected {1}".format(sorted(d.added), sorted(expect_added))
        )
    if d.removed != set(expect_removed):
        problems.append(
            "removed parts {0} != expected {1}".format(sorted(d.removed), sorted(expect_removed))
        )
    assert not problems, "part budget violated: " + "; ".join(problems) + " ({0})".format(d)
    return d


def assert_refusal_atomic(input_path, tmp_path, mutate, exc_type, save_kw=None, load_kw=None):
    """Assertion-5 helper.

    Copies ``input_path`` to a work file, loads it, applies ``mutate(wb)``
    (which is expected to raise ``exc_type`` either immediately or at save),
    and asserts the work file's bytes are untouched afterwards.

    ``mutate`` receives the loaded workbook and the save target path; it must
    attempt the full operation including save if the refusal is save-time.
    Returns the exception for message assertions.
    """
    import pytest

    work = os.path.join(str(tmp_path), "refusal_input" + os.path.splitext(input_path)[1])
    shutil.copyfile(input_path, work)
    with open(work, "rb") as f:
        before = f.read()

    wb = load_workbook(work, **(load_kw or {}))
    with pytest.raises(exc_type) as excinfo:
        mutate(wb, work)

    with open(work, "rb") as f:
        after = f.read()
    assert after == before, (
        "refusal was not atomic: file bytes changed after {0}".format(exc_type.__name__)
    )
    return excinfo.value


def assert_sheet_xml_equivalent(before, after, part_name):
    """Semantic-equivalence assertion for one XML part across two packages."""
    from .partdiff import part_payloads

    pa = part_payloads(before)
    pb = part_payloads(after)
    diffs = xml_semantic_diff(pa[part_name], pb[part_name])
    assert not diffs, "{0} not semantically equivalent: {1}".format(part_name, diffs[:10])
