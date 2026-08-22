"""PR 11: shared-cache isolation and adoption."""
from __future__ import annotations

import io
import os
import re
import zipfile
from dataclasses import replace
from xml.etree.ElementTree import fromstring

import pytest

from openpyxl import load_workbook
from openpyxl.errors import (
    BoundaryViolationError,
    RelationshipPolicyError,
    UnsupportedStructureError,
)
from openpyxl.pivot.graph import load_workbook_pivot_graph
from openpyxl.pivot.qualify import PAPER_TAG
from openpyxl.preserve.lifecycle import PartPlan, _rels_path
from openpyxl.preserve.pivotgraph import drop_workbook_caches
from openpyxl.utils import get_column_letter
from openpyxl.xml.constants import PKG_REL_NS
from .support.harness import save_and_reopen
from .support.partdiff import part_payloads
from .test_pivot_adopt_dedicated import (
    _enable_evidence,
    _fingerprint,
    _pivot_events,
)
from .test_pivot_adoption_qualification import _strip_paper_tag
from .test_pivot_create_package import (
    _create_by_region,
    _expected_grid,
    _output_grid,
)
from .test_pivot_graph import _basic_package, _cache_xml, _records_xml
from .test_pivot_adoption_qualification import _codes, _load_payload


_TABLE = "features/tables.xlsx"
_ORIGINAL_CACHE = {
    "xl/pivotCache/pivotCacheDefinition1.xml",
    "xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels",
    "xl/pivotCache/pivotCacheRecords1.xml",
}


def _grid_at(ws, destination="E3"):
    from openpyxl.utils.cell import coordinate_to_tuple

    row, column = coordinate_to_tuple(destination)
    return {
        (r, c): ws.cell(r, c).value
        for r in range(row, row + 7)
        for c in range(column, column + 2)
    }


def _shift_expected(destination="J3"):
    from openpyxl.utils.cell import coordinate_to_tuple

    row, column = coordinate_to_tuple(destination)
    base_row, base_col = 3, 5
    return {
        (row + (r - base_row), column + (c - base_col)): value
        for (r, c), value in _expected_grid().items()
    }


def _rewrite_archive(source, destination, transform):
    with zipfile.ZipFile(source) as before:
        output = io.BytesIO()
        with zipfile.ZipFile(output, "w") as after:
            for info in before.infolist():
                body = transform(info.filename, before.read(info.filename))
                if body is None:
                    continue
                after.writestr(info.filename, body)
        with open(destination, "wb") as handle:
            handle.write(output.getvalue())
    return destination


def _remove_relationships_to(payload, target_part):
    filename = re.escape(target_part.rsplit("/", 1)[-1].encode("ascii"))
    pattern = re.compile(
        br"<Relationship\b[^>]*\bTarget=\"(?:[^\"]*/)?%s\"[^>]*/>" % filename
    )
    updated, count = pattern.subn(b"", payload, count=1)
    if count != 1:
        raise AssertionError(
            "expected one relationship targeting %s" % target_part)
    return updated


def _remove_overrides(payload, parts):
    updated = payload
    for part in parts:
        if part.endswith(".rels"):
            continue
        pattern = re.compile(
            br"<Override\b[^>]*\bPartName=\"/%s\"[^>]*/>"
            % re.escape(part.encode("ascii"))
        )
        updated, count = pattern.subn(b"", updated, count=1)
        if count != 1:
            raise AssertionError("expected one content-type override for %s" % part)
    return updated


def _share_cache_and_strip(source, destination, keep, drop):
    drop_parts = {
        drop.cache_definition_part,
        _rels_path(drop.cache_definition_part),
        drop.cache_records_part,
    }
    keep_id = str(keep.cache_id)
    drop_id = str(drop.cache_id)
    marker = (' tag="%s"' % PAPER_TAG).encode("ascii")
    drop_rels = _rels_path(drop.identity.pivot_part)

    def transform(name, body):
        if name in drop_parts:
            return None
        if name in (keep.identity.pivot_part, drop.identity.pivot_part):
            body = body.replace(marker, b"")
            if name == drop.identity.pivot_part:
                body = body.replace(
                    ('cacheId="%s"' % drop_id).encode("ascii"),
                    ('cacheId="%s"' % keep_id).encode("ascii"),
                    1,
                )
            return body
        if name == drop_rels:
            return body.replace(
                drop.cache_definition_part.rsplit("/", 1)[-1].encode("ascii"),
                keep.cache_definition_part.rsplit("/", 1)[-1].encode("ascii"),
            )
        if name == "xl/workbook.xml":
            return drop_workbook_caches(body, [int(drop_id)])
        if name == "xl/_rels/workbook.xml.rels":
            return _remove_relationships_to(body, drop.cache_definition_part)
        if name == "[Content_Types].xml":
            return _remove_overrides(body, drop_parts)
        return body

    return _rewrite_archive(source, destination, transform)


def _paper_then_shared_foreign(
        fixture_copy, tmp_path,
        names=("ByRegion", "ByRegionCopy"),
        destinations=("E3", "J3")):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    worksheet = wb["Data"]
    for name, destination in zip(names, destinations):
        _create_by_region(worksheet, name=name, destination=destination)
    created = str(tmp_path / ("two-paper-%s.xlsx" % "-".join(names)))
    save_and_reopen(wb, created, preserve=True)
    wb = load_workbook(created, preserve=True)
    graph = load_workbook_pivot_graph(wb)
    keep = next(node for node in graph.pivots if node.identity.name == names[0])
    drop = next(node for node in graph.pivots if node.identity.name == names[1])
    shared = str(tmp_path / ("shared-foreign-%s.xlsx" % "-".join(names)))
    _share_cache_and_strip(created, shared, keep, drop)
    return load_workbook(shared, preserve=True), shared


def _add_unadoptable_sibling(source, destination, name="Unadoptable"):
    with zipfile.ZipFile(source) as archive:
        template = archive.read("xl/pivotTables/pivotTable1.xml")
        rels = archive.read("xl/pivotTables/_rels/pivotTable1.xml.rels")
    sibling = "xl/pivotTables/pivotTable3.xml"
    sibling_rels = "xl/pivotTables/_rels/pivotTable3.xml.rels"
    body = template.replace(
        b'name="ByRegion"', ('name="%s"' % name).encode("ascii"), 1)
    body = body.replace(b'location ref="E3:F9"', b'location ref="N3:O9"', 1)
    pivot_rel = (
        b'<Relationship Id="rIdAdopt3" Type="http://schemas.openxmlformats.'
        b'org/officeDocument/2006/relationships/pivotTable" '
        b'Target="../pivotTables/pivotTable3.xml"/>'
    )
    override = (
        b'<Override PartName="/xl/pivotTables/pivotTable3.xml" '
        b'ContentType="application/vnd.openxmlformats-officedocument.'
        b'spreadsheetml.pivotTable+xml"/>'
    )

    def transform(part, payload):
        if part == "xl/worksheets/_rels/sheet1.xml.rels":
            return payload.replace(
                b"</Relationships>", pivot_rel + b"</Relationships>")
        if part == "[Content_Types].xml":
            return payload.replace(b"</Types>", override + b"</Types>")
        return payload

    _rewrite_archive(source, destination, transform)
    with zipfile.ZipFile(destination, "a") as archive:
        archive.writestr(sibling, body)
        archive.writestr(sibling_rels, rels)
    return destination


def _inject_gapped_cache(source, destination, cache_id="10"):
    part = "xl/pivotCache/pivotCacheDefinition%s.xml" % cache_id
    records = "xl/pivotCache/pivotCacheRecords%s.xml" % cache_id
    rid = "rIdGap%s" % cache_id
    cache_xml = _cache_xml(
        {"kind": "range", "sheet": "Data", "ref": "A1:B5"},
        [{"name": "Region", "items": ['<s v="East"/>']},
         {"name": "Amount", "items": []}],
        record_count=1,
    )
    records_xml = _records_xml(['<s v="East"/><n v="1"/>'])
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="%s">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/'
        'officeDocument/2006/relationships/pivotCacheRecords" '
        'Target="pivotCacheRecords%s.xml"/></Relationships>'
        % (PKG_REL_NS, cache_id)
    )
    override = (
        '<Override PartName="/%s" ContentType="application/vnd.openxmlformats'
        '-officedocument.spreadsheetml.pivotCacheDefinition+xml"/>'
        '<Override PartName="/%s" ContentType="application/vnd.openxmlformats'
        '-officedocument.spreadsheetml.pivotCacheRecords+xml"/>'
        % (part, records)
    )
    wb_rel = (
        '<Relationship Id="%s" Type="http://schemas.openxmlformats.org/'
        'officeDocument/2006/relationships/pivotCacheDefinition" '
        'Target="/%s"/>' % (rid, part)
    )

    def transform(name, body):
        if name == "xl/workbook.xml":
            return body.replace(
                b"</pivotCaches>",
                ('<pivotCache cacheId="%s" xmlns:r="http://schemas.'
                 'openxmlformats.org/officeDocument/2006/relationships" '
                 'r:id="%s"/></pivotCaches>' % (cache_id, rid)).encode("ascii"),
            )
        if name == "xl/_rels/workbook.xml.rels":
            return body.replace(
                b"</Relationships>",
                wb_rel.encode("ascii") + b"</Relationships>")
        if name == "[Content_Types].xml":
            return body.replace(
                b"</Types>", override.encode("ascii") + b"</Types>")
        return body

    _rewrite_archive(source, destination, transform)
    with zipfile.ZipFile(destination, "a") as archive:
        archive.writestr(part, cache_xml)
        archive.writestr(records, records_xml)
        archive.writestr(_rels_path(part), rels)
    return destination


def _tamper_part(package, part, transform):
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(package)) as before, \
            zipfile.ZipFile(output, "w") as after:
        for info in before.infolist():
            body = before.read(info.filename)
            if info.filename == part:
                body = transform(body)
            after.writestr(info, body)
    return output.getvalue()


def test_adopt_one_of_two_preserves_sibling_and_original_cache(
        fixture_copy, tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    wb, path = _paper_then_shared_foreign(fixture_copy, tmp_path)
    before = part_payloads(path)
    wb["Data"]["M1"] = "keep-me"
    pivot = wb["Data"].pivots["ByRegion"].adopt()
    assert pivot.origin == "paper"
    dest = str(tmp_path / "isolate-one.xlsx")
    receipt = wb.save(dest, receipt=True)
    after = part_payloads(dest)
    for part in _ORIGINAL_CACHE:
        assert after[part] == before[part]
    assert after["xl/pivotTables/pivotTable2.xml"] == before[
        "xl/pivotTables/pivotTable2.xml"]
    assert after["xl/pivotTables/_rels/pivotTable2.xml.rels"] == before[
        "xl/pivotTables/_rels/pivotTable2.xml.rels"]
    assert after["xl/worksheets/_rels/sheet1.xml.rels"] == before[
        "xl/worksheets/_rels/sheet1.xml.rels"]
    assert "xl/pivotCache/pivotCacheDefinition2.xml" in after
    assert "xl/pivotCache/pivotCacheRecords2.xml" in after
    assert PAPER_TAG.encode("ascii") in after["xl/pivotTables/pivotTable1.xml"]
    assert PAPER_TAG.encode("ascii") not in after["xl/pivotTables/pivotTable2.xml"]
    events = _pivot_events(receipt)
    assert [item["kind"] for item in events] == ["pivot_adopted"]
    assert events[0]["strategy"] == "shared-isolation"
    assert events[0]["original_cache_id"] != events[0]["managed_cache_id"]
    siblings = events[0]["cache_siblings_unchanged"]
    assert any(item["part"] == "xl/pivotTables/pivotTable2.xml"
               for item in siblings)
    assert "xl/pivotCache/pivotCacheDefinition2.xml" in events[0][
        "selected_parts_added"]
    reopened = load_workbook(dest, preserve=True)
    assert reopened["Data"].pivots["ByRegion"].origin == "paper"
    assert reopened["Data"].pivots["ByRegionCopy"].origin == "foreign"
    graph = load_workbook_pivot_graph(reopened)
    selected = next(node for node in graph.pivots if node.identity.name == "ByRegion")
    sibling = next(node for node in graph.pivots if node.identity.name == "ByRegionCopy")
    assert selected.cache_definition_part != sibling.cache_definition_part
    assert _output_grid(reopened["Data"]) == _expected_grid()
    assert _grid_at(reopened["Data"], "J3") == _shift_expected("J3")
    assert reopened["Data"]["M1"].value == "keep-me"


def test_same_sheet_xml_outside_selected_cells_is_unchanged(
        fixture_copy, tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    wb, path = _paper_then_shared_foreign(fixture_copy, tmp_path)
    wb["Data"]["M2"] = "adjacent"
    dest = str(tmp_path / "isolate-sheet.xml.xlsx")
    wb["Data"].pivots["ByRegion"].adopt()
    receipt = wb.save(dest, receipt=True)
    allowed = {
        "%s%s" % (get_column_letter(column), row)
        for row in range(3, 10)
        for column in range(5, 7)
    }
    changed = set(receipt.cells_changed.get("xl/worksheets/sheet1.xml", {}))
    assert "J3" not in changed
    assert "K9" not in changed
    assert changed <= allowed | {"M2"}
    reopened = load_workbook(dest, preserve=True)
    assert reopened["Data"]["M2"].value == "adjacent"
    assert _grid_at(reopened["Data"], "J3") == _shift_expected("J3")


def test_adopt_one_of_three_when_sibling_is_unadoptable(
        fixture_copy, tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    wb, path = _paper_then_shared_foreign(fixture_copy, tmp_path)
    triple = _add_unadoptable_sibling(
        path, str(tmp_path / "three-siblings.xlsx"))
    wb = load_workbook(triple, preserve=True)
    blocked = wb["Data"].pivots["Unadoptable"].qualify_adoption()
    assert blocked.eligible is False
    assert "foreign-output-unproved" in _codes(blocked)
    before = part_payloads(triple)
    wb["Data"].pivots["ByRegion"].adopt()
    dest = str(tmp_path / "isolate-unadoptable-sibling.xlsx")
    wb.save(dest)
    after = part_payloads(dest)
    assert after["xl/pivotTables/pivotTable2.xml"] == before[
        "xl/pivotTables/pivotTable2.xml"]
    assert after["xl/pivotTables/pivotTable3.xml"] == before[
        "xl/pivotTables/pivotTable3.xml"]
    for part in _ORIGINAL_CACHE:
        assert after[part] == before[part]
    reopened = load_workbook(dest, preserve=True)
    assert reopened["Data"].pivots["ByRegion"].origin == "paper"
    assert reopened["Data"].pivots["ByRegionCopy"].origin == "foreign"
    assert reopened["Data"].pivots["Unadoptable"].origin == "foreign"


@pytest.mark.parametrize("order", [
    ("ByRegion", "ByRegionCopy"),
    ("ByRegionCopy", "ByRegion"),
])
def test_adopt_both_siblings_is_order_independent(
        fixture_copy, tmp_path, monkeypatch, order):
    _enable_evidence(monkeypatch)
    first = _adopt_both_to_payloads(
        fixture_copy, tmp_path, order, suffix="a")
    second = _adopt_both_to_payloads(
        fixture_copy, tmp_path, order, suffix="b")
    assert first == second


def _adopt_both_to_payloads(fixture_copy, tmp_path, order, suffix):
    wb, path = _paper_then_shared_foreign(fixture_copy, tmp_path)
    for name in order:
        wb["Data"].pivots[name].adopt()
    dest = str(tmp_path / ("both-%s-%s.xlsx" % ("-".join(order), suffix)))
    receipt = wb.save(dest, receipt=True)
    events = _pivot_events(receipt)
    assert [item["kind"] for item in events] == ["pivot_adopted", "pivot_adopted"]
    reopened = load_workbook(dest, preserve=True)
    assert reopened["Data"].pivots["ByRegion"].origin == "paper"
    assert reopened["Data"].pivots["ByRegionCopy"].origin == "paper"
    assert _output_grid(reopened["Data"]) == _expected_grid()
    assert _grid_at(reopened["Data"], "J3") == _shift_expected("J3")
    after = part_payloads(dest)
    assert "xl/pivotCache/pivotCacheDefinition1.xml" not in after
    assert "xl/pivotCache/pivotCacheRecords1.xml" not in after
    caches = [
        name for name in after
        if name.startswith("xl/pivotCache/pivotCacheDefinition")
        and not name.endswith(".rels")
    ]
    assert len(caches) == 2
    return {
        "names": set(after),
        "by_region": _output_grid(reopened["Data"]),
        "copy": _grid_at(reopened["Data"], "J3"),
        "removed_original": "xl/pivotCache/pivotCacheDefinition1.xml" not in after,
    }


def test_adopt_every_sibling_removes_original_cache(
        fixture_copy, tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    wb, path = _paper_then_shared_foreign(fixture_copy, tmp_path)
    before = part_payloads(path)
    wb["Data"].pivots["ByRegion"].adopt()
    wb["Data"].pivots["ByRegionCopy"].adopt()
    dest = str(tmp_path / "all-isolated.xlsx")
    receipt = wb.save(dest, receipt=True)
    after = part_payloads(dest)
    assert "xl/pivotCache/pivotCacheDefinition1.xml" not in after
    assert "xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels" not in after
    assert "xl/pivotCache/pivotCacheRecords1.xml" not in after
    workbook = fromstring(after["xl/workbook.xml"])
    cache_ids = [
        child.get("cacheId")
        for child in workbook.iter()
        if child.tag.endswith("pivotCache")
    ]
    assert "1" not in cache_ids
    assert len(cache_ids) == 2
    events = _pivot_events(receipt)
    assert any(event.get("selected_parts_removed") for event in events)
    assert before["xl/pivotTables/pivotTable1.xml"] != after[
        "xl/pivotTables/pivotTable1.xml"]
    reopened = load_workbook(dest, preserve=True)
    graph = load_workbook_pivot_graph(reopened)
    caches = {node.cache_definition_part for node in graph.pivots}
    assert len(caches) == 2
    assert "xl/pivotCache/pivotCacheDefinition1.xml" not in caches


def test_gapped_cache_ids_are_not_reused(
        fixture_copy, tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    wb, path = _paper_then_shared_foreign(fixture_copy, tmp_path)
    gapped = _inject_gapped_cache(path, str(tmp_path / "gapped.xlsx"))
    wb = load_workbook(gapped, preserve=True)
    wb["Data"].pivots["ByRegion"].adopt()
    dest = str(tmp_path / "gapped-isolated.xlsx")
    receipt = wb.save(dest, receipt=True)
    event = _pivot_events(receipt)[0]
    assert event["original_cache_id"] == 1
    assert event["managed_cache_id"] not in (1, 10)
    after = part_payloads(dest)
    assert "xl/pivotCache/pivotCacheDefinition10.xml" in after
    assert after["xl/pivotCache/pivotCacheDefinition10.xml"] == part_payloads(
        gapped)["xl/pivotCache/pivotCacheDefinition10.xml"]


def test_allocation_collision_refuses_before_publication(
        fixture_copy, tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    wb, path = _paper_then_shared_foreign(fixture_copy, tmp_path)
    source = open(path, "rb").read()
    wb["Data"].pivots["ByRegion"].adopt()
    staged = next(iter(wb._paper_ledger.pivot_operations.values()))
    object.__setattr__(
        staged, "allocation",
        replace(staged.allocation,
                cache_part="xl/pivotCache/pivotCacheDefinition1.xml"),
    )
    dest = str(tmp_path / "collision.xlsx")
    with pytest.raises(RelationshipPolicyError) as exc:
        wb.save(dest)
    assert exc.value.kind == "invalid-pivot-graph"
    assert not os.path.exists(dest)
    assert open(path, "rb").read() == source


def test_sibling_graph_change_after_plan_refuses(
        fixture_copy, tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    wb, path = _paper_then_shared_foreign(fixture_copy, tmp_path)
    wb["Data"].pivots["ByRegion"].adopt()
    wb._paper_source = _tamper_part(
        wb._paper_source, "xl/pivotTables/pivotTable2.xml",
        lambda body: body + b"\n")
    dest = str(tmp_path / "sibling-drift.xlsx")
    with pytest.raises(UnsupportedStructureError) as exc:
        wb.save(dest)
    assert exc.value.kind == "stale-pivot"
    assert not os.path.exists(dest)


def test_new_sibling_after_plan_refuses(
        fixture_copy, tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    wb, path = _paper_then_shared_foreign(fixture_copy, tmp_path)
    wb["Data"].pivots["ByRegion"].adopt()
    extra = _add_unadoptable_sibling(
        path, str(tmp_path / "late-sibling.xlsx"))
    wb._paper_source = open(extra, "rb").read()
    dest = str(tmp_path / "late-sibling-save.xlsx")
    with pytest.raises(UnsupportedStructureError) as exc:
        wb.save(dest)
    assert exc.value.kind == "stale-pivot"
    assert not os.path.exists(dest)


def test_slicer_and_chart_dependencies_refuse(
        fixture_copy, tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    wb, path = _paper_then_shared_foreign(fixture_copy, tmp_path)
    slicer = str(tmp_path / "shared-slicer.xlsx")
    with zipfile.ZipFile(path) as before, zipfile.ZipFile(slicer, "w") as after:
        for info in before.infolist():
            after.writestr(info, before.read(info.filename))
        after.writestr(
            "xl/slicerCaches/slicerCache1.xml",
            '<?xml version="1.0"?><slicerCacheDefinition '
            'name="Slicer_Region" sourceName="ByRegion">'
            '<slicerCache pivotCacheId="1"/></slicerCacheDefinition>',
        )
    wb = load_workbook(slicer, preserve=True)
    before = _fingerprint(wb)
    with pytest.raises(UnsupportedStructureError) as exc:
        wb["Data"].pivots["ByRegion"].adopt()
    assert exc.value.kind == "unsupported-pivot-feature"
    assert _fingerprint(wb) == before

    chart = str(tmp_path / "shared-chart.xlsx")
    with zipfile.ZipFile(path) as before, zipfile.ZipFile(chart, "w") as after:
        for info in before.infolist():
            after.writestr(info, before.read(info.filename))
        after.writestr(
            "xl/charts/chart1.xml",
            '<?xml version="1.0"?>'
            '<c:chartSpace xmlns:c="http://schemas.openxmlformats.org/'
            'drawingml/2006/chart"><c:pivotSource>'
            '<c:name>ByRegion</c:name></c:pivotSource></c:chartSpace>',
        )
    wb = load_workbook(chart, preserve=True)
    with pytest.raises(UnsupportedStructureError):
        wb["Data"].pivots["ByRegion"].adopt()


def test_unexpected_incoming_relationship_refuses(tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    payload = _basic_package(
        extra_pivots=(
            ("MarginByRegion", "xl/pivotTables/pivotTable2.xml",
             "rIdPivot2", "1"),
        ),
    )
    payload = _rewrite_incoming(payload)
    wb = _load_payload(tmp_path, payload, name="incoming.xlsx")
    before = _fingerprint(wb)
    with pytest.raises(
            (UnsupportedStructureError, RelationshipPolicyError)):
        wb["Summary"].pivots["SalesByRegion"].adopt()
    assert _fingerprint(wb) == before


def _rewrite_incoming(payload):
    output = io.BytesIO()
    extra = (
        b'<Relationship Id="rIdEvil" Type="http://example.invalid/evil" '
        b'Target="/xl/pivotCache/pivotCacheDefinition1.xml"/>'
    )
    with zipfile.ZipFile(io.BytesIO(payload)) as before, \
            zipfile.ZipFile(output, "w") as after:
        for info in before.infolist():
            body = before.read(info.filename)
            if info.filename == "xl/_rels/workbook.xml.rels":
                body = body.replace(
                    b"</Relationships>", extra + b"</Relationships>")
            after.writestr(info, body)
    return output.getvalue()


def test_missing_records_refuse(tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    payload = _basic_package(
        extra_pivots=(
            ("MarginByRegion", "xl/pivotTables/pivotTable2.xml",
             "rIdPivot2", "1"),
        ),
        include_records=False,
    )
    wb = _load_payload(tmp_path, payload, name="norecords.xlsx")
    before = _fingerprint(wb)
    with pytest.raises(UnsupportedStructureError) as exc:
        wb["Summary"].pivots["SalesByRegion"].adopt()
    assert exc.value.kind == "unsupported-pivot-operation"
    assert _fingerprint(wb) == before


def test_selected_refresh_does_not_alter_sibling_cells(
        fixture_copy, tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    wb, path = _paper_then_shared_foreign(fixture_copy, tmp_path)
    sibling_before = _grid_at(wb["Data"], "J3")
    wb["Data"].pivots["ByRegion"].adopt().refresh()
    dest = str(tmp_path / "refresh-selected.xlsx")
    wb.save(dest)
    reopened = load_workbook(dest, preserve=True)
    assert _grid_at(reopened["Data"], "J3") == sibling_before
    assert part_payloads(dest)["xl/pivotTables/pivotTable2.xml"] == (
        part_payloads(path)["xl/pivotTables/pivotTable2.xml"])


def test_source_edit_plus_one_isolation_leaves_sibling_stale(
        fixture_copy, tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    wb, path = _paper_then_shared_foreign(fixture_copy, tmp_path)
    wb["Data"]["B2"] = 99
    with pytest.raises(UnsupportedStructureError) as stale:
        wb.save(str(tmp_path / "stale-shared.xlsx"))
    assert stale.value.kind in ("stale-pivot", "stale-pivot-cache")
    wb["Data"].pivots["ByRegion"].adopt()
    dest = str(tmp_path / "isolate-stale-sibling.xlsx")
    with pytest.raises(UnsupportedStructureError) as still_stale:
        wb.save(dest)
    assert still_stale.value.kind in ("stale-pivot", "stale-pivot-cache")
    assert not os.path.exists(dest)
    wb.set_pivot_refresh_on_load(pivots=["ByRegionCopy"])
    receipt = wb.save(dest, receipt=True)
    events = _pivot_events(receipt)
    assert any(item["kind"] == "pivot_adopted" for item in events)
    reopened = load_workbook(dest, preserve=True)
    assert reopened["Data"]["F5"].value == 99
    assert part_payloads(dest)["xl/pivotCache/pivotCacheRecords1.xml"] == (
        part_payloads(path)["xl/pivotCache/pivotCacheRecords1.xml"])


@pytest.mark.parametrize("verb", [
    "update", "refresh", "repoint", "move", "rename", "delete",
])
def test_composed_verbs_after_shared_isolation(
        fixture_copy, tmp_path, monkeypatch, verb):
    _enable_evidence(monkeypatch)
    wb, _path = _paper_then_shared_foreign(fixture_copy, tmp_path)
    pivot = wb["Data"].pivots["ByRegion"].adopt()
    if verb == "update":
        pivot = pivot.update(layout="outline")
        assert pivot.spec.layout == "outline"
    elif verb == "refresh":
        pivot = pivot.refresh()
    elif verb == "repoint":
        pivot = pivot.repoint_source("Data!A1:B5")
        assert pivot.source.kind == "range"
    elif verb == "move":
        pivot = pivot.move("N3")
        assert pivot.destination == "N3"
    elif verb == "rename":
        pivot = pivot.rename("IsolatedSales")
        assert pivot.name == "IsolatedSales"
    else:
        pivot.delete()
        dest = str(tmp_path / "isolate-delete.xlsx")
        receipt = wb.save(dest, receipt=True)
        events = _pivot_events(receipt)
        assert [item["kind"] for item in events] == ["pivot_deleted"]
        assert events[0]["origin_before"] == "foreign"
        reopened = load_workbook(dest, preserve=True)
        assert "ByRegion" not in [item.name for item in reopened["Data"].pivots]
        assert reopened["Data"].pivots["ByRegionCopy"].origin == "foreign"
        assert "xl/pivotCache/pivotCacheDefinition1.xml" in part_payloads(dest)
        return

    staged = next(
        operation for operation in wb._paper_ledger.pivot_operations.values()
        if operation.name in ("ByRegion", "IsolatedSales"))
    assert staged.origin_before == "foreign"
    assert staged.publication_strategy == "shared-isolation"
    dest = str(tmp_path / ("isolate-%s.xlsx" % verb))
    receipt = wb.save(dest, receipt=True)
    assert [item["kind"] for item in _pivot_events(receipt)] == ["pivot_adopted"]
    reopened = load_workbook(dest, preserve=True)
    name = "IsolatedSales" if verb == "rename" else "ByRegion"
    assert reopened["Data"].pivots[name].origin == "paper"
    assert reopened["Data"].pivots["ByRegionCopy"].origin == "foreign"
    if verb == "move":
        assert reopened["Data"].pivots[name].destination == "N3"


def test_exact_workbook_registry_and_relationship_edits(
        fixture_copy, tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    wb, path = _paper_then_shared_foreign(fixture_copy, tmp_path)
    before = part_payloads(path)
    wb["Data"].pivots["ByRegion"].adopt()
    dest = str(tmp_path / "registry.xlsx")
    wb.save(dest)
    after = part_payloads(dest)
    pivot_rels = fromstring(after["xl/pivotTables/_rels/pivotTable1.xml.rels"])
    rel = next(
        child for child in pivot_rels
        if child.get("Id") == fromstring(
            before["xl/pivotTables/_rels/pivotTable1.xml.rels"]
        )[0].get("Id")
    )
    assert rel.get("Type") == fromstring(
        before["xl/pivotTables/_rels/pivotTable1.xml.rels"])[0].get("Type")
    assert "pivotCacheDefinition2.xml" in rel.get("Target")
    assert after["xl/pivotTables/_rels/pivotTable2.xml.rels"] == before[
        "xl/pivotTables/_rels/pivotTable2.xml.rels"]
    workbook = fromstring(after["xl/workbook.xml"])
    cache_ids = [
        child.get("cacheId")
        for child in workbook.iter()
        if child.tag.endswith("pivotCache")
    ]
    assert "1" in cache_ids
    assert any(item != "1" for item in cache_ids)
    assert b"pivotCacheDefinition2.xml" in after["xl/_rels/workbook.xml.rels"]
    assert b"/xl/pivotCache/pivotCacheDefinition2.xml" in after[
        "[Content_Types].xml"]
    assert b"pivotCacheRecords2.xml" in after["[Content_Types].xml"]


def test_retarget_missing_or_conflicting_relationship_refuses():
    plan = PartPlan({"xl/pivotTables/_rels/pivotTable1.xml.rels"})
    payload = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        b'<Relationships xmlns="%s">'
        b'<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/'
        b'officeDocument/2006/relationships/pivotCacheDefinition" '
        b'Target="../pivotCache/pivotCacheDefinition1.xml"/>'
        b'</Relationships>' % PKG_REL_NS.encode("ascii")
    )
    plan.retarget_rel(
        "xl/pivotTables/_rels/pivotTable1.xml.rels",
        "rIdMissing",
        "../pivotCache/pivotCacheDefinition2.xml",
    )
    with pytest.raises(RelationshipPolicyError):
        plan.apply_rels("xl/pivotTables/_rels/pivotTable1.xml.rels", payload)
    plan = PartPlan({"xl/pivotTables/_rels/pivotTable1.xml.rels"})
    plan.retarget_rel(
        "xl/pivotTables/_rels/pivotTable1.xml.rels",
        "rId1",
        "../pivotCache/pivotCacheDefinition2.xml",
        expected_type="http://example.invalid/wrong",
        expected_target="xl/pivotCache/pivotCacheDefinition1.xml",
    )
    with pytest.raises(RelationshipPolicyError):
        plan.apply_rels("xl/pivotTables/_rels/pivotTable1.xml.rels", payload)


def test_stdlib_xml_parses_isolated_parts(
        fixture_copy, tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    wb, _path = _paper_then_shared_foreign(fixture_copy, tmp_path)
    wb["Data"].pivots["ByRegion"].adopt()
    dest = str(tmp_path / "isolated-parse.xlsx")
    wb.save(dest)
    payloads = part_payloads(dest)
    fromstring(payloads["xl/pivotTables/pivotTable1.xml"])
    fromstring(payloads["xl/pivotCache/pivotCacheDefinition2.xml"])
    fromstring(payloads["xl/pivotTables/_rels/pivotTable1.xml.rels"])
    fromstring(payloads["xl/workbook.xml"])


def test_libreoffice_loads_isolated_workbook(
        fixture_copy, tmp_path, monkeypatch, lo):
    _enable_evidence(monkeypatch)
    wb, _path = _paper_then_shared_foreign(fixture_copy, tmp_path)
    wb["Data"].pivots["ByRegion"].adopt()
    dest = str(tmp_path / "isolated-lo.xlsx")
    wb.save(dest)
    converted = lo.lo_convert(dest, fmt="xlsx")
    assert converted[:2] == b"PK"


def test_excel_transcript_remains_a_stub_for_shared(
        fixture_copy, tmp_path, monkeypatch):
    import sys

    from .support.excel_pivot import excel_available, run_transcript

    try:
        _enable_evidence(monkeypatch)
        wb, _path = _paper_then_shared_foreign(fixture_copy, tmp_path)
        wb["Data"].pivots["ByRegion"].adopt()
        dest = str(tmp_path / "isolated-excel.xlsx")
        wb.save(dest)
        if not excel_available():
            with pytest.raises(RuntimeError):
                run_transcript(dest, expected={})
            return
        with pytest.raises(NotImplementedError):
            run_transcript(dest, expected={})
    finally:
        for name in list(sys.modules):
            if "excel_pivot" in name:
                sys.modules.pop(name, None)


def test_second_adoption_failure_keeps_first(
        fixture_copy, tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    wb, path = _paper_then_shared_foreign(fixture_copy, tmp_path)
    triple = _add_unadoptable_sibling(path, str(tmp_path / "keep-first.xlsx"))
    wb = load_workbook(triple, preserve=True)
    first = wb["Data"].pivots["ByRegion"].adopt()
    assert first.origin == "paper"
    before = dict(wb._paper_ledger.pivot_operations)
    with pytest.raises((BoundaryViolationError, UnsupportedStructureError)):
        wb["Data"].pivots["Unadoptable"].adopt()
    after = dict(wb._paper_ledger.pivot_operations)
    assert list(after) == list(before)
    dest = str(tmp_path / "kept-first.xlsx")
    receipt = wb.save(dest, receipt=True)
    assert [item["kind"] for item in _pivot_events(receipt)] == ["pivot_adopted"]
    reopened = load_workbook(dest, preserve=True)
    assert reopened["Data"].pivots["ByRegion"].origin == "paper"
    assert reopened["Data"].pivots["ByRegionCopy"].origin == "foreign"
    assert reopened["Data"].pivots["Unadoptable"].origin == "foreign"
