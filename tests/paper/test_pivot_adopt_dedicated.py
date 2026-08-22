"""PR 10: dedicated-cache foreign PivotTable adoption."""
from __future__ import annotations

import io
import os
import zipfile
from xml.etree.ElementTree import fromstring

import pytest

from openpyxl import load_workbook
from openpyxl.errors import (
    BoundaryViolationError,
    RelationshipPolicyError,
    TargetNotFoundError,
    UnsupportedStructureError,
)
from openpyxl.pivot import create as create_mod
from openpyxl.pivot.api_types import PivotItemFilter
from openpyxl.pivot.qualify import PAPER_TAG
from .support.harness import save_and_reopen
from .support.partdiff import diff_parts, part_payloads
from .test_pivot_adoption_qualification import (
    _paper_then_foreign,
    _strip_paper_tag,
)
from .test_pivot_create_package import _create_by_region, _expected_grid, _output_grid
from .test_pivot_graph import _basic_package, _write_package
from .test_pivot_relationship_identity import (
    _CACHE_PART,
    _CACHE_RELS,
    _PIVOT_PART,
    _PIVOT_RELS,
    _REL_ID,
    _renumber_relationships,
)


_TABLE = "features/tables.xlsx"
_DEDICATED_BUDGET = {
    "xl/pivotTables/pivotTable1.xml",
    "xl/pivotCache/pivotCacheDefinition1.xml",
    "xl/pivotCache/pivotCacheRecords1.xml",
    "xl/worksheets/sheet1.xml",
    "xl/styles.xml",
}
_FORBIDDEN_ADOPT_PARTS = {
    "xl/worksheets/_rels/sheet1.xml.rels",
    "xl/workbook.xml",
    "xl/_rels/workbook.xml.rels",
    "[Content_Types].xml",
    "xl/pivotTables/_rels/pivotTable1.xml.rels",
    "xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels",
}


def _enable_evidence(monkeypatch):
    monkeypatch.setattr(
        "openpyxl.pivot.adopt_qualify.excel_equivalence_proved",
        lambda: True,
    )
    monkeypatch.setattr(
        "openpyxl.pivot.adopt_inventory.excel_equivalence_proved",
        lambda: True,
    )


def _adopted(fixture_copy, tmp_path, monkeypatch, name="ByRegion"):
    _enable_evidence(monkeypatch)
    wb, path = _paper_then_foreign(fixture_copy, tmp_path, name=name)
    pivot = wb["Data"].pivots[name].adopt()
    return wb, path, pivot


def _fingerprint(wb):
    ledger = wb._paper_ledger
    cells = {}
    for ws in wb.worksheets:
        cells[ws.title] = {
            coord: (cell.value, cell._style, cell.data_type)
            for coord, cell in ws._cells.items()
        }
    return {
        "cells": cells,
        "dirty": {
            ws.title: frozenset(ledger.dirty_coordinates(ws))
            for ws in wb.worksheets
        },
        "operations": tuple(sorted(ledger.pivot_operations)),
        "styles": len(wb._cell_styles),
        "formats": len(wb._number_formats),
    }


def _pivot_events(receipt):
    return [
        item for item in receipt.derived_effects
        if str(item.get("kind", "")).startswith("pivot_")
    ]


def _assert_dedicated_budget(before_path, after_path):
    diff = diff_parts(before_path, after_path)
    assert diff.added == set()
    assert diff.removed == set()
    assert diff.changed <= _DEDICATED_BUDGET
    assert "xl/pivotTables/pivotTable1.xml" in diff.changed
    assert not (diff.changed & _FORBIDDEN_ADOPT_PARTS)
    after = part_payloads(after_path)
    assert PAPER_TAG.encode("ascii") in after["xl/pivotTables/pivotTable1.xml"]
    return diff


def test_adopt_is_gated_until_excel_evidence(fixture_copy, tmp_path):
    wb, path = _paper_then_foreign(fixture_copy, tmp_path)
    before = _fingerprint(wb)
    source = open(path, "rb").read()
    pivot = wb["Data"].pivots["ByRegion"]
    with pytest.raises(UnsupportedStructureError) as exc:
        pivot.adopt()
    assert exc.value.kind == "unsupported-pivot-operation"
    assert "foreign-managed-equivalence-unproved" in (exc.value.options or [])
    assert _fingerprint(wb) == before
    assert open(path, "rb").read() == source
    dest = str(tmp_path / "gated.xlsx")
    wb.save(dest)
    assert part_payloads(path) == part_payloads(dest)


def test_adopt_save_reopen_is_managed(fixture_copy, tmp_path, monkeypatch):
    wb, path, pivot = _adopted(fixture_copy, tmp_path, monkeypatch)
    assert pivot.origin == "paper"
    assert pivot.capabilities.can_edit_layout is True
    assert pivot.capabilities.can_headless_refresh is True
    assert pivot.qualify_adoption().reasons[0].code == "already-managed"
    dest = str(tmp_path / "adopted.xlsx")
    receipt = wb.save(dest, receipt=True)
    _assert_dedicated_budget(path, dest)
    events = _pivot_events(receipt)
    assert [item["kind"] for item in events] == ["pivot_adopted"]
    assert events[0]["strategy"] == "dedicated-replacement"
    assert events[0]["name"] == "ByRegion"
    assert events[0]["original_cache_id"] == events[0]["managed_cache_id"]
    reopened = load_workbook(dest, preserve=True)
    managed = reopened["Data"].pivots["ByRegion"]
    assert managed.origin == "paper"
    assert managed.capabilities.can_edit_layout is True
    assert managed.capabilities.can_delete is True
    assert _output_grid(reopened["Data"]) == _expected_grid()
    assert managed.qualify_adoption().eligible is False


def test_managed_adopt_is_noop(fixture_copy, tmp_path):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    handle = _create_by_region(wb["Data"])
    again = handle.adopt()
    assert again is handle
    dest = str(tmp_path / "managed-noop.xlsx")
    receipt = wb.save(dest, receipt=True)
    assert [item["kind"] for item in _pivot_events(receipt)] == ["pivot_created"]


def test_original_handle_is_stale_after_adopt(
        fixture_copy, tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    wb, _path = _paper_then_foreign(fixture_copy, tmp_path)
    old = wb["Data"].pivots["ByRegion"]
    new = old.adopt()
    assert new.origin == "paper"
    with pytest.raises(TargetNotFoundError) as exc:
        old.origin
    assert exc.value.kind == "stale-pivot-handle"
    with pytest.raises(TargetNotFoundError):
        old.adopt()


def test_foreign_mutators_still_refuse_until_adopt(fixture_copy, tmp_path):
    wb, _path = _paper_then_foreign(fixture_copy, tmp_path)
    pivot = wb["Data"].pivots["ByRegion"]
    before = _fingerprint(wb)
    with pytest.raises(UnsupportedStructureError) as exc:
        pivot.update(layout="outline")
    assert exc.value.kind == "unsupported-pivot-operation"
    assert "adopt" in (exc.value.options or [])
    assert _fingerprint(wb) == before


def test_ineligible_foreign_adopt_is_atomic(tmp_path):
    src = _write_package(tmp_path, "foreign.xlsx", _basic_package())
    wb = load_workbook(src, preserve=True)
    before = _fingerprint(wb)
    with pytest.raises(
            (UnsupportedStructureError, BoundaryViolationError,
             RelationshipPolicyError)):
        wb["Summary"].pivots["SalesByRegion"].adopt()
    assert _fingerprint(wb) == before


def test_source_edit_then_adopt_discharges_selected_cache(
        fixture_copy, tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    wb, path = _paper_then_foreign(fixture_copy, tmp_path)
    wb["Data"]["B2"] = 99
    with pytest.raises(UnsupportedStructureError) as stale:
        wb.save(str(tmp_path / "stale-before-adopt.xlsx"))
    assert stale.value.kind in ("stale-pivot", "stale-pivot-cache")
    pivot = wb["Data"].pivots["ByRegion"].adopt()
    dest = str(tmp_path / "source-then-adopt.xlsx")
    wb.save(dest)
    reopened = load_workbook(dest, preserve=True)
    assert reopened["Data"].pivots["ByRegion"].origin == "paper"
    assert reopened["Data"]["F5"].value == 99


def test_adopt_then_source_edit_refuses_until_refresh(
        fixture_copy, tmp_path, monkeypatch):
    wb, _path, pivot = _adopted(fixture_copy, tmp_path, monkeypatch)
    wb["Data"]["B2"] = 99
    dest = str(tmp_path / "stale-after-adopt.xlsx")
    with pytest.raises(UnsupportedStructureError) as exc:
        wb.save(dest)
    assert exc.value.kind in ("stale-pivot", "stale-pivot-cache")
    assert not os.path.exists(dest)
    pivot.refresh()
    wb.save(dest)
    reopened = load_workbook(dest, preserve=True)
    assert reopened["Data"]["F5"].value == 99


def test_custom_relationship_ids_are_preserved(
        fixture_copy, tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    created = tmp_path / "created.xlsx"
    wb = load_workbook(fixture_copy(_TABLE), preserve=True)
    _create_by_region(wb["Data"])
    wb.save(created)
    renumbered = tmp_path / "renumbered.xlsx"
    _renumber_relationships(created, renumbered)
    _strip_paper_tag(str(renumbered))
    wb = load_workbook(renumbered, preserve=True)
    wb["Data"].pivots["ByRegion"].adopt()
    dest = tmp_path / "adopted-rids.xlsx"
    wb.save(dest)
    payloads = part_payloads(dest)
    pivot_root = fromstring(payloads[_PIVOT_PART])
    cache_root = fromstring(payloads[_CACHE_PART])
    assert pivot_root.attrib[_REL_ID] == "rId7"
    assert cache_root.attrib[_REL_ID] == "rId8"
    assert b'Id="rId7"' in payloads[_PIVOT_RELS]
    assert b'Id="rId8"' in payloads[_CACHE_RELS]
    assert payloads[_PIVOT_RELS] == part_payloads(renumbered)[_PIVOT_RELS]
    assert payloads[_CACHE_RELS] == part_payloads(renumbered)[_CACHE_RELS]


def _relocate_dedicated_graph(source, destination):
    mapping = {
        "xl/pivotTables/pivotTable1.xml": "xl/custom/report.xml",
        "xl/pivotTables/_rels/pivotTable1.xml.rels":
            "xl/custom/_rels/report.xml.rels",
        "xl/pivotCache/pivotCacheDefinition1.xml": "xl/custom/cache-a.xml",
        "xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels":
            "xl/custom/_rels/cache-a.xml.rels",
        "xl/pivotCache/pivotCacheRecords1.xml": "xl/custom/records-a.xml",
    }
    rewrites = (
        (b"/xl/pivotTables/pivotTable1.xml", b"/xl/custom/report.xml"),
        (b"/xl/pivotCache/pivotCacheDefinition1.xml",
         b"/xl/custom/cache-a.xml"),
        (b"/xl/pivotCache/pivotCacheRecords1.xml",
         b"/xl/custom/records-a.xml"),
        (b"xl/pivotTables/pivotTable1.xml", b"xl/custom/report.xml"),
        (b"xl/pivotCache/pivotCacheDefinition1.xml",
         b"xl/custom/cache-a.xml"),
        (b"xl/pivotCache/pivotCacheRecords1.xml",
         b"xl/custom/records-a.xml"),
        (b"../pivotTables/pivotTable1.xml", b"../custom/report.xml"),
        (b"../pivotCache/pivotCacheDefinition1.xml",
         b"../custom/cache-a.xml"),
        (b"pivotCache/pivotCacheDefinition1.xml", b"custom/cache-a.xml"),
        (b"pivotTables/pivotTable1.xml", b"custom/report.xml"),
        (b"Target=\"pivotCacheRecords1.xml\"",
         b"Target=\"../custom/records-a.xml\""),
    )
    with zipfile.ZipFile(source) as archive:
        output = io.BytesIO()
        with zipfile.ZipFile(output, "w") as after:
            for info in archive.infolist():
                body = archive.read(info.filename)
                for before, rewritten in rewrites:
                    body = body.replace(before, rewritten)
                after.writestr(mapping.get(info.filename, info.filename), body)
        with open(destination, "wb") as handle:
            handle.write(output.getvalue())
    return destination


def test_custom_part_names_and_cache_id_are_reused(
        fixture_copy, tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    _wb, path = _paper_then_foreign(fixture_copy, tmp_path)
    custom = _relocate_dedicated_graph(
        path, str(tmp_path / "custom-names.xlsx"))
    wb = load_workbook(custom, preserve=True)
    wb["Data"].pivots["ByRegion"].adopt()
    dest = str(tmp_path / "custom-adopted.xlsx")
    wb.save(dest)
    names = set(part_payloads(dest))
    assert "xl/custom/report.xml" in names
    assert "xl/custom/cache-a.xml" in names
    assert "xl/custom/records-a.xml" in names
    assert "xl/pivotTables/pivotTable1.xml" not in names
    table = fromstring(part_payloads(dest)["xl/custom/report.xml"])
    assert table.attrib.get("cacheId") == "1"
    assert table.attrib.get("tag") == PAPER_TAG


@pytest.mark.parametrize("verb", [
    "update", "refresh", "repoint", "move", "rename", "delete",
])
def test_adopt_then_managed_verb(fixture_copy, tmp_path, monkeypatch, verb):
    wb, _path, pivot = _adopted(fixture_copy, tmp_path, monkeypatch)
    if verb == "update":
        pivot = pivot.update(layout="outline")
        assert pivot.spec.layout == "outline"
    elif verb == "refresh":
        pivot = pivot.refresh()
    elif verb == "repoint":
        pivot = pivot.repoint_source("Data!A1:B5")
        assert pivot.source.kind == "range"
    elif verb == "move":
        pivot = pivot.move("J3")
        assert pivot.destination == "J3"
    elif verb == "rename":
        pivot = pivot.rename("RegionalSales")
        assert pivot.name == "RegionalSales"
    else:
        pivot.delete()
        assert list(wb["Data"].pivots) == []
        dest = str(tmp_path / "adopt-delete.xlsx")
        receipt = wb.save(dest, receipt=True)
        events = _pivot_events(receipt)
        assert [item["kind"] for item in events] == ["pivot_deleted"]
        assert events[0]["origin_before"] == "foreign"
        reopened = load_workbook(dest, preserve=True)
        assert list(reopened["Data"].pivots) == []
        return

    staged = next(iter(wb._paper_ledger.pivot_operations.values()))
    assert staged.origin_before == "foreign"
    assert staged.publication_strategy == "dedicated-replacement"
    assert staged.cache_rebuild is True
    assert staged.semantic_effects == ("adopt",)
    dest = str(tmp_path / ("adopt-%s.xlsx" % verb))
    receipt = wb.save(dest, receipt=True)
    assert [item["kind"] for item in _pivot_events(receipt)] == ["pivot_adopted"]
    reopened = load_workbook(dest, preserve=True)
    name = "RegionalSales" if verb == "rename" else "ByRegion"
    managed = reopened["Data"].pivots[name]
    assert managed.origin == "paper"
    if verb == "move":
        assert managed.destination == "J3"
    if verb == "repoint":
        assert managed.source.kind == "range"


@pytest.mark.parametrize("verb", ["move", "rename", "repoint", "update"])
def test_inverse_after_adopt_keeps_only_adoption_receipt(
        fixture_copy, tmp_path, monkeypatch, verb):
    wb, _path, pivot = _adopted(fixture_copy, tmp_path, monkeypatch)
    if verb == "move":
        pivot.move("J3").move("E3")
    elif verb == "rename":
        pivot.rename("Temporary").rename("ByRegion")
    elif verb == "repoint":
        pivot.repoint_source("Data!A1:B5").repoint_source("RegionTable")
    else:
        pivot.update(layout="outline").update(layout="tabular")
    dest = str(tmp_path / ("inverse-%s.xlsx" % verb))
    receipt = wb.save(dest, receipt=True)
    events = _pivot_events(receipt)
    assert [item["kind"] for item in events] == ["pivot_adopted"]
    assert events[0]["name"] == "ByRegion"
    assert events[0]["output_range"] == "E3:F9"


def test_later_operation_failure_keeps_staged_adoption(
        fixture_copy, tmp_path, monkeypatch):
    wb, _path, pivot = _adopted(fixture_copy, tmp_path, monkeypatch)
    before = dict(wb._paper_ledger.pivot_operations)
    with pytest.raises(UnsupportedStructureError):
        pivot.update(rows=[])
    after = dict(wb._paper_ledger.pivot_operations)
    assert list(after) == list(before)
    staged = next(iter(after.values()))
    assert staged.kind == "adopt"
    assert staged.origin_before == "foreign"
    dest = str(tmp_path / "kept-adopt.xlsx")
    receipt = wb.save(dest, receipt=True)
    assert [item["kind"] for item in _pivot_events(receipt)] == ["pivot_adopted"]


def test_checkpoint_failures_restore_workbook(
        fixture_copy, tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    for point in ("start", "validated", "planned", "ledger"):
        wb, _path = _paper_then_foreign(fixture_copy, tmp_path, name="ByRegion")
        before = _fingerprint(wb)

        def hook(name, workbook, expected=point):
            if name == expected:
                raise RuntimeError("injected %s" % expected)

        create_mod.CREATE_CHECKPOINT = hook
        try:
            with pytest.raises(RuntimeError):
                wb["Data"].pivots["ByRegion"].adopt()
        finally:
            create_mod.CREATE_CHECKPOINT = None
        assert _fingerprint(wb) == before
        assert wb["Data"].pivots["ByRegion"].origin == "foreign"


def test_prior_output_edit_refuses_adoption(
        fixture_copy, tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    wb, _path = _paper_then_foreign(fixture_copy, tmp_path)
    wb["Data"]["E5"] = "tampered"
    before = _fingerprint(wb)
    with pytest.raises((BoundaryViolationError, UnsupportedStructureError)):
        wb["Data"].pivots["ByRegion"].adopt()
    assert _fingerprint(wb)["operations"] == before["operations"]


def test_report_filter_cells_do_not_clear_adjacent_user_values(
        tmp_path, monkeypatch):
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table

    _enable_evidence(monkeypatch)
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Region"
    ws["B1"] = "Status"
    ws["C1"] = "Amount"
    ws["A2"] = "East"
    ws["B2"] = "Open"
    ws["C2"] = 10
    ws["A3"] = "West"
    ws["B3"] = "Open"
    ws["C3"] = 7
    ws.add_table(Table(displayName="SalesData", ref="A1:C3"))
    summary = wb.create_sheet("Summary")
    path = str(tmp_path / "filters.xlsx")
    wb.save(path)
    wb = load_workbook(path, preserve=True)
    wb["Summary"].pivots.create(
        name="Filtered",
        source="SalesData",
        destination="E5",
        rows=["Region"],
        filters=[PivotItemFilter("Status")],
        values=["Amount"],
    )
    dest = str(tmp_path / "filtered-created.xlsx")
    save_and_reopen(wb, dest, preserve=True)
    _strip_paper_tag(dest)
    wb = load_workbook(dest, preserve=True)
    wb["Summary"]["G3"] = "keep-me"
    wb["Summary"]["D3"] = "also-keep"
    wb["Summary"].pivots["Filtered"].adopt()
    out = str(tmp_path / "filtered-adopted.xlsx")
    wb.save(out)
    reopened = load_workbook(out, preserve=True)
    assert reopened["Summary"]["G3"].value == "keep-me"
    assert reopened["Summary"]["D3"].value == "also-keep"
    assert reopened["Summary"].pivots["Filtered"].origin == "paper"
    assert reopened["Summary"]["E3"].value == "Status"


def test_getpivotdata_blocks_rename_and_delete_after_adopt(
        fixture_copy, tmp_path, monkeypatch):
    wb, _path, pivot = _adopted(fixture_copy, tmp_path, monkeypatch)
    wb["Data"]["H1"] = '=GETPIVOTDATA("Amount","ByRegion")'
    before = _fingerprint(wb)
    with pytest.raises(UnsupportedStructureError) as renamed:
        pivot.rename("Other")
    assert renamed.value.kind == "unsupported-pivot-operation"
    assert "pivot-dependent-reference" in (renamed.value.options or [])
    with pytest.raises(UnsupportedStructureError) as deleted:
        pivot.delete()
    assert deleted.value.kind == "unsupported-pivot-operation"
    assert _fingerprint(wb)["operations"] == before["operations"]


def _paper_then_shared_foreign(fixture_copy, tmp_path):
    wb, path = _paper_then_foreign(fixture_copy, tmp_path)
    sibling = str(tmp_path / "shared-foreign.xlsx")
    pivot_rel = (
        b'<Relationship Id="rIdAdopt2" Type="http://schemas.openxmlformats.'
        b'org/officeDocument/2006/relationships/pivotTable" '
        b'Target="../pivotTables/pivotTable2.xml"/>'
    )
    override = (
        b'<Override PartName="/xl/pivotTables/pivotTable2.xml" '
        b'ContentType="application/vnd.openxmlformats-officedocument.'
        b'spreadsheetml.pivotTable+xml"/>'
    )
    with zipfile.ZipFile(path) as archive:
        output = io.BytesIO()
        with zipfile.ZipFile(output, "w") as after:
            for info in archive.infolist():
                body = archive.read(info.filename)
                if info.filename == "xl/pivotTables/pivotTable1.xml":
                    after.writestr(info, body)
                    after.writestr(
                        "xl/pivotTables/pivotTable2.xml",
                        body.replace(b'name="ByRegion"',
                                     b'name="ByRegionCopy"', 1),
                    )
                    continue
                if info.filename == "xl/worksheets/_rels/sheet1.xml.rels":
                    body = body.replace(
                        b"</Relationships>", pivot_rel + b"</Relationships>")
                if info.filename == "[Content_Types].xml":
                    body = body.replace(b"</Types>", override + b"</Types>")
                after.writestr(info, body)
        with open(sibling, "wb") as handle:
            handle.write(output.getvalue())
    return load_workbook(sibling, preserve=True), sibling


def test_shared_cache_is_not_dedicated_replacement(
        fixture_copy, tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    wb, _path = _paper_then_shared_foreign(fixture_copy, tmp_path)
    qualification = wb["Data"].pivots["ByRegion"].qualify_adoption()
    assert qualification.strategy == "shared-isolation"
    assert qualification.eligible is True


def test_formula_source_adopt_uses_oracle_and_does_not_publish_lo(
        fixture_copy, tmp_path, monkeypatch):
    from openpyxl.pivot.calculate import PivotCalculationArtifact

    _enable_evidence(monkeypatch)
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    _create_by_region(wb["Data"])
    created = str(tmp_path / "formula-created.xlsx")
    save_and_reopen(wb, created, preserve=True)
    _strip_paper_tag(created)
    wb = load_workbook(created, preserve=True)
    wb["Data"]["B2"] = "=20+0"

    def fake_calculate(workbook, snapshot):
        return PivotCalculationArtifact(
            candidate_sha256="abc",
            engine="libreoffice",
            engine_version="test",
            source=snapshot.source,
            source_identity=snapshot.identity,
            values_by_coordinate={addr: 20 for addr in snapshot.formula_coordinates},
            excluded_coordinates={},
            errors=(),
        )

    monkeypatch.setattr(
        "openpyxl.pivot.calculate.calculate_pivot_source", fake_calculate)
    wb["Data"].pivots["ByRegion"].adopt()
    dest = str(tmp_path / "formula-adopted.xlsx")
    receipt = wb.save(dest, receipt=True)
    events = _pivot_events(receipt)
    assert events[0]["kind"] == "pivot_adopted"
    assert events[0]["calculation"]["engine"] == "libreoffice"
    payloads = part_payloads(dest)
    joined = b"".join(payloads.values())
    assert b"PK\x03\x04lo-rewritten" not in joined
    reopened = load_workbook(dest, preserve=True)
    assert reopened["Data"]["F5"].value == 20


def test_missing_libreoffice_on_formula_adopt_is_atomic(
        fixture_copy, tmp_path, monkeypatch):
    _enable_evidence(monkeypatch)
    monkeypatch.setattr("openpyxl.oracle.find_soffice", lambda: None)
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    _create_by_region(wb["Data"])
    created = str(tmp_path / "lo-missing-created.xlsx")
    save_and_reopen(wb, created, preserve=True)
    _strip_paper_tag(created)
    wb = load_workbook(created, preserve=True)
    wb["Data"]["B2"] = "=20+0"
    before = _fingerprint(wb)
    with pytest.raises(UnsupportedStructureError) as exc:
        wb["Data"].pivots["ByRegion"].adopt()
    assert exc.value.kind == "unsupported-pivot-source"
    assert _fingerprint(wb) == before


def test_libreoffice_loads_adopted_workbook(
        fixture_copy, tmp_path, monkeypatch, lo):
    wb, _path, _pivot = _adopted(fixture_copy, tmp_path, monkeypatch)
    dest = str(tmp_path / "adopted-lo.xlsx")
    wb.save(dest)
    converted = lo.lo_convert(dest, fmt="xlsx")
    assert converted[:2] == b"PK"


def test_excel_transcript_remains_a_stub(fixture_copy, tmp_path, monkeypatch):
    import sys

    from .support.excel_pivot import excel_available, run_transcript

    try:
        wb, _path, _pivot = _adopted(fixture_copy, tmp_path, monkeypatch)
        dest = str(tmp_path / "adopted-excel.xlsx")
        wb.save(dest)
        if not excel_available():
            with pytest.raises(RuntimeError):
                run_transcript(dest, expected={})
            return
        with pytest.raises(NotImplementedError):
            run_transcript(dest, expected={})
    finally:
        for name in list(sys.modules):
            if "excel_pivot" in name:
                sys.modules.pop(name, None)


def test_graph_hash_drift_refuses_at_save(
        fixture_copy, tmp_path, monkeypatch):
    wb, _path, pivot = _adopted(fixture_copy, tmp_path, monkeypatch)
    staged = next(iter(wb._paper_ledger.pivot_operations.values()))
    object.__setattr__(
        staged, "original_payload_hashes", (("pivot", "0" * 64),))
    dest = str(tmp_path / "hash-drift.xlsx")
    with pytest.raises(UnsupportedStructureError) as exc:
        wb.save(dest)
    assert exc.value.kind == "stale-pivot"
    assert not os.path.exists(dest)
