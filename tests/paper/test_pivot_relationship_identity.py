"""Internal PivotTable relationship identity stays closed across rebuilds."""
from __future__ import annotations

import io
import zipfile
from xml.etree.ElementTree import fromstring

import pytest

from openpyxl import load_workbook
from openpyxl.errors import UnsupportedStructureError
from openpyxl.pivot.graph import load_pivot_graph
from openpyxl.xml.constants import REL_NS


_TABLE = "features/tables.xlsx"
_PIVOT_PART = "xl/pivotTables/pivotTable1.xml"
_PIVOT_RELS = "xl/pivotTables/_rels/pivotTable1.xml.rels"
_CACHE_PART = "xl/pivotCache/pivotCacheDefinition1.xml"
_CACHE_RELS = "xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels"
_REL_ID = "{%s}id" % REL_NS
_CACHE_REL_TYPE = (
    b"http://schemas.openxmlformats.org/officeDocument/2006/relationships/"
    b"pivotCacheDefinition"
)
_RECORDS_REL_TYPE = (
    b"http://schemas.openxmlformats.org/officeDocument/2006/relationships/"
    b"pivotCacheRecords"
)
_PIVOT_DECOY = (
    b'<Relationship Id="rId2" Type="' + _CACHE_REL_TYPE
    + b'" Target="../pivotCache/missing-cache.xml"/>'
)
_CACHE_DECOY = (
    b'<Relationship Id="rId2" Type="' + _RECORDS_REL_TYPE
    + b'" Target="missing-records.xml"/>'
)


def _rewrite_package(source, destination, replacements):
    with zipfile.ZipFile(source) as zin, zipfile.ZipFile(destination, "w") as zout:
        for info in zin.infolist():
            payload = zin.read(info.filename)
            for before, after in replacements.get(info.filename, ()):
                assert payload.count(before) == 1, (info.filename, before)
                payload = payload.replace(before, after)
            zout.writestr(info, payload)


def _create_pivot(source, destination):
    wb = load_workbook(source, preserve=True)
    wb["Data"].pivots.create(
        name="ByRegion",
        source="RegionTable",
        destination="E3",
        rows=["Region"],
        values=["Amount"],
    )
    wb.save(destination)


def _renumber_relationships(source, destination):
    _rewrite_package(source, destination, {
        _PIVOT_PART: ((b'r:id="rId1"', b'r:id="rId7"'),),
        _PIVOT_RELS: (
            (b'Id="rId1"', b'Id="rId7"'),
            (b'<Relationship Id="rId7"',
             _PIVOT_DECOY + b'<Relationship Id="rId7"'),
        ),
        _CACHE_PART: ((b'r:id="rId1"', b'r:id="rId8"'),),
        _CACHE_RELS: (
            (b'Id="rId1"', b'Id="rId8"'),
            (b'<Relationship Id="rId8"',
             _CACHE_DECOY + b'<Relationship Id="rId8"'),
        ),
    })


@pytest.fixture
def renumbered_pivot(fixture_copy, tmp_path):
    created = tmp_path / "created.xlsx"
    renumbered = tmp_path / "renumbered.xlsx"
    _create_pivot(fixture_copy(_TABLE), created)
    _renumber_relationships(created, renumbered)
    return renumbered


def _assert_relationship_closure(path):
    payload = path.read_bytes()
    graph = load_pivot_graph(payload)
    pivot = graph.pivots[0]
    cache = graph.caches_by_part[pivot.cache_definition_part]
    assert pivot.valid
    assert cache.valid
    assert pivot.cache_relationship_id == "rId7"
    assert cache.records_relationship_id == "rId8"

    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        pivot_root = fromstring(archive.read(_PIVOT_PART))
        cache_root = fromstring(archive.read(_CACHE_PART))
        assert pivot_root.attrib[_REL_ID] == "rId7"
        assert cache_root.attrib[_REL_ID] == "rId8"
        pivot_rels = fromstring(archive.read(_PIVOT_RELS))
        cache_rels = fromstring(archive.read(_CACHE_RELS))
        pivot_links = {
            node.attrib["Id"]: (node.attrib["Type"], node.attrib["Target"])
            for node in pivot_rels
        }
        cache_links = {
            node.attrib["Id"]: (node.attrib["Type"], node.attrib["Target"])
            for node in cache_rels
        }
        assert pivot_links["rId7"][0] == _CACHE_REL_TYPE.decode()
        assert pivot_links["rId2"][1].endswith("missing-cache.xml")
        assert cache_links["rId8"][0] == _RECORDS_REL_TYPE.decode()
        assert cache_links["rId2"][1] == "missing-records.xml"


@pytest.mark.parametrize("verb", ("refresh", "update", "move", "rename"))
def test_existing_mutators_preserve_internal_relationship_ids(
        renumbered_pivot, tmp_path, verb):
    wb = load_workbook(renumbered_pivot, preserve=True)
    pivot = wb["Data"].pivots["ByRegion"]
    assert pivot.valid

    if verb == "refresh":
        wb["Data"]["B2"] = 21
        pivot.refresh()
    elif verb == "update":
        pivot.update(layout="outline")
    elif verb == "move":
        pivot.move("G3")
    else:
        pivot.rename("RegionalSales")

    output = tmp_path / (verb + ".xlsx")
    wb.save(output)
    _assert_relationship_closure(output)
    reopened = load_workbook(output, preserve=True)
    assert list(reopened["Data"].pivots)[0].valid


def _duplicate_relationship(payload, relationship_id):
    marker = b'<Relationship Id="' + relationship_id + b'"'
    start = payload.index(marker)
    end = payload.index(b"/>", start) + 2
    relationship = payload[start:end]
    return payload[:end] + relationship + payload[end:]


@pytest.mark.parametrize(
    "hop, defect, reason",
    (
        ("pivot", "missing", "missing-internal-relationship"),
        ("pivot", "duplicate", "duplicate-relationship-id"),
        ("pivot", "type", "relationship-type-mismatch"),
        ("cache", "missing", "missing-internal-relationship"),
        ("cache", "duplicate", "duplicate-relationship-id"),
        ("cache", "type", "relationship-type-mismatch"),
    ),
)
def test_malformed_internal_relationship_disables_mutation(
        renumbered_pivot, tmp_path, hop, defect, reason):
    rels_part = _PIVOT_RELS if hop == "pivot" else _CACHE_RELS
    relationship_id = b"rId7" if hop == "pivot" else b"rId8"
    if defect == "missing":
        replacements = {
            rels_part: ((
                b'Id="' + relationship_id + b'"',
                b'Id="rId99"',
            ),),
        }
    elif defect == "type":
        current = _CACHE_REL_TYPE if hop == "pivot" else _RECORDS_REL_TYPE
        wrong = _RECORDS_REL_TYPE if hop == "pivot" else _CACHE_REL_TYPE
        prefix = b'Id="' + relationship_id + b'" Type="'
        replacements = {rels_part: ((prefix + current, prefix + wrong),)}
    else:
        with zipfile.ZipFile(renumbered_pivot) as archive:
            original = archive.read(rels_part)
        duplicated = _duplicate_relationship(original, relationship_id)
        replacements = {rels_part: ((original, duplicated),)}

    malformed = tmp_path / ("%s-%s.xlsx" % (hop, defect))
    _rewrite_package(renumbered_pivot, malformed, replacements)
    graph = load_pivot_graph(malformed.read_bytes())
    pivot_node = graph.pivots[0]
    cache_node = graph.caches_by_part[pivot_node.cache_definition_part]
    affected = pivot_node if hop == "pivot" else cache_node
    assert affected.valid is False
    assert reason in {item.code for item in affected.reasons}

    if hop == "cache":
        return

    wb = load_workbook(malformed, preserve=True)
    pivot = wb["Data"].pivots["ByRegion"]
    assert pivot.valid is False
    with pytest.raises(UnsupportedStructureError) as exc:
        pivot.rename("RegionalSales")
    assert exc.value.kind == "unsupported-pivot-operation"
    assert pivot.name == "ByRegion"
