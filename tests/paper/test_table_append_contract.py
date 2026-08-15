"""Closed success/refusal contract for ``Worksheet.append_table_row()``."""
from __future__ import annotations

import zipfile
from datetime import date

import pytest

from openpyxl import Workbook, load_workbook
from openpyxl.errors import TargetNotFoundError, UnsupportedStructureError
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import Table, TableFormula
from openpyxl.xml.functions import tostring

from .support.partdiff import diff_parts


TOTALS_COMMIT_STAGES = (
    "totals-style-A4",
    "totals-value-A4",
    "totals-format-A4",
    "totals-style-B4",
    "totals-value-B4",
    "totals-format-B4",
    "data-style-A3",
    "data-value-A3",
    "data-format-A3",
    "data-style-B3",
    "data-value-B3",
    "data-format-B3",
    "table-ref",
    "auto-filter-ref",
)


def _make_source(tmp_path, *, totals=False, totals_shown=False,
                 filter_includes_totals=True, protected=False):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Region", "Amount"])
    ws.append(["West", 10])
    ws["A2"].fill = PatternFill("solid", fgColor="FFF2CC")
    ws["B2"].number_format = "$#,##0.00"
    end_row = 2
    if totals:
        ws.append(["Total", "=SUBTOTAL(109,T[Amount])"])
        ws["A3"].font = Font(bold=True)
        ws["B3"].font = Font(bold=True)
        ws["B3"].number_format = "$#,##0.00"
        end_row = 3
    table = Table(displayName="T", ref="A1:B{0}".format(end_row))
    if totals_shown:
        table.totalsRowShown = True
    elif totals:
        table.totalsRowCount = 1
    ws.add_table(table)
    if totals and not filter_includes_totals:
        table.autoFilter = AutoFilter(ref="A1:B2")
    if protected:
        for coordinate in ("A2", "B2"):
            ws[coordinate].protection = Protection(locked=False)
        if totals:
            for coordinate in ("A3", "B3"):
                ws[coordinate].protection = Protection(locked=False)
        ws.protection.sheet = True
    source = tmp_path / "source.xlsx"
    wb.save(source)
    return source


def _preserved_table(tmp_path, **kwargs):
    return load_workbook(_make_source(tmp_path, **kwargs), preserve=True)


def _append_state(ws):
    wb = ws.parent
    table = ws.tables["T"]
    ledger = getattr(wb, "_paper_ledger", None)
    cells = []
    for coordinate, cell in sorted(ws._cells.items()):
        hyperlink = getattr(cell, "_hyperlink", None)
        comment = getattr(cell, "_comment", None)
        cells.append((
            coordinate,
            id(cell),
            cell._value,
            getattr(cell, "_data_type", getattr(cell, "data_type", None)),
            id(getattr(cell, "_style", None)),
            tuple(cell._style) if getattr(cell, "_style", None) is not None
            else None,
            id(hyperlink) if hyperlink is not None else None,
            getattr(hyperlink, "ref", None),
            id(comment) if comment is not None else None,
            id(getattr(comment, "_parent", None)) if comment is not None
            else None,
        ))
    ledger_state = None
    if ledger is not None and ledger.armed:
        ledger_state = (
            id(ledger.cells.get(ws)) if ws in ledger.cells else None,
            frozenset(ledger.cells.get(ws, set())),
            id(ledger.value_overwrites.get(ws))
            if ws in ledger.value_overwrites else None,
            frozenset(ledger.value_overwrites.get(ws, set())),
            id(ledger.cache_writes.get(ws))
            if ws in ledger.cache_writes else None,
            tuple(sorted(ledger.cache_writes.get(ws, {}).items())),
            ledger.formulas_changed,
            ws in ledger.protection_warned,
        )
    formats = wb._number_formats
    return (
        tuple(cells),
        ws._current_row,
        tostring(table.to_tree()),
        id(table.tableColumns),
        tuple(id(column) for column in table.tableColumns),
        id(table.autoFilter) if table.autoFilter is not None else None,
        ledger_state,
        tuple(formats),
        id(formats._dict),
        tuple(sorted(formats._dict.items())),
        formats.clean,
    )


@pytest.mark.parametrize("filter_includes_totals", [False, True])
@pytest.mark.parametrize("totals_shown", [False, True])
def test_totals_styles_formats_and_filter_convention_survive_round_trip(
        tmp_path, totals_shown, filter_includes_totals):
    wb = _preserved_table(
        tmp_path,
        totals=True,
        totals_shown=totals_shown,
        filter_includes_totals=filter_includes_totals,
    )
    ws = wb["Data"]
    data_style = ws["B2"]._style
    totals_style = ws["B3"]._style

    ws.append_table_row("T", {"Region": "East", "Amount": 20})

    table = ws.tables["T"]
    assert table.ref == "A1:B4"
    assert table.autoFilter.ref == (
        "A1:B4" if filter_includes_totals else "A1:B3")
    assert (ws["A3"].value, ws["B3"].value) == ("East", 20)
    assert (ws["A4"].value, ws["B4"].value) == (
        "Total", "=SUBTOTAL(109,T[Amount])")
    assert ws["B3"]._style == data_style
    assert ws["B4"]._style == totals_style
    assert ws["B3"].number_format == "$#,##0.00"

    output = tmp_path / "totals-output.xlsx"
    wb.save(output)
    reopened = load_workbook(output)
    sheet = reopened["Data"]
    assert sheet.tables["T"].ref == "A1:B4"
    assert sheet.tables["T"].autoFilter.ref == table.autoFilter.ref
    assert sheet["B3"].number_format == "$#,##0.00"
    assert sheet["B4"].value == "=SUBTOTAL(109,T[Amount])"


def test_declared_calculated_column_is_populated_without_guessing(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Region", "Amount", "Double"])
    ws.append(["West", 10, "=[@Amount]*2"])
    ws.add_table(Table(displayName="T", ref="A1:C2"))
    source = tmp_path / "calculated.xlsx"
    wb.save(source)

    wb = load_workbook(source, preserve=True)
    ws = wb.active
    ws.tables["T"].tableColumns[2].calculatedColumnFormula = TableFormula(
        attr_text="=[@Amount]*2")
    ws.append_table_row("T", {"Region": "East", "Amount": 20})
    assert ws["C3"].value == "=[@Amount]*2"

    output = tmp_path / "calculated-output.xlsx"
    wb.save(output)
    assert load_workbook(output, data_only=False).active["C3"].value == \
        "=[@Amount]*2"


def test_array_calculated_column_refuses_before_mutation(tmp_path):
    wb = _preserved_table(tmp_path)
    ws = wb["Data"]
    ws.tables["T"].tableColumns[1].calculatedColumnFormula = TableFormula(
        array=True, attr_text="=[Amount]*2")
    before = _append_state(ws)
    with pytest.raises(UnsupportedStructureError, match="array"):
        ws.append_table_row("T", {"Region": "East"})
    assert _append_state(ws) == before


def test_inherited_style_keeps_automatic_date_number_format(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["When"])
    ws.append(["pending"])
    ws["A2"].fill = PatternFill("solid", fgColor="FFF2CC")
    ws.add_table(Table(displayName="T", ref="A1:A2"))
    source = tmp_path / "dates.xlsx"
    wb.save(source)

    wb = load_workbook(source, preserve=True)
    ws = wb.active
    inherited_fill_id = ws["A2"]._style.fillId
    ws.append_table_row("T", [date(2026, 8, 15)])

    assert ws["A3"]._style.fillId == inherited_fill_id
    assert ws["A3"].is_date
    assert ws["A3"].number_format != "General"


def test_date_format_registry_rolls_back_after_injected_failure(
        tmp_path, monkeypatch):
    from openpyxl.preserve import tables

    wb = Workbook()
    ws = wb.active
    ws.append(["When"])
    ws.append(["pending"])
    ws.add_table(Table(displayName="T", ref="A1:A2"))
    source = tmp_path / "date-rollback.xlsx"
    wb.save(source)
    wb = load_workbook(source, preserve=True)
    ws = wb.active
    before = _append_state(ws)

    def fail_at(name):
        if name == "data-value-A3":
            raise RuntimeError("injected")

    monkeypatch.setattr(tables, "_append_commit_point", fail_at)
    with pytest.raises(RuntimeError, match="injected"):
        ws.append_table_row("T", [date(2026, 8, 15)])
    assert _append_state(ws) == before


def test_new_unsaved_table_supports_mapping_and_initializes_columns(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Region", "Amount"])
    ws.append(["West", 10])
    table = Table(displayName="T", ref="A1:B2")
    ws.add_table(table)
    assert not table.tableColumns

    ws.append_table_row("T", {"Region": "East", "Amount": 20})

    assert [column.name for column in table.tableColumns] == [
        "Region", "Amount"]
    assert table.ref == "A1:B3"
    assert table.autoFilter.ref == "A1:B3"
    output = tmp_path / "new-table.xlsx"
    wb.save(output)
    reopened = load_workbook(output)
    assert reopened.active["A3"].value == "East"
    assert reopened.active.tables["T"].column_names == ["Region", "Amount"]


def test_stock_loaded_table_requires_retained_source(tmp_path):
    source = _make_source(tmp_path)
    wb = load_workbook(source, preserve=False)
    ws = wb["Data"]
    before = _append_state(ws)

    with pytest.raises(
            UnsupportedStructureError, match="reopen with preserve=True"):
        ws.append_table_row("T", ["East", 20])

    assert _append_state(ws) == before


def test_missing_named_table_is_typed_and_does_not_create_cells(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "sentinel"
    before = dict(ws._cells)
    with pytest.raises(TargetNotFoundError, match="no table named"):
        ws.append_table_row("Missing", [1])
    assert dict(ws._cells) == before


@pytest.mark.parametrize(
    ("values", "message"),
    (
        ({"Unknown": 1}, "unknown column"),
        (["East", 20, "extra"], "received 3 values"),
    ),
)
def test_ambiguous_value_mapping_refuses_without_mutation(
        tmp_path, values, message):
    wb = _preserved_table(tmp_path)
    ws = wb["Data"]
    before = _append_state(ws)
    with pytest.raises(UnsupportedStructureError, match=message):
        ws.append_table_row("T", values)
    assert _append_state(ws) == before


def test_repeated_appends_expand_only_the_named_table(tmp_path):
    wb = _preserved_table(tmp_path)
    ws = wb["Data"]
    ws["A8"] = "unrelated"

    ws.append_table_row("T", ["East", 20])
    ws.append_table_row("T", ["North", 30])

    assert ws.tables["T"].ref == "A1:B4"
    assert [ws.cell(row, 1).value for row in (2, 3, 4)] == [
        "West", "East", "North"]
    assert ws["A8"].value == "unrelated"


def test_success_changes_only_sheet_and_table_parts(tmp_path):
    source = _make_source(tmp_path)
    wb = load_workbook(source, preserve=True)
    wb["Data"].append_table_row("T", ["East", 20])
    output = tmp_path / "exact-delta.xlsx"
    wb.save(output)

    diff = diff_parts(source, output)
    assert not diff.added
    assert not diff.removed
    assert diff.changed == {
        "xl/worksheets/sheet1.xml",
        "xl/tables/table1.xml",
    }
    reopened = load_workbook(output)
    assert reopened["Data"].tables["T"].ref == "A1:B3"
    assert reopened["Data"]["A3"].value == "East"


def test_commit_point_inventory_is_complete(tmp_path, monkeypatch):
    from openpyxl.preserve import tables

    wb = _preserved_table(tmp_path, totals=True, totals_shown=True)
    seen = []
    monkeypatch.setattr(tables, "_append_commit_point", seen.append)
    wb["Data"].append_table_row("T", ["East", 20])
    assert tuple(seen) == TOTALS_COMMIT_STAGES


@pytest.mark.parametrize("stage", TOTALS_COMMIT_STAGES)
def test_every_totals_commit_failure_rolls_back_exactly(
        tmp_path, monkeypatch, stage):
    from openpyxl.preserve import tables

    wb = _preserved_table(tmp_path, totals=True, totals_shown=True)
    ws = wb["Data"]
    before = _append_state(ws)

    def fail_at(name):
        if name == stage:
            raise RuntimeError("injected at {0}".format(name))

    monkeypatch.setattr(tables, "_append_commit_point", fail_at)
    with pytest.raises(RuntimeError, match="injected"):
        ws.append_table_row("T", ["East", 20])
    assert _append_state(ws) == before


def test_column_initialization_failure_restores_exact_table_state(
        tmp_path, monkeypatch):
    from openpyxl.preserve import tables

    wb = Workbook()
    ws = wb.active
    ws.append(["Region", "Amount"])
    ws.append(["West", 10])
    ws.add_table(Table(displayName="T", ref="A1:B2"))
    before = _append_state(ws)

    def fail_at(name):
        if name == "columns-initialized":
            raise RuntimeError("injected")

    monkeypatch.setattr(tables, "_append_commit_point", fail_at)
    with pytest.raises(RuntimeError, match="injected"):
        ws.append_table_row("T", {"Region": "East", "Amount": 20})
    assert _append_state(ws) == before


def test_base_exception_during_commit_rolls_back_exactly(
        tmp_path, monkeypatch):
    from openpyxl.preserve import tables

    wb = _preserved_table(tmp_path, totals=True, totals_shown=True)
    ws = wb["Data"]
    before = _append_state(ws)

    def interrupt(name):
        if name == "table-ref":
            raise KeyboardInterrupt("injected")

    monkeypatch.setattr(tables, "_append_commit_point", interrupt)
    with pytest.raises(KeyboardInterrupt, match="injected"):
        ws.append_table_row("T", ["East", 20])
    assert _append_state(ws) == before


@pytest.mark.parametrize(
    ("attribute", "value", "message"),
    (
        ("tableType", "queryTable", "query|connected"),
        ("connectionId", 7, "query|connected"),
        ("insertRow", True, "insert-row"),
    ),
)
def test_unsupported_model_states_refuse_without_mutation(
        tmp_path, attribute, value, message):
    wb = _preserved_table(tmp_path)
    ws = wb["Data"]
    setattr(ws.tables["T"], attribute, value)
    before = _append_state(ws)
    with pytest.raises(UnsupportedStructureError, match=message):
        ws.append_table_row("T", ["East", 20])
    assert _append_state(ws) == before


def test_duplicate_table_column_ids_refuse_without_mutation(tmp_path):
    wb = _preserved_table(tmp_path)
    ws = wb["Data"]
    ws.tables["T"].tableColumns[1].id = \
        ws.tables["T"].tableColumns[0].id
    before = _append_state(ws)
    with pytest.raises(UnsupportedStructureError, match="tableColumn ids"):
        ws.append_table_row("T", ["East", 20])
    assert _append_state(ws) == before


@pytest.mark.parametrize(
    ("malform", "message"),
    (
        ("duplicate-header", "duplicate"),
        ("missing-column", "defines 1 tableColumns"),
        ("bad-filter", "autoFilter range"),
        ("no-data-row", "no data row"),
    ),
)
def test_malformed_table_geometry_refuses_during_call(
        tmp_path, malform, message):
    wb = _preserved_table(tmp_path)
    ws = wb["Data"]
    table = ws.tables["T"]
    if malform == "duplicate-header":
        ws["B1"] = "Region"
    elif malform == "missing-column":
        table.tableColumns.pop()
    elif malform == "bad-filter":
        table.autoFilter.ref = "A1:A2"
    else:
        table.ref = "A1:B1"
        table.autoFilter.ref = "A1:B1"
    before = _append_state(ws)
    with pytest.raises(UnsupportedStructureError, match=message):
        ws.append_table_row("T", ["East", 20])
    assert _append_state(ws) == before


def test_preexisting_anchor_move_refuses_during_append_not_save(tmp_path):
    wb = _preserved_table(tmp_path)
    ws = wb["Data"]
    ws["D1"] = "Region"
    ws["E1"] = "Amount"
    ws["D2"] = "North"
    ws["E2"] = 30
    ws.tables["T"].ref = "D1:E2"
    ws.tables["T"].autoFilter.ref = "D1:E2"
    before = _append_state(ws)
    with pytest.raises(UnsupportedStructureError, match="anchor"):
        ws.append_table_row("T", ["East", 20])
    assert _append_state(ws) == before


@pytest.mark.parametrize("surface", ["merged", "array"])
def test_destination_formula_or_merge_surface_refuses_atomically(
        tmp_path, surface):
    wb = _preserved_table(tmp_path)
    ws = wb["Data"]
    if surface == "merged":
        ws.merge_cells("A3:B3")
    else:
        ws["A3"] = ArrayFormula(ref="A3:B3", text="=A1:B1")
    before = _append_state(ws)
    with pytest.raises(UnsupportedStructureError, match="merged|array|spill"):
        ws.append_table_row("T", ["East", 20])
    assert _append_state(ws) == before


def test_table_part_relationship_refuses_before_mutation(tmp_path):
    source = _make_source(tmp_path)
    connected = tmp_path / "connected.xlsx"
    relationship = (
        b'<?xml version="1.0" encoding="UTF-8"?>'
        b'<Relationships xmlns="http://schemas.openxmlformats.org/'
        b'package/2006/relationships">'
        b'<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/'
        b'officeDocument/2006/relationships/queryTable" '
        b'Target="../queryTables/queryTable1.xml"/>'
        b'</Relationships>'
    )
    with zipfile.ZipFile(source) as before, zipfile.ZipFile(
            connected, "w") as after:
        for item in before.infolist():
            after.writestr(item, before.read(item.filename))
        after.writestr("xl/tables/_rels/table1.xml.rels", relationship)

    wb = load_workbook(connected, preserve=True)
    ws = wb["Data"]
    state = _append_state(ws)
    with pytest.raises(UnsupportedStructureError, match="relationships"):
        ws.append_table_row("T", ["East", 20])
    assert _append_state(ws) == state


def test_strict_protection_uses_planned_inherited_style(tmp_path):
    wb = _preserved_table(tmp_path, protected=True)
    wb.strict_protection = True
    ws = wb["Data"]

    ws.append_table_row("T", ["East", 20])

    assert ws["A3"].protection.locked is False
    assert ws["B3"].protection.locked is False


def test_strict_protection_refuses_locked_inherited_style(tmp_path):
    wb = _preserved_table(tmp_path)
    wb.strict_protection = True
    ws = wb["Data"]
    ws.protection.sheet = True
    before = _append_state(ws)
    with pytest.raises(UnsupportedStructureError, match="locked data cell"):
        ws.append_table_row("T", ["East", 20])
    assert _append_state(ws) == before


@pytest.mark.lo_smoke
def test_appended_table_loads_in_libreoffice(tmp_path, lo):
    wb = _preserved_table(tmp_path, totals=True, totals_shown=True)
    wb["Data"].append_table_row("T", ["East", 20])
    output = tmp_path / "libreoffice-table.xlsx"
    wb.save(output)
    assert lo.lo_loads(output)
