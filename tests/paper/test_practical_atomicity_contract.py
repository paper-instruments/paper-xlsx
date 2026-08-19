from copy import copy
import datetime
import io
import warnings

import pytest

from openpyxl import Workbook, load_workbook
from openpyxl.errors import UnsupportedStructureError


class _Interrupt(BaseException):
    pass


def _preserved_workbook():
    workbook = Workbook()
    workbook.active.title = "Data"
    workbook.active["A1"] = 1
    workbook.active["B1"] = 2
    workbook.create_sheet("Other")
    target = io.BytesIO()
    workbook.save(target)
    return load_workbook(io.BytesIO(target.getvalue()), preserve=True)


def _ledger_cells(workbook):
    return {
        sheet: set(values)
        for sheet, values in workbook._paper_ledger.cells.items()}


def _ledger_state(workbook):
    ledger = workbook._paper_ledger
    return {slot: repr(getattr(ledger, slot)) for slot in ledger.__slots__}


def test_interrupt_during_cell_ledger_mark_rolls_back_value(monkeypatch):
    from openpyxl.cell import cell as cell_module

    workbook = _preserved_workbook()
    cell = workbook["Data"]["A1"]
    before = (cell.value, cell.data_type, _ledger_cells(workbook))
    real_mark = cell_module._mark_cell_dirty

    def interrupt(*args, **kwargs):
        real_mark(*args, **kwargs)
        raise KeyboardInterrupt("injected")

    monkeypatch.setattr(cell_module, "_mark_cell_dirty", interrupt)
    with pytest.raises(KeyboardInterrupt):
        cell.value = "changed"

    assert (cell.value, cell.data_type, _ledger_cells(workbook)) == before


@pytest.mark.parametrize("mutation", ["data_type", "hyperlink", "comment"])
def test_interrupt_during_cell_metadata_mark_rolls_back(
        monkeypatch, mutation):
    from openpyxl.cell import cell as cell_module
    from openpyxl.comments import Comment

    workbook = _preserved_workbook()
    cell = workbook["Data"]["A1"]
    before = (cell.data_type, cell.hyperlink, cell.comment,
              _ledger_cells(workbook))
    real_mark = cell_module._mark_cell_dirty

    def interrupt(*args, **kwargs):
        real_mark(*args, **kwargs)
        raise KeyboardInterrupt("injected")

    monkeypatch.setattr(cell_module, "_mark_cell_dirty", interrupt)
    with pytest.raises(KeyboardInterrupt):
        if mutation == "data_type":
            cell.data_type = "s"
        elif mutation == "hyperlink":
            cell.hyperlink = "https://example.com"
        else:
            cell.comment = Comment("note", "author")

    assert (cell.data_type, cell.hyperlink, cell.comment,
            _ledger_cells(workbook)) == before


def test_noop_data_type_assignment_does_not_mark_value_overwrite():
    workbook = _preserved_workbook()
    cell = workbook["Data"]["A1"]

    cell.data_type = cell.data_type

    assert workbook._paper_ledger.cells.get(cell.parent, set()) == set()
    assert workbook._paper_ledger.value_overwrites.get(
        cell.parent, set()) == set()


def test_strict_protection_refuses_data_type_change_atomically():
    workbook = _preserved_workbook()
    sheet = workbook["Data"]
    sheet["A1"] = "=1+1"
    sheet.protection.sheet = True
    workbook.strict_protection = True
    cell = sheet["A1"]
    before = (cell.value, cell.data_type, _ledger_state(workbook))

    with pytest.raises(UnsupportedStructureError):
        cell.data_type = "s"

    assert (cell.value, cell.data_type, _ledger_state(workbook)) == before


def test_append_failure_restores_partial_row_and_ledger():
    workbook = _preserved_workbook()
    sheet = workbook["Data"]
    before_package = io.BytesIO()
    workbook.save(before_package)
    before_cells = dict(sheet._cells)
    before_row = sheet._current_row
    before_ledger = _ledger_cells(workbook)

    with pytest.raises(ValueError):
        sheet.append(["=1+1", object()])

    assert sheet._cells == before_cells
    assert sheet._current_row == before_row
    assert _ledger_cells(workbook) == before_ledger
    assert not workbook._paper_ledger.formulas_changed
    after_package = io.BytesIO()
    workbook.save(after_package)
    assert after_package.getvalue() == before_package.getvalue()


def test_append_past_last_row_refuses_without_consuming_generator():
    from openpyxl.errors import BoundaryViolationError

    workbook = _preserved_workbook()
    sheet = workbook["Data"]
    sheet._current_row = 1048576
    consumed = []

    def values():
        consumed.append(True)
        yield 1

    with pytest.raises(BoundaryViolationError, match="1048577"):
        sheet.append(values())

    assert consumed == []
    assert sheet._current_row == 1048576
    assert (1048577, 1) not in sheet._cells


@pytest.mark.parametrize("column", [0, -1, 16385, 1.5])
def test_append_invalid_numeric_column_rolls_back_partial_dict(column):
    from openpyxl.errors import BoundaryViolationError

    workbook = _preserved_workbook()
    sheet = workbook["Data"]
    before_cells = dict(sheet._cells)
    before_row = sheet._current_row
    before_ledger = _ledger_state(workbook)

    with pytest.raises(BoundaryViolationError, match="column keys"):
        sheet.append({1: "partial", column: "invalid"})

    assert sheet._cells == before_cells
    assert sheet._current_row == before_row
    assert _ledger_state(workbook) == before_ledger


def test_append_failure_restores_prebuilt_cell_binding():
    from openpyxl.cell import Cell

    workbook = _preserved_workbook()
    sheet = workbook["Data"]
    cell = Cell(sheet, value="prebuilt")
    before = (cell.parent, cell.row, cell.column)

    with pytest.raises(ValueError):
        sheet.append([cell, object()])

    assert (cell.parent, cell.row, cell.column) == before


def test_structural_interrupt_restores_cells_and_ledger(monkeypatch):
    from openpyxl.preserve import structural

    workbook = _preserved_workbook()
    sheet = workbook["Data"]
    before_cells = dict(sheet._cells)
    before_ledger = _ledger_cells(workbook)

    def interrupt(*_args, **_kwargs):
        raise KeyboardInterrupt("injected")

    monkeypatch.setattr(structural, "apply_model_shift", interrupt)
    with pytest.raises(KeyboardInterrupt):
        sheet.insert_rows(1)

    assert sheet._cells == before_cells
    assert _ledger_cells(workbook) == before_ledger


def test_sheet_create_and_remove_failures_restore_workbook(monkeypatch):
    from openpyxl.preserve import ledger

    workbook = _preserved_workbook()
    sheets = workbook._sheets
    before = list(sheets)
    before_ledger = _ledger_state(workbook)
    real_added = ledger.mark_sheet_added

    def fail_added(*args, **kwargs):
        real_added(*args, **kwargs)
        raise RuntimeError("injected create failure")

    monkeypatch.setattr(ledger, "mark_sheet_added", fail_added)
    with pytest.raises(RuntimeError, match="create failure"):
        workbook.create_sheet("Failed")
    assert workbook._sheets is sheets
    assert list(sheets) == before
    assert _ledger_state(workbook) == before_ledger

    monkeypatch.setattr(ledger, "mark_sheet_added", real_added)
    real_removed = ledger.record_sheet_removal

    def fail_removed(*args, **kwargs):
        real_removed(*args, **kwargs)
        raise RuntimeError("injected remove failure")

    monkeypatch.setattr(ledger, "record_sheet_removal", fail_removed)
    with pytest.raises(RuntimeError, match="remove failure"):
        workbook.remove(workbook["Other"])
    assert workbook._sheets is sheets
    assert list(sheets) == before
    assert _ledger_state(workbook) == before_ledger


def test_table_append_invalid_value_happens_before_commit(tmp_path):
    from openpyxl.worksheet.table import Table

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Amount", "Formula"])
    sheet.append([1, "=SUM(A2)"])
    sheet.add_table(Table(displayName="Inputs", ref="A1:B2"))
    source = tmp_path / "table.xlsx"
    workbook.save(source)
    workbook = load_workbook(source, preserve=True)
    before = io.BytesIO()
    workbook.save(before)

    with pytest.raises(ValueError, match="Cannot convert"):
        workbook.active.append_table_row("Inputs", [3, object()])

    after = io.BytesIO()
    workbook.save(after)
    assert after.getvalue() == before.getvalue()


@pytest.mark.parametrize("failure", [RuntimeError, _Interrupt])
def test_datetime_bind_failure_restores_cell_registry_and_ledger(
        monkeypatch, failure):
    import openpyxl.cell.cell as cell_module

    workbook = _preserved_workbook()
    sheet = workbook["Data"]
    cell = sheet["A1"]
    ledger = workbook._paper_ledger
    ledger.cells[sheet] = {(9, 9)}
    ledger.value_overwrites[sheet] = {(8, 8)}
    ledger.cache_writes[sheet] = {(1, 1): 42, (7, 7): 7}
    cells_bucket = ledger.cells[sheet]
    overwrite_bucket = ledger.value_overwrites[sheet]
    cache_bucket = ledger.cache_writes[sheet]
    before_style = copy(cell._style)
    registry = workbook._number_formats
    before_registry = list(registry)
    before_index = dict(registry._dict)
    index_identity = registry._dict
    before_clean = registry.clean
    real_mark = cell_module._mark_cell_dirty

    def fail_after_mark(*args, **kwargs):
        real_mark(*args, **kwargs)
        raise failure("injected")

    monkeypatch.setattr(cell_module, "_mark_cell_dirty", fail_after_mark)
    with pytest.raises(failure):
        cell.value = datetime.datetime(2026, 8, 13, 12, 30)

    assert (cell.value, cell.data_type, cell._style) == (1, "n", before_style)
    assert cell.number_format == "General"
    assert list(registry) == before_registry
    assert registry._dict is index_identity
    assert dict(registry._dict) == before_index
    assert registry.clean == before_clean
    assert ledger.cells[sheet] is cells_bucket
    assert ledger.value_overwrites[sheet] is overwrite_bucket
    assert ledger.cache_writes[sheet] is cache_bucket
    assert cells_bucket == {(9, 9)}
    assert overwrite_bucket == {(8, 8)}
    assert cache_bucket == {(1, 1): 42, (7, 7): 7}


def test_successful_cell_edit_invalidates_only_its_staged_cache():
    workbook = _preserved_workbook()
    sheet = workbook["Data"]
    cache = {(1, 1): 42, (7, 7): 7}
    workbook._paper_ledger.cache_writes[sheet] = cache

    sheet["A1"] = 2

    assert workbook._paper_ledger.cache_writes[sheet] is cache
    assert cache == {(7, 7): 7}


def test_explicit_number_format_failure_restores_registry(monkeypatch):
    import openpyxl.styles.styleable as styleable

    workbook = _preserved_workbook()
    cell = workbook["Data"]["A1"]
    before_style = copy(cell._style)
    registry = workbook._number_formats
    before = (list(registry), dict(registry._dict), registry.clean)
    real_mark = styleable._mark_styleable_dirty

    def fail_after_mark(*args, **kwargs):
        real_mark(*args, **kwargs)
        raise RuntimeError("injected")

    monkeypatch.setattr(styleable, "_mark_styleable_dirty", fail_after_mark)
    with pytest.raises(RuntimeError, match="injected"):
        cell.number_format = '0.0000 "units"'

    assert cell._style == before_style
    assert cell.number_format == "General"
    assert (list(registry), dict(registry._dict), registry.clean) == before


def test_merged_interior_number_format_assignment_is_atomic(tmp_path,
                                                            monkeypatch):
    import openpyxl.styles.styleable as styleable

    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:B1")
    source = tmp_path / "merged.xlsx"
    workbook.save(source)

    workbook = load_workbook(source, preserve=True)
    sheet = workbook.active
    interior = sheet["B1"]
    registry = workbook._number_formats
    before_style = copy(interior._style)
    before_registry = (list(registry), dict(registry._dict), registry.clean)
    real_mark = styleable._mark_styleable_dirty

    def fail_after_mark(*args, **kwargs):
        real_mark(*args, **kwargs)
        raise RuntimeError("injected")

    monkeypatch.setattr(styleable, "_mark_styleable_dirty", fail_after_mark)
    with pytest.raises(RuntimeError, match="injected"):
        interior.number_format = '0.0000 "units"'

    assert interior._style == before_style
    assert interior.number_format == "General"
    assert (list(registry), dict(registry._dict), registry.clean) \
        == before_registry
    assert workbook._paper_ledger.cells.get(sheet, set()) == set()


def test_merged_interior_number_format_assignment_saves(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.merge_cells("A1:B1")
    source = tmp_path / "merged.xlsx"
    workbook.save(source)

    workbook = load_workbook(source, preserve=True)
    interior = workbook.active["B1"]
    interior.number_format = "0.00"
    assert interior.number_format == "0.00"
    output = tmp_path / "formatted.xlsx"
    workbook.save(output)
    assert output.exists()


def test_append_refuses_reentrancy_and_rolls_back_partial_row():
    workbook = _preserved_workbook()
    sheet = workbook["Data"]
    before_cells = dict(sheet._cells)
    before_row = sheet._current_row

    def values():
        yield 2
        sheet.append([99])

    with pytest.raises(RuntimeError, match="re-entrant"):
        sheet.append(values())
    assert sheet._cells == before_cells
    assert sheet._current_row == before_row


def test_append_rollback_does_not_revert_generator_edit_on_other_sheet():
    workbook = _preserved_workbook()
    target = workbook["Data"]
    other = workbook["Other"]
    before_cells = dict(target._cells)

    def values():
        yield 2
        other["A1"] = "unrelated"
        raise RuntimeError("generator failed")

    with pytest.raises(RuntimeError, match="generator failed"):
        target.append(values())
    assert target._cells == before_cells
    assert other["A1"].value == "unrelated"
