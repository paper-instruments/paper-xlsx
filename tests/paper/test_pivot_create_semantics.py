"""PR 5: full v1 pivot creation vocabulary and refusals."""
from __future__ import annotations

import os
from datetime import date, datetime

import pytest

from openpyxl import Workbook, load_workbook
from openpyxl.errors import BoundaryViolationError, UnsupportedStructureError
from openpyxl.pivot.api_types import (
    PivotAxisField,
    PivotItemFilter,
    PivotMeasure,
)
from openpyxl.pivot.cache import CacheDefinition
from openpyxl.pivot.table import TableDefinition
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, range_boundaries
from openpyxl.utils.datetime import MAC_EPOCH, WINDOWS_EPOCH
from openpyxl.worksheet.table import Table
from openpyxl.xml.functions import fromstring

from .conftest import FIXTURES_DIR
from .support.harness import assert_part_budget, save_and_reopen
from .support.partdiff import part_payloads


_PIVOTS = os.path.join(FIXTURES_DIR, "pivots")
_ADDED = {
    "xl/pivotTables/pivotTable1.xml",
    "xl/pivotTables/_rels/pivotTable1.xml.rels",
    "xl/pivotCache/pivotCacheDefinition1.xml",
    "xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels",
    "xl/pivotCache/pivotCacheRecords1.xml",
}
_HEADERS = [
    "Region", "Product", "Quarter", "Status", "Amount", "Units", "Flag", "Day", "Code",
]
_ROWS = [
    ["East", "A", "Q1", "Closed", 10, 1, True, date(2024, 1, 1), "001"],
    ["East", "A", "Q2", "Closed", 4, 1, False, date(2024, 1, 2), "002"],
    ["East", "B", "Q1", "Pending", 7, 2, True, date(2024, 1, 3), "001"],
    ["West", "A", "Q1", "Closed", 6, 1, False, date(2024, 1, 4), "002"],
    ["West", "B", "Q2", "Closed", 3, 1, True, date(2024, 1, 5), "001"],
    ["東", "東京", "Q2", "Closed", 5, 1, False, date(2024, 1, 6), "001"],
]


def _write_matrix(ws, headers, rows, origin="A1"):
    column_letter, row = coordinate_from_string(origin)
    column = column_index_from_string(column_letter)
    for offset, header in enumerate(headers):
        ws.cell(row, column + offset, header)
    for row_offset, record in enumerate(rows, start=1):
        for col_offset, value in enumerate(record):
            ws.cell(row + row_offset, column + col_offset, value)
    last = "%s%s" % (
        get_column_letter(column + len(headers) - 1),
        row + len(rows),
    )
    return "%s:%s" % (origin, last)


def _preserved(tmp_path, headers=None, rows=None, sheet="Data", table=None,
               origin="A1", epoch=None, name="src.xlsx", summary="Summary"):
    wb = Workbook()
    if epoch is not None:
        wb.epoch = epoch
    ws = wb.active
    ws.title = sheet
    ref = _write_matrix(ws, headers or _HEADERS, rows or _ROWS, origin)
    if table:
        ws.add_table(Table(displayName=table, ref=ref))
    if summary and summary != sheet:
        wb.create_sheet(summary)
    path = str(tmp_path / name)
    wb.save(path)
    return path, load_workbook(path, preserve=True)


def _grid(ws, ref):
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    return {
        (row, column): ws.cell(row, column).value
        for row in range(min_row, max_row + 1)
        for column in range(min_col, max_col + 1)
    }


def _parse_created(path):
    parts = part_payloads(path)
    cache = CacheDefinition.from_tree(fromstring(
        parts["xl/pivotCache/pivotCacheDefinition1.xml"]))
    table = TableDefinition.from_tree(fromstring(
        parts["xl/pivotTables/pivotTable1.xml"]))
    return cache, table, parts


def test_nested_rows_columns_and_two_measures(tmp_path):
    src, wb = _preserved(tmp_path, table="SalesData")
    ws = wb["Summary"]
    handle = ws.pivots.create(
        name="SalesByRegion",
        source="SalesData",
        destination="B4",
        rows=[PivotAxisField("Region"), PivotAxisField("Product")],
        columns=[PivotAxisField("Quarter")],
        values=[
            PivotMeasure("Amount", aggregate="sum", caption="Revenue"),
            PivotMeasure("Units", aggregate="sum", caption="Units"),
        ],
        layout="tabular",
        values_axis="columns",
        style="PivotStyleMedium9",
    )
    assert handle.valid is True
    assert handle.origin == "paper"
    payload = handle.to_dict()
    assert payload["rows"] == ["Region", "Product"]
    assert payload["columns"] == ["Quarter"]
    assert payload["values"] == [
        {"field": "Amount", "aggregate": "sum"},
        {"field": "Units", "aggregate": "sum"},
    ]
    dest = src + ".nested.xlsx"
    reopened = save_and_reopen(wb, dest, preserve=True)
    pivot = reopened["Summary"].pivots["SalesByRegion"]
    assert pivot.valid is True
    assert pivot.to_dict()["columns"] == ["Quarter"]
    assert pivot.spec.layout == "tabular"
    assert pivot.spec.style == "PivotStyleMedium9"
    cache, table, _parts = _parse_created(dest)
    assert cache.recordCount == 6
    assert len(table.rowFields) == 2
    assert any(field.x == -2 for field in table.colFields)
    assert len(table.dataFields) == 2
    assert table.dataFields[0].name == "Revenue"
    grid = _grid(reopened["Summary"], pivot.output_range)
    assert "East" in grid.values()
    assert "Revenue" in grid.values()
    assert_part_budget(
        src, dest,
        expect_changed={
            "xl/worksheets/sheet2.xml",
            "xl/workbook.xml",
            "xl/_rels/workbook.xml.rels",
            "[Content_Types].xml",
        },
        expect_added=_ADDED | {"xl/worksheets/_rels/sheet2.xml.rels"},
    )


def test_values_on_rows_and_filter_include(tmp_path):
    src, wb = _preserved(tmp_path, table="SalesData")
    handle = wb["Summary"].pivots.create(
        name="ClosedOnly",
        source="SalesData",
        destination="B4",
        rows=["Region"],
        filters=[PivotItemFilter("Status", include=["Closed", "Pending"])],
        values=[
            PivotMeasure("Amount", aggregate="sum", caption="Sum"),
            PivotMeasure("Amount", aggregate="count", caption="Count"),
        ],
        values_axis="rows",
        row_grand_totals=False,
    )
    dest = src + ".rows.xlsx"
    reopened = save_and_reopen(wb, dest, preserve=True)
    pivot = reopened["Summary"].pivots["ClosedOnly"]
    assert pivot.spec.values_axis == "rows"
    assert pivot.to_dict()["filters"][0]["field"] == "Status"
    assert set(pivot.to_dict()["filters"][0]["include"]) == {"Closed", "Pending"}
    cache, table, _parts = _parse_created(dest)
    assert table.dataOnRows is True
    assert table.pageFields[0].fld == 3
    assert table.location.rowPageCount == 1
    assert table.location.colPageCount == 1
    assert reopened["Summary"]["B2"].value == "Status"
    assert reopened["Summary"]["C2"].value == "(Multiple Items)"
    assert "Sum" in _grid(reopened["Summary"], pivot.output_range).values()
    assert handle.output_range == pivot.output_range


def test_duplicate_filter_items_are_collapsed_before_serialization(tmp_path):
    src, wb = _preserved(tmp_path, table="SalesData")
    wb["Summary"].pivots.create(
        name="ClosedOnly",
        source="SalesData",
        destination="B4",
        rows=["Region"],
        filters=[PivotItemFilter("Status", include=["Closed", "Closed"])],
        values=["Amount"],
    )
    dest = src + ".deduped-filter.xlsx"
    reopened = save_and_reopen(wb, dest, preserve=True)
    pivot = reopened["Summary"].pivots["ClosedOnly"]
    assert pivot.spec.filters[0].include == ("Closed",)
    assert reopened["Summary"]["C2"].value == "Closed"
    _cache, table, _parts = _parse_created(dest)
    assert table.pageFields[0].item >= 0


@pytest.mark.parametrize("aggregate,caption", [
    ("sum", "Sum of Amount"),
    ("count", "Count of Amount"),
    ("count_numbers", "Count of Amount"),
    ("average", "Average of Amount"),
    ("min", "Min of Amount"),
    ("max", "Max of Amount"),
])
def test_each_supported_aggregate(tmp_path, aggregate, caption):
    src, wb = _preserved(tmp_path, table="SalesData", name="%s.xlsx" % aggregate)
    values = [PivotMeasure("Amount", aggregate=aggregate)]
    if aggregate in ("count", "count_numbers") and aggregate == "count_numbers":
        values = [PivotMeasure("Amount", aggregate=aggregate, caption="N")]
        caption = "N"
    elif aggregate == "count":
        values = [PivotMeasure("Amount", aggregate=aggregate, caption="Count")]
        caption = "Count"
    handle = wb["Summary"].pivots.create(
        name="Agg",
        source="SalesData",
        destination="L2",
        rows=["Region"],
        values=values,
        row_grand_totals=False,
    )
    dest = src + ".agg.xlsx"
    reopened = save_and_reopen(wb, dest, preserve=True)
    pivot = reopened["Summary"].pivots["Agg"]
    assert pivot.spec.values[0].aggregate == aggregate
    assert caption in _grid(reopened["Summary"], pivot.output_range).values()
    _cache, table, _parts = _parse_created(dest)
    xml_name = {
        "sum": "sum",
        "count": "count",
        "count_numbers": "countNums",
        "average": "average",
        "min": "min",
        "max": "max",
    }[aggregate]
    assert table.dataFields[0].subtotal == xml_name
    assert handle.valid is True


def test_compact_outline_tabular_geometry(tmp_path):
    layouts = {}
    for layout in ("compact", "outline", "tabular"):
        src, wb = _preserved(tmp_path, table="SalesData", name="%s.xlsx" % layout)
        handle = wb["Summary"].pivots.create(
            name="ByLayout",
            source="SalesData",
            destination="B4",
            rows=["Region", "Product"],
            values=["Amount"],
            layout=layout,
            row_grand_totals=False,
        )
        dest = src + ".out.xlsx"
        reopened = save_and_reopen(wb, dest, preserve=True)
        pivot = reopened["Summary"].pivots["ByLayout"]
        layouts[layout] = pivot.spec.layout, pivot.output_range
        _cache, table, _parts = _parse_created(dest)
        assert pivot.spec.layout == layout
        if layout == "compact":
            assert table.compact is True
            assert table.location.firstDataCol == 1
        elif layout == "outline":
            assert table.outline is True
            assert table.compact is False
        else:
            assert table.compact is False
            assert table.outline is False
            assert table.location.firstDataCol == 2
        assert handle.output_range == pivot.output_range
    assert layouts["compact"][0] != layouts["tabular"][0]


def test_subtotals_and_grand_total_toggles(tmp_path):
    src, wb = _preserved(tmp_path, table="SalesData")
    handle = wb["Summary"].pivots.create(
        name="Totals",
        source="SalesData",
        destination="B4",
        rows=["Region", "Product"],
        values=["Amount"],
        subtotals=True,
        row_grand_totals=True,
        column_grand_totals=False,
    )
    dest = src + ".totals.xlsx"
    reopened = save_and_reopen(wb, dest, preserve=True)
    pivot = reopened["Summary"].pivots["Totals"]
    assert pivot.spec.subtotals is True
    assert pivot.spec.row_grand_totals is True
    grid = _grid(reopened["Summary"], pivot.output_range)
    assert "Grand Total" in grid.values()
    assert any(
        isinstance(value, str) and value.endswith(" Total")
        for value in grid.values() if value is not None)
    assert handle.valid is True


def test_captions_number_formats_and_typed_items(tmp_path):
    src, wb = _preserved(tmp_path, table="SalesData")
    wb["Summary"].pivots.create(
        name="Formats",
        source="SalesData",
        destination="B4",
        rows=["Flag"],
        columns=["Code"],
        values=[
            PivotMeasure(
                "Amount", aggregate="sum", caption="Revenue",
                number_format="$#,##0",
            ),
        ],
        row_grand_totals=False,
        column_grand_totals=False,
    )
    dest = src + ".fmt.xlsx"
    reopened = save_and_reopen(wb, dest, preserve=True)
    pivot = reopened["Summary"].pivots["Formats"]
    assert True in _grid(reopened["Summary"], pivot.output_range).values()
    assert "001" in _grid(reopened["Summary"], pivot.output_range).values()
    data_cell = None
    for row in reopened["Summary"].iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.number_format == "$#,##0":
                data_cell = cell
                break
    assert data_cell is not None
    _cache, table, _parts = _parse_created(dest)
    assert table.dataFields[0].name == "Revenue"
    assert table.dataFields[0].numFmtId is not None


@pytest.mark.parametrize("epoch", [WINDOWS_EPOCH, MAC_EPOCH])
def test_date_epochs(tmp_path, epoch):
    src, wb = _preserved(
        tmp_path, table="SalesData", epoch=epoch,
        name="epoch-%s.xlsx" % epoch.year)
    assert wb.epoch == epoch
    wb["Summary"].pivots.create(
        name="ByDay",
        source="SalesData",
        destination="B4",
        rows=["Day"],
        values=["Amount"],
        row_grand_totals=False,
    )
    dest = src + ".dates.xlsx"
    reopened = save_and_reopen(wb, dest, preserve=True)
    assert reopened.epoch == epoch
    pivot = reopened["Summary"].pivots["ByDay"]
    values = _grid(reopened["Summary"], pivot.output_range).values()
    assert any(isinstance(value, (date, datetime)) for value in values)


def test_table_not_at_a1_quoted_sheet_and_non_ascii(tmp_path):
    src, wb = _preserved(
        tmp_path, sheet="Raw Data", table="SalesOffset", origin="C5",
        name="offset.xlsx")
    handle = wb["Summary"].pivots.create(
        name="Offset",
        source="SalesOffset",
        destination="B4",
        rows=["Region"],
        values=["Amount"],
        row_grand_totals=False,
    )
    dest = src + ".offset.xlsx"
    reopened = save_and_reopen(wb, dest, preserve=True)
    pivot = reopened["Summary"].pivots["Offset"]
    assert "東" in _grid(reopened["Summary"], pivot.output_range).values()
    src2, wb2 = _preserved(
        tmp_path, sheet="Raw Data", origin="A1", name="quoted.xlsx")
    wb2["Summary"].pivots.create(
        name="Quoted",
        source="'Raw Data'!A1:I7",
        destination="B4",
        rows=["Region"],
        values=["Amount"],
        row_grand_totals=False,
    )
    dest2 = src2 + ".quoted.xlsx"
    reopened2 = save_and_reopen(wb2, dest2, preserve=True)
    assert reopened2["Summary"].pivots["Quoted"].source.sheet == "Raw Data"
    assert handle.valid is True


def test_first_seen_order_and_duplicate_source_rows(tmp_path):
    rows = [
        ["West", "A", 1],
        ["East", "A", 2],
        ["West", "A", 3],
    ]
    src, wb = _preserved(
        tmp_path, headers=["Region", "Product", "Amount"], rows=rows,
        table="Dupes")
    wb["Summary"].pivots.create(
        name="Order",
        source="Dupes",
        destination="E2",
        rows=["Region"],
        values=["Amount"],
        row_grand_totals=False,
    )
    dest = src + ".order.xlsx"
    reopened = save_and_reopen(wb, dest, preserve=True)
    grid = _grid(reopened["Summary"], reopened["Summary"].pivots["Order"].output_range)
    labels = [value for (row, _col), value in sorted(grid.items())
              if isinstance(value, str) and value in ("West", "East")]
    assert labels == ["West", "East"]


def test_libreoffice_loads_nested_create(tmp_path, lo):
    src, wb = _preserved(tmp_path, table="SalesData")
    wb["Summary"].pivots.create(
        name="SalesByRegion",
        source="SalesData",
        destination="B4",
        rows=["Region", "Product"],
        columns=["Quarter"],
        values=["Amount"],
    )
    dest = src + ".lo.xlsx"
    wb.save(dest)
    converted = lo.lo_convert(dest, fmt="xlsx")
    assert converted[:2] == b"PK"


def test_excel_sidecar_comparison_is_skipped_until_transcripts():
    expected = []
    if os.path.isdir(_PIVOTS):
        for name in os.listdir(_PIVOTS):
            if name.endswith(".json"):
                expected.append(name)
    pytest.skip("no approved Excel-authored creation sidecar")


def test_refusals_are_typed_and_atomic(tmp_path):
    src, wb = _preserved(tmp_path, table="SalesData")
    ws = wb["Summary"]
    before = {
        title: {
            coord: (cell.value, cell._style)
            for coord, cell in sheet._cells.items()
        }
        for title, sheet in ((s.title, s) for s in wb.worksheets)
    }
    ops = dict(wb._paper_ledger.pivot_operations)

    with pytest.raises(TypeError):
        ws.pivots.create(
            name="Bad", source="SalesData", destination="B4",
            rows=["Region"], values=["Amount"], grouping=True)
    with pytest.raises(TypeError):
        ws.pivots.create(
            name="Bad", source="SalesData", destination="B4",
            rows=["Region"], values=["Amount"], show_data_as="percent")
    with pytest.raises(ValueError):
        ws.pivots.create(
            name="Bad", source="SalesData", destination="B4",
            rows=["Region"], values=["Amount"], layout="calendar")
    with pytest.raises(UnsupportedStructureError) as style:
        ws.pivots.create(
            name="Bad", source="SalesData", destination="B4",
            rows=["Region"], values=["Amount"], style="NotAPivotStyle")
    assert style.value.kind == "unsupported-pivot-feature"
    with pytest.raises(BoundaryViolationError) as captions:
        ws.pivots.create(
            name="Bad", source="SalesData", destination="B4",
            rows=["Region"],
            values=[
                PivotMeasure("Amount", aggregate="sum", caption="Same"),
                PivotMeasure("Units", aggregate="sum", caption="Same"),
            ])
    assert captions.value.kind == "invalid-pivot-source"
    with pytest.raises(BoundaryViolationError) as repeated:
        ws.pivots.create(
            name="Bad", source="SalesData", destination="B4",
            rows=["Region"], columns=["Region"], values=["Amount"])
    assert repeated.value.kind == "invalid-pivot-source"
    with pytest.raises(BoundaryViolationError) as items:
        ws.pivots.create(
            name="Bad", source="SalesData", destination="B4",
            rows=[PivotAxisField("Region", items=["Missing"])],
            values=["Amount"])
    assert items.value.kind == "invalid-pivot-source"
    with pytest.raises(BoundaryViolationError) as unknown:
        ws.pivots.create(
            name="Bad", source="SalesData", destination="B4",
            rows=["Region"],
            filters=[PivotItemFilter("Status", include=["Never"])],
            values=["Amount"])
    assert unknown.value.kind == "invalid-pivot-source"
    with pytest.raises(BoundaryViolationError) as empty:
        ws.pivots.create(
            name="Bad", source="SalesData", destination="B4",
            rows=["Region"],
            filters=[PivotItemFilter(
                "Region", exclude=["East", "West", "東"])],
            values=["Amount"])
    assert empty.value.kind == "invalid-pivot-source"
    with pytest.raises(BoundaryViolationError) as mixed:
        ws.pivots.create(
            name="Bad", source="SalesData", destination="B4",
            rows=["Region"],
            values=[PivotMeasure("Status", aggregate="sum")])
    assert mixed.value.kind == "invalid-pivot-source"
    with pytest.raises(BoundaryViolationError) as overlap:
        wb["Data"].pivots.create(
            name="Bad", source="SalesData", destination="A1",
            rows=["Region"], values=["Amount"])
    assert overlap.value.kind in (
        "pivot-source-output-overlap", "pivot-output-collision")

    after = {
        title: {
            coord: (cell.value, cell._style)
            for coord, cell in sheet._cells.items()
        }
        for title, sheet in ((s.title, s) for s in wb.worksheets)
    }
    assert after == before
    assert dict(wb._paper_ledger.pivot_operations) == ops
    with open(src, "rb") as handle:
        original = handle.read()
    assert open(src, "rb").read() == original
