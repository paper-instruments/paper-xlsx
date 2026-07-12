import io
import os

import pytest

from openpyxl import Workbook, load_workbook
from openpyxl.errors import UnsupportedStructureError
from openpyxl.preserve import zipio


def _workbook_bytes(value):
    target = io.BytesIO()
    workbook = Workbook()
    workbook.active["A1"] = value
    workbook.save(target)
    return target.getvalue()


@pytest.mark.parametrize("receipt", [False, True])
def test_same_path_save_refuses_replacement_since_load(tmp_path, receipt):
    source = tmp_path / "source.xlsx"
    source.write_bytes(_workbook_bytes("original"))
    workbook = load_workbook(source, preserve=True)
    workbook.active["B2"] = "edited"

    replacement = tmp_path / "replacement.xlsx"
    intruder = _workbook_bytes("intruder")
    replacement.write_bytes(intruder)
    os.replace(replacement, source)

    with pytest.raises(UnsupportedStructureError) as refusal:
        workbook.save(source, receipt=receipt)

    assert refusal.value.kind == "destination-identity-changed"
    assert source.read_bytes() == intruder


def test_validation_race_preserves_racing_occupant(tmp_path):
    destination = tmp_path / "destination.xlsx"
    destination.write_bytes(b"original")

    def race(_staged):
        replacement = tmp_path / "replacement"
        replacement.write_bytes(b"intruder")
        os.replace(replacement, destination)

    with pytest.raises(UnsupportedStructureError) as refusal:
        zipio.deliver(b"candidate", destination, validator=race)

    assert refusal.value.kind == "destination-identity-changed"
    assert destination.read_bytes() == b"intruder"


@pytest.mark.skipif(os.name == "nt", reason="POSIX exchange hook")
def test_exchange_race_never_reports_success_for_an_intruder(
        tmp_path, monkeypatch):
    destination = tmp_path / "destination.xlsx"
    destination.write_bytes(b"original")
    real_exchange = zipio._posix_exchange

    def race(first, second):
        real_exchange(first, second)
        replacement = tmp_path / "replacement"
        replacement.write_bytes(b"intruder")
        os.replace(replacement, destination)

    monkeypatch.setattr(zipio, "_posix_exchange", race)
    with pytest.raises(UnsupportedStructureError) as refusal:
        zipio.deliver(b"candidate", destination)

    assert refusal.value.kind == "destination-identity-changed"
    assert destination.read_bytes() == b"intruder"


def test_save_as_refuses_when_loaded_source_changes(tmp_path):
    source = tmp_path / "source.xlsx"
    source.write_bytes(_workbook_bytes("original"))
    workbook = load_workbook(source, preserve=True)
    workbook.active["B2"] = "edited"
    intruder = _workbook_bytes("intruder")
    source.write_bytes(intruder)
    destination = tmp_path / "copy.xlsx"

    with pytest.raises(UnsupportedStructureError) as refusal:
        workbook.save(destination)

    assert refusal.value.kind == "destination-identity-changed"
    assert source.read_bytes() == intruder
    assert not destination.exists()


def test_receipt_save_rechecks_source_before_delivery(tmp_path, monkeypatch):
    source = tmp_path / "source.xlsx"
    source.write_bytes(_workbook_bytes("original"))
    workbook = load_workbook(source, preserve=True)
    workbook.active["B2"] = "edited"
    destination = tmp_path / "copy.xlsx"
    from openpyxl.preserve import receipts

    real_receipt = receipts.receipt

    def replace_source(*args, **kwargs):
        result = real_receipt(*args, **kwargs)
        source.write_bytes(b"concurrent replacement")
        return result

    monkeypatch.setattr(receipts, "receipt", replace_source)
    with pytest.raises(UnsupportedStructureError) as refusal:
        workbook.save(destination, receipt=True)

    assert refusal.value.kind == "destination-identity-changed"
    assert not destination.exists()


@pytest.mark.skipif(os.name == "nt", reason="POSIX permission semantics")
def test_new_destination_uses_normal_creation_permissions(tmp_path):
    destination = tmp_path / "new.xlsx"
    old_umask = os.umask(0o027)
    try:
        zipio.deliver(b"candidate", destination)
    finally:
        os.umask(old_umask)

    assert destination.stat().st_mode & 0o777 == 0o640


def test_repeated_same_path_saves_advance_custody(tmp_path):
    source = tmp_path / "source.xlsx"
    source.write_bytes(_workbook_bytes("original"))
    workbook = load_workbook(source, preserve=True)
    loaded = workbook._paper_source_identity

    workbook.active["B2"] = "first"
    workbook.save(source)
    first = workbook._paper_source_identity
    assert first != loaded

    workbook.active["C3"] = "second"
    workbook.save(source)
    reopened = load_workbook(source)
    assert reopened.active["B2"].value == "first"
    assert reopened.active["C3"].value == "second"
    assert workbook._paper_source_identity != first


def test_source_handle_save_works_without_ledger_crosscheck(
        tmp_path, monkeypatch):
    source = tmp_path / "source.xlsx"
    source.write_bytes(_workbook_bytes("original"))
    workbook = load_workbook(source, preserve=True)
    workbook.active["B2"] = "edited"
    monkeypatch.setenv("PAPER_LEDGER_CROSSCHECK", "0")

    with open(source, "r+b") as target:
        workbook.save(target)
        assert not target.closed

    assert load_workbook(source).active["B2"].value == "edited"


@pytest.mark.skipif(os.name == "nt", reason="POSIX permission semantics")
def test_unwritable_destination_refuses_without_replacement(tmp_path):
    destination = tmp_path / "source.xlsx"
    original = _workbook_bytes("original")
    destination.write_bytes(original)
    workbook = load_workbook(destination, preserve=True)
    workbook.active["B2"] = "edited"
    destination.chmod(0o444)
    try:
        with pytest.raises(UnsupportedStructureError) as refusal:
            workbook.save(destination)
        assert refusal.value.kind == "destination-not-writable"
        assert destination.read_bytes() == original
    finally:
        destination.chmod(0o644)
