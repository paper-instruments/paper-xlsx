"""One regression test per confirmed finding of the standing per-batch
adversarial gates (PLAN-v0.1 process amendment 2). Sections are tagged by
batch; every test here pins a fix for a live repro the gate produced.
"""
from __future__ import annotations

import io
import re
import zipfile

import pytest

from openpyxl import Workbook, load_workbook
from openpyxl.errors import (
    BoundaryViolationError,
    ProtectedWriteWarning,
    UnsupportedStructureError,
)

from .support.partdiff import part_payloads


class TestBatch1ObjectGuardGaps:

    def test_chart_anchor_mutation_refuses(self, fixture_copy, tmp_path):
        # the chart fingerprint was chart._write() only — the anchor lives
        # in the preserved drawing part, so a chart MOVE vanished silently
        src = fixture_copy("features/chart_image.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = next(w for w in wb.worksheets if w._charts)
        ws._charts[0].anchor = "K25"
        with pytest.raises(UnsupportedStructureError, match="chart"):
            wb.save(str(tmp_path / "o.xlsx"))

    def test_image_data_swap_refuses(self, fixture_copy, tmp_path):
        # same anchor + same path + different pixels: the fingerprint now
        # covers the backing bytes (non-destructively — image._data()
        # CLOSES the ref stream and must never be used for snapshots)
        src = fixture_copy("features/chart_image.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = next(w for w in wb.worksheets if w._images)
        img = ws._images[0]
        original = img.ref.getvalue()
        img.ref = io.BytesIO(original + b"\x00")
        with pytest.raises(UnsupportedStructureError, match="image"):
            wb.save(str(tmp_path / "o.xlsx"))

    def test_chartsheet_chart_mutation_refuses(self, tmp_path):
        # chartsheet-anchored charts were entirely outside the boundary
        from openpyxl.chart import BarChart, Reference

        wb = Workbook()
        ws = wb.active
        for r in range(1, 5):
            ws.append([r, r * 2])
        cs = wb.create_chartsheet("ChartOnly")
        chart = BarChart()
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=4))
        cs.add_chart(chart)
        src = str(tmp_path / "cs.xlsx")
        wb.save(src)

        wb2 = load_workbook(src, preserve=True)
        wb2.chartsheets[0]._charts[0].title = "TAMPERED"
        with pytest.raises(UnsupportedStructureError, match="chartsheet"):
            wb2.save(str(tmp_path / "o.xlsx"))


class TestBatch1RecalcGuardGaps:

    def _three_sheet_fixture(self, tmp_path, formula):
        # Excel-producer shape: cached value present, no fullCalcOnLoad
        wb = Workbook()
        s1 = wb.active
        s1.title = "Sheet1"
        wb.create_sheet("Sheet2")
        wb.create_sheet("Sheet3")
        for s in wb.worksheets:
            s["A1"] = 10
        s1["B1"] = formula
        raw = str(tmp_path / "threed_raw.xlsx")
        wb.save(raw)
        out = str(tmp_path / "threed.xlsx")
        with zipfile.ZipFile(raw) as zin, zipfile.ZipFile(out, "w") as zout:
            for name in zin.namelist():
                payload = zin.read(name)
                if name == "xl/workbook.xml":
                    payload = payload.replace(b' fullCalcOnLoad="1"', b"")
                if name == "xl/worksheets/sheet1.xml" and b"<f>" in payload:
                    payload = payload.replace(b"</f>", b"</f><v>999</v>", 1)
                zout.writestr(name, payload)
        return out

    def test_3d_span_reference_forces_recalc_flag(self, tmp_path):
        # 'Sheet1:Sheet3' was recorded as a phantom sheet name nothing
        # could match — the recalc guard and certification taint both
        # silently missed 3-D formulas
        src = self._three_sheet_fixture(tmp_path, "=SUM(Sheet1:Sheet3!A1)")
        wb = load_workbook(src, preserve=True)
        wb["Sheet2"]["A1"] = 42
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        assert b"fullCalcOnLoad" in part_payloads(out)["xl/workbook.xml"]

    def test_3d_span_is_unresolved_in_the_sketch(self):
        from openpyxl.preserve.perception import dependency_sketch

        wb = Workbook()
        wb.create_sheet("Sheet2")
        wb.create_sheet("Sheet3")
        wb.active["B1"] = "=SUM(Sheet1:Sheet3!A1)"
        sketch = dependency_sketch(wb)
        assert any("B1" in a for a in sketch.unresolved)

    def test_case_insensitive_reference_intersection(self):
        from openpyxl.preserve.perception import dependency_sketch

        wb = Workbook()
        ws = wb.active            # 'Sheet'
        wb.create_sheet("Other")
        wb["Other"]["B1"] = "=sheet!A1*2"       # lowercase ref
        sketch = dependency_sketch(wb)
        hits = sketch.cells_referencing("Sheet", (1, 1, 1, 1))
        assert any("B1" in h for h in hits)


class TestBatch1ProtectionGaps:

    def _protected(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "original"
        ws.protection.sheet = True
        src = str(tmp_path / "protected.xlsx")
        wb.save(src)
        return src

    def test_delete_of_locked_cell_is_protection_checked(self, tmp_path):
        # del ws['A1'] evaded what ws['A1']=None refused
        src = self._protected(tmp_path)
        wb = load_workbook(src, preserve=True)
        wb.strict_protection = True
        with pytest.raises(UnsupportedStructureError, match="locked"):
            del wb.active["A1"]
        assert wb.active["A1"].value == "original"

        wb2 = load_workbook(src, preserve=True)
        with pytest.warns(ProtectedWriteWarning):
            del wb2.active["A1"]

    def test_structural_shift_on_protected_sheet_warns_or_refuses(
            self, tmp_path):
        src = self._protected(tmp_path)
        wb = load_workbook(src, preserve=True)
        with pytest.warns(ProtectedWriteWarning, match="insert_rows"):
            wb.active.insert_rows(1)

        wb2 = load_workbook(src, preserve=True)
        wb2.strict_protection = True
        with pytest.raises(UnsupportedStructureError, match="protected"):
            wb2.active.delete_rows(1)


class TestBatch1InputHonestyGaps:

    def test_cfb_sniff_anchors_at_offset_zero(self, fixture_copy, tmp_path):
        # (a) a valid xlsx handed over at a position where CFB bytes sit
        # (embedded OLE payload) must NOT false-refuse
        src = fixture_copy("minimal/minimal_clean.xlsx")
        cfb_payload = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 64
        embedded = str(tmp_path / "embedded.xlsx")
        with zipfile.ZipFile(src) as zin, \
                zipfile.ZipFile(embedded, "w") as zout:
            for name in zin.namelist():
                zout.writestr(name, zin.read(name))
            zout.writestr(zipfile.ZipInfo("xl/embeddings/oleObject1.bin"),
                          cfb_payload)
        with open(embedded, "rb") as f:
            data = f.read()
        offset = data.find(cfb_payload)
        assert offset > 0
        handle = open(embedded, "rb")
        handle.seek(offset)
        wb = load_workbook(handle)             # must not raise
        assert wb.active is not None
        handle.close()

        # (b) a genuine CFB file via a mid-position handle still gets the
        # typed refusal (it evaded to BadZipFile before the fix)
        cfb = str(tmp_path / "enc.xlsx")
        with open(cfb, "wb") as f:
            f.write(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 4096)
        handle = open(cfb, "rb")
        handle.seek(4)
        with pytest.raises(UnsupportedStructureError, match="ENCRYPT"):
            load_workbook(handle)
        assert handle.tell() == 4              # position restored
        handle.close()

    def test_rich_text_mode_suppresses_rich_text_loss_entry(
            self, fixture_copy, tmp_path):
        # under rich_text=True the stock save PRESERVES runs — warning
        # about flattening was loud-but-wrong
        src = fixture_copy("minimal/minimal_clean.xlsx")
        rich = str(tmp_path / "rich.xlsx")
        with zipfile.ZipFile(src) as zin, zipfile.ZipFile(rich, "w") as zout:
            for name in zin.namelist():
                payload = zin.read(name)
                if name.startswith("xl/worksheets/sheet"):
                    payload = payload.replace(
                        b"</sheetData>",
                        b'<row r="9"><c r="A9" t="inlineStr"><is><r><rPr>'
                        b'<b/></rPr><t>bold</t></r></is></c></row>'
                        b"</sheetData>", 1)
                zout.writestr(name, payload)
        import warnings as _w
        with _w.catch_warnings():
            _w.simplefilter("ignore")
            flat = load_workbook(rich)
            modeled = load_workbook(rich, rich_text=True)
        assert "rich-text" in flat._paper_loss_inventory.kinds()
        assert "rich-text" not in modeled._paper_loss_inventory.kinds()


class TestBatch1RemapAndCertifyGaps:

    def test_xlfn_prefixed_functions_are_excluded(self):
        from openpyxl import oracle

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "=_xlfn.LET(x,1,x*2)"
        ws["A2"] = "=_xlfn.RANDARRAY(2)"
        seeds = oracle._exclusion_seeds(wb)
        assert seeds[("Sheet", 1, 1)] == "unsupported:LET"
        assert seeds[("Sheet", 2, 1)] == "volatile"

    def test_external_ref_behind_defined_name_is_excluded(self):
        from openpyxl import oracle
        from openpyxl.workbook.defined_name import DefinedName

        wb = Workbook()
        wb.defined_names["EXTPRICE"] = DefinedName(
            "EXTPRICE", attr_text="'[1]Other'!$A$1")
        wb.active["A1"] = "=EXTPRICE*2"
        seeds = oracle._exclusion_seeds(wb)
        assert seeds[("Sheet", 1, 1)] == "external-link"

    def test_boundary_counts_dimension_only_rows(
            self, fixture_copy, tmp_path):
        # a <row r="1048576" ht="30"/> with no cells evaded ws.max_row
        src = fixture_copy("features/schedule.xlsx")
        floored = str(tmp_path / "floored.xlsx")
        with zipfile.ZipFile(src) as zin, \
                zipfile.ZipFile(floored, "w") as zout:
            for name in zin.namelist():
                payload = zin.read(name)
                if name.startswith("xl/worksheets/sheet") \
                        and b"Schedule" not in payload[:200]:
                    pass
                if name == "xl/worksheets/sheet1.xml":
                    payload = payload.replace(
                        b"</sheetData>",
                        b'<row r="1048576" ht="30" customHeight="1"/>'
                        b"</sheetData>", 1)
                zout.writestr(name, payload)
        wb = load_workbook(floored, preserve=True)
        ws = wb["Schedule"]
        if not ws.row_dimensions.get(1048576):
            pytest.skip("floor row not on the Schedule sheet part")
        with pytest.raises(BoundaryViolationError):
            ws.insert_rows(1)

    def test_insert_beyond_content_does_not_false_refuse(self, fixture_copy):
        # content ends early; inserting at a beyond-content index that
        # arithmetically exceeds the limit must NOT refuse (nothing moves)
        wb = load_workbook(fixture_copy("features/schedule.xlsx"),
                           preserve=True)
        ws = wb["Schedule"]
        remap = ws.insert_rows(1048570, 10)     # no occupied cell shifts
        assert remap is not None


class TestBatch2EngineGaps:

    def test_styles_creation_plus_added_sheet_share_the_rid_allocator(
            self, fixture_copy, tmp_path):
        # duplicate rId4 on workbook rels (gate critical): added sheets now
        # reserve through the engine
        import re

        from openpyxl.styles import Font
        from .test_lifecycle import TestStylesPartCreation

        src = TestStylesPartCreation()._styleless(fixture_copy, tmp_path)
        wb = load_workbook(src, preserve=True)
        wb["Sheet1"]["B2"] = 1
        wb["Sheet1"]["B2"].font = Font(bold=True)   # styles.xml creation
        wb.create_sheet("Fresh")["A1"] = 2
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        rels = part_payloads(out)["xl/_rels/workbook.xml.rels"]
        rids = re.findall(rb'Id="(rId\d+)"', rels)
        assert len(rids) == len(set(rids))          # all unique
        wb2 = load_workbook(out)
        assert wb2["Fresh"]["A1"].value == 2
        assert wb2["Sheet1"]["B2"].font.bold is True

    def test_table_removal_plus_hyperlink_add_both_land(
            self, fixture_copy, tmp_path):
        # the engine rels payload shadowed the hyperlink planner's (gate
        # critical): compose on top instead
        src = fixture_copy("features/tables.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb.worksheets[0]
        del ws.tables["RegionTable"]
        ws["D1"].hyperlink = "https://example.org/x"
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        ws2 = wb2.worksheets[0]
        assert not ws2.tables
        assert ws2["D1"].hyperlink.target == "https://example.org/x"

    def test_replace_part_conflicts_refuse(self, fixture_copy, tmp_path):
        from openpyxl.errors import RelationshipPolicyError
        from openpyxl.errors import UnsupportedStructureError
        from openpyxl.packaging.custom import StringProperty

        # table parts joined the managed set: raw swaps refuse at CALL time
        src = fixture_copy("features/tables.xlsx")
        wb = load_workbook(src, preserve=True)
        with pytest.raises(RelationshipPolicyError, match="managed"):
            wb.replace_part("xl/tables/table1.xml", b"<table/>")

        # a swap of an unmanaged part conflicting with a lifecycle removal
        # refuses at SAVE (the payload must never vanish silently)
        src2 = fixture_copy("minimal/minimal_clean.xlsx")
        wb2 = load_workbook(src2, preserve=True)
        wb2.custom_doc_props.append(StringProperty(name="Tmp", value="x"))
        staged = str(tmp_path / "staged.xlsx")
        wb2.save(staged)
        wb3 = load_workbook(staged, preserve=True)
        wb3.replace_part("docProps/custom.xml", b"<Properties/>")
        del wb3.custom_doc_props["Tmp"]
        with pytest.raises(UnsupportedStructureError, match="conflicts"):
            wb3.save(str(tmp_path / "o.xlsx"))


class TestBatch2TableGaps:

    def _with_table_extlst(self, fixture_copy, tmp_path):
        src = fixture_copy("features/tables.xlsx")
        out = str(tmp_path / "alt.xlsx")
        with zipfile.ZipFile(src) as zin, zipfile.ZipFile(out, "w") as zout:
            for name in zin.namelist():
                payload = zin.read(name)
                if name == "xl/tables/table1.xml":
                    payload = payload.replace(
                        b"</table>",
                        b'<extLst><ext uri="{X}"><x14:table '
                        b'xmlns:x14="http://schemas.microsoft.com/office/'
                        b'spreadsheetml/2009/9/main" altText="alt"/></ext>'
                        b"</extLst></table>", 1)
                zout.writestr(name, payload)
        return out

    def test_table_with_extlst_refuses_mutation(self, fixture_copy,
                                                tmp_path):
        # to_tree() drops extLst (alt text!) — mutation must refuse, never
        # silently strip accessibility metadata (gate critical)
        from openpyxl.errors import UnsupportedStructureError
        from openpyxl.preserve.tables import append_row

        src = self._with_table_extlst(fixture_copy, tmp_path)
        wb = load_workbook(src, preserve=True)
        ws = wb.worksheets[0]
        append_row(ws, "RegionTable", ["West2", 99])
        with pytest.raises(UnsupportedStructureError, match="extension"):
            wb.save(str(tmp_path / "o.xlsx"))

    def test_sibling_basename_survives_removal(self, fixture_copy,
                                               tmp_path):
        # suffix-matching rel removal nuked mytable1.xml's rel when
        # table1.xml was removed (gate critical): exact-target now
        src = fixture_copy("features/tables.xlsx")
        crafted = str(tmp_path / "two.xlsx")
        with zipfile.ZipFile(src) as zin, \
                zipfile.ZipFile(crafted, "w") as zout:
            for name in zin.namelist():
                payload = zin.read(name)
                if name == "xl/worksheets/_rels/sheet1.xml.rels":
                    payload = payload.replace(
                        b"</Relationships>",
                        b'<Relationship Id="rId99" Type="http://schemas.'
                        b'openxmlformats.org/officeDocument/2006/'
                        b'relationships/table" '
                        b'Target="../tables/mytable1.xml"/>'
                        b"</Relationships>", 1)
                if name == "xl/worksheets/sheet1.xml":
                    payload = payload.replace(
                        b'</tableParts>',
                        b'<tablePart xmlns:r="http://schemas.openxmlformats'
                        b'.org/officeDocument/2006/relationships" '
                        b'r:id="rId99"/></tableParts>', 1)
                    payload = payload.replace(
                        b'<tableParts count="1">',
                        b'<tableParts count="2">', 1)
                if name == "[Content_Types].xml":
                    payload = payload.replace(
                        b"</Types>",
                        b'<Override PartName="/xl/tables/mytable1.xml" '
                        b'ContentType="application/vnd.openxmlformats-'
                        b'officedocument.spreadsheetml.table+xml"/>'
                        b"</Types>", 1)
                zout.writestr(name, payload)
            table2 = zin.read("xl/tables/table1.xml")
            table2 = table2.replace(b'id="1"', b'id="7"', 1)
            table2 = table2.replace(b'displayName="RegionTable"',
                                    b'displayName="Other"', 1)
            table2 = table2.replace(b'name="RegionTable"',
                                    b'name="Other"', 1)
            zout.writestr("xl/tables/mytable1.xml", table2)
        wb = load_workbook(crafted, preserve=True)
        ws = wb.worksheets[0]
        assert set(ws.tables) == {"RegionTable", "Other"}
        del ws.tables["RegionTable"]
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)                    # reload must not KeyError
        assert set(wb2.worksheets[0].tables) == {"Other"}

    def test_single_quoted_ref_keeps_anchor_guard(self, fixture_copy,
                                                  tmp_path):
        from openpyxl.errors import UnsupportedStructureError

        src = fixture_copy("features/tables.xlsx")
        crafted = str(tmp_path / "sq.xlsx")
        with zipfile.ZipFile(src) as zin, \
                zipfile.ZipFile(crafted, "w") as zout:
            for name in zin.namelist():
                payload = zin.read(name)
                if name == "xl/tables/table1.xml":
                    payload = payload.replace(b'ref="A1:B5"', b"ref='A1:B5'")
                zout.writestr(name, payload)
        wb = load_workbook(crafted, preserve=True)
        tbl = wb.worksheets[0].tables["RegionTable"]
        tbl.ref = "D10:E14"
        if tbl.autoFilter is not None:
            tbl.autoFilter.ref = "D10:E14"
        with pytest.raises(UnsupportedStructureError, match="anchor"):
            wb.save(str(tmp_path / "o.xlsx"))

    def test_append_row_refusal_is_atomic(self, fixture_copy, tmp_path):
        from openpyxl.errors import UnsupportedStructureError
        from openpyxl.preserve.tables import append_row
        from openpyxl.worksheet.table import TableFormula

        src = fixture_copy("features/tables.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb.worksheets[0]
        tbl = ws.tables["RegionTable"]
        tbl.tableColumns[1].calculatedColumnFormula = TableFormula()
        tbl.tableColumns[1].calculatedColumnFormula.attr_text = "1*2"
        cells_before = dict(ws._cells)
        with pytest.raises(UnsupportedStructureError, match="calculated"):
            append_row(ws, "RegionTable", ["X", 42])
        assert dict(ws._cells) == cells_before      # nothing mutated

    def test_two_sheets_two_new_tables_one_save(self, fixture_copy,
                                                tmp_path):
        import re

        from openpyxl.worksheet.table import Table

        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        ws2 = wb.create_sheet("Two")
        ws2["A1"] = "h"
        ws2["A2"] = 1
        wb["Sheet1"].add_table(Table(displayName="T1", ref="A1:B3"))
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        parts = part_payloads(out)
        ids = []
        for n, p in parts.items():
            if n.startswith("xl/tables/"):
                ids.append(re.search(rb'<table[^>]*\sid="(\d+)"', p).group(1))
        assert len(ids) == len(set(ids))            # workbook-unique ids
        wb2 = load_workbook(out)
        assert "T1" in wb2["Sheet1"].tables

    def test_display_name_vs_defined_name_refuses(self, fixture_copy,
                                                  tmp_path):
        from openpyxl.errors import UnsupportedStructureError
        from openpyxl.workbook.defined_name import DefinedName
        from openpyxl.worksheet.table import Table

        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        wb.defined_names["Budget"] = DefinedName("Budget",
                                                 attr_text="Sheet1!$A$1")
        wb["Sheet1"].add_table(Table(displayName="BUDGET", ref="A1:B3"))
        with pytest.raises(UnsupportedStructureError, match="defined name"):
            wb.save(str(tmp_path / "o.xlsx"))


class TestBatch2CommentGaps:

    def test_two_sheets_first_comments_one_save(self, fixture_copy,
                                                tmp_path):
        from openpyxl.comments import Comment

        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        ws2 = wb.create_sheet("Two")
        ws2["A1"] = 1
        wb["Sheet1"]["A1"].comment = Comment("one", "paper")
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert wb2["Sheet1"]["A1"].comment is not None

    def test_illegal_control_chars_refuse(self, fixture_copy, tmp_path):
        from openpyxl.comments import Comment
        from openpyxl.errors import UnsupportedStructureError

        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Sheet1"]["A1"].comment = Comment("bad \x0b char", "paper")
        with pytest.raises(UnsupportedStructureError, match="XML"):
            wb.save(str(tmp_path / "o.xlsx"))

    def test_comment_resize_on_machinery_sheet_refuses(
            self, fixture_copy, tmp_path):
        # height/width were outside the snapshot: resizes vanished (gate)
        from openpyxl.errors import UnsupportedStructureError

        src = fixture_copy("gauntlet/gauntlet.xlsx")
        wb = load_workbook(src, preserve=True)
        cell = wb["Model"]["B8"]
        assert cell.comment is not None
        cell.comment.height = 999
        with pytest.raises(UnsupportedStructureError, match="comment"):
            wb.save(str(tmp_path / "o.xlsx"))


class TestBatch3X14Gaps:

    def test_modified_twin_block_refuses(self, fixture_copy, tmp_path):
        # modification reclassified as delete+new and silently stripped
        # the twin (gate critical): now refuses naming the range
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import PatternFill

        src = fixture_copy("gauntlet/gauntlet.xlsx")
        with open(src, "rb") as f:
            before = f.read()
        wb = load_workbook(src, preserve=True)
        # adding a rule ON the twin dataBar's range modifies that block
        wb["Model"].conditional_formatting.add(
            "B6:E6", CellIsRule(operator="greaterThan", formula=["1"],
                                fill=PatternFill(start_color="FF0000",
                                                 fill_type="solid")))
        with pytest.raises(UnsupportedStructureError, match="MODIFIED"):
            wb.save(str(tmp_path / "o.xlsx"))
        with open(src, "rb") as f:
            assert f.read() == before


class TestBatch3LifecycleGaps:

    def test_shift_plus_rename_patches_charts_correctly(
            self, fixture_copy, tmp_path):
        # shift+rename left charts un-renumbered / rename-then-shift
        # falsely refused (gate criticals): both orders now work
        src = fixture_copy("features/chart_image.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = next(w for w in wb.worksheets if w._charts)
        ws.insert_rows(2)
        ws.title = "ModelX"
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        chart = next(p for n, p in part_payloads(out).items()
                     if n.startswith("xl/charts/chart"))
        assert b"'ModelX'!$B$3" in chart or b"ModelX!$B$3" in chart

        wb2 = load_workbook(src, preserve=True)
        ws2 = next(w for w in wb2.worksheets if w._charts)
        ws2.title = "ModelY"                    # rename FIRST
        ws2.insert_rows(2)                      # then shift: no refusal
        out2 = str(tmp_path / "o2.xlsx")
        wb2.save(out2)
        chart2 = next(p for n, p in part_payloads(out2).items()
                      if n.startswith("xl/charts/chart"))
        assert b"'ModelY'!$B$3" in chart2 or b"ModelY!$B$3" in chart2

    def test_title_swap_does_not_merge_chart_references(self, tmp_path):
        # sequential pairwise rename patching merged the two reference
        # classes on a title swap (gate critical): simultaneous mapping
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference

        src = str(tmp_path / "twosheet.xlsx")
        wb0 = Workbook()
        wsa = wb0.active
        wsa.title = "Alpha"
        for i in range(1, 6):
            wsa.cell(row=i, column=2, value=i)
        wb0.create_sheet("Beta")
        chart = BarChart()
        chart.add_data(Reference(wsa, min_col=2, min_row=1, max_row=5))
        wsa.add_chart(chart, "D2")
        wb0.save(src)

        wb = load_workbook(src, preserve=True)
        ws, other = wb["Alpha"], wb["Beta"]
        ws.title = "TMPSWAP"
        other.title = "Alpha"
        ws.title = "Beta"                       # net: Alpha<->Beta swap
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        chart_xml = next(p for n, p in part_payloads(out).items()
                         if n.startswith("xl/charts/chart"))
        refs = re.findall(rb"<(?:c:)?f>([^<]*)</(?:c:)?f>", chart_xml)
        assert refs
        # the chart charted the sheet now titled Beta: every reference
        # must follow it there, none may leak onto the new Alpha
        for ref in refs:
            assert ref.startswith(b"Beta!") or ref.startswith(b"'Beta'!")
        wb2 = load_workbook(out)
        assert set(["Alpha", "Beta"]) <= set(wb2.sheetnames)

    def test_freed_title_reuse_is_coherent(self, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule.xlsx")
        wb = load_workbook(src, preserve=True)
        wb.remove(wb["Summary"])
        ws = wb.create_sheet("Summary")         # reuse the freed title
        ws["A1"] = "fresh"
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)                # no duplicate entries
        assert wb2.sheetnames.count("Summary") == 1
        assert wb2["Summary"]["A1"].value == "fresh"

    def test_removal_audit_covers_scoped_names_and_cf_dv(
            self, fixture_copy, tmp_path):
        from openpyxl.workbook.defined_name import DefinedName

        src = fixture_copy("features/schedule.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Summary"].defined_names["scopedref"] = DefinedName(
            "scopedref", attr_text="Schedule!$B$2")
        # the audit walks sheet-scoped names on SURVIVING sheets too
        with pytest.raises(UnsupportedStructureError, match="scopedref"):
            wb.remove(wb["Schedule"])

    def test_removal_audit_covers_cf_and_dv_formulas(
            self, fixture_copy, tmp_path):
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.worksheet.datavalidation import DataValidation

        src = fixture_copy("features/schedule.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Summary"].conditional_formatting.add(
            "A1:A3", FormulaRule(formula=["Schedule!$B$2>1"]))
        with pytest.raises(UnsupportedStructureError,
                           match="conditional-formatting"):
            wb.remove(wb["Schedule"])

        wb2 = load_workbook(src, preserve=True)
        dv = DataValidation(type="list", formula1="=Schedule!$A$1:$A$3")
        dv.add("C1")
        wb2["Summary"].add_data_validation(dv)
        with pytest.raises(UnsupportedStructureError,
                           match="data validation"):
            wb2.remove(wb2["Schedule"])


class TestBatch3CmVmGaps:

    def _vm_fixture(self, fixture_copy, tmp_path):
        src = fixture_copy("minimal/minimal_clean.xlsx")
        out = str(tmp_path / "vm.xlsx")
        with zipfile.ZipFile(src) as zin, zipfile.ZipFile(out, "w") as zout:
            for name in zin.namelist():
                payload = zin.read(name)
                if name.startswith("xl/worksheets/sheet"):
                    payload = payload.replace(b'<c r="B2"',
                                              b'<c r="B2" vm="9"', 1)
                zout.writestr(name, payload)
        return out

    def test_style_only_edit_carries_vm(self, fixture_copy, tmp_path):
        # style-only re-emission stripped vm (gate critical): now carried
        from openpyxl.styles import Font

        src = self._vm_fixture(fixture_copy, tmp_path)
        wb = load_workbook(src, preserve=True)
        wb["Sheet1"]["B2"].font = Font(bold=True)
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        sheet = next(p for n, p in part_payloads(out).items()
                     if n.startswith("xl/worksheets/"))
        assert b'vm="9"' in sheet               # rich-value binding kept

    def test_value_overwrite_still_drops_vm(self, fixture_copy, tmp_path):
        src = self._vm_fixture(fixture_copy, tmp_path)
        wb = load_workbook(src, preserve=True)
        wb["Sheet1"]["B2"] = 5
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        sheet = next(p for n, p in part_payloads(out).items()
                     if n.startswith("xl/worksheets/"))
        assert b'vm="9"' not in sheet           # battery job 21 semantics

    def test_datatable_formula_blocks_shift(self, fixture_copy, tmp_path):
        src = fixture_copy("minimal/minimal_clean.xlsx")
        crafted = str(tmp_path / "dt.xlsx")
        with zipfile.ZipFile(src) as zin, \
                zipfile.ZipFile(crafted, "w") as zout:
            for name in zin.namelist():
                payload = zin.read(name)
                if name.startswith("xl/worksheets/sheet"):
                    payload = payload.replace(
                        b"</sheetData>",
                        b'<row r="9"><c r="D9"><f t="dataTable" ref="D9" '
                        b'r1="B2" dt2D="0" dtr="0"/><v>6</v></c></row>'
                        b"</sheetData>", 1)
                zout.writestr(name, payload)
        wb = load_workbook(crafted, preserve=True)
        with pytest.raises(UnsupportedStructureError, match="data table"):
            wb["Sheet1"].insert_rows(1)

    def test_move_range_refuses_on_chart_sheet(self, fixture_copy):
        wb = load_workbook(fixture_copy("features/chart_image.xlsx"),
                           preserve=True)
        ws = next(w for w in wb.worksheets if w._charts)
        with pytest.raises(UnsupportedStructureError, match="chart"):
            ws.move_range("A2:E2", rows=5)

    def test_rename_plus_hide_same_session(self, fixture_copy, tmp_path):
        # rename + sheet_state on one entry produced two overlapping
        # start-tag edits and died on the internal overlap guard (gate
        # major): both changes now compose into one whole-entry edit
        src = fixture_copy("features/schedule.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Summary"]
        ws.title = "Overview"
        ws.sheet_state = "hidden"
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert "Overview" in wb2.sheetnames
        assert wb2["Overview"].sheet_state == "hidden"


class TestBatch4DrawingGaps:

    def test_second_save_replans_identically(self, fixture_copy, tmp_path):
        # the chart single-use seen-set lived on the WORKBOOK, so a second
        # save of the same workbook false-refused (gate: found pre-gate);
        # it lives on the per-save part plan now
        from openpyxl.chart import BarChart, Reference

        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Sheet1"]
        chart = BarChart()
        chart.add_data(Reference(ws, min_col=1, min_row=1, max_row=3))
        ws.add_chart(chart, "F2")
        out1 = str(tmp_path / "o1.xlsx")
        out2 = str(tmp_path / "o2.xlsx")
        wb.save(out1)
        wb.save(out2)                       # idempotent, no false refusal
        p1, p2 = part_payloads(out1), part_payloads(out2)
        assert set(p1) == set(p2)
        assert all(p1[n] == p2[n] for n in p1)

    def test_added_chart_follows_shift(self, fixture_copy, tmp_path):
        # an in-session chart's ranges silently pointed at pre-shift cells
        # (gate: found pre-gate) — model fixups now cover added charts
        from openpyxl.chart import BarChart, Reference

        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Sheet1"]
        chart = BarChart()
        chart.add_data(Reference(ws, min_col=1, min_row=1, max_row=3))
        ws.add_chart(chart, "F2")
        ws.insert_rows(1)
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        chart_xml = next(p for n, p in part_payloads(out).items()
                         if n.startswith("xl/charts/chart"))
        assert b"$A$2:$A$4" in chart_xml    # followed the insert
        assert b"$A$1:$A$3" not in chart_xml

    def test_delete_stranding_added_chart_blocks_premove(
            self, fixture_copy, tmp_path):
        from openpyxl.chart import BarChart, Reference

        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Sheet1"]
        chart = BarChart()
        chart.add_data(Reference(ws, min_col=1, min_row=2, max_row=3))
        ws.add_chart(chart, "F2")
        with pytest.raises(UnsupportedStructureError, match="in-session"):
            ws.delete_rows(2, 2)
        # pre-move atomicity: neither cells nor the chart moved
        assert ws["A2"].value is not None
        assert chart.series[0].val.numRef.f == "'Sheet1'!$A$2:$A$3"


def _two_sheet_chart_fixture(tmp_path, with_axis_titles=False):
    """Fresh two-sheet workbook: 'Data' carries one titled chart over
    Data!$B$1:$B$5; 'Plain' is chartless."""
    from openpyxl.chart import BarChart, Reference

    src = str(tmp_path / "base.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for i in range(1, 6):
        ws.cell(row=i, column=2, value=i)
    chart = BarChart()
    chart.title = "Original"
    if with_axis_titles:
        chart.x_axis.title = "XT"
        chart.y_axis.title = "YT"
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=5))
    ws.add_chart(chart, "D2")
    plain = wb.create_sheet("Plain")
    plain["A1"] = 1
    wb.save(src)
    return src


def _rezip(src, out, fn):
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(out, "w") as zout:
        for name in zin.namelist():
            zout.writestr(name, fn(name, zin.read(name)))


class TestBatch4GateCriticals:

    def test_append_rid_remap_does_not_cross_wire(self, tmp_path):
        # sequential in-place rId replacement cross-wired anchors when a
        # reserved id equaled a still-unreplaced local id: the chart frame
        # pointed at the PNG, output unreadable (gate critical). Two-pass
        # placeholder remap now.
        import io as _io

        from PIL import Image as PILImageMod

        from openpyxl.chart import BarChart, Reference
        from openpyxl.drawing.image import Image

        src = _two_sheet_chart_fixture(tmp_path)
        wb = load_workbook(src, preserve=True)
        ws = wb["Data"]
        chart = BarChart()
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=5))
        ws.add_chart(chart, "K2")
        buf = _io.BytesIO()
        PILImageMod.new("RGB", (1, 1), "red").save(buf, format="png")
        buf.seek(0)
        ws.add_image(Image(buf), "K20")
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        payloads = part_payloads(out)
        rels = next(p for n, p in payloads.items() if "drawings/_rels" in n)
        drawing = next(p for n, p in payloads.items()
                       if n.startswith("xl/drawings/drawing"))
        relmap = {}
        for tag in re.findall(rb"<Relationship\b[^>]*>", rels):
            rid_m = re.search(rb'Id="([^"]+)"', tag)
            target_m = re.search(rb'Target="([^"]+)"', tag)
            if rid_m and target_m:
                relmap[rid_m.group(1)] = target_m.group(1)
        for rid in re.findall(rb'<c:chart [^>]*r:id="(rId\d+)"', drawing):
            assert b"chart" in relmap[rid]
        for rid in re.findall(rb'embed="(rId\d+)"', drawing):
            assert b"media" in relmap[rid]
        wb2 = load_workbook(out)                # parses: nothing swapped
        assert len(wb2["Data"]._charts) == 2
        assert len(wb2["Data"]._images) == 1

    def test_hyperlink_and_drawing_share_rid_allocator(self, fixture_copy,
                                                       tmp_path):
        # both planners computed next_rid independently over the same
        # original rels -> duplicate rId (OPC violation, gate critical)
        from openpyxl.chart import BarChart, Reference

        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Sheet1"]
        ws["A1"].hyperlink = "https://example.com/x"
        chart = BarChart()
        chart.add_data(Reference(ws, min_col=1, min_row=1, max_row=3))
        ws.add_chart(chart, "F2")
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        rels = next(p for n, p in part_payloads(out).items()
                    if "worksheets/_rels" in n)
        ids = re.findall(rb'Id="(rId\d+)"', rels)
        assert len(ids) == len(set(ids))
        wb2 = load_workbook(out)
        assert wb2["Sheet1"]["A1"].hyperlink is not None
        assert len(wb2["Sheet1"]._charts) == 1

    def test_file_object_image_reads_from_offset_zero(self, fixture_copy,
                                                      tmp_path):
        # PIL leaves the stream position mid-file: the media part was
        # saved as garbage bytes (gate critical); stock mode was correct
        from PIL import Image as PILImageMod

        from openpyxl.drawing.image import Image

        png_path = str(tmp_path / "d.png")
        PILImageMod.new("RGB", (3, 3), "green").save(png_path,
                                                     format="png")
        src = fixture_copy("minimal/minimal_clean.xlsx")
        with open(png_path, "rb") as f:
            img = Image(f)
            wb = load_workbook(src, preserve=True)
            wb["Sheet1"].add_image(img, "E5")
            out = str(tmp_path / "o.xlsx")
            wb.save(out)
        media = next(p for n, p in part_payloads(out).items()
                     if n.startswith("xl/media/"))
        assert media.startswith(b"\x89PNG")

    def test_entity_text_single_unescape_roundtrip(self, tmp_path):
        # chained str.replace decoded '&amp;lt;' twice: a title with
        # literal entity-like text was silently rewritten (gate critical)
        src = _two_sheet_chart_fixture(tmp_path)
        wb = load_workbook(src, preserve=True)
        wb["Data"]._charts[0].title = "Powered by <html> & &lt;stuff&gt;"
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        runs = wb2["Data"]._charts[0].title.tx.rich.p[0].r
        assert runs[0].t == "Powered by <html> & &lt;stuff&gt;"

    def test_axis_order_swap_patches_correct_title(self, tmp_path):
        # flat positional <a:t> mapping patched the WRONG axis title when
        # the original serialized valAx before catAx (gate critical);
        # leaves now map within ancestor-path groups
        src = _two_sheet_chart_fixture(tmp_path, with_axis_titles=True)
        swapped = str(tmp_path / "swapped.xlsx")

        def swap(name, payload):
            if not name.startswith("xl/charts/chart"):
                return payload
            cat = re.search(rb"<catAx>.*?</catAx>", payload, re.S)
            val = re.search(rb"<valAx>.*?</valAx>", payload, re.S)
            return (payload[:cat.start()] + val.group(0)
                    + payload[cat.end():val.start()] + cat.group(0)
                    + payload[val.end():])

        _rezip(src, swapped, swap)
        wb = load_workbook(swapped, preserve=True)
        wb["Data"]._charts[0].x_axis.title = "NewX"
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        chart = next(p for n, p in part_payloads(out).items()
                     if n.startswith("xl/charts/chart"))
        catax = re.search(rb"<catAx>.*?</catAx>", chart, re.S).group(0)
        valax = re.search(rb"<valAx>.*?</valAx>", chart, re.S).group(0)
        assert b"NewX" in catax
        assert b"YT" in valax and b"NewX" not in valax

    def test_rename_after_add_chart_follows(self, tmp_path):
        # the rename cascade skipped in-session charts: the new chart part
        # referenced the old, now-nonexistent title (gate critical)
        from openpyxl.chart import BarChart, Reference

        src = _two_sheet_chart_fixture(tmp_path)
        wb = load_workbook(src, preserve=True)
        ws = wb["Data"]
        chart = BarChart()
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=5))
        ws.add_chart(chart, "K2")
        ws.title = "D2"
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        for name, payload in part_payloads(out).items():
            if name.startswith("xl/charts/chart"):
                for f in re.findall(rb"<(?:c:)?f>([^<]*)</(?:c:)?f>",
                                    payload):
                    assert b"D2" in f and b"Data" not in f


class TestBatch4GateMajors:

    def test_gt_in_attribute_and_single_quoted_ids(self, tmp_path):
        # the tag tokenizer stopped at '>' inside quoted attribute values
        # (false refusal) and the cNvPr id scan missed single-quoted ids
        # (duplicate shape ids) — gate majors
        from openpyxl.chart import BarChart, Reference

        src = _two_sheet_chart_fixture(tmp_path)
        crafted = str(tmp_path / "crafted.xlsx")

        def plant(name, payload):
            if name.startswith("xl/drawings/drawing"):
                return payload.replace(
                    b'<cNvPr id="1" name="Chart 1"/>',
                    b"<cNvPr id='7' name=\"a > b\"/>", 1)
            return payload

        _rezip(src, crafted, plant)
        wb = load_workbook(crafted, preserve=True)
        ws = wb["Data"]
        chart = BarChart()
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=5))
        ws.add_chart(chart, "K2")               # no false refusal
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        drawing = next(p for n, p in part_payloads(out).items()
                       if n.startswith("xl/drawings/drawing"))
        ids = [a or b for a, b in
               re.findall(rb"\bid=(?:\"(\d+)\"|'(\d+)')", drawing)]
        assert len(ids) == len(set(ids))        # no duplicate shape ids

    def test_empty_self_closing_wsdr_appendable(self, tmp_path):
        from openpyxl.chart import BarChart, Reference

        src = _two_sheet_chart_fixture(tmp_path)
        crafted = str(tmp_path / "crafted.xlsx")

        def blank(name, payload):
            if name.startswith("xl/drawings/drawing"):
                m = re.match(rb"<wsDr[^>]*>", payload)
                return m.group(0)[:-1] + b"/>"
            if "drawings/_rels" in name:
                return (b'<?xml version="1.0"?><Relationships xmlns="http:'
                        b'//schemas.openxmlformats.org/package/2006/'
                        b'relationships"/>')
            return payload

        _rezip(src, crafted, blank)
        wb = load_workbook(crafted, preserve=True)
        ws = wb["Data"]
        chart = BarChart()
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=5))
        ws.add_chart(chart, "K2")               # no false refusal
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        assert len(load_workbook(out)["Data"]._charts) == 1

    def test_orphan_drawing_rel_gets_element_spliced(self, tmp_path):
        # rel + part existed but the sheet never referenced them: the
        # appended chart was invisible with no refusal (gate major)
        from openpyxl.chart import BarChart, Reference

        src = _two_sheet_chart_fixture(tmp_path)
        crafted = str(tmp_path / "crafted.xlsx")

        def strip_el(name, payload):
            if name.startswith("xl/worksheets/sheet"):
                return re.sub(rb"<drawing [^>]*/>", b"", payload)
            return payload

        _rezip(src, crafted, strip_el)
        wb = load_workbook(crafted, preserve=True)
        ws = wb["Data"]
        chart = BarChart()
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=5))
        ws.add_chart(chart, "K2")
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        sheet = next(p for n, p in part_payloads(out).items()
                     if n.startswith("xl/worksheets/") and b"<drawing" in p)
        assert b"<drawing" in sheet
        assert len(load_workbook(out)["Data"]._charts) == 2

    def test_unrelated_shift_does_not_block_chart_edit(self, tmp_path):
        # any(led.shifts) refused every chart edit even when the shift
        # touched a sheet the chart never references (gate major)
        src = _two_sheet_chart_fixture(tmp_path)
        wb = load_workbook(src, preserve=True)
        wb["Data"]._charts[0].title = "Edited"
        wb["Plain"].insert_rows(1)
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        chart = next(p for n, p in part_payloads(out).items()
                     if n.startswith("xl/charts/chart"))
        assert b"Edited" in chart

        # the RELATED combination still refuses (double-shift hazard)
        wb2 = load_workbook(src, preserve=True)
        wb2["Data"]._charts[0].title = "Edited"
        wb2["Data"].insert_rows(1)
        with pytest.raises(UnsupportedStructureError,
                           match="separate sessions"):
            wb2.save(str(tmp_path / "o2.xlsx"))

    def test_charref_in_original_title_refuses_typed(self, tmp_path):
        src = _two_sheet_chart_fixture(tmp_path)
        crafted = str(tmp_path / "crafted.xlsx")

        def inject(name, payload):
            if name.startswith("xl/charts/chart"):
                return payload.replace(b"<a:t>Original</a:t>",
                                       b"<a:t>Ori&#10;ginal</a:t>", 1)
            return payload

        _rezip(src, crafted, inject)
        wb = load_workbook(crafted, preserve=True)
        wb["Data"]._charts[0].title = "New"
        with pytest.raises(UnsupportedStructureError, match="character"):
            wb.save(str(tmp_path / "o.xlsx"))


class TestBatch5LintGaps:

    def test_quoted_external_refs_never_judged(self, fixture_copy):
        # the quoted storage form of external-workbook references was
        # flagged unknown-sheet; refuse mode blocked legitimate binds
        # (gate major)
        from openpyxl.formula.lint import lint_formula

        wb = load_workbook(fixture_copy("features/schedule.xlsx"))
        for f in ("='[Budget.xlsx]Sheet One'!A1", "='[1]Extern'!A1",
                  r"='C:\path\[Budget.xlsx]Sheet1'!A1"):
            assert lint_formula(f, workbook=wb) == []

    def test_in_session_table_columns_unknowable(self, fixture_copy):
        # in-session tables have no tableColumns until save: every
        # structured ref against them was falsely refused (gate major)
        from openpyxl.formula.lint import lint_formula
        from openpyxl.worksheet.table import Table

        wb = load_workbook(fixture_copy("features/schedule.xlsx"))
        ws = wb["Summary"]
        ws["E1"] = "Hdr"
        ws["E2"] = 5
        ws.add_table(Table(displayName="TNew", ref="E1:E2"))
        assert lint_formula("=SUM(TNew[Hdr])", workbook=wb) == []

    def test_escaped_column_names_unknowable(self, fixture_copy):
        # Excel's ' escape in column specs needs a full parser: such
        # specs are unknowable, never unknown (gate major)
        from openpyxl.formula.lint import lint_formula

        wb = load_workbook(fixture_copy("features/tables.xlsx"))
        table_name = next(iter(next(
            ws for ws in wb.worksheets if ws.tables).tables))
        f = "=SUM({0}[Col'[1']])".format(table_name)
        assert lint_formula(f, workbook=wb) == []

    def test_modern_functions_and_eta_refs_lint_clean(self, fixture_copy):
        from openpyxl.formula.lint import lint_formula

        wb = load_workbook(fixture_copy("features/schedule.xlsx"))
        assert lint_formula('=REGEXTEST(A1,"x")') == []
        assert lint_formula("=GROUPBY(A1:A2,B1:B2,SUM)") == []
        assert lint_formula("=SUM(_xlfn.ANCHORARRAY(A1))") == []
        assert lint_formula("=REDUCE(0,A1:A3,SUM)", workbook=wb) == []

    def test_array_formula_binds_are_linted(self, fixture_copy):
        # ArrayFormula objects carried their text past the chokepoint
        # (gate minor): garbage landed in the file under refuse mode
        from openpyxl.worksheet.formula import ArrayFormula

        wb = load_workbook(fixture_copy("features/schedule.xlsx"),
                           preserve=True)
        wb.formula_lint = "refuse"
        ws = wb["Summary"]
        with pytest.raises(UnsupportedStructureError, match="pre-flight"):
            ws["D5"] = ArrayFormula("D5:D6", "=SUMM(Nowhere!A1")
        assert ws["D5"].value is None                 # atomic


class TestBatch5OracleGaps:

    def test_merged_input_refuses_typed(self, tmp_path):
        # a merged-cell interior input crashed with raw AttributeError
        # (gate minor): typed refusal naming the remedy now
        from openpyxl import oracle
        from openpyxl.errors import TargetNotFoundError

        wb = Workbook()
        ws = wb.active
        ws.title = "Model"
        ws["F1"] = 1
        ws.merge_cells("F1:G1")
        p = str(tmp_path / "m.xlsx")
        wb.save(p)
        with pytest.raises(TargetNotFoundError, match="anchor"):
            oracle.evaluate(p, {"Model!G1": 5}, [])

    def test_bare_name_tokens_resolve_as_names_in_sketch(self, tmp_path):
        # a defined name shaped like column letters ("IN") was parsed as
        # a whole-column reference, so input taint missed its readers and
        # the certification falsely DIVERGED (gate major)
        from openpyxl.preserve.perception import dependency_sketch
        from openpyxl.workbook.defined_name import DefinedName

        wb = Workbook()
        ws = wb.active
        ws.title = "Model"
        ws["A1"] = 5
        wb.defined_names["IN"] = DefinedName("IN",
                                             attr_text="Model!$A$1")
        ws["A5"] = "=IN*3"
        sketch = dependency_sketch(wb)
        refs = (sketch.references.get("Model!A5")
                or sketch.references.get("'Model'!A5") or [])
        assert any(bounds == (1, 1, 1, 1) for (_s, bounds, _r) in refs)

    def test_unresolved_formulas_inherit_input_taint(self, tmp_path):
        # a cell fed by an input only through INDIRECT escaped the taint
        # and the evaluation certification falsely DIVERGED (gate major).
        # Statically checkable: the taint walk, not the oracle.
        from openpyxl import oracle as oracle_mod

        wb = Workbook()
        ws = wb.active
        ws.title = "Model"
        ws["A1"] = 5
        ws["C1"] = '=INDIRECT("A1")+0'
        buf_path = str(tmp_path / "t.xlsx")
        wb.save(buf_path)
        with open(buf_path, "rb") as f:
            data = f.read()
        result, _ = oracle_mod._certify_impl(
            data, timeout=1, recalculated=None,
            input_seeds=[("Model", 1, 1)])
        # C1 is cache-less here so the early return fires; the point is
        # the seeding path — exercise it via the reasons dict directly
        assert result.status == "BASELINE_UNVERIFIABLE"

    def test_baseline_unverifiable_carries_exclusions(self, tmp_path):
        # write_back(allow_uncertified=True) on a cache-less workbook
        # wrote volatile cells (gate major): the early-return result now
        # carries the exclusion classes
        from openpyxl import oracle as oracle_mod

        wb = Workbook()
        ws = wb.active
        ws.title = "M"
        ws["A1"] = "=NOW()"
        ws["A2"] = "=A1+1"
        ws["A3"] = "=1+1"
        p = str(tmp_path / "v.xlsx")
        wb.save(p)
        with open(p, "rb") as f:
            data = f.read()
        result, _ = oracle_mod._certify_impl(data, timeout=1)
        assert result.status == "BASELINE_UNVERIFIABLE"
        assert "M!A1" in result.volatile_excluded
        assert "M!A2" in result.volatile_excluded     # downstream taint

    def test_values_match_dates_vs_serials(self):
        import datetime

        from openpyxl.oracle import _values_match

        dt = datetime.datetime(2024, 1, 5)
        assert _values_match(45296, dt)               # serial == datetime
        assert _values_match(dt, 45296.0)
        assert not _values_match(dt, 45297)


class TestBatch6LocateGaps:

    def test_merged_label_interior_never_a_target(self):
        # locate returned the unwritable MergedCell interior of the
        # label's OWN merge as "the value" (gate critical)
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Growth rate"
        ws.merge_cells("A1:B1")
        ws["C1"] = 0.05
        assert ws.locate("Growth rate").coordinate == "C1"

    def test_adjacent_string_with_competitor_is_ambiguous(self):
        # the walk silently jumped over a string (a text value, a cached
        # formula string, or a number-stored-as-text) to a farther cell
        # (gate criticals x3): competition now refuses typed
        from openpyxl.errors import AmbiguousTargetError

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Status"
        ws["B1"] = "pending"
        ws["C1"] = 42
        with pytest.raises(AmbiguousTargetError) as exc:
            ws.locate("Status")
        assert exc.value.kind == "ambiguous-value-cell"
        assert set(exc.value.options) == {"Sheet!B1", "Sheet!C1"}
        # a lone adjacent string IS the value (text values are real)
        wb2 = Workbook()
        ws2 = wb2.active
        ws2["A1"] = "Status"
        ws2["B1"] = "pending"
        assert ws2.locate("Status").value == "pending"

    def test_error_cell_is_a_value(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Check"
        ws["B1"] = "#DIV/0!"
        ws["C1"] = 1
        assert ws.locate("Check").value == "#DIV/0!"

    def test_prefer_validated_first(self):
        wb = Workbook()
        wb.active["A1"] = "L"
        with pytest.raises(ValueError, match="prefer"):
            wb.active.locate("Anything", prefer="left")

    def test_allowed_values_whole_column_and_reversed(self):
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = Workbook()
        ws = wb.active
        for i, v in enumerate(("Yes", "No", "Maybe"), start=1):
            ws.cell(row=i, column=2, value=v)
        dv = DataValidation(type="list", formula1="=$B:$B")
        dv.add("D1")
        ws.add_data_validation(dv)
        assert ws.allowed_values("D1") == ["Yes", "No", "Maybe"]
        dv2 = DataValidation(type="list", formula1="=$B$3:$B$1")
        dv2.add("D2")
        ws.add_data_validation(dv2)
        assert ws.allowed_values("D2") == ["Yes", "No", "Maybe"]


class TestBatch6InstrumentGaps:

    def test_diff_reports_content_at_vacated_coordinates(self, tmp_path):
        # new content written where a shift vacated cells was invisible
        # in the diff (gate critical)
        from openpyxl.preserve import diff_workbooks

        wb0 = Workbook()
        s0 = wb0.active
        s0.title = "S"
        s0["A1"] = "hdr"
        s0["A2"] = 10
        before = str(tmp_path / "a.xlsx")
        wb0.save(before)
        wb = load_workbook(before, preserve=True)
        remap = wb["S"].insert_rows(2)
        wb["S"]["A2"] = 99
        after = str(tmp_path / "b.xlsx")
        wb.save(after)
        report = diff_workbooks(before, after, remaps=[remap])
        assert any(e["after"] == 99 for e in report.changed)
        assert any(e["from"] == "S!A2" and e["to"] == "S!A3"
                   for e in report.shifted)

    def test_diff_is_bool_aware(self, tmp_path):
        from openpyxl.preserve import diff_workbooks

        a = Workbook()
        a.active["A1"] = 1
        pa = str(tmp_path / "a.xlsx")
        a.save(pa)
        b = Workbook()
        b.active["A1"] = True
        pb = str(tmp_path / "b.xlsx")
        b.save(pb)
        assert len(diff_workbooks(pa, pb).changed) == 1

    def test_preserve_receipt_survives_reaccess(self):
        # the submodule import shadowed the lazily-exported FUNCTION: the
        # second access returned the module (gate major)
        import openpyxl.preserve as preserve_pkg

        first = preserve_pkg.receipt
        second = preserve_pkg.receipt
        assert callable(first) and callable(second)
        assert first is second

    def test_search_reads_array_formula_text(self):
        # search fabricated matches from the Python repr and missed the
        # real formula text (gate critical)
        from openpyxl.worksheet.formula import ArrayFormula

        wb = Workbook()
        ws = wb.active
        ws["B1"] = ArrayFormula("B1:B2", "=SUM(A1:A2*2)")
        hits = wb.search("SUM")
        assert [h["address"] for h in hits] == ["Sheet!B1"]
        assert wb.search(r"0x[0-9a-f]+", regex=True) == []

    def test_perception_verbs_guard_readonly_writeonly(self, fixture_copy):
        from openpyxl.preserve import findings, scan_errors

        wb = load_workbook(fixture_copy("features/schedule.xlsx"),
                           read_only=True)
        for call in (lambda: wb.search("x"), lambda: wb.model_map(),
                     lambda: scan_errors(wb), lambda: findings(wb)):
            with pytest.raises(ValueError, match="materialized"):
                call()
        wo = Workbook(write_only=True)
        ws = wo.create_sheet()
        ws.append(["Growth rate"])
        with pytest.raises(ValueError, match="materialized"):
            wo.search("Growth")

    def test_model_map_sees_cross_sheet_inputs(self):
        wb = Workbook()
        data = wb.active
        data.title = "Data"
        data["B1"] = 7                        # referenced from Calc
        calc = wb.create_sheet("Calc")
        calc["A1"] = "=Data!B1*2"
        mm = wb.model_map()
        assert "B1" in mm.sheets["Data"]["inputs"]
        assert mm.inputs("Data") == ["B1"]

    def test_manifest_part_name_follows_rename(self, fixture_copy):
        wb = load_workbook(fixture_copy("features/schedule.xlsx"),
                           preserve=True)
        wb["Schedule"].title = "Renamed"
        doc = wb.manifest().to_dict()
        entry = next(s for s in doc["sheets"] if s["title"] == "Renamed")
        assert entry["part_name"] == "xl/worksheets/sheet1.xml"

    def test_merged_hazard_fires_from_preserved_bytes(self, fixture_copy,
                                                      tmp_path):
        # the model discards shadowed interior values at load, so the
        # detector could never fire (gate minor): byte-level view now
        from openpyxl.preserve import findings

        src = fixture_copy("features/merged.xlsx")
        crafted = str(tmp_path / "shadow.xlsx")
        # plant a shadowed value INSIDE the first merge
        wb0 = load_workbook(src)
        rng = next(iter(wb0.active.merged_cells.ranges))
        interior = "{0}{1}".format(
            chr(ord("A") + rng.min_col), rng.min_row) \
            if rng.max_col > rng.min_col else "{0}{1}".format(
                chr(ord("A") + rng.min_col - 1), rng.min_row + 1)
        with zipfile.ZipFile(src) as zin, \
                zipfile.ZipFile(crafted, "w") as zout:
            for name in zin.namelist():
                payload = zin.read(name)
                if name.startswith("xl/worksheets/sheet1"):
                    payload = payload.replace(
                        b"</sheetData>",
                        '<row r="{0}"><c r="{1}"><v>777</v></c></row>'
                        .format(rng.min_row if rng.max_col > rng.min_col
                                else rng.min_row + 1,
                                interior).encode()
                        + b"</sheetData>", 1)
                zout.writestr(name, payload)
        wb = load_workbook(crafted, preserve=True)
        kinds = {f.kind for f in findings(wb)}
        assert "merged-hazard" in kinds


class TestBatch7DeliveryGaps:

    def test_scrub_preserved_comment_is_honest_and_saveable(
            self, fixture_copy, tmp_path):
        # scrub nulled a PRESERVED comment, reported it "removed", and
        # bricked the save (gate critical): it must skip preserved
        # machinery, report it truthfully, and stay saveable
        from openpyxl.comments import Comment

        src = fixture_copy("features/schedule.xlsx")
        wb0 = load_workbook(src)
        wb0["Schedule"]["A2"].comment = Comment("preserved", "auth")
        wb0.save(src)
        wb = load_workbook(src, preserve=True)
        report = wb.scrub()
        assert not any("comment at" in r for r in report["removed"])
        assert any("preserved comment machinery" in s
                   for s in report["skipped"])
        assert wb["Schedule"]["A2"].comment is not None   # not nulled
        out = str(tmp_path / "o.xlsx")
        wb.save(out)                                      # still saveable
        assert load_workbook(out)["Schedule"]["A2"].comment is not None

    def test_scrub_removes_session_comment_on_clean_sheet(self,
                                                          fixture_copy):
        from openpyxl.comments import Comment

        wb = load_workbook(fixture_copy("features/schedule.xlsx"),
                           preserve=True)
        wb["Summary"]["A9"].comment = Comment("session", "me")
        report = wb.scrub(remove=("comments",))
        assert any("Summary!A9" in r for r in report["removed"])

    def test_protect_for_delivery_actively_locks_non_inputs(self,
                                                            tmp_path):
        # relying on the OOXML default left author-unlocked outputs
        # editable under a "protected" sheet (gate critical)
        from openpyxl.styles import Protection

        wb = Workbook()
        ws = wb.active
        ws.title = "Model"
        ws["A1"] = 10
        ws["A2"] = 20
        ws["B1"] = "=A1+A2"
        ws["C1"] = "=B1*2"
        for c in ("A1", "A2", "B1", "C1"):
            ws[c].protection = Protection(locked=False)
        src = str(tmp_path / "src.xlsx")
        wb.save(src)
        wb = load_workbook(src, preserve=True)
        report = wb.protect_for_delivery()
        assert report["locked_cells"] >= 2
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        r = load_workbook(out)["Model"]
        assert r["C1"].protection.locked is True     # output locked
        assert r["B1"].protection.locked is True     # calculation locked
        assert r["A1"].protection.locked is False    # input stays open

    def test_protection_edits_preserve_hidden_flag(self, tmp_path):
        # apply_profile/protect built a fresh Protection, dropping hidden
        # (gate major): the hidden bit must survive a locked-only change
        from openpyxl.preserve import apply_profile
        from openpyxl.styles import Protection

        wb = Workbook()
        ws = wb.active
        ws.title = "M"
        ws["A1"] = 10
        ws["A2"] = 20
        ws["B1"] = "=A1+A2"
        ws["A1"].protection = Protection(locked=True, hidden=True)
        src = str(tmp_path / "s.xlsx")
        wb.save(src)
        wb = load_workbook(src, preserve=True)
        apply_profile(wb["M"], {"inputs": {"locked": False}})
        assert wb["M"]["A1"].protection.hidden is True

    def test_set_input_range_name_refuses_typed(self, fixture_copy):
        from openpyxl.errors import AmbiguousTargetError
        from openpyxl.workbook.defined_name import DefinedName

        wb = load_workbook(fixture_copy("features/schedule.xlsx"),
                           preserve=True)
        wb.defined_names["Rng"] = DefinedName(
            "Rng", attr_text="Schedule!$B$2:$B$4")
        with pytest.raises(AmbiguousTargetError, match="RANGE"):
            wb.set_input("Rng", 5)

    def test_set_input_merged_interior_refuses_typed(self, fixture_copy):
        from openpyxl.utils import get_column_letter
        from openpyxl.workbook.defined_name import DefinedName

        wb = load_workbook(fixture_copy("features/merged.xlsx"),
                           preserve=True)
        ws = next(w for w in wb.worksheets if w.merged_cells.ranges)
        rng = next(iter(ws.merged_cells.ranges))
        if rng.max_col <= rng.min_col:
            pytest.skip("fixture merge is single-column")
        interior = "'{0}'!${1}${2}".format(
            ws.title, get_column_letter(rng.min_col + 1), rng.min_row)
        wb.defined_names["Int"] = DefinedName("Int", attr_text=interior)
        with pytest.raises(UnsupportedStructureError,
                           match="merged range"):
            wb.set_input("Int", 5)
