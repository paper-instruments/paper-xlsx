"""Relationship-resolved pivot graph inventory."""
from __future__ import annotations

import io
import json
import os
import zipfile

import pytest

from openpyxl import load_workbook
from openpyxl.errors import (
    AmbiguousTargetError,
    RelationshipPolicyError,
    TargetNotFoundError,
)
from openpyxl.pivot.graph import load_pivot_graph, load_workbook_pivot_graph
from openpyxl.xml.constants import REL_NS, SHEET_MAIN_NS

from .conftest import FIXTURES_DIR
from .support.harness import assert_part_budget
from .support.partdiff import part_payloads

_NS = SHEET_MAIN_NS
_RNS = REL_NS
_PIVOTS = os.path.join(FIXTURES_DIR, "pivots")


def _ct_overrides(parts):
    mapping = {
        "pivotTable": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml."
            "pivotTable+xml"),
        "pivotCacheDefinition": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml."
            "pivotCacheDefinition+xml"),
        "pivotCacheRecords": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml."
            "pivotCacheRecords+xml"),
        "worksheet": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml."
            "worksheet+xml"),
        "table": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml."
            "table+xml"),
    }
    overrides = [
        '<Override PartName="/xl/workbook.xml" ContentType="application/'
        'vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
    ]
    for name, kind in parts:
        overrides.append(
            '<Override PartName="/%s" ContentType="%s"/>'
            % (name, mapping[kind]))
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/'
        'content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats'
        '-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        + "".join(overrides)
        + "</Types>"
    )


def _rels(entries, xmlns="http://schemas.openxmlformats.org/package/2006/relationships"):
    body = "".join(
        '<Relationship Id="%s" Type="%s" Target="%s"/>' % entry
        for entry in entries
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="%s">%s</Relationships>' % (xmlns, body)
    )


def _sheet_xml(cells=""):
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="%s"><sheetData>%s</sheetData></worksheet>'
        % (_NS, cells)
    )


def _workbook_xml(sheets, caches=()):
    sheet_xml = "".join(
        '<sheet name="%s" sheetId="%s" r:id="%s"/>' % item for item in sheets)
    cache_xml = ""
    if caches:
        cache_xml = "<pivotCaches>" + "".join(
            '<pivotCache cacheId="%s" r:id="%s"/>' % item for item in caches
        ) + "</pivotCaches>"
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="%s" xmlns:r="%s"><sheets>%s</sheets>%s</workbook>'
        % (_NS, _RNS, sheet_xml, cache_xml)
    )


def _cache_xml(source, fields, record_count=0, records_rid="rId1",
               grouping=False, calculated=False, ext_uri=None,
               prefixed=False):
    if source["kind"] == "range":
        ws = '<worksheetSource sheet="%s" ref="%s"/>' % (
            source["sheet"], source["ref"])
        source_type = source.get("type", "worksheet")
    elif source["kind"] in ("table", "defined-name", "named"):
        name = source["name"]
        sheet = source.get("sheet")
        ws = ('<worksheetSource name="%s"%s/>'
              % (name, ' sheet="%s"' % sheet if sheet else ""))
        source_type = source.get("type", "worksheet")
    else:
        ws = ""
        source_type = source.get("type", source["kind"])
    field_xml = "".join(
        '<cacheField name="%s"><sharedItems count="%s">%s</sharedItems>'
        '%s%s</cacheField>' % (
            field["name"],
            len(field.get("items", ())),
            "".join(field.get("items", ())),
            '<fieldGroup><discretePr count="0"/></fieldGroup>'
            if grouping and field is fields[0] else "",
            ' formula="1+1"' if False else "",
        )
        for field in fields
    )
    if calculated:
        field_xml += (
            '<cacheField name="Calc" formula="Revenue*2">'
            '<sharedItems/></cacheField>'
        )
    extras = "<calculatedItems/>" if calculated else ""
    ext = ""
    if ext_uri:
        ext = (
            '<extLst><ext uri="%s" xmlns:x14="http://schemas.microsoft.com/'
            'office/spreadsheetml/2009/9/main"><x14:pivotCacheDefinition/>'
            '</ext></extLst>' % ext_uri
        )
    root = "x:pivotCacheDefinition" if prefixed else "pivotCacheDefinition"
    source_tag = "x:cacheSource" if prefixed else "cacheSource"
    fields_tag = "x:cacheFields" if prefixed else "cacheFields"
    ns = (' xmlns:x="%s" xmlns:r="%s"' % (_NS, _RNS)) if prefixed \
        else ' xmlns="%s" xmlns:r="%s"' % (_NS, _RNS)
    rid = ' r:id="%s"' % records_rid if records_rid else ""
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<%s%s recordCount="%s"%s>'
        '<%s type="%s">%s</%s>'
        '<%s count="%s">%s</%s>%s%s</%s>'
        % (root, ns, record_count, rid, source_tag, source_type, ws,
           source_tag, fields_tag, len(fields) + (1 if calculated else 0),
           field_xml, fields_tag, extras, ext, root)
    )


def _records_xml(rows):
    body = "".join("<r>%s</r>" % row for row in rows)
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<pivotCacheRecords xmlns="%s" count="%s">%s</pivotCacheRecords>'
        % (_NS, len(rows), body)
    )


def _pivot_xml(name, cache_id, location="A3:B6", rows=(), columns=(),
               pages=(), values=(), tag=None, ext_uri=None,
               compact=None, outline=None, data_on_rows=None,
               cache_relationship_id="rId1"):
    row_xml = ""
    if rows:
        row_xml = '<rowFields count="%s">%s</rowFields>' % (
            len(rows), "".join('<field x="%s"/>' % item for item in rows))
    col_xml = ""
    if columns:
        col_xml = '<colFields count="%s">%s</colFields>' % (
            len(columns),
            "".join('<field x="%s"/>' % item for item in columns))
    page_xml = ""
    if pages:
        page_bits = []
        for item in pages:
            if isinstance(item, tuple):
                fld = item[0]
                extra = item[1] if len(item) > 1 else None
                if extra is None:
                    page_bits.append('<pageField fld="%s"/>' % fld)
                else:
                    page_bits.append(
                        '<pageField fld="%s" item="%s"/>' % (fld, extra))
            else:
                page_bits.append('<pageField fld="%s"/>' % item)
        page_xml = '<pageFields count="%s">%s</pageFields>' % (
            len(pages), "".join(page_bits))
    data_xml = ""
    if values:
        data_xml = '<dataFields count="%s">%s</dataFields>' % (
            len(values),
            "".join(
                '<dataField name="%s" fld="%s" subtotal="%s"/>' % item
                for item in values))
    page_indexes = [
        item[0] if isinstance(item, tuple) else item for item in pages]
    field_count = 1 + max(
        [0] + list(rows) + list(columns) + page_indexes
        + [item[1] for item in values])
    fields = "".join('<pivotField showAll="0"/>' for _ in range(field_count))
    tag_attr = ' tag="%s"' % tag if tag else ""
    layout_attr = ""
    if compact is not None:
        layout_attr += ' compact="%s"' % (1 if compact else 0)
    if outline is not None:
        layout_attr += ' outline="%s"' % (1 if outline else 0)
    if data_on_rows is not None:
        layout_attr += ' dataOnRows="%s"' % (1 if data_on_rows else 0)
    ext = ""
    if ext_uri:
        ext = (
            '<extLst><ext uri="%s" xmlns:x14="http://schemas.microsoft.com/'
            'office/spreadsheetml/2009/9/main"/></extLst>' % ext_uri
        )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<pivotTableDefinition xmlns="%s" xmlns:r="%s" name="%s" cacheId="%s"'
        ' dataCaption="Values" r:id="%s"%s%s>'
        '<location ref="%s" firstHeaderRow="1" firstDataRow="1"'
        ' firstDataCol="1"/>'
        '<pivotFields count="%s">%s</pivotFields>'
        '%s%s%s%s%s</pivotTableDefinition>'
        % (_NS, _RNS, name, cache_id, cache_relationship_id, tag_attr,
           layout_attr, location, field_count, fields, row_xml, col_xml,
           page_xml, data_xml, ext)
    )


def _package(members):
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zout:
        for name, payload in members:
            if isinstance(payload, str):
                payload = payload.encode("utf-8")
            zout.writestr(name, payload)
    return buf.getvalue()


def _basic_package(
        *,
        cache_part="xl/pivotCache/pivotCacheDefinition1.xml",
        records_part="xl/pivotCache/pivotCacheRecords1.xml",
        pivot_part="xl/pivotTables/pivotTable1.xml",
        cache_target=None,
        pivot_target=None,
        cache_rid="rIdCache1",
        pivot_rid="rIdPivot1",
        records_rid="rId1",
        pivot_cache_rid="rId1",
        source=None,
        extra_pivots=(),
        extra_caches=(),
        include_records=True,
        include_workbook_cache=True,
        include_pivot_cache_rel=True,
        pivot_name="SalesByRegion",
        cache_id="1",
        fields=None,
        records=None,
        location="B3:C8",
        rows=(0,),
        columns=(),
        pages=(),
        values=(("Sum of Amount", 1, "sum"),),
        tag=None,
        compact=None,
        outline=None,
        data_on_rows=None,
        ext_uri=None,
        cache_ext_uri=None,
        grouping=False,
        calculated=False,
        prefixed_cache=False,
        table_part=None,
        defined_name=None,
        dangling_pivot=False,
        dangling_cache=False,
        dangling_records=False,
        duplicate_cache_id=False,
        record_count=None,
):
    source = source or {"kind": "range", "sheet": "Data", "ref": "A1:B5"}
    fields = fields or [
        {"name": "Region", "items": ['<s v="East"/>', '<s v="West"/>']},
        {"name": "Amount", "items": []},
    ]
    records = records or ['<s v="East"/><n v="10"/>', '<s v="West"/><n v="7"/>']
    cache_target = cache_target or ("/" + cache_part)
    pivot_target = pivot_target or "../pivotTables/" + pivot_part.rsplit("/", 1)[-1]
    overrides = [
        ("xl/worksheets/sheet1.xml", "worksheet"),
        ("xl/worksheets/sheet2.xml", "worksheet"),
        (pivot_part, "pivotTable"),
        (cache_part, "pivotCacheDefinition"),
    ]
    if include_records:
        overrides.append((records_part, "pivotCacheRecords"))
    members = [
        ("[Content_Types].xml", _ct_overrides(overrides)),
        ("_rels/.rels", _rels([
            ("rId1",
             "http://schemas.openxmlformats.org/officeDocument/2006/"
             "relationships/officeDocument",
             "xl/workbook.xml"),
        ])),
    ]
    caches = [("1", cache_rid)] if include_workbook_cache else []
    if duplicate_cache_id:
        caches.append(("1", "rIdCacheDup"))
    extra_cache_rels = []
    for extra_id, extra_part, extra_rid in extra_caches:
        caches.append((extra_id, extra_rid))
        extra_cache_rels.append((
            extra_rid,
            "http://schemas.openxmlformats.org/officeDocument/2006/"
            "relationships/pivotCacheDefinition",
            "/" + extra_part,
        ))
        overrides.append((extra_part, "pivotCacheDefinition"))
    if defined_name:
        defined = (
            '<definedNames><definedName name="%s">Data!A1:B5'
            '</definedName></definedNames>' % defined_name
        )
    else:
        defined = ""
    workbook = _workbook_xml(
        (("Summary", "1", "rId1"), ("Data", "2", "rId2")),
        caches,
    )
    if defined:
        workbook = workbook.replace("</workbook>", defined + "</workbook>")
    members.append(("xl/workbook.xml", workbook))
    wb_rels = [
        ("rId1",
         "http://schemas.openxmlformats.org/officeDocument/2006/"
         "relationships/worksheet",
         "worksheets/sheet1.xml"),
        ("rId2",
         "http://schemas.openxmlformats.org/officeDocument/2006/"
         "relationships/worksheet",
         "worksheets/sheet2.xml"),
    ]
    if include_workbook_cache:
        wb_rels.append((
            cache_rid,
            "http://schemas.openxmlformats.org/officeDocument/2006/"
            "relationships/pivotCacheDefinition",
            cache_target,
        ))
    if duplicate_cache_id:
        wb_rels.append((
            "rIdCacheDup",
            "http://schemas.openxmlformats.org/officeDocument/2006/"
            "relationships/pivotCacheDefinition",
            "/xl/custom/other-cache.xml",
        ))
        overrides.append(("xl/custom/other-cache.xml", "pivotCacheDefinition"))
    wb_rels.extend(extra_cache_rels)
    members.append(("xl/_rels/workbook.xml.rels", _rels(wb_rels)))
    members.append((
        "xl/worksheets/sheet1.xml",
        _sheet_xml('<row r="3"><c r="B3"><v>10</v></c></row>'),
    ))
    members.append((
        "xl/worksheets/sheet2.xml",
        _sheet_xml(
            '<row r="1"><c r="A1" t="inlineStr"><is><t>Region</t></is></c>'
            '<c r="B1" t="inlineStr"><is><t>Amount</t></is></c></row>'
            '<row r="2"><c r="A2" t="inlineStr"><is><t>East</t></is></c>'
            '<c r="B2"><v>10</v></c></row>'
        ),
    ))
    sheet_rels = []
    if not dangling_pivot:
        sheet_rels.append((
            pivot_rid,
            "http://schemas.openxmlformats.org/officeDocument/2006/"
            "relationships/pivotTable",
            pivot_target if not dangling_pivot else "../pivotTables/missing.xml",
        ))
    else:
        sheet_rels.append((
            pivot_rid,
            "http://schemas.openxmlformats.org/officeDocument/2006/"
            "relationships/pivotTable",
            "../pivotTables/missing.xml",
        ))
    for extra_name, extra_part, extra_rid, extra_cache in extra_pivots:
        sheet_rels.append((
            extra_rid,
            "http://schemas.openxmlformats.org/officeDocument/2006/"
            "relationships/pivotTable",
            "../pivotTables/" + extra_part.rsplit("/", 1)[-1],
        ))
    if table_part:
        sheet_rels.append((
            "rIdTable1",
            "http://schemas.openxmlformats.org/officeDocument/2006/"
            "relationships/table",
            "../tables/" + table_part.rsplit("/", 1)[-1],
        ))
        members.append((
            "xl/worksheets/_rels/sheet2.xml.rels",
            _rels([(
                "rIdTable1",
                "http://schemas.openxmlformats.org/officeDocument/2006/"
                "relationships/table",
                "../tables/" + table_part.rsplit("/", 1)[-1],
            )]),
        ))
        members.append((
            table_part,
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<table xmlns="%s" name="SalesData" displayName="SalesData"'
            ' ref="A1:B5" headerRowCount="1">'
            '<tableColumns count="2">'
            '<tableColumn id="1" name="Region"/>'
            '<tableColumn id="2" name="Amount"/>'
            '</tableColumns></table>' % _NS,
        ))
        overrides.append((table_part, "table"))
    members.append(("xl/worksheets/_rels/sheet1.xml.rels", _rels(sheet_rels)))
    if not dangling_pivot:
        members.append((
            pivot_part,
            _pivot_xml(
                pivot_name, cache_id, location=location, rows=rows,
                columns=columns, pages=pages, values=values, tag=tag,
                ext_uri=ext_uri, compact=compact, outline=outline,
                data_on_rows=data_on_rows,
                cache_relationship_id=pivot_cache_rid),
        ))
        pivot_rels = []
        if include_pivot_cache_rel:
            pivot_rels.append((
                pivot_cache_rid,
                "http://schemas.openxmlformats.org/officeDocument/2006/"
                "relationships/pivotCacheDefinition",
                cache_target if not cache_target.startswith("..")
                else cache_target,
            ))
        if pivot_rels:
            folder, _, base = pivot_part.rpartition("/")
            members.append((
                "%s/_rels/%s.rels" % (folder, base),
                _rels(pivot_rels),
            ))
    for extra_name, extra_part, extra_rid, extra_cache in extra_pivots:
        members.append((
            extra_part,
            _pivot_xml(extra_name, extra_cache, location="E3:F8", rows=(0,),
                       values=(("Sum of Amount", 1, "sum"),)),
        ))
        if extra_cache == cache_id:
            folder, _, base = extra_part.rpartition("/")
            members.append((
                "%s/_rels/%s.rels" % (folder, base),
                _rels([(
                    "rId1",
                    "http://schemas.openxmlformats.org/officeDocument/2006/"
                    "relationships/pivotCacheDefinition",
                    cache_target,
                )]),
            ))
        overrides.append((extra_part, "pivotTable"))
    members[0] = ("[Content_Types].xml", _ct_overrides(overrides))
    declared = len(records) if record_count is None else record_count
    cache_payload = _cache_xml(
        source, fields, record_count=declared,
        records_rid=records_rid,
        grouping=grouping, calculated=calculated, ext_uri=cache_ext_uri,
        prefixed=prefixed_cache,
    )
    members.append((cache_part, cache_payload))
    if include_records and not dangling_records:
        members.append((records_part, _records_xml(records)))
        folder, _, base = cache_part.rpartition("/")
        members.append((
            "%s/_rels/%s.rels" % (folder, base),
            _rels([(
                records_rid,
                "http://schemas.openxmlformats.org/officeDocument/2006/"
                "relationships/pivotCacheRecords",
                records_part.rsplit("/", 1)[-1]
                if records_part.rsplit("/", 1)[0] == cache_part.rsplit("/", 1)[0]
                else "/" + records_part,
            )]),
        ))
    elif dangling_records:
        folder, _, base = cache_part.rpartition("/")
        members.append((
            "%s/_rels/%s.rels" % (folder, base),
            _rels([(
                records_rid,
                "http://schemas.openxmlformats.org/officeDocument/2006/"
                "relationships/pivotCacheRecords",
                "missing-records.xml",
            )]),
        ))
    if dangling_cache:
        members = [
            (name, payload) for name, payload in members
            if name != cache_part
        ]
    if duplicate_cache_id:
        members.append((
            "xl/custom/other-cache.xml",
            _cache_xml(
                {"kind": "range", "sheet": "Data", "ref": "A1:B5"},
                fields, record_count=0, records_rid=None),
        ))
    return _package(members)


def _write_package(tmp_path, name, payload):
    path = tmp_path / name
    path.write_bytes(payload)
    return str(path)


def test_graph_resolves_custom_filenames_and_relative_targets():
    payload = _basic_package(
        cache_part="xl/custom/cache-a.xml",
        records_part="xl/custom/records-a.xml",
        pivot_part="xl/custom/report.xml",
        cache_target="/xl/custom/cache-a.xml",
        pivot_target="../custom/report.xml",
    )
    graph = load_pivot_graph(payload)
    assert len(graph.pivots) == 1
    pivot = graph.pivots[0]
    assert pivot.identity.pivot_part == "xl/custom/report.xml"
    assert pivot.identity.relationship_id == "rIdPivot1"
    assert pivot.cache_definition_part == "xl/custom/cache-a.xml"
    assert pivot.cache_records_part == "xl/custom/records-a.xml"
    assert pivot.valid
    assert graph.caches_by_part["xl/custom/cache-a.xml"].valid


def test_graph_indexes_duplicate_names_by_sheet_identity():
    payload = _basic_package(
        extra_pivots=(
            ("SalesByRegion", "xl/pivotTables/pivotTable2.xml",
             "rIdPivot2", "1"),
        ),
    )
    graph = load_pivot_graph(payload)
    names = [node.identity.name for node in graph.pivots]
    assert names == ["SalesByRegion", "SalesByRegion"]
    identities = list(graph.pivots_by_identity)
    assert identities[0].pivot_part != identities[1].pivot_part


def test_graph_detects_shared_cache_from_relationships():
    payload = _basic_package(
        extra_pivots=(
            ("MarginByRegion", "xl/pivotTables/pivotTable2.xml",
             "rIdPivot2", "1"),
        ),
    )
    graph = load_pivot_graph(payload)
    cache = graph.caches_by_id["1"]
    assert cache.referenced_by == (
        ("Summary", "MarginByRegion"),
        ("Summary", "SalesByRegion"),
    )
    assert not any(reason.code == "duplicate-incoming"
                   for reason in graph.reasons)


def test_graph_reads_range_table_and_defined_name_sources(tmp_path):
    range_graph = load_pivot_graph(_basic_package())
    assert range_graph.pivots[0].source_descriptor.kind == "range"
    assert range_graph.pivots[0].source_descriptor.sheet == "Data"
    assert range_graph.pivots[0].source_descriptor.ref == "A1:B5"

    table_graph = load_pivot_graph(_basic_package(
        source={"kind": "table", "name": "SalesData"},
        table_part="xl/tables/table1.xml",
    ))
    assert table_graph.caches_by_id["1"].source_descriptor.kind == "table"
    assert table_graph.caches_by_id["1"].source_descriptor.name == "SalesData"

    named_graph = load_pivot_graph(_basic_package(
        source={"kind": "defined-name", "name": "SalesRange"},
        defined_name="SalesRange",
    ))
    assert named_graph.caches_by_id["1"].source_descriptor.kind == "defined-name"


def test_graph_reports_output_range_and_field_assignments():
    graph = load_pivot_graph(_basic_package(
        location="B4:H28",
        rows=(0, 2),
        values=(("Revenue", 1, "sum"), ("Units", 3, "count")),
    ))
    pivot = graph.pivots[0]
    assert pivot.output_range == "B4:H28"
    assert pivot.row_fields == (0, 2)
    assert [item["aggregate"] for item in pivot.data_fields] == ["sum", "count"]
    assert graph.caches_by_id["1"].field_names[:2] == ("Region", "Amount")


def test_graph_preserves_extension_identity_without_granting_capability():
    graph = load_pivot_graph(_basic_package(
        ext_uri="{725AE2AE-9491-48be-B2B4-4EB974FC3084}",
        cache_ext_uri="{725AE2AE-9491-48be-B2B4-4EB974FC3084}",
    ))
    pivot = graph.pivots[0]
    cache = graph.caches_by_id["1"]
    assert pivot.extension_fingerprints[0].uri == (
        "{725AE2AE-9491-48be-B2B4-4EB974FC3084}")
    assert "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main" \
        in cache.extension_fingerprints[0].namespaces
    assert cache.payload_sha256
    assert pivot.payload_sha256


def test_graph_detects_dangling_and_count_errors():
    dangling_pivot = load_pivot_graph(_basic_package(dangling_pivot=True))
    assert dangling_pivot.pivots[0].valid is False
    assert any(reason.code == "dangling-sheet-pivot"
               for reason in dangling_pivot.pivots[0].reasons)

    dangling_cache = load_pivot_graph(_basic_package(dangling_cache=True))
    assert any(reason.code in ("dangling-workbook-cache", "missing-part",
                               "dangling-pivot-cache")
               for reason in dangling_cache.reasons + dangling_cache.pivots[0].reasons)

    dangling_records = load_pivot_graph(_basic_package(dangling_records=True))
    assert any(reason.code == "dangling-cache-records"
               for reason in dangling_records.caches[0].reasons)

    mismatch = load_pivot_graph(_basic_package(record_count=9))
    assert any(reason.code == "record-count-mismatch"
               for reason in mismatch.caches[0].reasons)
    assert mismatch.pivots[0].identity.name == "SalesByRegion"


def test_graph_detects_duplicate_cache_ids():
    graph = load_pivot_graph(_basic_package(duplicate_cache_id=True))
    assert any(reason.code == "duplicate-cache-id" for reason in graph.reasons)
    assert "1" not in graph.caches_by_id


def test_graph_keeps_valid_sibling_when_one_pivot_is_invalid():
    payload = _basic_package(
        extra_pivots=(
            ("Broken", "xl/pivotTables/pivotTable2.xml", "rIdPivot2", "99"),
        ),
    )
    graph = load_pivot_graph(payload)
    by_name = {node.identity.name: node for node in graph.pivots}
    assert by_name["SalesByRegion"].valid
    assert by_name["Broken"].valid is False
    assert any(reason.code == "dangling-pivot-cache"
               for reason in by_name["Broken"].reasons)


def test_refresh_on_load_uses_graph_and_custom_cache_part(
        fixture_copy, tmp_path):
    wb = load_workbook(fixture_copy("minimal/minimal_clean.xlsx"),
                       preserve=True)
    from tests.paper.test_delivery import _with_pivot_graph

    custom = "xl/custom/pivot-cache-a.xml"
    _with_pivot_graph(
        wb, [("SalesPivot", "1")],
        cache_parts={"1": custom},
        orphan_parts=("xl/pivotCache/pivotCacheDefinition99.xml",),
    )
    graph = load_workbook_pivot_graph(wb)
    assert graph.registered_cache_parts == (custom,)
    assert graph.pivots[0].cache_definition_part == custom
    assert wb.set_pivot_refresh_on_load(pivots=["SalesPivot"]) == [custom]
    before = wb._paper_source
    out = tmp_path / "graph-refresh.xlsx"
    wb.save(out)
    after = out.read_bytes()
    changed = {
        name for name, payload in part_payloads(after).items()
        if part_payloads(before).get(name) != payload
    }
    assert changed == {custom}


def test_refresh_on_load_shared_cache_still_deduplicates(fixture_copy):
    wb = load_workbook(fixture_copy("minimal/minimal_clean.xlsx"),
                       preserve=True)
    from tests.paper.test_delivery import _with_pivot_graph

    _with_pivot_graph(wb, [("SalesPivot", "1"), ("MarginPivot", "1")])
    assert wb.set_pivot_refresh_on_load(
        pivots=["Sheet1!SalesPivot", "MarginPivot"]
    ) == ["xl/pivotCache/pivotCacheDefinition1.xml"]


def test_refresh_on_load_ambiguity_and_missing_names_unchanged(fixture_copy):
    wb = load_workbook(fixture_copy("minimal/minimal_clean.xlsx"),
                       preserve=True)
    from tests.paper.test_delivery import _with_pivot_graph

    _with_pivot_graph(wb, [("Duplicate", "1"), ("Duplicate", "2")])
    with pytest.raises(AmbiguousTargetError):
        wb.set_pivot_refresh_on_load(pivots=["Duplicate"])
    with pytest.raises(TargetNotFoundError):
        wb.set_pivot_refresh_on_load(pivots=["Missing"])


def test_read_only_inspection_and_noop_save_are_byte_identical(
        fixture_copy, tmp_path):
    src = fixture_copy("minimal/minimal_clean.xlsx")
    wb = load_workbook(src, preserve=True)
    from tests.paper.test_delivery import _with_pivot_graph

    _with_pivot_graph(wb, [("ReadyPivot", "1")], refresh_on_load=True)
    source = wb._paper_source
    graph = load_pivot_graph(source)
    assert graph.pivots[0].identity.name == "ReadyPivot"
    path = tmp_path / "noop-pivot.xlsx"
    path.write_bytes(source)
    wb = load_workbook(path, preserve=True)
    out = tmp_path / "noop-out.xlsx"
    wb.save(out)
    assert_part_budget(source, out.read_bytes())


def test_graph_does_not_mutate_ledger_or_cells(fixture_copy):
    wb = load_workbook(fixture_copy("minimal/minimal_clean.xlsx"),
                       preserve=True)
    from tests.paper.test_delivery import _with_pivot_graph

    _with_pivot_graph(wb, [("SalesPivot", "1")])
    before_cells = {
        ws.title: set(ws._cells) for ws in wb.worksheets
    }
    before_dirty = {
        ws.title: set(wb._paper_ledger.dirty_coordinates(ws))
        for ws in wb.worksheets
    }
    load_workbook_pivot_graph(wb)
    assert {
        ws.title: set(ws._cells) for ws in wb.worksheets
    } == before_cells
    assert {
        ws.title: set(wb._paper_ledger.dirty_coordinates(ws))
        for ws in wb.worksheets
    } == before_dirty


def _sidecar_binaries():
    if not os.path.isdir(_PIVOTS):
        return []
    found = []
    for name in sorted(os.listdir(_PIVOTS)):
        if name.endswith((".xlsx", ".xlsm", ".xltx")):
            found.append(name)
    return found


@pytest.mark.parametrize("filename", _sidecar_binaries())
def test_real_fixture_graph_matches_sidecar(filename):
    path = os.path.join(_PIVOTS, filename)
    sidecar_path = path + ".json"
    with open(sidecar_path) as handle:
        sidecar = json.load(handle)
    expected = sidecar.get("expected_pivot_graph")
    if expected is None:
        pytest.skip("sidecar has no expected_pivot_graph")
    with open(path, "rb") as handle:
        payload = handle.read()
    wb = None
    try:
        wb = load_workbook(path, preserve=True)
    except (RelationshipPolicyError, ValueError, OSError):
        graph = load_pivot_graph(payload)
    else:
        graph = load_workbook_pivot_graph(wb)
    assert graph.to_dict()["pivots"] == expected["pivots"]
    assert graph.to_dict()["caches"] == expected["caches"]


def test_excel_fixture_absence_is_documented():
    manifest = os.path.join(_PIVOTS, "EXTERNAL_CORPUS.md")
    assert os.path.exists(manifest)
    required = [
        "excel_basic_table.xlsx",
        "excel_cross_tab.xlsx",
        "excel_filtered.xlsx",
        "excel_shared_cache.xlsx",
        "excel_formula_source.xlsx",
        "excel_semantic_edges.xlsx",
        "excel_extension_pivot.xlsx",
        "excel_unsupported_grouping.xlsx",
        "excel_data_model.xlsx",
        "excel_macro_pivot.xlsm",
        "excel_template_pivot.xltx",
        "excel_strict_basic.xlsx",
        "libreoffice_basic_pivot.xlsx",
    ]
    with open(manifest) as handle:
        text = handle.read()
    for name in required:
        assert name in text
