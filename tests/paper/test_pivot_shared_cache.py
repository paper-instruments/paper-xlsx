"""PR 7: shared-cache isolation for managed lifecycle verbs."""
from __future__ import annotations

import pytest

from openpyxl import load_workbook
from openpyxl.errors import RelationshipPolicyError
from openpyxl.pivot.qualify import PAPER_TAG

from .test_pivot_graph import _basic_package, _write_package


def test_shared_cache_refuses_lifecycle_with_siblings(tmp_path):
    src = _write_package(tmp_path, "shared.xlsx", _basic_package(
        tag=PAPER_TAG,
        extra_pivots=(
            ("MarginByRegion", "xl/pivotTables/pivotTable2.xml",
             "rIdPivot2", "1"),
        ),
    ))
    wb = load_workbook(src, preserve=True)
    pivot = wb["Summary"].pivots["SalesByRegion"]
    sibling = wb["Summary"].pivots["MarginByRegion"]
    assert pivot.capabilities.can_headless_refresh is False
    assert pivot.capabilities.can_edit_layout is False
    assert pivot.capabilities.can_delete is False
    with pytest.raises(RelationshipPolicyError) as refresh:
        pivot.refresh()
    assert refresh.value.kind == "pivot-cache-shared"
    assert "Summary!MarginByRegion" in refresh.value.options
    with pytest.raises(RelationshipPolicyError) as updated:
        pivot.update(layout="compact")
    assert updated.value.kind == "pivot-cache-shared"
    with pytest.raises(RelationshipPolicyError) as deleted:
        pivot.delete()
    assert deleted.value.kind == "pivot-cache-shared"
    assert "Summary!MarginByRegion" in deleted.value.options
    assert sibling.name == "MarginByRegion"


def test_rename_does_not_upgrade_shared_cache_capabilities(tmp_path):
    src = _write_package(tmp_path, "shared-rename.xlsx", _basic_package(
        tag=PAPER_TAG,
        extra_pivots=((
            "MarginByRegion", "xl/pivotTables/pivotTable2.xml",
            "rIdPivot2", "1",
        ),),
    ))
    workbook = load_workbook(src, preserve=True)
    renamed = workbook["Summary"].pivots["SalesByRegion"].rename(
        "RegionalSales")

    assert renamed.capabilities.can_repoint_source is False
    with pytest.raises(RelationshipPolicyError) as exc:
        renamed.repoint_source("Data!A1:B5")
    assert exc.value.kind == "pivot-cache-shared"
