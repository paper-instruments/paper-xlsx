"""Release-cut contracts for the intentionally small Paper API surface."""
from __future__ import annotations

from copy import copy
import datetime
import io
import json
import re
import zipfile

import pytest

from openpyxl import Workbook, load_workbook


def _preserved_workbook():
    workbook = Workbook()
    workbook.active.title = "Data"
    workbook.active["A1"] = 1
    stream = io.BytesIO()
    workbook.save(stream)
    return load_workbook(io.BytesIO(stream.getvalue()), preserve=True)


class _Interrupt(BaseException):
    pass


@pytest.mark.parametrize("failure", [RuntimeError, _Interrupt])
def test_datetime_bind_failure_restores_cell_registry_and_ledger(
        monkeypatch, failure):
    import openpyxl.cell.cell as cell_module

    workbook = _preserved_workbook()
    sheet = workbook["Data"]
    cell = sheet["A1"]
    ledger = workbook._paper_ledger
    ledger.cells[sheet] = {(9, 9)}
    ledger.value_overwrites[sheet] = {(8, 8)}
    ledger.cache_writes[sheet] = {(1, 1): 42, (7, 7): 7}
    cells_bucket = ledger.cells[sheet]
    overwrite_bucket = ledger.value_overwrites[sheet]
    cache_bucket = ledger.cache_writes[sheet]
    before_style = copy(cell._style)
    registry = workbook._number_formats
    before_registry = list(registry)
    before_index = dict(registry._dict)
    index_identity = registry._dict
    before_clean = registry.clean
    real_mark = cell_module._mark_cell_dirty

    def fail_after_mark(*args, **kwargs):
        real_mark(*args, **kwargs)
        raise failure("injected")

    monkeypatch.setattr(cell_module, "_mark_cell_dirty", fail_after_mark)
    with pytest.raises(failure):
        cell.value = datetime.datetime(2026, 8, 13, 12, 30)

    assert (cell.value, cell.data_type, cell._style) == (1, "n", before_style)
    assert cell.number_format == "General"
    assert list(registry) == before_registry
    assert registry._dict is index_identity
    assert dict(registry._dict) == before_index
    assert registry.clean == before_clean
    assert ledger.cells[sheet] is cells_bucket
    assert ledger.value_overwrites[sheet] is overwrite_bucket
    assert ledger.cache_writes[sheet] is cache_bucket
    assert cells_bucket == {(9, 9)}
    assert overwrite_bucket == {(8, 8)}
    assert cache_bucket == {(1, 1): 42, (7, 7): 7}


def test_successful_cell_edit_invalidates_only_its_staged_cache():
    workbook = _preserved_workbook()
    sheet = workbook["Data"]
    cache = {(1, 1): 42, (7, 7): 7}
    workbook._paper_ledger.cache_writes[sheet] = cache

    sheet["A1"] = 2

    assert workbook._paper_ledger.cache_writes[sheet] is cache
    assert cache == {(7, 7): 7}


def test_explicit_number_format_failure_restores_registry(monkeypatch):
    import openpyxl.styles.styleable as styleable

    workbook = _preserved_workbook()
    cell = workbook["Data"]["A1"]
    before_style = copy(cell._style)
    registry = workbook._number_formats
    before = (list(registry), dict(registry._dict), registry.clean)
    real_mark = styleable._mark_styleable_dirty

    def fail_after_mark(*args, **kwargs):
        real_mark(*args, **kwargs)
        raise RuntimeError("injected")

    monkeypatch.setattr(styleable, "_mark_styleable_dirty", fail_after_mark)
    with pytest.raises(RuntimeError, match="injected"):
        cell.number_format = '0.0000 "units"'

    assert cell._style == before_style
    assert cell.number_format == "General"
    assert (list(registry), dict(registry._dict), registry.clean) == before


def test_append_refuses_reentrancy_and_rolls_back_partial_row():
    workbook = _preserved_workbook()
    sheet = workbook["Data"]
    before_cells = dict(sheet._cells)
    before_row = sheet._current_row

    def values():
        yield 2
        sheet.append([99])

    with pytest.raises(RuntimeError, match="re-entrant"):
        sheet.append(values())
    assert sheet._cells == before_cells
    assert sheet._current_row == before_row


def test_append_rollback_does_not_revert_generator_edit_on_other_sheet():
    workbook = _preserved_workbook()
    target = workbook["Data"]
    other = workbook.create_sheet("Other")
    before_cells = dict(target._cells)

    def values():
        yield 2
        other["A1"] = "unrelated"
        raise RuntimeError("generator failed")

    with pytest.raises(RuntimeError, match="generator failed"):
        target.append(values())
    assert target._cells == before_cells
    assert other["A1"].value == "unrelated"


def test_validate_runs_the_save_plan_without_building_an_archive(
        fixture_copy, monkeypatch):
    from openpyxl.preserve import zipio

    workbook = load_workbook(
        fixture_copy("features/schedule.xlsx"), preserve=True)
    workbook["Schedule"]["A2"] = "edited"

    def unexpected(*args, **kwargs):
        raise AssertionError("validation assembled an archive")

    monkeypatch.setattr(zipio, "build_archive_bytes", unexpected)
    monkeypatch.setattr(zipio, "build_and_deliver", unexpected)
    assert workbook.validate() is None


def test_file_like_default_sniffs_ooxml_and_restores_position(
        fixture_copy):
    from openpyxl.reader.excel import _preserve_by_default

    with open(fixture_copy("minimal/minimal_clean.xlsx"), "rb") as handle:
        stream = io.BytesIO(handle.read())
    stream.name = "misleading.xlsb"
    stream.seek(37)
    assert _preserve_by_default(stream, False) is True
    assert stream.tell() == 37
    assert not stream.closed

    non_ooxml = io.BytesIO(b"not a zip")
    non_ooxml.name = "misleading.xlsx"
    non_ooxml.seek(3)
    assert _preserve_by_default(non_ooxml, False) is False
    assert non_ooxml.tell() == 3


def test_image_replacement_retargets_one_relationship_and_keeps_old_media(
        fixture_copy, tmp_path):
    source = fixture_copy("features/chart_image.xlsx")
    with zipfile.ZipFile(source) as archive:
        old_media = archive.read("xl/media/image1.png")
        old_rels = archive.read("xl/drawings/_rels/drawing1.xml.rels")
    replacement = tmp_path / "replacement.png"
    with open(replacement, "wb") as handle:
        handle.write(old_media)

    workbook = load_workbook(source, preserve=True)
    selected = workbook["Model"].replace_image("H20", replacement)
    assert selected is workbook["Model"]._images[0]
    output = tmp_path / "replaced.xlsx"
    workbook.save(output)

    with zipfile.ZipFile(output) as archive:
        assert archive.read("xl/media/image1.png") == old_media
        assert archive.read("xl/media/image2.png") == old_media
        new_rels = archive.read("xl/drawings/_rels/drawing1.xml.rels")
    assert new_rels != old_rels
    assert b'Target="/xl/media/image2.png"' in new_rels


def test_image_replacement_preserves_relative_relationship_form(
        fixture_copy, tmp_path):
    source = fixture_copy("features/chart_image.xlsx")
    relative = tmp_path / "relative-image.xlsx"
    with zipfile.ZipFile(source) as zin, zipfile.ZipFile(relative, "w") as zout:
        for info in zin.infolist():
            payload = zin.read(info.filename)
            if info.filename == "xl/drawings/_rels/drawing1.xml.rels":
                payload = payload.replace(
                    b'Target="/xl/media/image1.png"',
                    b'Target="../media/image1.png"')
            zout.writestr(info, payload)

    with zipfile.ZipFile(relative) as archive:
        media = archive.read("xl/media/image1.png")
    replacement = tmp_path / "replacement-relative.png"
    with open(replacement, "wb") as handle:
        handle.write(media)

    workbook = load_workbook(relative, preserve=True)
    workbook["Model"].replace_image("H20", replacement)
    output = tmp_path / "relative-replaced.xlsx"
    workbook.save(output)

    with zipfile.ZipFile(output) as archive:
        rels = archive.read("xl/drawings/_rels/drawing1.xml.rels")
    assert b'Target="../media/image2.png"' in rels


def test_image_replacement_receipt_reports_specific_cause(
        fixture_copy, tmp_path):
    source = fixture_copy("features/chart_image.xlsx")
    with zipfile.ZipFile(source) as archive:
        media = archive.read("xl/media/image1.png")
    replacement = tmp_path / "replacement.png"
    replacement.write_bytes(media)

    workbook = load_workbook(source, preserve=True)
    workbook["Model"].replace_image("H20", replacement)
    receipt = workbook.save(tmp_path / "replaced.xlsx", receipt=True)

    assert any(
        effect == {
            "kind": "relationship_changed",
            "part": "xl/drawings/_rels/drawing1.xml.rels",
            "cause": "image_replaced",
        }
        for effect in receipt.derived_effects
    )


def test_image_replacement_refuses_shared_drawing_relationship(
        fixture_copy, tmp_path):
    from openpyxl.errors import UnsupportedStructureError

    source = fixture_copy("features/chart_image.xlsx")
    shared = tmp_path / "shared-image-relationship.xlsx"
    drawing_part = "xl/drawings/drawing1.xml"
    with zipfile.ZipFile(source) as zin, zipfile.ZipFile(shared, "w") as zout:
        for info in zin.infolist():
            payload = zin.read(info.filename)
            if info.filename == drawing_part:
                start = payload.index(
                    b"<oneCellAnchor><from><col>7</col><colOff>0</colOff>"
                    b"<row>19</row>")
                end = payload.index(b"</oneCellAnchor>", start) \
                    + len(b"</oneCellAnchor>")
                duplicate = payload[start:end]
                duplicate = duplicate.replace(
                    b"<row>19</row>", b"<row>29</row>", 1)
                duplicate = duplicate.replace(
                    b'id="2" name="Image 2"',
                    b'id="3" name="Image 3"', 1)
                payload = payload.replace(
                    b"</wsDr>", duplicate + b"</wsDr>")
            zout.writestr(info, payload)

    with zipfile.ZipFile(shared) as archive:
        media = archive.read("xl/media/image1.png")
    replacement = tmp_path / "replacement-shared.png"
    replacement.write_bytes(media)
    workbook = load_workbook(shared, preserve=True)
    images = workbook["Model"]._images
    assert len(images) == 2
    assert {image._paper_rel_id for image in images} == {"rId2"}

    with pytest.raises(
            UnsupportedStructureError,
            match="shares drawing relationship") as caught:
        workbook["Model"].replace_image("H20", replacement)

    assert caught.value.kind == "shared-image-relationship"
    assert not workbook._paper_ledger.image_replacements


def test_receipt_reports_deterministic_derived_effects(
        fixture_copy, tmp_path):
    workbook = load_workbook(
        fixture_copy("features/schedule_calc.xlsx"), preserve=True)
    workbook["Schedule"]["B2"] = 123
    receipt = workbook.save(tmp_path / "receipt.xlsx", receipt=True)
    payload = receipt.to_dict()

    assert payload["derived_effects_version"] == 1
    kinds = [effect["kind"] for effect in payload["derived_effects"]]
    assert "formula_cache_removed" in kinds
    assert "recalculation_metadata_changed" in kinds
    assert json.loads(json.dumps(payload))["derived_effects"] == \
        payload["derived_effects"]


def test_receipt_does_not_report_unchanged_calc_properties(
        fixture_copy, tmp_path):
    workbook = load_workbook(
        fixture_copy("minimal/minimal_clean.xlsx"), preserve=True)
    workbook.create_sheet("Added")

    receipt = workbook.save(tmp_path / "added-sheet.xlsx", receipt=True)

    assert not any(
        effect["kind"] == "recalculation_metadata_changed"
        for effect in receipt.derived_effects
    )


def test_chart_repoint_removes_only_its_matching_cache(
        fixture_copy, tmp_path):
    source = fixture_copy("features/chart_image.xlsx")
    cached = tmp_path / "chart-caches.xlsx"
    first = b"<f>'Model'!$B$2</f>"
    second = b"<f>'Model'!$C$2</f>"
    cache_10 = (
        b"<numCache><formatCode>General</formatCode><ptCount val=\"1\"/>"
        b"<pt idx=\"0\"><v>10</v></pt></numCache>")
    cache_20 = cache_10.replace(b">10<", b">20<")
    with zipfile.ZipFile(source) as zin, zipfile.ZipFile(cached, "w") as zout:
        for info in zin.infolist():
            payload = zin.read(info.filename)
            if info.filename.startswith("xl/charts/chart"):
                payload = payload.replace(first, first + cache_10, 1)
                payload = payload.replace(second, second + cache_20, 1)
            zout.writestr(info, payload)

    workbook = load_workbook(cached, preserve=True)
    workbook["Model"]._charts[0].repoint(0, "Model!$D$1:$D$4")
    output = tmp_path / "chart-repoint.xlsx"
    receipt = workbook.save(output, receipt=True)
    with zipfile.ZipFile(output) as archive:
        chart = archive.read("xl/charts/chart1.xml")

    assert chart.count(b"<numCache>") == 1
    assert b"<v>20</v>" in chart
    assert b"<v>10</v>" not in chart
    assert any(effect["kind"] == "chart_cache_removed"
               for effect in receipt.derived_effects)


def test_sheet_rename_composes_with_internal_hyperlinks(tmp_path):
    from openpyxl.worksheet.hyperlink import Hyperlink

    workbook = Workbook()
    target = workbook.active
    target.title = "Old Name"
    target["A1"] = "target"
    summary = workbook.create_sheet("Summary")
    summary["A1"] = "jump"
    summary["A1"].hyperlink = Hyperlink(
        ref="A1", location="'Old Name'!A1", display="jump")
    summary["A2"] = "external"
    summary["A2"].hyperlink = "https://example.com/path"
    source = tmp_path / "hyperlink-rename.xlsx"
    workbook.save(source)

    workbook = load_workbook(source, preserve=True)
    workbook["Old Name"].title = "New Name"
    output = tmp_path / "renamed.xlsx"
    workbook.save(output)
    reopened = load_workbook(output)

    assert reopened["Summary"]["A1"].hyperlink.location == \
        "'New Name'!A1"
    assert reopened["Summary"]["A2"].hyperlink.target == \
        "https://example.com/path"


def test_data_validation_range_edit_preserves_omitted_defaults(tmp_path):
    from openpyxl.worksheet.datavalidation import DataValidation

    workbook = Workbook()
    validation = DataValidation(type="whole", formula1="1")
    workbook.active.add_data_validation(validation)
    validation.add("A1")
    generated = tmp_path / "generated.xlsx"
    workbook.save(generated)

    source = tmp_path / "producer.xlsx"
    omitted = (b"allowBlank", b"showDropDown", b"showInputMessage",
               b"showErrorMessage")
    with zipfile.ZipFile(generated) as zin, zipfile.ZipFile(source, "w") as zout:
        for info in zin.infolist():
            payload = zin.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                for name in omitted:
                    payload = re.sub(
                        rb"\s+" + name + rb'=(?:"[^"]*"|\'[^\']*\')',
                        b"", payload)
            zout.writestr(info, payload)

    with zipfile.ZipFile(source) as archive:
        before_sheet = archive.read("xl/worksheets/sheet1.xml")
    before = before_sheet[
        before_sheet.index(b"<dataValidations"):
        before_sheet.index(b"</dataValidations>") + 18]

    workbook = load_workbook(source, preserve=True)
    workbook.active.data_validations.dataValidation[0].sqref = "B2"
    output = tmp_path / "edited.xlsx"
    workbook.save(output)
    with zipfile.ZipFile(output) as archive:
        after_sheet = archive.read("xl/worksheets/sheet1.xml")
    after = after_sheet[
        after_sheet.index(b"<dataValidations"):
        after_sheet.index(b"</dataValidations>") + 18]

    assert after == before.replace(b'sqref="A1"', b'sqref="B2"')
    for name in omitted:
        assert name + b"=" not in after
