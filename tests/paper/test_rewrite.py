"""Phase 6b: Excel-insert-semantics reference rewriting (PLAN Phase 6b).

Property tests per PLAN: insert-then-delete round-trips to the original;
sums recomputed by the oracle match the pre-edit values.
"""
from __future__ import annotations

import pytest

from openpyxl import load_workbook
from openpyxl.errors import UnsupportedStructureError
from openpyxl.package import diff_cells
from openpyxl.preserve.rewrite import shift_formula, shift_ref


class TestShiftRef:
    """Excel semantics, not fill semantics: absolutes move too; ranges
    spanning the edit point expand; deleted references become #REF!."""

    @pytest.mark.parametrize("ref,axis,idx,amt,delete,expected", [
        # inserts: endpoints at/after the index shift — absolutes included
        ("B2:B11", "rows", 5, 1, False, "B2:B12"),      # spanning: expands
        ("$B$2:$B$11", "rows", 5, 1, False, "$B$2:$B$12"),
        ("B2:B11", "rows", 2, 1, False, "B3:B12"),      # fully below: shifts
        ("B2:B11", "rows", 12, 1, False, "B2:B11"),     # above: untouched
        ("B15", "rows", 5, 1, False, "B16"),
        ("$B$15", "rows", 5, 1, False, "$B$16"),        # $ does NOT pin
        ("B4", "rows", 5, 1, False, "B4"),
        ("B:B", "rows", 5, 1, False, "B:B"),            # whole column
        ("5:9", "rows", 6, 2, False, "5:11"),           # whole-row range
        ("C3", "cols", 2, 1, False, "D3"),              # column insert
        ("$C$3:$E$3", "cols", 4, 2, False, "$C$3:$G$3"),
        # deletes: shrink, shift up, or #REF!
        ("B2:B11", "rows", 3, 2, True, "B2:B9"),
        ("B15", "rows", 5, 1, True, "B14"),
        ("B5", "rows", 5, 1, True, "#REF!"),            # deleted outright
        ("B5:B6", "rows", 5, 2, True, "#REF!"),         # fully deleted
        ("B4:B6", "rows", 5, 2, True, "B4:B4"),         # clamped
        ("D1", "cols", 4, 1, True, "#REF!"),
    ])
    def test_shift_ref_table(self, ref, axis, idx, amt, delete, expected):
        assert shift_ref(ref, axis, idx, amt, delete) == expected

    def test_cross_sheet_and_quoted_prefixes(self):
        new, changed = shift_formula(
            "=SUM(Data!B2:B5)+'My Sheet'!C7", None, "Data", "rows", 3, 1,
            False)
        assert changed and new == "=SUM(Data!B2:B6)+'My Sheet'!C7"
        new, changed = shift_formula(
            "='My Sheet'!C7*2", None, "My Sheet", "rows", 3, 1, False)
        assert changed and new == "='My Sheet'!C8*2"

    def test_other_sheet_references_untouched(self):
        _new, changed = shift_formula(
            "=SUM(Data!B2:B5)", "Model", "Model", "rows", 3, 1, False)
        assert not changed

    def test_defined_names_left_to_the_name_level(self):
        _new, changed = shift_formula(
            "=B12*(1+Growth)", "Schedule", "Schedule", "rows", 20, 1, False)
        assert not changed   # B12 above idx 20; Growth is not an A1 ref


class TestInsertRows:

    def test_full_insert_semantics(self, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Schedule"]
        ws.insert_rows(5)
        ws["A5"] = "Item new"
        ws["B5"] = 999
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        ws2 = wb2["Schedule"]
        assert ws2["B13"].value == "=SUM(B2:B12)"          # expanded
        assert ws2["B14"].value == "=B13*(1+Growth)"        # shifted
        assert wb2.defined_names["Growth"].value == "Schedule!$B$16"
        assert wb2["Summary"]["B1"].value == "=Schedule!B13"
        assert ws2["B16"].value == 0.05                     # input moved
        assert ws2["B5"].value == 999                       # new row content
        assert ws2["B6"].value == 500                       # shifted data
        # recalc-on-load set (formulas changed)
        from .support.partdiff import part_payloads
        assert b'fullCalcOnLoad="1"' in part_payloads(out)["xl/workbook.xml"]

    @pytest.mark.lo_smoke
    def test_oracle_confirms_correct_sums(self, lo, fixture_copy, tmp_path):
        from openpyxl import oracle

        src = fixture_copy("features/schedule.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Schedule"]
        ws.insert_rows(5)
        ws["B5"] = 999
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        calced = str(tmp_path / "calc.xlsx")
        oracle.recalc(out, output_path=calced)
        wb3 = load_workbook(calced, data_only=True)
        # the Q11 justification numbers, now CORRECT instead of silent-wrong
        assert wb3["Schedule"]["B13"].value == 7499
        assert wb3["Summary"]["B1"].value == 7499
        assert abs(wb3["Schedule"]["B14"].value - 7873.95) < 1e-9

    def test_insert_then_delete_round_trips(self, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule.xlsx")
        mid = str(tmp_path / "mid.xlsx")
        out = str(tmp_path / "out.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Schedule"].insert_rows(5)
        wb.save(mid)
        wb = load_workbook(mid, preserve=True)
        wb["Schedule"].delete_rows(5)
        wb.save(out)
        assert diff_cells(src, out).clean

    def test_shared_group_on_shifted_sheet_dissolves_correctly(
            self, fixture_copy, tmp_path):
        src = fixture_copy("gauntlet/gauntlet.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Calc"]                       # shared group B2:B6, no traps
        ws.insert_rows(3)
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        ws2 = wb2["Calc"]
        assert ws2["B2"].value == "=A2*2"      # above: untouched
        assert ws2["B4"].value == "=A4*2"      # shifted follower
        assert ws2["B7"].value == "=A7*2"
        assert ws2["A3"].value is None         # the inserted row

    def test_hyperlink_anchor_tracks_its_cell(self, fixture_copy, tmp_path):
        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Sheet1"]
        ws["A3"].hyperlink = "https://example.org/plums"
        mid = str(tmp_path / "mid.xlsx")
        wb.save(mid)
        wb = load_workbook(mid, preserve=True)
        wb["Sheet1"].insert_rows(2)
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert wb2["Sheet1"]["A4"].hyperlink.target == "https://example.org/plums"
        assert wb2["Sheet1"]["A4"].value == "pears"


class TestDeleteRows:

    def test_delete_shrinks_and_refs_out(self, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Schedule"]
        ws["D1"] = "=B5*10"                    # references the doomed row
        ws.delete_rows(5)
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        ws2 = wb2["Schedule"]
        assert ws2["D1"].value == "=#REF!*10"  # Excel semantics, loud
        assert ws2["B11"].value == "=SUM(B2:B10)"
        assert wb2.defined_names["Growth"].value == "Schedule!$B$14"

    def test_column_shift(self, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Schedule"]
        ws.insert_cols(2)                       # push B -> C
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        ws2 = wb2["Schedule"]
        assert ws2["C12"].value == "=SUM(C2:C11)"
        assert wb2.defined_names["Growth"].value == "Schedule!$C$15"
        assert ws2["C2"].value == 200           # data moved
        assert ws2["B2"].value is None


class TestShiftRefusals:

    def test_trap_sheet_still_refuses_with_blockers(self, fixture_copy):
        src = fixture_copy("gauntlet/gauntlet.xlsx")
        wb = load_workbook(src, preserve=True)
        with pytest.raises(UnsupportedStructureError) as exc:
            wb["Model"].insert_rows(3)
        msg = str(exc.value)
        assert "extLst" in msg or "extension" in msg
        assert "chart" in msg.lower()
        assert "Nothing was changed" in msg

    def test_array_formula_sheet_refuses(self, fixture_copy):
        src = fixture_copy("features/shared_formulas.xlsx")
        wb = load_workbook(src, preserve=True)
        with pytest.raises(UnsupportedStructureError, match="array"):
            wb["Calc"].insert_rows(3)

    def test_second_shift_same_session_refuses(self, fixture_copy):
        src = fixture_copy("features/schedule.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Schedule"].insert_rows(5)
        with pytest.raises(UnsupportedStructureError, match="save the workbook"):
            wb["Schedule"].insert_rows(7)

    def test_move_range_still_refuses(self, fixture_copy):
        src = fixture_copy("features/schedule.xlsx")
        wb = load_workbook(src, preserve=True)
        with pytest.raises(UnsupportedStructureError, match="move_range"):
            wb["Schedule"].move_range("B2:B4", rows=2)
