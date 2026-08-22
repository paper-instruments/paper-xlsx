"""PR 6: formula-backed pivot sources via the existing oracle seam."""
from __future__ import annotations

import pytest

from openpyxl import Workbook, load_workbook
from openpyxl.errors import UnsupportedStructureError
from openpyxl.pivot.calculate import (
    PivotCalculationArtifact,
    apply_calculated_values,
    snapshot_for_pivot,
)
from openpyxl.pivot.source import snapshot_from_workbook
from openpyxl.worksheet.table import Table

from .support.harness import save_and_reopen


def _formula_workbook(tmp_path, amount_formula="=10+1"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Region"
    ws["B1"] = "Amount"
    ws["A2"] = "East"
    ws["B2"] = amount_formula
    ws["A3"] = "West"
    ws["B3"] = 7
    ws.add_table(Table(displayName="Sales", ref="A1:B3"))
    wb.create_sheet("Summary")
    path = str(tmp_path / "formula.xlsx")
    wb.save(path)
    return path, load_workbook(path, preserve=True)


def _fake_artifact(snapshot, values):
    return PivotCalculationArtifact(
        candidate_sha256="abc",
        engine="libreoffice",
        engine_version="test",
        source=snapshot.source,
        source_identity=snapshot.identity,
        values_by_coordinate=values,
        excluded_coordinates={},
        errors=(),
    )


def test_formula_source_uses_calculated_values(tmp_path, monkeypatch):
    path, wb = _formula_workbook(tmp_path)

    def fake_calculate(workbook, snapshot):
        return _fake_artifact(snapshot, {addr: 11 for addr in snapshot.formula_coordinates})

    monkeypatch.setattr(
        "openpyxl.pivot.calculate.calculate_pivot_source", fake_calculate)
    wb["Summary"].pivots.create(
        name="ByRegion", source="Sales", destination="A1",
        rows=["Region"], values=["Amount"])
    destination = str(tmp_path / "out.xlsx")
    receipt = wb.save(destination, receipt=True)
    created = [
        item for item in receipt.derived_effects
        if item["kind"] == "pivot_created"
    ]
    calculation = created[0]["calculation"]
    assert calculation["candidate_sha256"] == "abc"
    assert calculation["engine"] == "libreoffice"
    assert calculation["engine_version"] == "test"
    assert calculation["source_identity"]
    assert calculation["value_count"] == 1
    assert calculation["excluded"] == []
    assert calculation["errors"] == []
    wb = load_workbook(destination, preserve=True)
    assert wb["Summary"]["B3"].value == 11


def test_formula_source_includes_unsaved_edits(tmp_path, monkeypatch):
    path, wb = _formula_workbook(tmp_path, amount_formula="=C2")
    wb["Data"]["C2"] = 21

    seen = {}

    def fake_calculate(workbook, snapshot):
        seen["value"] = workbook["Data"]["C2"].value
        return _fake_artifact(snapshot, {"Data!B2": workbook["Data"]["C2"].value})

    monkeypatch.setattr(
        "openpyxl.pivot.calculate.calculate_pivot_source", fake_calculate)
    wb["Summary"].pivots.create(
        name="ByRegion", source="Sales", destination="A1",
        rows=["Region"], values=["Amount"])
    assert seen["value"] == 21


def test_missing_libreoffice_refuses_before_mutation(tmp_path, monkeypatch):
    monkeypatch.setattr("openpyxl.oracle.find_soffice", lambda: None)
    path, wb = _formula_workbook(tmp_path)
    with pytest.raises(UnsupportedStructureError) as exc:
        wb["Summary"].pivots.create(
            name="ByRegion", source="Sales", destination="A1",
            rows=["Region"], values=["Amount"])
    assert exc.value.kind == "unsupported-pivot-source"
    assert "ByRegion" not in [p.name for p in wb["Summary"].pivots]


def test_calculation_artifact_hash_mismatch_refuses(tmp_path):
    path, wb = _formula_workbook(tmp_path)
    snapshot = snapshot_from_workbook(wb, "Sales")
    artifact = _fake_artifact(snapshot, {"Data!B2": 11})
    artifact = PivotCalculationArtifact(
        candidate_sha256=artifact.candidate_sha256,
        engine=artifact.engine,
        engine_version=artifact.engine_version,
        source=artifact.source,
        source_identity="not-the-snapshot",
        values_by_coordinate=artifact.values_by_coordinate,
        excluded_coordinates={},
        errors=(),
    )
    with pytest.raises(UnsupportedStructureError):
        apply_calculated_values(snapshot, artifact)
    matching = _fake_artifact(snapshot, {"Data!B2": 11})
    wb._paper_pivot_candidate_sha256 = "not-the-candidate"
    with pytest.raises(UnsupportedStructureError) as hashed:
        apply_calculated_values(snapshot, matching, workbook=wb)
    assert hashed.value.kind == "unsupported-pivot-source"


def test_libreoffice_bytes_are_not_published(tmp_path, monkeypatch):
    lo_bytes = b"PK\x03\x04lo-rewritten-not-a-real-package"

    def fake_recalc(data, timeout=120.0):
        return lo_bytes

    monkeypatch.setattr("openpyxl.oracle.find_soffice", lambda: "/fake/soffice")
    monkeypatch.setattr("openpyxl.oracle._recalculate_bytes", fake_recalc)

    path, wb = _formula_workbook(tmp_path)
    with pytest.raises(Exception):
        snapshot_for_pivot(wb, "Sales")
    published = str(tmp_path / "published.xlsx")
    wb.save(published)
    with open(published, "rb") as handle:
        assert handle.read() != lo_bytes
