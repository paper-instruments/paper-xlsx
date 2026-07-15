import io
import re
import zipfile

import pytest

from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.preserve import emit, saver, splice
from openpyxl.preserve.xmlscan import scan_sheet


def _formula_package():
    workbook = Workbook()
    workbook.active["A1"] = "=1+1"
    target = io.BytesIO()
    workbook.save(target)
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(target.getvalue())) as source, \
            zipfile.ZipFile(output, "w") as destination:
        for info in source.infolist():
            payload = source.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                payload, count = re.subn(
                    br'<c r="A1"><f>1\+1</f>(?:<v\s*/>|<v></v>)</c>',
                    b'<c r="A1" xmlns:x="urn:vendor" x:flag="keep">'
                    b' <f t="normal">1+1</f>\r\n<v>2</v> </c>', payload)
                assert count == 1
            destination.writestr(info, payload)
    return output.getvalue()


def _sheet_xml(package):
    with zipfile.ZipFile(io.BytesIO(package)) as archive:
        return archive.read("xl/worksheets/sheet1.xml")


def test_attribute_carry_preserves_exact_foreign_tokens_and_namespace():
    original = (
        b"<c r='A1' xmlns:x='urn:vendor' "
        b"x:hint='text s=&quot;9&quot;' custom = 'keep'><v>1</v></c>")
    rendered = b'<c r="A1" t="n"><v>2</v></c>'

    carried = emit.carry_attributes(rendered, original)

    assert b" xmlns:x='urn:vendor'" in carried
    assert b" x:hint='text s=&quot;9&quot;'" in carried
    assert b" custom = 'keep'" in carried
    assert carried.count(b" s=") == 1  # quoted text was not parsed as markup


def test_cache_patch_changes_only_type_and_direct_value_span():
    original = (
        b"<c r='A1' t='str' xmlns:x='urn:vendor' x:hint='keep'> "
        b"<f t='shared' si='2'>A2</f>\r\n"
        b"<v xml:space='preserve'> old </v>\n</c>")

    patched = splice._patch_cached_value(original, 2, None, "Sheet!A1")

    assert patched == (
        b"<c r='A1' xmlns:x='urn:vendor' x:hint='keep'> "
        b"<f t='shared' si='2'>A2</f>\r\n<v>2</v>\n</c>")


def test_cache_patch_preserves_boundary_whitespace_in_string_value():
    original = b'<c r="A1"><f>1</f><v>old</v></c>'

    patched = splice._patch_cached_value(
        original, " updated ", None, "Sheet!A1")

    assert patched == (
        b'<c r="A1" t="str"><f>1</f>'
        b'<v xml:space="preserve"> updated </v></c>')


def test_cache_patch_accepts_prefixed_main_namespace_children():
    original = (
        b'<c r="A1" xmlns:m="http://schemas.openxmlformats.org/'
        b'spreadsheetml/2006/main"><m:f>1+1</m:f><m:v>2</m:v></c>')

    patched = splice._patch_cached_value(
        original, 3, None, "Sheet!A1",
        formula_names=(b"m:f",), cache_names=(b"m:v",))

    assert patched == (
        b'<c r="A1" xmlns:m="http://schemas.openxmlformats.org/'
        b'spreadsheetml/2006/main"><m:f>1+1</m:f><v>3</v></c>')


def test_cache_patch_refuses_duplicate_logical_formula_children():
    original = (
        b'<c r="A1" xmlns:m="http://schemas.openxmlformats.org/'
        b'spreadsheetml/2006/main"><f>1+1</f><m:f>1+1</m:f><v>2</v></c>')

    with pytest.raises(splice.SpliceRefusal, match="duplicate formula/cache"):
        splice._patch_cached_value(
            original, 3, None, "Sheet!A1", formula_names=(b"m:f",))


def test_cache_invalidation_changes_only_type_and_direct_value_span():
    original = (
        b"<c r='A1' t='str' xmlns:x='urn:vendor' x:hint='keep'> "
        b"<f t='shared' si='2'>A2</f>\r\n"
        b"<v xml:space='preserve'> stale </v>\n</c>")

    patched = splice._patch_formula_cache_invalidation(
        original, "Sheet!A1")

    assert patched == (
        b"<c r='A1' xmlns:x='urn:vendor' x:hint='keep'> "
        b"<f t='shared' si='2'>A2</f>\r\n\n</c>")


def test_cache_invalidation_removes_every_duplicate_cache():
    original = b'<c r="A1"><f>1</f><v>stale</v> <v></v></c>'

    patched = splice._patch_formula_cache_invalidation(
        original, "Sheet!A1")

    assert patched == b'<c r="A1"><f>1</f> </c>'


def test_in_place_rich_text_edit_reemits_modeled_value():
    workbook = Workbook()
    workbook.active["A1"] = CellRichText(
        "old", TextBlock(InlineFont(b=True), " bold"))
    source = io.BytesIO()
    workbook.save(source)
    workbook = load_workbook(source, preserve=True, rich_text=True)

    workbook.active["A1"].value[0] = "new"
    output = io.BytesIO()
    workbook.save(output)

    value = load_workbook(output, rich_text=True).active["A1"].value
    assert value[0] == "new"
    assert value[1].text == " bold"
    assert value[1].font.b is True


@pytest.mark.parametrize("markup", [b"<!--keep-->", b"<?vendor keep?>"])
def test_value_edit_refuses_unowned_non_element_markup(markup):
    workbook = Workbook()
    workbook.active["A1"] = 1
    source = io.BytesIO()
    workbook.save(source)
    package = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(source.getvalue())) as zin, \
            zipfile.ZipFile(package, "w") as zout:
        for info in zin.infolist():
            payload = zin.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                payload, count = re.subn(
                    br'(<c r="A1"[^>]*>.*?<v>1</v>)(</c>)',
                    lambda match: match.group(1) + markup + match.group(2),
                    payload, flags=re.S)
                assert count == 1
            zout.writestr(info, payload)
    workbook = load_workbook(
        io.BytesIO(package.getvalue()), preserve=True)
    workbook.active["A1"] = 2

    with pytest.raises(splice.SpliceRefusal, match="unowned direct-child"):
        workbook.save(io.BytesIO())


def test_style_only_formula_edit_preserves_formula_cache_and_foreign_xml():
    source = _formula_package()
    workbook = load_workbook(io.BytesIO(source), preserve=True)
    workbook.active["A1"].font = workbook.active["A1"].font.copy(bold=True)
    output = io.BytesIO()
    workbook.save(output)

    xml = _sheet_xml(output.getvalue())
    cell = re.search(br'<c r="A1".*?</c>', xml, re.S).group(0)
    assert b'xmlns:x="urn:vendor" x:flag="keep"' in cell
    assert b' <f t="normal">1+1</f>\r\n<v>2</v> ' in cell


def test_formula_edit_replaces_formula_and_removes_old_cache():
    workbook = load_workbook(io.BytesIO(_formula_package()), preserve=True)
    workbook.active["A1"] = "=2+2"
    output = io.BytesIO()
    workbook.save(output)

    cell = re.search(
        br'<c r="A1".*?</c>', _sheet_xml(output.getvalue()), re.S).group(0)
    assert b"<f>2+2</f>" in cell
    assert b"<v>2</v>" not in cell
    assert b'x:flag="keep"' in cell


def test_direct_child_scanner_ignores_markup_inside_cdata():
    fragment = (
        b'<c r="A1"><f><![CDATA[IF(A1<2,"<f>","")]]></f><v>1</v></c>')

    children = splice._direct_cell_children(fragment)

    assert [child.name for child in children] == [b"f", b"v"]


def test_shared_formula_spacing_dissolves_group(
        fixture_copy, tmp_path):
    source = fixture_copy("features/shared_formulas.xlsx")
    rewritten = tmp_path / "spaced-shared.xlsx"
    with zipfile.ZipFile(source) as archive, zipfile.ZipFile(
            rewritten, "w") as output:
        for info in archive.infolist():
            payload = archive.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                payload = payload.replace(b't="shared"', b't = "shared"')
            output.writestr(info, payload)

    workbook = load_workbook(rewritten, preserve=True)
    workbook["Calc"]["B2"].font = workbook["Calc"]["B2"].font.copy(
        bold=True)
    destination = tmp_path / "dissolved.xlsx"
    workbook.save(destination)
    xml = _sheet_xml(destination.read_bytes())

    assert b't = "shared"' not in xml
    assert b't="shared"' not in xml


def test_style_only_prefixed_formula_preserves_formula_bytes():
    source = _formula_package()
    formula = b'<m:f xmlns:m="http://schemas.openxmlformats.org/' \
        b'spreadsheetml/2006/main">1+1</m:f>'
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(source)) as archive, \
            zipfile.ZipFile(output, "w") as destination:
        for info in archive.infolist():
            payload = archive.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                payload = payload.replace(b'<f t="normal">1+1</f>', formula)
            destination.writestr(info, payload)
    workbook = load_workbook(io.BytesIO(output.getvalue()), preserve=True)
    workbook.active["A1"].font = workbook.active["A1"].font.copy(bold=True)
    saved = io.BytesIO()
    workbook.save(saved)

    assert formula in _sheet_xml(saved.getvalue())
    assert b"<v>2</v>" in _sheet_xml(saved.getvalue())


def test_cache_invalidation_recognizes_prefixed_formula_and_cache_children():
    source = _formula_package()
    main = b"http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    prefixed = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(source)) as archive, \
            zipfile.ZipFile(prefixed, "w") as destination:
        for info in archive.infolist():
            payload = archive.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                payload = payload.replace(
                    b'<c r="A1"',
                    b'<c r="A1" xmlns:m="' + main + b'"', 1)
                payload = payload.replace(
                    b'<f t="normal">1+1</f>',
                    b'<m:f t="normal">1+1</m:f>', 1)
                payload = payload.replace(
                    b"<v>2</v>", b"<m:v>2</m:v>", 1)
            destination.writestr(info, payload)

    workbook = load_workbook(
        io.BytesIO(prefixed.getvalue()), preserve=True)
    workbook.active["B1"] = "=3+3"
    saved = io.BytesIO()
    workbook.save(saved)

    xml = _sheet_xml(saved.getvalue())
    assert b'<m:f t="normal">1+1</m:f>' in xml
    assert b"<m:v>2</m:v>" not in xml


@pytest.mark.parametrize("array_ref, coordinate", [
    ("A:A", (1, 1)),
    ("1:1", (1, 2)),
])
def test_open_array_refs_have_finite_cache_invalidation_bounds(
        array_ref, coordinate):
    row, col = coordinate
    column = chr(ord("A") + col - 1)
    main = b"http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    sheet = (
        b'<worksheet xmlns="' + main + b'"><sheetData><row r="' +
        str(row).encode("ascii") + b'"><c r="' + column.encode("ascii") +
        str(row).encode("ascii") + b'"><f t="array" ref="' +
        array_ref.encode("ascii") + b'">1+1</f><v>2</v></c></row>'
        b'</sheetData></worksheet>')

    scan = scan_sheet(sheet)

    assert coordinate in saver._formula_cache_invalidations(scan)
    with pytest.raises(splice.SpliceRefusal, match="array/spill range"):
        splice.resolve_dirty_cells(Workbook().active, {coordinate}, scan)


def test_prefixed_array_follower_before_anchor_is_recorded():
    main = b"http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    sheet = (
        b'<worksheet xmlns="' + main + b'"><sheetData><row r="1">'
        b'<c r="A1" xmlns:m="' + main + b'"><m:v>2</m:v></c>'
        b'<c r="B1"><f t="array" ref="A1:B1">1+1</f><v>2</v></c>'
        b'</row></sheetData></worksheet>')

    scan = scan_sheet(sheet)
    follower = scan.rows[1].cells[1]
    patched = splice._patch_formula_cache_invalidation(
        sheet[follower.start:follower.end], "Sheet!A1",
        allow_cache_only=True, cache_names=scan.cache_names[(1, 1)])

    assert scan.cache_names[(1, 1)] == (b"m:v",)
    assert b"m:v" not in patched


def test_self_closing_array_follower_needs_no_cache_invalidation_edit():
    main = b"http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    follower = b'<c r="A1" s="1"/>'
    sheet = (
        b'<worksheet xmlns="' + main + b'"><sheetData><row r="1">' +
        follower +
        b'<c r="B1"><f t="array" ref="A1:B1">1+1</f><v>2</v></c>'
        b'</row></sheetData></worksheet>')
    scan = scan_sheet(sheet)

    edits = splice._formula_cache_invalidation_edits(
        Workbook().active, scan, sheet, {(1, 1), (1, 2)})

    assert edits[0][2] == follower
    assert b"<v>2</v>" not in edits[1][2]


def test_style_only_scalar_preserves_foreign_child_markup():
    workbook = Workbook()
    workbook.active["A1"] = 1
    source = io.BytesIO()
    workbook.save(source)
    foreign = (
        b'<mc:AlternateContent xmlns:mc="http://schemas.openxmlformats.org/'
        b'markup-compatibility/2006"><mc:Fallback/></mc:AlternateContent>')
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(source.getvalue())) as archive, \
            zipfile.ZipFile(output, "w") as destination:
        for info in archive.infolist():
            payload = archive.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                payload = payload.replace(b"<v>1</v>", b"<v>1</v>" + foreign)
            destination.writestr(info, payload)
    workbook = load_workbook(io.BytesIO(output.getvalue()), preserve=True)
    workbook.active["A1"].font = workbook.active["A1"].font.copy(bold=True)
    saved = io.BytesIO()
    workbook.save(saved)

    assert foreign in _sheet_xml(saved.getvalue())
