"""Phase 2c: the splice writer (CONVENTIONS §3.4; PR-0 D6/D7/D15).

Every assertion follows the reopen rule (save → reopen → assert) and the
part budget is checked literally. The ledger cross-check runs on every
preserve save in this suite (PAPER_LEDGER_CROSSCHECK=1 via conftest).
"""
from __future__ import annotations

import io
import time
import zipfile

import pytest

from openpyxl import load_workbook
from openpyxl.errors import PaperRefusal, UnsupportedStructureError
from openpyxl.package import diff_package

from .support.harness import assert_part_budget
from .support.partdiff import part_payloads

GAUNTLET = "gauntlet/gauntlet.xlsx"


def _model_sheet(path):
    for name, payload in part_payloads(path).items():
        if name.startswith("xl/worksheets/") and b"Quarterly Model" in payload:
            return name, payload
    raise AssertionError("no Model sheet found")


class TestNoOpRoundTrip:
    """The no-op invariant: load(preserve=True) + save == byte-identical
    part payloads, on every fixture class (CONVENTIONS §4)."""

    @pytest.mark.parametrize("fixture", [
        "gauntlet/gauntlet.xlsx",
        "features/lo_authored.xlsx",       # LO producer: sst, declarations
        "features/schedule_calc.xlsx",     # cached values
        "features/shared_formulas.xlsx",   # shared groups + array formula
        "features/macro_stub.xlsm",        # binary vba part
        "minimal/minimal_clean.xlsx",
    ])
    def test_noop_is_byte_identical(self, fixture_copy, tmp_path, fixture):
        src = fixture_copy(fixture)
        wb = load_workbook(src, preserve=True)
        out = str(tmp_path / ("noop" + fixture[-5:]))
        wb.save(out)
        d = diff_package(src, out)
        assert d.clean and not d.equivalent, d

    def test_noop_after_reads_and_materialization(self, fixture_copy, tmp_path):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        _ = wb["Model"]["Z99"]
        _ = wb["Model"].row_dimensions[55]
        for _row in wb["Model"].iter_rows():
            pass
        out = str(tmp_path / "noop.xlsx")
        wb.save(out)
        assert diff_package(src, out).clean


class TestSpliceCompletenessTrap:
    """THE signature test (CONVENTIONS §4): a one-cell edit on a sheet
    carrying sparklines, x14 CF, and a drawing reference — everything
    survives, and exactly one part changes."""

    def test_one_cell_edit_preserves_everything(self, fixture_copy, tmp_path):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb["Model"]["B8"] = 0.15
        out = str(tmp_path / "edit.xlsx")
        wb.save(out)

        sheet_name, _ = _model_sheet(src)
        assert_part_budget(src, out, expect_changed={sheet_name})

        _, sheet_after = _model_sheet(out)
        assert b"sparklineGroups" in sheet_after
        assert b"x14:conditionalFormattings" in sheet_after
        assert b"<x14:id>" in sheet_after            # CF twin pointer
        assert b"<drawing" in sheet_after            # chart attachment
        assert b"<legacyDrawing" in sheet_after

        wb2 = load_workbook(out)
        assert wb2["Model"]["B8"].value == 0.15
        assert len(wb2["Model"].merged_cells.ranges) == 1
        assert len(wb2["Model"].conditional_formatting) == 3
        assert wb2["Model"]["B6"].value == "=B3-B4-B5"

    @pytest.mark.lo_smoke
    def test_spliced_output_loads_in_libreoffice(self, fixture_copy, tmp_path, lo):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb["Model"]["B8"] = 0.15
        out = str(tmp_path / "edit.xlsx")
        wb.save(out)
        assert lo.lo_loads(out)


class TestCellEdits:

    def test_type_change_number_to_string(self, fixture_copy, tmp_path):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb["Model"]["B8"] = "TBD"
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert wb2["Model"]["B8"].value == "TBD"
        _, sheet = _model_sheet(out)
        assert b't="inlineStr"' in sheet             # PR-0 D1

    def test_new_cell_new_row_and_delete(self, fixture_copy, tmp_path):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb["Model"]["G3"] = 123                       # new cell, existing row
        wb["Model"]["B40"] = "=SUM(B3:B5)"            # new row
        del wb["Model"]["B9"]                         # delete a cell
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert wb2["Model"]["G3"].value == 123
        assert wb2["Model"]["B40"].value == "=SUM(B3:B5)"
        assert wb2["Model"]["B9"].value is None
        # untouched neighbours intact
        assert wb2["Model"]["B3"].value == 100
        assert wb2["Model"]["A9"].value == "Scenario"

    def test_formula_edit_without_calcchain(self, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Schedule"]["B2"] = 999
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert wb2["Schedule"]["B2"].value == 999
        assert wb2["Schedule"]["B12"].value == "=SUM(B2:B11)"

    def test_style_reuse_is_spliceable(self, fixture_copy, tmp_path):
        # assigning a style that already exists in the stylesheet works
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb["Model"]["B9"].style = "paper_input"       # existing named style
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert wb2["Model"]["B9"].style == "paper_input"



class TestSharedAndArrayFormulas:

    def test_shared_group_dissolves_on_touch(self, fixture_copy, tmp_path):
        src = fixture_copy("features/shared_formulas.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Calc"]["B3"] = 999                        # member of B2:B6 group
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        parts = part_payloads(out)
        sheet = next(p for n, p in parts.items()
                     if n.startswith("xl/worksheets/") and b"A2*2" in p)
        assert b't="shared"' not in sheet             # group dissolved whole
        wb2 = load_workbook(out)
        assert wb2["Calc"]["B3"].value == 999
        assert wb2["Calc"]["B4"].value == "=A4*2"     # follower kept meaning
        assert wb2["Calc"]["B6"].value == "=A6*2"
        # the array formula's bytes were never touched
        assert b't="array"' in sheet

    def test_untouched_shared_group_passes_through_verbatim(
            self, fixture_copy, tmp_path):
        src = fixture_copy("features/shared_formulas.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Calc"]["A10"] = "note"                    # outside the group
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        parts = part_payloads(out)
        sheet = next(p for n, p in parts.items()
                     if n.startswith("xl/worksheets/") and b"A2*2" in p)
        assert sheet.count(b'<f t="shared" si="0"/>') == 4   # untouched

    def test_array_formula_edit_refuses_atomically(self, fixture_copy, tmp_path):
        src = fixture_copy("features/shared_formulas.xlsx")
        with open(src, "rb") as f:
            before = f.read()
        wb = load_workbook(src, preserve=True)
        wb["Calc"]["D3"] = 5                          # inside D2:D4 array
        with pytest.raises(UnsupportedStructureError, match="in_spill"):
            wb.save(str(tmp_path / "o.xlsx"))
        with open(src, "rb") as f:
            assert f.read() == before


class TestRegionEdits:

    def test_merge_cells_updates_only_the_sheet(self, fixture_copy, tmp_path):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb["Model"].merge_cells("A14:C15")
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        sheet_name, _ = _model_sheet(src)
        assert_part_budget(src, out, expect_changed={sheet_name})
        wb2 = load_workbook(out)
        assert {"A1:F1", "A14:C15"} == {
            str(r) for r in wb2["Model"].merged_cells.ranges}
        _, sheet = _model_sheet(out)
        assert b"sparklineGroups" in sheet            # traps still intact

    def test_freeze_panes_change(self, fixture_copy, tmp_path):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb["Model"].freeze_panes = "B4"
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert wb2["Model"].freeze_panes == "B4"

    def test_column_width_change(self, fixture_copy, tmp_path):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb["Model"].column_dimensions["A"].width = 33.5
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert wb2["Model"].column_dimensions["A"].width == 33.5

    def test_row_height_change_keeps_cells(self, fixture_copy, tmp_path):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb["Model"].row_dimensions[3].height = 30
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert wb2["Model"].row_dimensions[3].height == 30
        assert wb2["Model"]["B3"].value == 100        # cells untouched

    def test_data_validation_add_on_plain_sheet(self, fixture_copy, tmp_path):
        from openpyxl.worksheet.datavalidation import DataValidation

        src = fixture_copy("features/datavalidation.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb.active
        dv = DataValidation(type="list", formula1='"Yes,No"')
        dv.add("C2")
        ws.add_data_validation(dv)
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert len(wb2.active.data_validations.dataValidation) == 3

    def test_new_region_inserted_at_schema_position(self, fixture_copy, tmp_path):
        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Sheet1"].merge_cells("F1:G2")             # sheet had no mergeCells
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert [str(r) for r in wb2["Sheet1"].merged_cells.ranges] == ["F1:G2"]

    def test_cf_twin_sync_sqref_patch_and_delete(self, fixture_copy,
                                                  tmp_path):
        # FLIPPED by v0.1 Batch 3 (was the blanket twin refusal): sqref
        # changes patch BOTH sides in lockstep; deleting a twin-bearing
        # rule removes its twin entry; untouched blocks stay verbatim
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        cfs = list(wb["Model"].conditional_formatting)
        databar = next(cf for cf in cfs
                       if any(r.type == "dataBar" for r in cf.rules))
        # sqref is the block's dict key upstream: re-key, never mutate
        cf_list = wb["Model"].conditional_formatting
        rules = cf_list._cf_rules.pop(databar)
        databar.sqref = "B6:F6"                    # sqref-only change
        cf_list._cf_rules[databar] = rules
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        _, sheet = _model_sheet(out)
        assert b'sqref="B6:F6"' in sheet           # classic side patched
        assert b"<xm:sqref>B6:F6</xm:sqref>" in sheet   # twin side synced
        assert b"<x14:id>" in sheet
        wb2 = load_workbook(out)
        assert any(str(cf.sqref) == "B6:F6"
                   for cf in wb2["Model"].conditional_formatting)

        # deletion: the twin entry goes with the rule
        wb3 = load_workbook(src, preserve=True)
        cf_list = wb3["Model"].conditional_formatting
        target = next(cf for cf in cf_list
                      if any(r.type == "dataBar" for r in cf.rules))
        twins_before = _model_sheet(src)[1].count(b"<x14:id>")
        cf_list._cf_rules.pop(target)
        out2 = str(tmp_path / "o2.xlsx")
        wb3.save(out2)
        _, sheet2 = _model_sheet(out2)
        assert sheet2.count(b"<x14:id>") == twins_before - 1
        wb4 = load_workbook(out2)
        assert len(wb4["Model"].conditional_formatting) == 2


class TestV0Refusals:
    """Operations outside the v0 write set refuse loudly and atomically —
    never a silent drop (PR-0 D8/D9/D15)."""

    def _assert_refuses(self, wb, src, tmp_path, match):
        with open(src, "rb") as f:
            before = f.read()
        out = str(tmp_path / "refused.xlsx")
        with pytest.raises(PaperRefusal, match=match):
            wb.save(out)
        import os
        assert not os.path.exists(out)
        with open(src, "rb") as f:
            assert f.read() == before

    def test_workbook_pr_change_refuses(self, fixture_copy, tmp_path):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb.code_name = "ThisWorkbookX"        # serializes into workbookPr
        self._assert_refuses(wb, src, tmp_path, "workbook-level")

    def test_comment_change_refuses(self, fixture_copy, tmp_path):
        from openpyxl.comments import Comment

        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb["Model"]["B9"].comment = Comment("new note", "tester")
        self._assert_refuses(wb, src, tmp_path, "comment")

    def test_data_only_save_refuses(self, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule_calc.xlsx")
        wb = load_workbook(src, preserve=True, data_only=True)
        self._assert_refuses(wb, src, tmp_path, "data_only")

    def test_style_registry_mutation_refuses(self, fixture_copy, tmp_path):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb["Model"]["B8"].fill.start_color.rgb = "FF12AB34"   # proxy leak
        wb["Model"]["A2"] = "trigger a save plan"
        self._assert_refuses(wb, src, tmp_path, "mutated in place")


class TestProducerGuards:

    def _surgery(self, fixture_copy, tmp_path, mutate_sheet):
        src = fixture_copy("minimal/minimal_clean.xlsx")
        out = str(tmp_path / "surgery.xlsx")
        with zipfile.ZipFile(src) as zin, zipfile.ZipFile(out, "w") as zout:
            for name in zin.namelist():
                payload = zin.read(name)
                if name.startswith("xl/worksheets/sheet"):
                    payload = mutate_sheet(payload)
                zout.writestr(name, payload)
        return out

    def test_ph_attribute_carried_over(self, fixture_copy, tmp_path):
        # PR-0 D6 carry rule: legal extra cell attributes survive the edit
        surgical = self._surgery(
            fixture_copy, tmp_path,
            lambda p: p.replace(b'<c r="B2"', b'<c r="B2" ph="1"', 1))
        wb = load_workbook(surgical, preserve=True)
        wb["Sheet1"]["B2"] = 42
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        parts = part_payloads(out)
        sheet = next(p for n, p in parts.items()
                     if n.startswith("xl/worksheets/"))
        assert b'ph="1"' in sheet
        assert load_workbook(out)["Sheet1"]["B2"].value == 42

    def test_cm_metadata_drops_on_overwrite(self, fixture_copy, tmp_path):
        # FLIPPED by v0.1 Batch 3 (was a refusal): a value overwrite ends
        # the cell's rich-value role; cm/vm never carry (battery job 21)
        surgical = self._surgery(
            fixture_copy, tmp_path,
            lambda p: p.replace(b'<c r="B2"', b'<c r="B2" cm="1"', 1))
        wb = load_workbook(surgical, preserve=True)
        wb["Sheet1"]["B2"] = 42
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        parts = part_payloads(out)
        sheet = next(p for n, p in parts.items()
                     if n.startswith("xl/worksheets/"))
        assert b'cm="1"' not in sheet
        assert load_workbook(out)["Sheet1"]["B2"].value == 42


    def test_rless_rows_refuse(self, fixture_copy, tmp_path):
        import re

        surgical = self._surgery(
            fixture_copy, tmp_path,
            lambda p: re.sub(br'<row r="\d+"', b"<row", p))
        wb = load_workbook(surgical, preserve=True)
        wb["Sheet1"]["B2"] = 42
        with pytest.raises(UnsupportedStructureError, match="no r attribute"):
            wb.save(str(tmp_path / "o.xlsx"))

    def test_attr_value_r_decoy_scans_true_column(
            self, fixture_copy, tmp_path):
        # a ' r="B9"' lookalike INSIDE another attribute's quoted value must
        # not hijack cell keying: openpyxl loads such files fine, and a
        # quote-blind r extraction would key the cell at the decoy's column
        # (silent duplicate-reference splice)
        surgical = self._surgery(
            fixture_copy, tmp_path,
            lambda p: p.replace(b'<c r="B2"',
                                b'<c r="B2" foo=\'x r="B9" y\'', 1))
        with zipfile.ZipFile(surgical) as z:
            payload = next(z.read(n) for n in z.namelist()
                           if n.startswith("xl/worksheets/sheet"))
        from openpyxl.preserve.xmlscan import scan_sheet

        scan = scan_sheet(payload)
        spans = [c for row in scan.rows.values()
                 for c in row.cells.values() if c.attrs.get("r") == "B2"]
        assert len(spans) == 1
        assert spans[0].column == 2

        wb = load_workbook(surgical, preserve=True)
        wb["Sheet1"]["B2"] = 42
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        sheet = next(p for n, p in part_payloads(out).items()
                     if n.startswith("xl/worksheets/"))
        assert sheet.count(b'<c r="B2"') == 1
        assert load_workbook(out)["Sheet1"]["B2"].value == 42

    def test_doctype_refuses(self, fixture_copy, tmp_path):
        surgical = self._surgery(
            fixture_copy, tmp_path,
            lambda p: b"<!DOCTYPE worksheet>" + p)
        wb = load_workbook(surgical, preserve=True)
        wb["Sheet1"]["B2"] = 42
        with pytest.raises(UnsupportedStructureError, match="DOCTYPE"):
            wb.save(str(tmp_path / "o.xlsx"))



# (region tag, self-closing original element, model edit, expected byte)
# every satellite region whose original can legally carry the self-closing
# form, both matrix arms exercised: no-op passthrough and model edit
_SELF_CLOSING_CASES = [
    ("sheetFormatPr", b'<sheetFormatPr defaultRowHeight="15"/>',
     lambda ws: setattr(ws.sheet_format, "defaultRowHeight", 22.5),
     b'defaultRowHeight="22.5"'),
    ("autoFilter", b'<autoFilter ref="A1:D5"/>',
     lambda ws: setattr(ws.auto_filter, "ref", "A1:D4"),
     b'ref="A1:D4"'),
    ("printOptions", b'<printOptions gridLines="1"/>',
     lambda ws: setattr(ws.print_options, "headings", True),
     b'headings="1"'),
    ("pageMargins", b'<pageMargins left="0.75" right="0.75" top="1" '
                    b'bottom="1" header="0.5" footer="0.5"/>',
     lambda ws: setattr(ws.page_margins, "left", 1.25),
     b'left="1.25"'),
    ("pageSetup", b'<pageSetup orientation="portrait"/>',
     lambda ws: setattr(ws.page_setup, "orientation", "landscape"),
     b'orientation="landscape"'),
    ("sheetProtection", b'<sheetProtection sheet="1" objects="1"/>',
     lambda ws: setattr(ws.protection, "formatCells", False),
     b'formatCells="0"'),
]

_REGION_INSERT_POINT = {
    # schema position: before which existing minimal_clean marker each
    # region element must be injected (CT_Worksheet sequence); None means
    # the fixture already carries the element in self-closing form
    "sheetFormatPr": None,
    "autoFilter": b"<pageMargins",
    "printOptions": b"<pageMargins",
    "pageMargins": None,
    "pageSetup": b"</worksheet>",
    "sheetProtection": b"<pageMargins",
}


class TestRegionSelfClosingMatrix:
    """PLAN-v0.1 §0.2: the region x self-closing matrix. The v0 scanner
    never set RegionSpan.end for self-closing top-level elements, so any
    splice touching one emitted malformed XML — silently."""

    def _build(self, fixture_copy, tmp_path, tag, element):
        src = fixture_copy("minimal/minimal_clean.xlsx")
        out = str(tmp_path / "{0}_sc.xlsx".format(tag))
        marker = _REGION_INSERT_POINT[tag]
        with zipfile.ZipFile(src) as zin, zipfile.ZipFile(out, "w") as zout:
            for name in zin.namelist():
                payload = zin.read(name)
                if name.startswith("xl/worksheets/sheet") \
                        and marker is not None:
                    assert marker in payload
                    payload = payload.replace(marker, element + marker, 1)
                zout.writestr(name, payload)
        return out

    @pytest.mark.parametrize(
        "tag,element,edit,expect",
        _SELF_CLOSING_CASES, ids=[c[0] for c in _SELF_CLOSING_CASES])
    def test_edit_of_self_closing_original(
            self, fixture_copy, tmp_path, tag, element, edit, expect):
        import xml.etree.ElementTree as ET

        src = self._build(fixture_copy, tmp_path, tag, element)
        wb = load_workbook(src, preserve=True)
        edit(wb["Sheet1"])
        out = str(tmp_path / "{0}_out.xlsx".format(tag))
        wb.save(out)
        sheet = next(p for n, p in part_payloads(out).items()
                     if n.startswith("xl/worksheets/"))
        ET.fromstring(sheet)                                  # well-formed
        assert sheet.count(b"<" + tag.encode()) == 1
        assert expect in sheet
        load_workbook(out)                                    # reloadable

    @pytest.mark.parametrize(
        "tag,element,edit,expect",
        _SELF_CLOSING_CASES, ids=[c[0] for c in _SELF_CLOSING_CASES])
    def test_noop_save_of_self_closing_original_is_byte_identical(
            self, fixture_copy, tmp_path, tag, element, edit, expect):
        src = self._build(fixture_copy, tmp_path, tag, element)
        wb = load_workbook(src, preserve=True)
        out = str(tmp_path / "{0}_noop.xlsx".format(tag))
        wb.save(out)
        assert part_payloads(src) == part_payloads(out)

    def test_removal_of_self_closing_original(self, fixture_copy, tmp_path):
        # the third matrix arm (gate finding: untested): the model renders
        # the region to None -> the splice excises exactly the element
        # bytes — a regressed end=None here resumes whole-document
        # corruption
        import xml.etree.ElementTree as ET

        src = self._build(fixture_copy, tmp_path, "autoFilter",
                          b'<autoFilter ref="A1:D5"/>')
        wb = load_workbook(src, preserve=True)
        wb["Sheet1"].auto_filter.ref = None
        out = str(tmp_path / "af_removed.xlsx")
        wb.save(out)
        sheet = next(p for n, p in part_payloads(out).items()
                     if n.startswith("xl/worksheets/"))
        ET.fromstring(sheet)
        assert b"<autoFilter" not in sheet
        assert sheet.count(b"<pageMargins") == 1      # neighbour intact
        assert load_workbook(out)["Sheet1"]["B2"].value is not None


class TestSelfClosingSheetData:
    """The 0.2 fix also repairs the self-closing <sheetData/> expansion
    path (splice.py builds its edit from span.end, which was None before
    the fix — the same document-duplication corruption)."""

    def test_cell_add_into_self_closing_sheetdata(
            self, fixture_copy, tmp_path):
        import xml.etree.ElementTree as ET

        src = fixture_copy("minimal/minimal_clean.xlsx")
        surgical = str(tmp_path / "sc_sheetdata.xlsx")
        with zipfile.ZipFile(src) as zin, \
                zipfile.ZipFile(surgical, "w") as zout:
            for name in zin.namelist():
                payload = zin.read(name)
                if name.startswith("xl/worksheets/sheet"):
                    i = payload.find(b"<sheetData>")
                    j = payload.find(b"</sheetData>") + len(b"</sheetData>")
                    payload = payload[:i] + b"<sheetData/>" + payload[j:]
                zout.writestr(name, payload)
        wb = load_workbook(surgical, preserve=True)
        wb["Sheet1"]["A1"] = 42
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        sheet = next(p for n, p in part_payloads(out).items()
                     if n.startswith("xl/worksheets/"))
        ET.fromstring(sheet)
        assert sheet.count(b"<sheetData") == 1
        assert load_workbook(out)["Sheet1"]["A1"].value == 42


class TestImpureSerializerPinning:
    """PLAN-v0.1 §0.3: a region whose serializer disagrees with itself at
    arm time is PINNED — no-op saves keep the original bytes (never false
    dirty), and USER edits to it refuse rather than splice an
    untrustworthy render. The one real instance (sheetFormatPr's outline
    sync reading DimensionHolder's render-time side effect) was fixed by
    making the render pure, so on the shipped corpus NOTHING pins; the
    guard stays armed for the next impure upstream serializer and is
    proven here with a synthetic one."""

    def test_no_region_pins_anywhere_on_the_corpus(self, fixture_copy):
        from .test_properties import ALL_LOADABLE

        for fixture in ALL_LOADABLE:
            wb = load_workbook(fixture_copy(fixture), preserve=True)
            for ws, pinned in wb._paper_ledger.pinned_regions.items():
                assert not pinned, (fixture, ws.title, pinned)

    def _make_impure(self, monkeypatch):
        from openpyxl.preserve import regions as regions_mod

        state = {"calls": 0}

        def impure(ws):
            # settles after the first render, like the DimensionHolder
            # instance did: render 1 disagrees with every later render
            from openpyxl.xml.functions import Element

            state["calls"] += 1
            el = Element("printOptions")
            if state["calls"] > 1:
                el.set("gridLines", "1")
            return el

        replacement = regions_mod.Region("printOptions", impure)
        patched = [replacement if r.tag == "printOptions" else r
                   for r in regions_mod.SPLICEABLE_REGIONS]
        monkeypatch.setattr(regions_mod, "SPLICEABLE_REGIONS", patched)
        monkeypatch.setitem(regions_mod.REGION_BY_TAG, "printOptions",
                            replacement)
        return state

    def test_synthetic_impure_serializer_pins_and_noop_stays_identical(
            self, fixture_copy, tmp_path, monkeypatch):
        self._make_impure(monkeypatch)
        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        assert "printOptions" in \
            wb._paper_ledger.pinned_regions[wb["Sheet1"]]
        out = str(tmp_path / "noop.xlsx")
        wb.save(out)
        assert part_payloads(src) == part_payloads(out)

    def test_synthetic_impure_serializer_refuses_edits(
            self, fixture_copy, tmp_path, monkeypatch):
        state = self._make_impure(monkeypatch)
        src = fixture_copy("minimal/minimal_clean.xlsx")
        with open(src, "rb") as f:
            before = f.read()
        wb = load_workbook(src, preserve=True)
        state["calls"] = -2          # next renders disagree with settled arm
        with pytest.raises(UnsupportedStructureError, match="impure"):
            wb.save(str(tmp_path / "pinned.xlsx"))
        with open(src, "rb") as f:
            assert f.read() == before

    def test_col_width_edit_does_not_drift_sheetformatpr(
            self, fixture_copy, tmp_path):
        # the gate's drift repro: a width-only edit must not rewrite
        # sheetFormatPr (no outlineLevelCol="0" appearing, no unmodeled
        # attribute loss)
        import re

        src = fixture_copy("features/hidden.xlsx")
        sheet_before = next(p for n, p in part_payloads(src).items()
                            if b"HiddenNotes" not in p
                            and n.startswith("xl/worksheets/"))
        fmt_before = re.search(rb"<sheetFormatPr[^>]*>", sheet_before)
        wb = load_workbook(src, preserve=True)
        wb["Visible"].column_dimensions["B"].width = 30
        out = str(tmp_path / "width.xlsx")
        wb.save(out)
        sheet_after = next(p for n, p in part_payloads(out).items()
                           if b"HiddenNotes" not in p
                           and n.startswith("xl/worksheets/"))
        fmt_after = re.search(rb"<sheetFormatPr[^>]*>", sheet_after)
        assert fmt_after.group(0) == fmt_before.group(0)
        wb2 = load_workbook(out)
        assert wb2["Visible"].column_dimensions["B"].width == 30
        assert wb2["Visible"].column_dimensions["C"].hidden is True

    def test_sheetformatpr_edit_on_cols_sheet_now_works(
            self, fixture_copy, tmp_path):
        # refused under the interim pin; the pure render makes it splice
        src = fixture_copy("features/hidden.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Visible"].sheet_format.defaultRowHeight = 22.5
        out = str(tmp_path / "fmt.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert wb2["Visible"].sheet_format.defaultRowHeight == 22.5

    def test_column_grouping_works_on_cols_sheet(
            self, fixture_copy, tmp_path):
        # the gate's collateral repro: grouping refused under the interim
        # pin; now splices, with the outline sync landing in BOTH cols and
        # sheetFormatPr exactly as a stock save would write them
        src = fixture_copy("features/hidden.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Visible"].column_dimensions.group("E", "F", outline_level=1)
        out = str(tmp_path / "grouped.xlsx")
        wb.save(out)
        sheet = next(p for n, p in part_payloads(out).items()
                     if b"outlineLevel" in p and n.startswith("xl/worksheets/"))
        assert b'outlineLevelCol="1"' in sheet
        wb2 = load_workbook(out)
        assert wb2["Visible"].column_dimensions["E"].outlineLevel == 1

    def test_explicit_outlinelevelcol_edit_matches_stock_semantics(
            self, fixture_copy, tmp_path):
        # outlineLevelCol is derived metadata: on a cols-bearing sheet the
        # writer's sync owns it (stock parity — the assignment is
        # normalized away); on a cols-free sheet it lands verbatim
        src = fixture_copy("features/hidden.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Visible"].sheet_format.outlineLevelCol = 5
        out = str(tmp_path / "derived.xlsx")
        wb.save(out)
        assert b'outlineLevelCol="5"' not in part_payloads(out)[
            "xl/worksheets/sheet1.xml"]

        src2 = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src2, preserve=True)
        wb["Sheet1"].sheet_format.outlineLevelCol = 5
        out2 = str(tmp_path / "derived2.xlsx")
        wb.save(out2)
        sheet = next(p for n, p in part_payloads(out2).items()
                     if n.startswith("xl/worksheets/"))
        assert b'outlineLevelCol="5"' in sheet


class TestSaveTargets:

    def test_save_to_file_like_target(self, fixture_copy):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb["Model"]["B8"] = 0.2
        buf = io.BytesIO(b"pre-existing garbage that must be truncated away")
        wb.save(buf)
        buf.seek(0)
        wb2 = load_workbook(buf)
        assert wb2["Model"]["B8"].value == 0.2


class TestPerformanceGuardrail:
    """PR-0 D4 (as amended in Phase 2c): preserve save within 2x stock save
    on the large fixture. Measured 1.82x-1.87x; see PR0-API-PROPOSAL.md D4
    for the amendment evidence."""

    def test_splice_save_within_budget(self, fixture_copy, tmp_path, monkeypatch):
        monkeypatch.delenv("PAPER_LEDGER_CROSSCHECK", raising=False)
        src = fixture_copy("large/large150k.xlsx")

        def best_of(n, fn):
            times = []
            for _ in range(n):
                t0 = time.perf_counter()
                fn()
                times.append(time.perf_counter() - t0)
            return min(times)

        wb_stock = load_workbook(src)
        t_stock = best_of(2, lambda: wb_stock.save(str(tmp_path / "stock.xlsx")))

        wb = load_workbook(src, preserve=True)
        wb["Big"]["A2"] = 424242
        t_preserve = best_of(2, lambda: wb.save(str(tmp_path / "preserve.xlsx")))

        assert load_workbook(str(tmp_path / "preserve.xlsx"))["Big"]["A2"].value == 424242
        assert t_preserve <= 2.0 * t_stock, (
            "splice save {0:.3f}s exceeded 2x stock save {1:.3f}s".format(
                t_preserve, t_stock))
