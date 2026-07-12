import io
import os
import stat
import struct
import zipfile
from types import SimpleNamespace

import pytest

from openpyxl import Workbook, load_workbook
from openpyxl.errors import (
    HandleRebindWarning,
    RelationshipPolicyError,
    UnsupportedStructureError,
)
from openpyxl.formatting.rule import CellIsRule
from openpyxl.comments import Comment
from openpyxl.preserve.inventory import scan_archive
from openpyxl.preserve.splice import SpliceRefusal, resolve_dirty_cells
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle


def _preserved(tmp_path, setup=None):
    path = tmp_path / "source.xlsx"
    wb = Workbook()
    wb.active["A1"] = "seed"
    if setup is not None:
        setup(wb)
    wb.save(path)
    return load_workbook(path, preserve=True), path


def _sheet_xml(path):
    with zipfile.ZipFile(path) as archive:
        return archive.read("xl/worksheets/sheet1.xml")


def _central_entry_offset(data, target):
    offset = 0
    while True:
        offset = data.find(b"PK\x01\x02", offset)
        if offset < 0:
            raise AssertionError("central entry not found: {0}".format(target))
        name_len, extra_len, comment_len = struct.unpack(
            "<HHH", data[offset + 28:offset + 34])
        name = bytes(data[offset + 46:offset + 46 + name_len]).decode(
            "utf-8")
        if name == target:
            return offset
        offset += 46 + name_len + extra_len + comment_len


def _rewrite_package(source, target, replacements, additions=()):
    with zipfile.ZipFile(source) as zin, zipfile.ZipFile(target, "w") as zout:
        for info in zin.infolist():
            zout.writestr(info, replacements.get(info.filename,
                                                  zin.read(info)))
        for name, payload in additions:
            zout.writestr(name, payload)


def test_dirty_sheet_dimension_tracks_add_delete_and_read_only(tmp_path):
    wb, _source = _preserved(tmp_path)
    ws = wb.active
    ws["Z40"] = "edge"
    grown = tmp_path / "grown.xlsx"
    wb.save(grown)

    assert load_workbook(grown).active.dimensions == "A1:Z40"
    ro = load_workbook(grown, read_only=True)
    assert ro.active.calculate_dimension() == "A1:Z40"
    ro.close()

    wb2 = load_workbook(grown, preserve=True)
    del wb2.active["Z40"]
    shrunk = tmp_path / "shrunk.xlsx"
    wb2.save(shrunk)
    assert load_workbook(shrunk).active.dimensions == "A1:A1"
    ro = load_workbook(shrunk, read_only=True)
    assert ro.active.calculate_dimension() == "A1:A1"
    ro.close()


def test_noop_save_retains_original_dimension_bytes(tmp_path):
    wb, source = _preserved(tmp_path)
    before = _sheet_xml(source)
    out = tmp_path / "noop.xlsx"
    wb.save(out)
    assert _sheet_xml(out) == before


def test_array_and_shared_formula_ranges_do_not_expand_the_grid():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "=1"
    scan = SimpleNamespace(
        array_refs=["A1:XFD1048576"],
        shared_groups={},
        shared_members={},
        rows={},
    )

    with pytest.raises(SpliceRefusal, match="array/spill range"):
        resolve_dirty_cells(ws, {(1, 1)}, scan)

    scan.array_refs = []
    scan.shared_groups = {0: "A1:XFD1048576"}
    scan.shared_members = {0: {(1, 1)}}
    assert resolve_dirty_cells(ws, {(1, 1)}, scan) == {(1, 1)}


def test_validate_restores_calculation_dxf_and_hyperlink_ids(tmp_path):
    red = PatternFill(fill_type="solid", fgColor="FF0000")

    def setup(wb):
        ws = wb.active
        ws["B1"] = "=A1"
        ws["C1"].hyperlink = "https://example.test"
        ws.conditional_formatting.add("A1", CellIsRule(
            operator="equal", formula=["1"], fill=red))

    wb, _source = _preserved(tmp_path, setup)
    ws = wb.active
    ws["A1"] = "changed"
    rule = next(iter(ws.conditional_formatting)).rules[0]
    rule.dxf = DifferentialStyle(fill=PatternFill(
        fill_type="solid", fgColor="00FF00"))
    before_dxfs = list(wb._differential_styles.styles)
    before_dxf_id = rule.dxfId
    before_id = ws["C1"].hyperlink.id
    before_calc = wb.calculation.fullCalcOnLoad
    before_cells = {sheet: set(coords)
                    for sheet, coords in wb._paper_ledger.cells.items()}

    wb.validate()

    assert list(wb._differential_styles.styles) == before_dxfs
    assert rule.dxfId == before_dxf_id
    assert ws["C1"].hyperlink.id == before_id
    assert wb.calculation.fullCalcOnLoad == before_calc
    assert wb._paper_ledger.cells == before_cells


def test_validate_restores_added_sheet_writer_state(tmp_path):
    wb, _source = _preserved(tmp_path)
    ws = wb.create_sheet("Added")
    ws["A1"] = "linked"
    ws["A1"].hyperlink = "https://example.test"
    ws["B2"].comment = Comment("note", "paper")
    ws.column_dimensions.group("B", "D", outline_level=2)
    original_links = ws._hyperlinks
    original_comments = ws._comments
    original_outline = ws.sheet_format.outlineLevelCol

    wb.validate()

    assert ws._hyperlinks is original_links
    assert ws._comments is original_comments
    assert ws._hyperlinks == []
    assert ws._comments == []
    assert ws.sheet_format.outlineLevelCol == original_outline
    assert ws["A1"].hyperlink.id is None


def test_late_save_refusal_rolls_back_planner_mutations(tmp_path):
    red = PatternFill(fill_type="solid", fgColor="FF0000")

    def setup(wb):
        ws = wb.active
        ws["B1"] = "=A1"
        ws["C1"].hyperlink = "https://example.test"
        ws.conditional_formatting.add("A1", CellIsRule(
            operator="equal", formula=["1"], fill=red))

    wb, _source = _preserved(tmp_path, setup)
    ws = wb.active
    ws["A1"] = "changed"
    rule = next(iter(ws.conditional_formatting)).rules[0]
    rule.dxf = DifferentialStyle(fill=PatternFill(
        fill_type="solid", fgColor="00FF00"))
    ws["C1"].hyperlink = None
    before_dxfs = list(wb._differential_styles.styles)
    before_dxf_id = rule.dxfId
    before_calc = wb.calculation.fullCalcOnLoad
    before_cells = {sheet: set(coords)
                    for sheet, coords in wb._paper_ledger.cells.items()}

    with pytest.raises(RelationshipPolicyError):
        wb.save(tmp_path / "refused.xlsx")

    assert list(wb._differential_styles.styles) == before_dxfs
    assert rule.dxfId == before_dxf_id
    assert wb.calculation.fullCalcOnLoad == before_calc
    assert wb._paper_ledger.cells == before_cells
    assert not (tmp_path / "refused.xlsx").exists()


def test_strict_hyperlink_refusal_does_not_attach_link(tmp_path):
    wb, _source = _preserved(tmp_path)
    ws = wb.active
    ws.protection.sheet = True
    wb.strict_protection = True
    cell = ws["D4"]
    dirty = set(wb._paper_ledger.dirty_coordinates(ws))

    with pytest.raises(UnsupportedStructureError):
        cell.hyperlink = "https://example.test"

    assert cell.hyperlink is None
    assert cell.value is None
    assert wb._paper_ledger.dirty_coordinates(ws) == dirty


class _PersistentFailureStream(io.BytesIO):
    def __init__(self, initial):
        super().__init__(initial)
        self.calls = 0

    def write(self, data):
        self.calls += 1
        super().write(bytes(data[:1]))
        raise OSError("persistent injected write failure")


def test_preserve_save_replaces_exact_bytesio(tmp_path):
    wb, _source = _preserved(tmp_path)
    wb.active["B2"] = "saved"

    target = io.BytesIO(b"old trailing bytes")
    wb.save(target)

    target.seek(0)
    assert load_workbook(target).active["B2"].value == "saved"


def test_preserve_save_rebinds_path_backed_buffered_random(tmp_path):
    wb, source = _preserved(tmp_path)
    wb.active["B2"] = "saved"

    with open(source, "r+b") as target:
        original_name = target.name
        wb.save(target)
        assert not target.closed
        assert target.name == original_name
        target.seek(0)
        assert load_workbook(target).active["B2"].value == "saved"


def test_path_backed_replace_failure_reopens_original(
        tmp_path, monkeypatch):
    from openpyxl.preserve import zipio

    wb, source = _preserved(tmp_path)
    wb.active["B2"] = "unsaved"
    original = source.read_bytes()

    def fail_replace(_source, _target):
        raise OSError("injected replace failure")

    monkeypatch.setattr(zipio, "_posix_exchange", fail_replace)
    with open(source, "r+b") as target:
        target.seek(7)
        with pytest.raises(OSError, match="injected replace failure"):
            wb.save(target)
        assert not target.closed
        assert target.tell() == 7
        target.seek(0)
        assert target.read() == original
    assert source.read_bytes() == original


@pytest.mark.skipif(os.name == "nt", reason="POSIX descriptor rebinding")
def test_path_backed_rebind_preflight_failure_does_not_commit(
        tmp_path, monkeypatch):
    from openpyxl.preserve import zipio

    wb, source = _preserved(tmp_path)
    wb.active["B2"] = "unsaved"
    original = source.read_bytes()
    real_fileio = zipio.io.FileIO

    def fail_temp(path, *args, **kwargs):
        if str(path).endswith(".tmp"):
            raise OSError("injected descriptor preflight failure")
        return real_fileio(path, *args, **kwargs)

    monkeypatch.setattr(zipio.io, "FileIO", fail_temp)
    with open(source, "r+b") as target:
        target.seek(11)
        with pytest.raises(OSError, match="descriptor preflight"):
            wb.save(target)
        assert not target.closed
        assert target.tell() == 11
        target.seek(0)
        assert target.read() == original
    assert source.read_bytes() == original


def test_windows_post_commit_rebind_failure_warns_instead_of_raising(
        tmp_path, monkeypatch):
    from openpyxl.preserve import zipio

    destination = tmp_path / "destination.xlsx"
    replacement = tmp_path / "replacement.tmp"
    destination.write_bytes(b"old")
    replacement.write_bytes(b"new")

    def fail_reopen(*_args, **_kwargs):
        raise OSError("injected reopen failure")

    monkeypatch.setattr(zipio, "_reopen_buffered_random", fail_reopen)
    with open(destination, "r+b") as target:
        with pytest.warns(HandleRebindWarning, match="saved correctly"):
            zipio._replace_for_open_path_windows(
                str(replacement), str(destination), target,
                str(destination))
        assert target.closed
    assert destination.read_bytes() == b"new"


def test_receipt_failure_precedes_destination_delivery(tmp_path, monkeypatch):
    from openpyxl.preserve import receipts

    wb, _source = _preserved(tmp_path)
    wb.active["B2"] = "unsaved"
    destination = tmp_path / "destination.xlsx"
    destination.write_bytes(b"original destination")

    def fail_receipt(*_args, **_kwargs):
        raise UnsupportedStructureError("injected receipt failure")

    monkeypatch.setattr(receipts, "receipt", fail_receipt)
    with pytest.raises(UnsupportedStructureError, match="receipt failure"):
        wb.save(destination, receipt=True)
    assert destination.read_bytes() == b"original destination"


def test_preserve_load_refuses_ascii_case_colliding_part_names(tmp_path):
    source = tmp_path / "source.xlsx"
    crafted = tmp_path / "case-collision.xlsx"
    Workbook().save(source)
    with zipfile.ZipFile(source) as zin:
        workbook_xml = zin.read("xl/workbook.xml")
    _rewrite_package(
        source, crafted, {}, additions=[("XL/WORKBOOK.XML", workbook_xml)])

    with pytest.raises(UnsupportedStructureError, match="case-colliding"):
        load_workbook(crafted, preserve=True)


def test_opc_name_check_is_ascii_only_not_unicode_casefold(tmp_path):
    source = tmp_path / "source.xlsx"
    crafted = tmp_path / "unicode-names.xlsx"
    Workbook().save(source)
    _rewrite_package(source, crafted, {}, additions=[
        ("custom/Ä.bin", b"upper"),
        ("custom/ä.bin", b"lower"),
    ])

    wb = load_workbook(crafted, preserve=True)
    output = tmp_path / "unicode-output.xlsx"
    wb.save(output)
    with zipfile.ZipFile(output) as archive:
        assert archive.read("custom/Ä.bin") == b"upper"
        assert archive.read("custom/ä.bin") == b"lower"


def test_preserve_load_refuses_forged_crc_before_model_construction(tmp_path):
    source = tmp_path / "source.xlsx"
    Workbook().save(source)
    data = bytearray(source.read_bytes())
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        info = archive.getinfo("docProps/core.xml")
    central = _central_entry_offset(data, info.filename)
    wrong_crc = (info.CRC ^ 0xFFFFFFFF) & 0xFFFFFFFF
    data[info.header_offset + 14:info.header_offset + 18] = struct.pack(
        "<L", wrong_crc)
    data[central + 16:central + 20] = struct.pack("<L", wrong_crc)
    crafted = tmp_path / "forged-crc.xlsx"
    crafted.write_bytes(data)

    with pytest.raises(UnsupportedStructureError, match="CRC"):
        load_workbook(crafted, preserve=True)


def test_preserve_load_refuses_encrypted_flag_before_model_construction(
        tmp_path):
    source = tmp_path / "source.xlsx"
    Workbook().save(source)
    data = bytearray(source.read_bytes())
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        info = archive.getinfo("docProps/core.xml")
    central = _central_entry_offset(data, info.filename)
    local_flags = struct.unpack(
        "<H", data[info.header_offset + 6:info.header_offset + 8])[0] | 1
    central_flags = struct.unpack(
        "<H", data[central + 8:central + 10])[0] | 1
    data[info.header_offset + 6:info.header_offset + 8] = struct.pack(
        "<H", local_flags)
    data[central + 8:central + 10] = struct.pack("<H", central_flags)
    crafted = tmp_path / "encrypted-flag.xlsx"
    crafted.write_bytes(data)

    with pytest.raises(UnsupportedStructureError, match="ZIP-encrypted"):
        load_workbook(crafted, preserve=True)


def test_preserve_load_refuses_missing_workbook_sheet_target(tmp_path):
    source = tmp_path / "source.xlsx"
    crafted = tmp_path / "missing-sheet.xlsx"
    wb = Workbook()
    wb.create_sheet("Second")
    wb.save(source)
    with zipfile.ZipFile(source) as archive:
        rels_name = "xl/_rels/workbook.xml.rels"
        rels = archive.read(rels_name).replace(
            b"/xl/worksheets/sheet2.xml", b"/xl/worksheets/missing.xml")
    _rewrite_package(source, crafted, {rels_name: rels})

    with pytest.raises(UnsupportedStructureError, match="targets missing part"):
        load_workbook(crafted, preserve=True)


@pytest.mark.parametrize("fixture", [
    "corrupt/truncated.xlsx",
    "corrupt/not_a_zip.xlsx",
])
def test_corrupt_packages_raise_typed_refusals(fixture, fixture_copy):
    with pytest.raises(UnsupportedStructureError):
        load_workbook(fixture_copy(fixture), preserve=True)


def test_preserve_source_read_is_bounded_and_cursor_safe():
    raw = io.BytesIO()
    Workbook().save(raw)

    class NoUnboundedRead(io.BytesIO):
        def read(self, size=-1):
            assert size >= 0
            return super().read(size)

    source = NoUnboundedRead(raw.getvalue())
    source.seek(7)
    assert load_workbook(source, preserve=True).sheetnames == ["Sheet"]
    assert source.tell() == 7


def test_persistent_failure_stream_refused_without_mutation(tmp_path):
    wb, _source = _preserved(tmp_path)
    wb.active["B1"] = "=A1"
    wb.calculation.fullCalcOnLoad = False

    original = b"destination bytes"
    failed = _PersistentFailureStream(original)
    failed.seek(5)

    with pytest.raises(TypeError, match="exact io.BytesIO"):
        wb.save(failed)

    assert failed.getvalue() == original
    assert failed.tell() == 5
    assert failed.calls == 0
    assert wb.calculation.fullCalcOnLoad is False


def test_exported_bytesio_refused_without_mutation(tmp_path):
    wb, _source = _preserved(tmp_path)

    exported = io.BytesIO(b"destination bytes")
    exported.seek(5)
    view = exported.getbuffer()
    try:
        with pytest.raises(TypeError, match="no exported buffer views"):
            wb.save(exported)
        assert bytes(view) == b"destination bytes"
        assert exported.tell() == 5
    finally:
        view.release()


def test_stock_save_keeps_custom_filelike_compatibility():
    class CustomBuffer(io.BytesIO):
        pass

    wb = Workbook()
    wb.active["A1"] = "stock"
    target = CustomBuffer()

    wb.save(target)

    target.seek(0)
    assert load_workbook(target).active["A1"].value == "stock"


def test_nonseekable_destination_refused_before_model_mutation(tmp_path):
    wb, _source = _preserved(tmp_path)
    wb.active["B1"] = "=A1"
    wb.calculation.fullCalcOnLoad = False

    class WriteOnly:
        def write(self, data):
            raise AssertionError("must not write")

    with pytest.raises(TypeError, match="exact io.BytesIO"):
        wb.save(WriteOnly())
    assert wb.calculation.fullCalcOnLoad is False


@pytest.mark.skipif(os.name == "nt", reason="POSIX modes and symlinks")
def test_path_save_preserves_mode_and_symlink(tmp_path):
    wb, source = _preserved(tmp_path)
    os.chmod(source, 0o640)
    wb.active["B2"] = "saved"
    wb.save(source)
    assert stat.S_IMODE(os.stat(source).st_mode) == 0o640

    target = tmp_path / "target.xlsx"
    target.write_bytes(source.read_bytes())
    os.chmod(target, 0o604)
    link = tmp_path / "linked.xlsx"
    link.symlink_to(target.name)
    wb.active["C3"] = "through-link"
    wb.save(link)
    assert link.is_symlink()
    assert stat.S_IMODE(os.stat(target).st_mode) == 0o604
    assert load_workbook(target).active["C3"].value == "through-link"


def test_unknown_relationship_is_in_stock_loss_inventory(tmp_path):
    path = tmp_path / "unknown.xlsx"
    rels = b'''<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
      <Relationship Id="rId1" Type="http://schemas.microsoft.com/office/2006/relationships/ui/extensibility" Target="customUI/customUI.xml"/>
    </Relationships>'''
    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr("_rels/.rels", rels)
        archive.writestr("customUI/customUI.xml", b"<customUI/>")
    with zipfile.ZipFile(path) as archive:
        inventory = scan_archive(archive, archive.namelist())
    assert "unmodeled-opc" in inventory.kinds()
    assert inventory.losses[0]["location"] == "customUI/customUI.xml"


@pytest.mark.parametrize("attribute", ["quotePrefix", "pivotButton"])
def test_public_style_flags_are_saved(attribute, tmp_path):
    wb, _source = _preserved(tmp_path)
    setattr(wb.active["A1"], attribute, True)
    out = tmp_path / (attribute + ".xlsx")
    wb.save(out)
    assert getattr(load_workbook(out).active["A1"], attribute) is True
