"""Adversarial publication and in-session composition regressions."""
from __future__ import annotations

import os
import re
import zipfile
from copy import copy
from datetime import date

import pytest

from openpyxl import load_workbook
from openpyxl.errors import BoundaryViolationError
from openpyxl.errors import UnsupportedStructureError
from openpyxl.pivot.api_types import (
    PivotAxisField,
    PivotItemFilter,
    PivotMeasure,
)
from openpyxl.worksheet.table import Table

from .support.harness import save_and_reopen
from .support.partdiff import part_payloads
from .test_pivot_create_package import _create_by_region
from .test_pivot_refresh import _preserved_matrix


_TABLE = "features/tables.xlsx"
_CACHE = "xl/pivotCache/pivotCacheDefinition1.xml"
_RECORDS = "xl/pivotCache/pivotCacheRecords1.xml"
_PIVOT = "xl/pivotTables/pivotTable1.xml"


@pytest.mark.parametrize("mutation", ["value", "style"])
def test_direct_staged_output_edit_refuses_save(fixture_copy, tmp_path, mutation):
    source = fixture_copy(_TABLE)
    workbook = load_workbook(source, preserve=True)
    _create_by_region(workbook["Data"])
    output = workbook["Data"]["F4"]
    if mutation == "value":
        output.value = 999
    else:
        output.number_format = "$#,##0"

    destination = str(tmp_path / (mutation + ".xlsx"))
    with pytest.raises(BoundaryViolationError) as exc:
        workbook.save(destination)
    assert exc.value.kind == "pivot-output-collision"
    assert not os.path.exists(destination)


def test_direct_reopened_output_edit_refuses_save(fixture_copy, tmp_path):
    source = fixture_copy(_TABLE)
    workbook = load_workbook(source, preserve=True)
    _create_by_region(workbook["Data"])
    created = str(tmp_path / "created.xlsx")
    workbook = save_and_reopen(workbook, created, preserve=True)
    workbook["Data"]["F4"] = True

    destination = str(tmp_path / "edited.xlsx")
    with pytest.raises(BoundaryViolationError) as exc:
        workbook.save(destination)
    assert exc.value.kind == "pivot-output-collision"
    assert not os.path.exists(destination)
    assert workbook["Data"].pivots["ByRegion"].capabilities.can_delete is False
    assert workbook["Data"].pivots["ByRegion"].capabilities.can_edit_layout \
        is False


def test_unsaved_create_delete_refuses_to_erase_later_blank_slot(tmp_path):
    _source, workbook = _preserved_matrix(
        tmp_path,
        ["Region", "Amount"],
        [["East", 10], ["West", 7]],
        table="Sales",
    )
    worksheet = workbook["Summary"]
    pivot = worksheet.pivots.create(
        name="ByRegion",
        source="Sales",
        destination="B4",
        rows=[],
        columns=["Region"],
        values=["Amount"],
    )
    assert worksheet["B5"].value is None
    worksheet["B5"] = "keep me"

    with pytest.raises(BoundaryViolationError) as exc:
        pivot.delete()
    assert exc.value.kind == "pivot-output-collision"
    assert worksheet["B5"].value == "keep me"


def test_create_then_delete_restores_custom_format_registry(tmp_path):
    source, workbook = _preserved_matrix(
        tmp_path,
        ["Region", "Amount"],
        [["East", 10], ["West", 7]],
        table="Sales",
    )
    pivot = workbook["Summary"].pivots.create(
        name="ByRegion", source="Sales", destination="A1",
        rows=["Region"],
        values=[PivotMeasure("Amount", number_format="USD #,##0")],
    )
    assert "USD #,##0" in workbook._number_formats
    pivot.delete()
    assert "USD #,##0" not in workbook._number_formats

    destination = str(tmp_path / "cancelled-format.xlsx")
    workbook.save(destination)
    assert part_payloads(destination) == part_payloads(source)


def test_refresh_then_rename_keeps_cache_and_output_in_sync(fixture_copy, tmp_path):
    source = fixture_copy(_TABLE)
    workbook = load_workbook(source, preserve=True)
    _create_by_region(workbook["Data"])
    created = str(tmp_path / "created.xlsx")
    workbook = save_and_reopen(workbook, created, preserve=True)
    before = part_payloads(created)

    workbook["Data"]["B2"] = 99
    renamed = workbook["Data"].pivots["ByRegion"].refresh().rename(
        "RegionalSales")
    receipt = workbook.save(str(tmp_path / "renamed.xlsx"), receipt=True)
    after = part_payloads(str(tmp_path / "renamed.xlsx"))

    assert renamed.name == "RegionalSales"
    assert after[_CACHE] != before[_CACHE]
    assert b' name="RegionalSales"' in after[_PIVOT]
    effects = [item for item in receipt.derived_effects
               if item["kind"] in ("pivot_refreshed", "pivot_renamed")]
    assert [item["kind"] for item in effects] == [
        "pivot_refreshed", "pivot_renamed"]
    assert all(item["name"] == "RegionalSales" for item in effects)
    assert all(item["sheet"] == "Data" for item in effects)
    assert all(item["output_range"] == "E3:F9" for item in effects)
    assert all(item["source"] == {
        "kind": "table", "name": "RegionTable"} for item in effects)
    assert effects[0]["source_identity"]
    assert effects[1]["cache_rebuilt"] is True
    reopened = load_workbook(str(tmp_path / "renamed.xlsx"), preserve=True)
    assert reopened["Data"]["F5"].value == 99
    assert reopened["Data"].pivots["RegionalSales"].capabilities.can_refresh_on_open


def test_repeated_move_clears_every_vacated_output(tmp_path):
    _source, workbook = _preserved_matrix(
        tmp_path,
        ["Region", "Amount"],
        [["East", 10], ["West", 7]],
        table="Sales",
    )
    pivot = workbook["Summary"].pivots.create(
        name="ByRegion", source="Sales", destination="A1",
        rows=["Region"], values=["Amount"])
    pivot = pivot.move("E1")
    pivot = pivot.move("I1")
    workbook = save_and_reopen(
        workbook, str(tmp_path / "moved-twice.xlsx"), preserve=True)

    assert workbook["Summary"].pivots["ByRegion"].destination == "I1"
    assert workbook["Summary"]["A1"].value is None
    assert workbook["Summary"]["E1"].value is None
    assert workbook["Summary"]["I1"].value is not None


def test_vacated_output_can_be_reused_before_save(tmp_path):
    _source, workbook = _preserved_matrix(
        tmp_path,
        ["Region", "Amount"],
        [["East", 10], ["West", 7]],
        table="Sales",
    )
    pivot = workbook["Summary"].pivots.create(
        name="ByRegion", source="Sales", destination="A1",
        rows=["Region"], values=["Amount"])
    pivot.move("E1")
    workbook["Summary"]["A1"] = "reused"

    workbook = save_and_reopen(
        workbook, str(tmp_path / "reused.xlsx"), preserve=True)
    assert workbook["Summary"]["A1"].value == "reused"
    assert workbook["Summary"].pivots["ByRegion"].destination == "E1"


def test_update_and_delete_remove_pivot_owned_number_formats(tmp_path):
    _source, workbook = _preserved_matrix(
        tmp_path,
        ["Region", "Amount"],
        [["East", 10], ["West", 7]],
        table="Sales",
    )
    pivot = workbook["Summary"].pivots.create(
        name="ByRegion", source="Sales", destination="A1",
        rows=["Region"],
        values=[PivotMeasure("Amount", number_format="$#,##0")],
    )
    pivot = pivot.update(values=[PivotMeasure("Amount")])
    numeric = [
        cell for row in workbook["Summary"].iter_rows()
        for cell in row if isinstance(cell.value, (int, float))
    ]
    assert numeric
    assert all(cell.number_format == "General" for cell in numeric)

    pivot.delete()
    replacement = workbook["Summary"].pivots.create(
        name="Replacement", source="Sales", destination="A1",
        rows=["Region"], values=["Amount"])
    workbook = save_and_reopen(
        workbook, str(tmp_path / "format-cleared.xlsx"), preserve=True)
    assert replacement.destination == "A1"
    assert workbook["Summary"].pivots["Replacement"].destination == "A1"


def test_staged_pivot_follows_worksheet_rename(tmp_path):
    _source, workbook = _preserved_matrix(
        tmp_path,
        ["Region", "Amount"],
        [["East", 10], ["West", 7]],
        table="Sales",
    )
    summary = workbook["Summary"]
    summary.pivots.create(
        name="ByRegion", source="Sales", destination="A1",
        rows=["Region"], values=["Amount"])
    summary.title = "Analysis"
    with pytest.raises(BoundaryViolationError) as overlap:
        summary.pivots.create(
            name="Overlap", source="Sales", destination="A1",
            rows=["Region"], values=["Amount"])
    assert overlap.value.kind == "pivot-output-collision"

    destination = str(tmp_path / "renamed-sheet.xlsx")
    receipt = workbook.save(destination, receipt=True)
    effects = [
        item for item in receipt.derived_effects
        if item["kind"] == "pivot_created"
    ]
    assert effects[0]["sheet"] == "Analysis"
    workbook = load_workbook(destination, preserve=True)
    assert workbook["Analysis"].pivots["ByRegion"].destination == "A1"


def test_reopened_pivot_keeps_capabilities_after_worksheet_rename(tmp_path):
    _source, workbook = _preserved_matrix(
        tmp_path,
        ["Region", "Amount"],
        [["East", 10], ["West", 7]],
        table="Sales",
    )
    workbook["Summary"].pivots.create(
        name="ByRegion", source="Sales", destination="A1",
        rows=["Region"], values=["Amount"])
    created = str(tmp_path / "created-before-sheet-rename.xlsx")
    workbook = save_and_reopen(workbook, created, preserve=True)
    workbook["Summary"].title = "Analysis"

    pivot = workbook["Analysis"].pivots["ByRegion"]
    assert pivot.capabilities.can_move is True
    pivot.move("E1")
    workbook = save_and_reopen(
        workbook, str(tmp_path / "renamed-and-moved.xlsx"), preserve=True)
    assert workbook["Analysis"].pivots["ByRegion"].destination == "E1"


def test_materialized_pivot_collection_tracks_worksheet_rename(tmp_path):
    _source, workbook = _preserved_matrix(
        tmp_path,
        ["Region", "Amount"],
        [["East", 10], ["West", 7]],
        table="Sales",
    )
    workbook["Summary"].pivots.create(
        name="ByRegion", source="Sales", destination="A1",
        rows=["Region"], values=["Amount"])
    created = str(tmp_path / "collection-before-rename.xlsx")
    workbook = save_and_reopen(workbook, created, preserve=True)
    summary = workbook["Summary"]
    assert len(summary.pivots) == 1

    summary.title = "Analysis"
    assert [pivot.name for pivot in summary.pivots] == ["ByRegion"]


def test_staged_range_source_blocks_source_sheet_rename(tmp_path):
    _source, workbook = _preserved_matrix(
        tmp_path,
        ["Region", "Amount"],
        [["East", 10], ["West", 7]],
    )
    workbook["Summary"].pivots.create(
        name="ByRegion", source="Data!A1:B3", destination="A1",
        rows=["Region"], values=["Amount"])

    with pytest.raises(UnsupportedStructureError) as exc:
        workbook["Data"].title = "RenamedData"
    assert exc.value.kind == "unsupported-pivot-operation"
    assert "Data" in workbook.sheetnames
    assert "RenamedData" not in workbook.sheetnames


def test_staged_pivot_blocks_output_sheet_removal(tmp_path):
    _source, workbook = _preserved_matrix(
        tmp_path,
        ["Region", "Amount"],
        [["East", 10], ["West", 7]],
        table="Sales",
    )
    summary = workbook["Summary"]
    summary.pivots.create(
        name="ByRegion", source="Sales", destination="A1",
        rows=["Region"], values=["Amount"])

    with pytest.raises(UnsupportedStructureError) as exc:
        workbook.remove(summary)
    assert exc.value.kind == "invalid-pivot-graph"
    assert summary in workbook.worksheets


def test_staged_pivot_blocks_table_source_sheet_removal(tmp_path):
    _source, workbook = _preserved_matrix(
        tmp_path,
        ["Region", "Amount"],
        [["East", 10], ["West", 7]],
        table="Sales",
    )
    workbook["Summary"].pivots.create(
        name="ByRegion", source="Sales", destination="A1",
        rows=["Region"], values=["Amount"])

    with pytest.raises(UnsupportedStructureError) as exc:
        workbook.remove(workbook["Data"])
    assert exc.value.kind == "invalid-pivot-graph"
    assert "Data" in workbook.sheetnames


def test_collision_preflight_uses_renamed_output_sheet(tmp_path):
    _source, workbook = _preserved_matrix(
        tmp_path,
        ["Region", "Amount"],
        [["East", 10], ["West", 7]],
        table="Sales",
    )
    workbook["Summary"].pivots.create(
        name="ByRegion", source="Sales", destination="A1",
        rows=["Region"], values=["Amount"])
    created = str(tmp_path / "collision-before-rename.xlsx")
    workbook = save_and_reopen(workbook, created, preserve=True)
    summary = workbook["Summary"]
    summary.title = "Analysis"

    with pytest.raises(BoundaryViolationError) as exc:
        summary.pivots.create(
            name="Overlapping", source="Sales", destination="A1",
            rows=["Region"], values=["Amount"])
    assert exc.value.kind == "pivot-output-collision"
    assert not workbook._paper_ledger.pivot_operations


def test_refresh_policy_composes_with_cache_rebuild(tmp_path):
    _source, workbook = _preserved_matrix(
        tmp_path,
        ["Region", "Amount"],
        [["East", 10], ["West", 7]],
        table="Sales",
    )
    workbook["Summary"].pivots.create(
        name="ByRegion", source="Sales", destination="A1",
        rows=["Region"], values=["Amount"])
    created = str(tmp_path / "created-for-refresh-policy.xlsx")
    workbook = save_and_reopen(workbook, created, preserve=True)
    workbook["Data"]["B2"] = 11
    workbook["Summary"].pivots["ByRegion"].refresh()
    workbook.set_pivot_refresh_on_load(pivots=["ByRegion"])
    destination = str(tmp_path / "rebuilt-refresh-policy.xlsx")
    workbook.save(destination)

    cache = part_payloads(destination)[_CACHE]
    assert b'refreshOnLoad="1"' in cache


def test_refresh_policy_resolves_staged_pivot_name(tmp_path):
    _source, workbook = _preserved_matrix(
        tmp_path,
        ["Region", "Amount"],
        [["East", 10], ["West", 7]],
        table="Sales",
    )
    workbook["Summary"].pivots.create(
        name="ByRegion", source="Sales", destination="A1",
        rows=["Region"], values=["Amount"])
    created = str(tmp_path / "created-before-pivot-rename.xlsx")
    workbook = save_and_reopen(workbook, created, preserve=True)
    workbook["Data"]["B2"] = 11
    renamed = workbook["Summary"].pivots["ByRegion"].refresh().rename(
        "RegionalSales")

    assert renamed.name == "RegionalSales"
    assert workbook.set_pivot_refresh_on_load(
        pivots=["RegionalSales"]
    ) == [_CACHE]
    destination = str(tmp_path / "renamed-refresh-policy.xlsx")
    workbook.save(destination)
    assert b'refreshOnLoad="1"' in part_payloads(destination)[_CACHE]


def test_receipt_reports_every_surviving_composed_pivot_effect(
        fixture_copy, tmp_path):
    source = fixture_copy(_TABLE)
    workbook = load_workbook(source, preserve=True)
    _create_by_region(workbook["Data"])
    created = str(tmp_path / "created-for-composed-receipt.xlsx")
    workbook = save_and_reopen(workbook, created, preserve=True)

    pivot = workbook["Data"].pivots["ByRegion"]
    pivot = pivot.repoint_source("Data!A1:B5")
    pivot = pivot.move("J3")
    pivot.rename("RegionalSales")
    receipt = workbook.save(
        str(tmp_path / "composed-receipt.xlsx"), receipt=True)

    events = [
        item for item in receipt.derived_effects
        if item["kind"] in (
            "pivot_repointed", "pivot_moved", "pivot_renamed")
    ]
    assert [item["kind"] for item in events] == [
        "pivot_repointed", "pivot_moved", "pivot_renamed"]
    assert all(item["name"] == "RegionalSales" for item in events)
    assert all(item["sheet"] == "Data" for item in events)
    assert all(item["output_range"] == "J3:K9" for item in events)
    assert all(item["source"] == {
        "kind": "range", "sheet": "Data", "ref": "A1:B5"
    } for item in events)


@pytest.mark.parametrize("verb", ["move", "rename", "repoint", "update"])
def test_receipt_omits_pivot_effect_that_is_reversed_before_save(
        fixture_copy, tmp_path, verb):
    source = fixture_copy(_TABLE)
    workbook = load_workbook(source, preserve=True)
    _create_by_region(workbook["Data"])
    created = str(tmp_path / ("created-for-reverse-%s.xlsx" % verb))
    workbook = save_and_reopen(workbook, created, preserve=True)
    pivot = workbook["Data"].pivots["ByRegion"]

    if verb == "move":
        pivot = pivot.move("J3")
        pivot.move("E3")
    elif verb == "rename":
        pivot = pivot.rename("Temporary")
        pivot.rename("ByRegion")
    elif verb == "repoint":
        pivot = pivot.repoint_source("Data!A1:B5")
        pivot.repoint_source("RegionTable")
    else:
        pivot = pivot.update(layout="outline")
        pivot.update(layout="tabular")

    destination = str(tmp_path / ("reversed-%s.xlsx" % verb))
    receipt = workbook.save(destination, receipt=True)
    pivot_events = [
        item for item in receipt.derived_effects
        if item["kind"] in (
            "pivot_refreshed", "pivot_repointed", "pivot_moved",
            "pivot_updated", "pivot_renamed")
    ]
    assert pivot_events == []
    assert part_payloads(destination) == part_payloads(created)


def test_repoint_move_rename_composition_keeps_one_coherent_graph(tmp_path):
    _source, workbook = _preserved_matrix(
        tmp_path,
        ["Region", "Amount"],
        [["East", 10], ["West", 7]],
        table="Sales",
    )
    data = workbook["Data"]
    data["D1"] = "Region"
    data["E1"] = "Amount"
    data["D2"] = "North"
    data["E2"] = 41
    data.add_table(Table(displayName="Other", ref="D1:E2"))
    pivot = workbook["Summary"].pivots.create(
        name="ByRegion", source="Sales", destination="A1",
        rows=["Region"], values=["Amount"])
    pivot = pivot.repoint_source("Other")
    pivot = pivot.move("D4")
    pivot = pivot.rename("NorthOnly")

    destination = str(tmp_path / "composed.xlsx")
    workbook = save_and_reopen(workbook, destination, preserve=True)
    reopened = workbook["Summary"].pivots["NorthOnly"]
    assert reopened.source.name == "Other"
    assert reopened.destination == "D4"
    assert workbook["Summary"]["D6"].value == "North"
    assert workbook["Summary"]["E6"].value == 41
    assert reopened.capabilities.can_headless_refresh is True


def test_rename_is_a_lexical_definition_patch(fixture_copy, tmp_path):
    source = fixture_copy(_TABLE)
    workbook = load_workbook(source, preserve=True)
    _create_by_region(workbook["Data"])
    created = str(tmp_path / "created.xlsx")
    workbook.save(created)
    before = part_payloads(created)

    workbook = load_workbook(created, preserve=True)
    workbook["Data"].pivots["ByRegion"].rename("RegionalSales")
    renamed = str(tmp_path / "renamed.xlsx")
    workbook.save(renamed)
    after = part_payloads(renamed)

    expected = before[_PIVOT].replace(
        b'name="ByRegion"', b'name="RegionalSales"', 1)
    assert after[_PIVOT] == expected
    assert after[_RECORDS] == before[_RECORDS]


def test_refresh_rebuilds_cache_when_aggregate_is_unchanged(tmp_path):
    _source, workbook = _preserved_matrix(
        tmp_path,
        ["Region", "Amount"],
        [["East", 20], ["East", 30]],
        table="Sales",
    )
    workbook["Summary"].pivots.create(
        name="ByRegion", source="Sales", destination="A1",
        rows=["Region"], values=[PivotMeasure("Amount", "sum")])
    created = str(tmp_path / "created.xlsx")
    workbook = save_and_reopen(workbook, created, preserve=True)
    before = part_payloads(created)[_CACHE]

    workbook["Data"]["B2"] = 21
    workbook["Data"]["B3"] = 29
    workbook["Summary"].pivots["ByRegion"].refresh()
    refreshed = str(tmp_path / "refreshed.xlsx")
    workbook.save(refreshed)

    assert part_payloads(refreshed)[_CACHE] != before
    assert load_workbook(refreshed, preserve=True)["Summary"]["B3"].value == 50


def test_custom_measure_format_survives_reopen_and_refresh(tmp_path):
    _source, workbook = _preserved_matrix(
        tmp_path,
        ["Region", "Amount"],
        [["East", 20], ["West", 30]],
        table="Sales",
    )
    workbook["Summary"].pivots.create(
        name="ByRegion", source="Sales", destination="A1",
        rows=["Region"],
        values=[PivotMeasure("Amount", "sum", number_format="$#,##0")],
    )
    created = str(tmp_path / "formatted.xlsx")
    workbook = save_and_reopen(workbook, created, preserve=True)
    pivot = workbook["Summary"].pivots["ByRegion"]
    assert pivot.spec.values[0].number_format == "$#,##0"

    workbook["Data"]["B2"] = 25
    pivot.refresh()
    refreshed = str(tmp_path / "formatted-refreshed.xlsx")
    workbook = save_and_reopen(workbook, refreshed, preserve=True)
    assert workbook["Summary"].pivots["ByRegion"].spec.values[0].number_format \
        == "$#,##0"
    assert b'numFmtId="164"' in part_payloads(refreshed)[_PIVOT]


def test_unmodeled_standard_semantics_disable_reserializing_mutators(
        fixture_copy, tmp_path):
    source = fixture_copy(_TABLE)
    workbook = load_workbook(source, preserve=True)
    _create_by_region(workbook["Data"])
    created = str(tmp_path / "created.xlsx")
    workbook.save(created)
    tampered = str(tmp_path / "tampered.xlsx")
    _rewrite_part(
        created,
        tampered,
        _PIVOT,
        lambda payload: payload.replace(
            b'showRowStripes="0"', b'showRowStripes="1"', 1),
    )

    workbook = load_workbook(tampered, preserve=True)
    pivot = workbook["Data"].pivots["ByRegion"]
    assert pivot.capabilities.can_edit_layout is False
    assert pivot.capabilities.can_headless_refresh is False
    assert pivot.capabilities.can_rename is True
    with pytest.raises(UnsupportedStructureError):
        pivot.update(layout="outline")
    renamed = pivot.rename("RegionalSales")
    destination = str(tmp_path / "renamed-tampered.xlsx")
    workbook.save(destination)
    assert renamed.name == "RegionalSales"
    assert b'showRowStripes="1"' in part_payloads(destination)[_PIVOT]


def test_removed_subtotal_item_type_disables_reserializing_mutators(
        tmp_path):
    _source, workbook = _preserved_matrix(
        tmp_path,
        ["Region", "Product", "Amount"],
        [["East", "A", 10], ["East", "B", 7], ["West", "A", 5]],
        table="Sales",
    )
    workbook["Summary"].pivots.create(
        name="ByProduct", source="Sales", destination="A1",
        rows=["Region", "Product"], values=["Amount"],
        layout="tabular", subtotals=True)
    created = str(tmp_path / "created-subtotals.xlsx")
    workbook.save(created)
    tampered = str(tmp_path / "subtotal-type-removed.xlsx")

    def remove_subtotal_type(payload):
        marker = b' t="default"'
        assert marker in payload
        return payload.replace(marker, b"", 1)

    _rewrite_part(created, tampered, _PIVOT, remove_subtotal_type)
    pivot = load_workbook(tampered, preserve=True)["Summary"].pivots[
        "ByProduct"]
    assert pivot.capabilities.can_headless_refresh is False
    assert pivot.capabilities.can_edit_layout is False
    assert any(
        reason.code == "paper-semantics-unproved"
        for reason in pivot.qualification_reasons
    )


def test_omitted_default_data_item_type_keeps_paper_capabilities(tmp_path):
    _source, workbook = _preserved_matrix(
        tmp_path,
        ["Region", "Amount", "Units"],
        [["East", 10, 1], ["West", 7, 2]],
        table="Sales",
    )
    workbook["Summary"].pivots.create(
        name="ValuesOnRows", source="Sales", destination="A1",
        rows=["Region"], values=["Amount", "Units"],
        values_axis="rows")
    created = str(tmp_path / "created-data-item.xlsx")
    workbook.save(created)
    normalized = str(tmp_path / "omitted-data-item.xlsx")

    def omit_default_type(payload):
        marker = b' t="data"'
        assert marker in payload
        return payload.replace(marker, b"", 1)

    _rewrite_part(created, normalized, _PIVOT, omit_default_type)
    pivot = load_workbook(normalized, preserve=True)["Summary"].pivots[
        "ValuesOnRows"]
    assert pivot.capabilities.can_headless_refresh is True
    assert pivot.capabilities.can_edit_layout is True
    assert pivot.capabilities.can_repoint_source is True
    assert pivot.capabilities.can_move is True
    assert pivot.capabilities.can_delete is True


def test_namespaced_lookalike_attribute_disables_reserialization(tmp_path):
    _source, workbook = _preserved_matrix(
        tmp_path,
        ["Region", "Amount"],
        [["East", 10], ["West", 7]],
        table="Sales",
    )
    workbook["Summary"].pivots.create(
        name="ByRegion", source="Sales", destination="A1",
        rows=["Region"], values=["Amount"])
    created = str(tmp_path / "created-for-namespace.xlsx")
    workbook.save(created)
    tampered = str(tmp_path / "namespaced-lookalike.xlsx")

    def replace_compact(payload):
        marker = b' dataOnRows="0"'
        assert marker in payload
        return payload.replace(
            marker,
            b' xmlns:evil="urn:paper-test:evil" evil:dataOnRows="0"',
            1,
        )

    _rewrite_part(created, tampered, _PIVOT, replace_compact)
    pivot = load_workbook(tampered, preserve=True)["Summary"].pivots[
        "ByRegion"]
    assert pivot.capabilities.can_headless_refresh is False
    assert pivot.capabilities.can_edit_layout is False
    assert any(
        reason.code == "paper-semantics-unproved"
        for reason in pivot.qualification_reasons
    )


def test_namespaced_grand_total_child_disables_reserialization(
        fixture_copy, tmp_path):
    source = fixture_copy(_TABLE)
    workbook = load_workbook(source, preserve=True)
    _create_by_region(workbook["Data"])
    created = str(tmp_path / "created-for-evil-grand-total.xlsx")
    workbook.save(created)
    tampered = str(tmp_path / "evil-grand-total.xlsx")

    def add_namespaced_child(payload):
        pattern = br'<i t="grand" r="0" i="0"\s*/>'
        assert re.search(pattern, payload)
        payload = payload.replace(
            b"<pivotTableDefinition ",
            b'<pivotTableDefinition xmlns:evil="urn:paper-test:evil" ',
            1,
        )
        return re.sub(
            pattern,
            b'<i t="grand" r="0" i="0"><evil:x/></i>',
            payload,
            count=1,
        )

    _rewrite_part(created, tampered, _PIVOT, add_namespaced_child)
    pivot = load_workbook(tampered, preserve=True)["Data"].pivots[
        "ByRegion"]
    assert pivot.capabilities.can_headless_refresh is False
    assert pivot.capabilities.can_edit_layout is False
    assert any(
        reason.code == "paper-semantics-unproved"
        for reason in pivot.qualification_reasons
    )


def test_excel_minus_one_page_item_is_a_benign_all_filter(tmp_path):
    _source, workbook = _preserved_matrix(
        tmp_path,
        ["Region", "Status", "Amount"],
        [["East", "Closed", 10], ["West", "Pending", 7]],
        table="Sales",
    )
    workbook["Summary"].pivots.create(
        name="ByRegion", source="Sales", destination="A4",
        rows=["Region"], filters=[PivotItemFilter("Status")],
        values=["Amount"])
    created = str(tmp_path / "created-all-filter.xlsx")
    workbook.save(created)
    normalized = str(tmp_path / "excel-all-filter.xlsx")

    def add_minus_one(payload):
        marker = b'<pageField fld="1" hier="-1"'
        assert marker in payload
        return payload.replace(
            marker, b'<pageField fld="1" item="-1" hier="-1"', 1)

    _rewrite_part(created, normalized, _PIVOT, add_minus_one)
    pivot = load_workbook(normalized, preserve=True)["Summary"].pivots[
        "ByRegion"]
    assert pivot.spec.filters[0].include is None
    assert pivot.capabilities.can_headless_refresh is True
    assert pivot.capabilities.can_edit_layout is True
    assert pivot.capabilities.can_delete is True


def test_exact_empty_excel_extensions_keep_paper_capabilities(
        fixture_copy, tmp_path):
    source = fixture_copy(_TABLE)
    workbook = load_workbook(source, preserve=True)
    _create_by_region(workbook["Data"])
    created = str(tmp_path / "created-for-excel-extension.xlsx")
    workbook.save(created)
    pivot_extended = str(tmp_path / "pivot-extended.xlsx")
    _rewrite_part(
        created, pivot_extended, _PIVOT,
        lambda payload: _add_excel_extension(payload, "pivot"),
    )
    excel_extended = str(tmp_path / "excel-extended.xlsx")
    _rewrite_part(
        pivot_extended, excel_extended, _CACHE,
        lambda payload: _add_excel_extension(payload, "cache"),
    )

    workbook = load_workbook(excel_extended, preserve=True)
    capabilities = workbook["Data"].pivots["ByRegion"].capabilities
    assert capabilities.can_headless_refresh is True
    assert capabilities.can_edit_layout is True
    assert capabilities.can_repoint_source is True
    assert capabilities.can_move is True
    assert capabilities.can_delete is True


def test_excel_extension_uri_with_nonempty_payload_disables_rebuild(
        fixture_copy, tmp_path):
    source = fixture_copy(_TABLE)
    workbook = load_workbook(source, preserve=True)
    _create_by_region(workbook["Data"])
    created = str(tmp_path / "created-for-bad-extension.xlsx")
    workbook.save(created)
    tampered = str(tmp_path / "nonempty-extension.xlsx")
    _rewrite_part(
        created, tampered, _PIVOT,
        lambda payload: _add_excel_extension(
            payload, "pivot", child_attributes=b' future="1"'),
    )

    workbook = load_workbook(tampered, preserve=True)
    pivot = workbook["Data"].pivots["ByRegion"]
    assert pivot.capabilities.can_headless_refresh is False
    assert pivot.capabilities.can_edit_layout is False
    assert pivot.capabilities.can_delete is False
    assert any(
        reason.code == "unsupported-extension"
        for reason in pivot.qualification_reasons
    )


def test_show_data_as_is_not_projected_as_an_ordinary_sum(fixture_copy, tmp_path):
    source = fixture_copy(_TABLE)
    workbook = load_workbook(source, preserve=True)
    _create_by_region(workbook["Data"])
    created = str(tmp_path / "created.xlsx")
    workbook.save(created)
    tampered = str(tmp_path / "show-data-as.xlsx")
    _rewrite_part(
        created,
        tampered,
        _PIVOT,
        lambda payload: payload.replace(
            b'showDataAs="normal"', b'showDataAs="percentOfTotal"', 1),
    )

    workbook = load_workbook(tampered, preserve=True)
    pivot = workbook["Data"].pivots["ByRegion"]
    assert pivot.spec is None
    assert pivot.capabilities.can_edit_layout is False
    assert pivot.capabilities.can_headless_refresh is False


def test_date_and_manual_item_order_reopen_with_lifecycle_capabilities(tmp_path):
    _source, workbook = _preserved_matrix(
        tmp_path,
        ["AsOf", "Region", "Amount"],
        [
            [date(2026, 1, 1), "East", 20],
            [date(2026, 2, 1), "West", 30],
        ],
        table="Sales",
    )
    workbook["Summary"].pivots.create(
        name="ByDate", source="Sales", destination="A1",
        rows=["AsOf"], values=["Amount"])
    workbook["Summary"].pivots.create(
        name="ByRegion", source="Sales", destination="E1",
        rows=[PivotAxisField("Region", items=["West", "East"])],
        values=["Amount"])
    created = str(tmp_path / "ordered.xlsx")
    workbook = save_and_reopen(workbook, created, preserve=True)

    by_date = workbook["Summary"].pivots["ByDate"]
    by_region = workbook["Summary"].pivots["ByRegion"]
    assert by_date.capabilities.can_headless_refresh is True
    assert by_date.capabilities.can_delete is True
    assert by_region.capabilities.can_headless_refresh is True
    assert by_region.capabilities.can_delete is True
    assert workbook["Summary"]["E3"].value == "West"
    assert workbook["Summary"]["E4"].value == "East"


def test_direct_edit_to_reopened_report_filter_cell_refuses_save(tmp_path):
    _source, workbook = _preserved_matrix(
        tmp_path,
        ["Region", "Status", "Amount"],
        [["East", "Closed", 20], ["West", "Open", 30]],
        table="Sales",
    )
    workbook["Summary"].pivots.create(
        name="ClosedOnly", source="Sales", destination="A3",
        rows=["Region"],
        filters=[PivotItemFilter("Status", include=["Closed"])],
        values=["Amount"],
    )
    created = str(tmp_path / "filtered.xlsx")
    workbook = save_and_reopen(workbook, created, preserve=True)
    workbook["Summary"]["B1"] = "tampered"

    destination = str(tmp_path / "filtered-tampered.xlsx")
    with pytest.raises(BoundaryViolationError) as exc:
        workbook.save(destination)
    assert exc.value.kind == "pivot-output-collision"
    assert not os.path.exists(destination)


def _rewrite_part(source, destination, part, transform):
    with zipfile.ZipFile(source) as before, zipfile.ZipFile(destination, "w") as after:
        for info in before.infolist():
            payload = before.read(info.filename)
            if info.filename == part:
                payload = transform(payload)
            after.writestr(copy(info), payload)


def _add_excel_extension(payload, kind, child_attributes=b""):
    if kind == "pivot":
        closing = b"</pivotTableDefinition>"
        extension = (
            b'<extLst><ext uri="{747A6164-185A-40DC-8AA5-F01512510D54}" '
            b'xmlns:xpdl="http://schemas.microsoft.com/office/spreadsheetml/'
            b'2016/pivotdefaultlayout"><xpdl:pivotTableDefinition16'
            + child_attributes + b'/></ext></extLst>'
        )
    else:
        closing = b"</pivotCacheDefinition>"
        extension = (
            b'<extLst><ext uri="{725AE2AE-9491-48be-B2B4-4EB974FC3084}" '
            b'xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/'
            b'2009/9/main"><x14:pivotCacheDefinition'
            + child_attributes + b'/></ext></extLst>'
        )
    return payload.replace(closing, extension + closing, 1)
