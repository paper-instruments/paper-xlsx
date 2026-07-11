from __future__ import annotations

import io
import hashlib
import os
import zipfile
from xml.etree import ElementTree

import pytest

from openpyxl import Workbook, load_workbook
from openpyxl import oracle
from openpyxl.errors import AmbiguousTargetError, TargetNotFoundError, UnsupportedStructureError
from openpyxl.preserve.splice import _serialize_cached_value
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
from openpyxl.workbook.defined_name import DefinedName


MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _rewrite_zip(data, replacements):
    out = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(data)) as zin, zipfile.ZipFile(out, "w") as zout:
        for info in zin.infolist():
            zout.writestr(info, replacements.get(info.filename, zin.read(info)))
    return out.getvalue()


def _array_package(values):
    ElementTree.register_namespace("", MAIN)
    wb = Workbook()
    ws = wb.active
    ws.title = "Array"
    ws["A1"] = ArrayFormula(ref="A1:A3", text="=ROW(A1:A3)")
    raw = io.BytesIO()
    wb.save(raw)

    with zipfile.ZipFile(io.BytesIO(raw.getvalue())) as zin:
        sheet_name = "xl/worksheets/sheet1.xml"
        root = ElementTree.fromstring(zin.read(sheet_name))
    sheet_data = root.find("{%s}sheetData" % MAIN)
    rows = {int(row.attrib["r"]): row for row in sheet_data}
    for row_number, value in enumerate(values, 1):
        row = rows.get(row_number)
        if row is None:
            row = ElementTree.SubElement(
                sheet_data, "{%s}row" % MAIN, {"r": str(row_number)})
        cell = next((c for c in row if c.attrib.get("r") == "A%d" % row_number), None)
        if cell is None:
            cell = ElementTree.SubElement(
                row, "{%s}c" % MAIN, {"r": "A%d" % row_number})
        cell.attrib.pop("t", None)
        cached = cell.find("{%s}v" % MAIN)
        if cached is None:
            cached = ElementTree.SubElement(cell, "{%s}v" % MAIN)
        cached.text = str(value)
    payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=False)
    return _rewrite_zip(raw.getvalue(), {sheet_name: payload})


def _template_package():
    wb = Workbook()
    wb.template = True
    raw = io.BytesIO()
    wb.save(raw)
    return raw.getvalue()


def _formula_cache_package(formulas, cached_values):
    wb = Workbook()
    ws = wb.active
    for address, formula in formulas.items():
        ws[address] = formula
    raw = io.BytesIO()
    wb.save(raw)
    with zipfile.ZipFile(io.BytesIO(raw.getvalue())) as zin:
        root = ElementTree.fromstring(
            zin.read("xl/worksheets/sheet1.xml"))
    cells = {cell.get("r"): cell for cell in root.iter("{%s}c" % MAIN)}
    for address, value in cached_values.items():
        cached = cells[address].find("{%s}v" % MAIN)
        cached.text = None if value is None else str(value)
    payload = ElementTree.tostring(root, encoding="utf-8")
    return _rewrite_zip(
        raw.getvalue(), {"xl/worksheets/sheet1.xml": payload})


def _formula_package(cached_values):
    return _formula_cache_package(
        {"A1": "=1", "A2": "=2"},
        dict(zip(("A1", "A2"), cached_values)),
    )


class TestArrayFormulaCoverage:

    def test_stale_follower_diverges_and_write_back_updates_it(
            self, tmp_path, monkeypatch):
        stale = _array_package([1, 999, 3])
        computed = _array_package([1, 2, 3])
        monkeypatch.setattr(
            oracle, "_recalculate_bytes",
            lambda data, timeout, suffix=".xlsx", profile_root=None: computed)

        certification = oracle.certify(stale)
        assert certification.status == "DIVERGED"
        assert certification.checked == 3
        assert [item["address"] for item in certification.divergences] == [
            "Array!A2"]

        path = tmp_path / "array.xlsx"
        path.write_bytes(stale)
        first = oracle.write_back(path, allow_uncertified=True)
        assert first.written == ["Array!A2"]
        assert first.cleared_fullcalc is False
        assert load_workbook(path, data_only=True)["Array"]["A2"].value == 2

        second = oracle.write_back(path)
        assert second.certification.status == "CERTIFIED"
        assert second.certification.checked == 3
        assert second.cells_written == 0
        assert second.cleared_fullcalc is True

    @pytest.mark.lo_smoke
    def test_libreoffice_array_results_are_all_certified(self, lo, tmp_path):
        path = tmp_path / "array.xlsx"
        path.write_bytes(_array_package([1, 2, 3]))
        result = oracle.certify(path)
        assert result.status == "CERTIFIED"
        assert result.checked == 3


def test_path_qualified_external_references_are_excluded():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "='C:\\Models\\[Budget.xlsx]Plan'!A1"
    ws["A2"] = "='https://example.test/models/[Budget.xlsx]Plan'!A1"
    ws["A3"] = "='\\\\server\\share\\[Budget.xlsx]Plan'!A1"
    ws["A4"] = "=SUM(Table1[Amount])"
    seeds = oracle._exclusion_seeds(wb)
    assert {seeds[("Sheet", row, 1)] for row in range(1, 4)} == {
        "external-link"}
    assert ("Sheet", 4, 1) not in seeds


def test_zero_checked_formulas_are_never_certified(monkeypatch):
    wb = Workbook()
    wb.active["A1"] = "=RAND()"
    raw = io.BytesIO()
    wb.save(raw)
    cached = raw.getvalue().replace(b"<v></v>", b"<v>0.5</v>", 1)
    monkeypatch.setattr(oracle, "_recalculate_bytes", lambda data, timeout: data)
    result = oracle.certify(cached)
    assert result.checked == 0
    assert result.status == "BASELINE_UNVERIFIABLE"


def test_unresolved_formula_and_downstream_are_excluded_from_certification(
        monkeypatch):
    baseline = _formula_cache_package(
        {
            "A1": "=RAND()",
            "B1": '=INDIRECT("A1")',
            "C1": "=B1",
        },
        {"A1": 0.5, "B1": 0.5, "C1": 0.5},
    )
    monkeypatch.setattr(
        oracle, "_recalculate_bytes",
        lambda data, timeout, suffix=".xlsx", profile_root=None: baseline)

    result = oracle.certify(baseline)

    assert result.status == "BASELINE_UNVERIFIABLE"
    assert result.checked == 0
    assert result.volatile_excluded == ["Sheet!A1"]
    assert result.unsupported_excluded == [
        "Sheet!B1 (unresolved-reference)",
        "Sheet!C1 (unresolved-reference)",
    ]


def test_unresolved_formula_without_other_exclusions_is_not_certified(
        monkeypatch):
    baseline = _formula_cache_package(
        {"B1": '=INDIRECT("A1")', "C1": "=B1+1"},
        {"B1": 2, "C1": 3},
    )
    monkeypatch.setattr(
        oracle, "_recalculate_bytes",
        lambda data, timeout, suffix=".xlsx", profile_root=None: baseline)

    result = oracle.certify(baseline)

    assert result.status == "BASELINE_UNVERIFIABLE"
    assert result.checked == 0
    assert result.unsupported_excluded == [
        "Sheet!B1 (unresolved-reference)",
        "Sheet!C1 (unresolved-reference)",
    ]


def test_unresolved_formula_and_downstream_inherit_scenario_input_taint():
    baseline = _formula_cache_package(
        {"B1": '=INDIRECT("A1")', "C1": "=B1+1"},
        {"B1": 2, "C1": 3},
    )

    result, _ = oracle._certify_impl(
        baseline,
        timeout=1,
        recalculated=baseline,
        input_seeds=[("Sheet", 1, 1)],
    )

    assert result.status == "BASELINE_UNVERIFIABLE"
    assert result.checked == 0
    assert result.input_excluded == ["Sheet!B1", "Sheet!C1"]
    assert result.unsupported_excluded == []


def test_partial_formula_baseline_is_never_certified(monkeypatch):
    baseline = _formula_package([1, None])
    computed = _formula_package([1, 2])
    monkeypatch.setattr(
        oracle, "_recalculate_bytes",
        lambda data, timeout, suffix=".xlsx", profile_root=None: computed)

    result = oracle.certify(baseline)
    assert result.status == "BASELINE_UNVERIFIABLE"
    assert result.checked == 1
    assert result.unverifiable == ["Sheet!A2"]
    assert result.artifact_sha256 == hashlib.sha256(baseline).hexdigest()


def test_oracle_defined_name_resolution_is_local_first_and_case_insensitive():
    wb = Workbook()
    ws = wb.active
    wb.defined_names["Rate"] = DefinedName(
        "Rate", attr_text="Sheet!$A$1")
    ws.defined_names["RATE"] = DefinedName(
        "RATE", attr_text="'[Budget.xlsx]Inputs'!$A$1")
    ws["B1"] = "=rate"

    reasons = oracle._exclusion_seeds(wb)
    assert reasons[("Sheet", 1, 2)] == "external-link"


def test_template_conversion_and_template_destination_refuse(tmp_path, monkeypatch):
    template = _template_package()
    monkeypatch.setattr(oracle, "find_soffice", lambda: "/fake/soffice")
    with pytest.raises(UnsupportedStructureError, match="templates"):
        oracle.recalc(template)

    wb = Workbook()
    raw = io.BytesIO()
    wb.save(raw)
    with pytest.raises(UnsupportedStructureError, match="template"):
        oracle.recalc(raw.getvalue(), output_path=tmp_path / "result.xltx")


@pytest.mark.parametrize("token", sorted(oracle.ERROR_TOKENS))
def test_all_oracle_error_tokens_serialize_as_errors(token):
    cell_type, payload = _serialize_cached_value(token, None)
    assert cell_type == b"e"
    assert payload == token.encode("ascii")


def test_file_like_read_starts_at_zero_and_restores_cursor():
    stream = io.BytesIO(b"complete workbook bytes")
    stream.seek(9)
    assert oracle._read_source(stream) == b"complete workbook bytes"
    assert stream.tell() == 9


def test_nonseekable_file_like_source_refuses_instead_of_reading_a_suffix():
    class NonSeekable:
        def read(self):
            return b"partial package"

    with pytest.raises(ValueError, match="must be seekable"):
        oracle._read_source(NonSeekable())


def test_oracle_input_assignment_never_overwrites_formula():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "=1+1"

    with pytest.raises(UnsupportedStructureError, match="never overwrite"):
        oracle._set_input_cell(ws, 1, 1, 5, "Sheet!A1")
    assert ws["A1"].value == "=1+1"


def test_set_input_resolves_unique_local_name_and_refuses_ambiguity():
    wb = Workbook()
    first = wb.active
    first.title = "First"
    second = wb.create_sheet("Second")
    first.defined_names.add(DefinedName(
        "LocalInput", attr_text="'First'!$A$1"))

    assert wb.set_input("LocalInput", 7) is first["A1"]
    assert first["A1"].value == 7

    second.defined_names.add(DefinedName(
        "LocalInput", attr_text="'Second'!$A$1"))
    with pytest.raises(AmbiguousTargetError, match="exists on 2 sheets"):
        wb.set_input("LocalInput", 9)
    with pytest.raises(AmbiguousTargetError, match="exists on 2 sheets"):
        oracle._resolve_single_cell(wb, "LocalInput")


def test_set_input_missing_defined_name_sheet_is_typed():
    wb = Workbook()
    wb.defined_names.add(DefinedName(
        "Missing", attr_text="'Gone'!$A$1"))

    with pytest.raises(TargetNotFoundError, match="missing sheet"):
        wb.set_input("Missing", 1)


def test_data_table_result_range_is_excluded_as_one_unsupported_unit(
        monkeypatch):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = DataTableFormula(ref="A1:A2", r1="B1")
    ws["A2"] = 2
    raw = io.BytesIO()
    wb.save(raw)
    data = raw.getvalue().replace(b"<v></v>", b"<v>1</v>", 1)
    monkeypatch.setattr(oracle, "_recalculate_bytes",
                        lambda package, timeout: package)

    result = oracle.certify(data)
    assert result.status == "BASELINE_UNVERIFIABLE"
    assert result.checked == 0
    assert result.unsupported_excluded == [
        "Sheet!A1 (data-table)", "Sheet!A2 (data-table)"]


def test_pathological_multi_cell_formula_range_refuses_before_expansion():
    wb = Workbook()
    wb.active["A1"] = ArrayFormula(
        ref="A1:XFD1048576", text="=ROW(A1:XFD1048576)")
    with pytest.raises(UnsupportedStructureError, match="safety cap"):
        oracle._formula_result_cells(wb, wb)


def test_profile_uri_is_absolute_file_uri(tmp_path):
    uri = oracle._profile_uri(tmp_path / "profile with spaces")
    assert uri.startswith("file:")
    assert "%20" in uri


def test_process_session_kwargs_are_platform_specific(monkeypatch):
    monkeypatch.setattr(os, "name", "nt")
    monkeypatch.setattr(oracle.subprocess, "CREATE_NEW_PROCESS_GROUP", 512,
                        raising=False)
    assert oracle._popen_session_kwargs() == {"creationflags": 512}
    monkeypatch.setattr(os, "name", "posix")
    assert oracle._popen_session_kwargs() == {"start_new_session": True}
