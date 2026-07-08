#!/usr/bin/env python3
"""Fixture-corpus generator for tests/paper/fixtures.

THE CORPUS IS FROZEN. This script exists to document provenance and to
regenerate the corpus ONLY via an explicit human decision (CONVENTIONS §4:
golden files update only via explicit command with human-reviewed diffs).
Code under test never runs this. After any regeneration, MANIFEST.sha256
must be rewritten (--write-manifest) and the diff human-reviewed.

Provenance honesty (CONVENTIONS §4): everything here is authored by stock
openpyxl from this checkout, by zip surgery on such files, or by LibreOffice
conversion — and each sidecar says which. Nothing here may ever be labeled
Excel-authored. Real-Excel fixtures arrive via FIXTURE-REQUESTS.md only.

Requires LibreOffice (soffice) for: schedule_calc, lo_authored, legacy.xls,
and load-verification of every fixture.
"""
from __future__ import annotations

import hashlib
import io
import json
import os
import shutil
import subprocess
import sys
import tempfile
import zipfile
from datetime import date

HERE = os.path.dirname(os.path.abspath(__file__))
FIXTURES = os.path.dirname(HERE)
sys.path.insert(0, os.path.abspath(os.path.join(FIXTURES, "..", "..", "..")))

from openpyxl import Workbook, load_workbook  # noqa: E402
from openpyxl.chart import BarChart, Reference  # noqa: E402
from openpyxl.comments import Comment  # noqa: E402
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule  # noqa: E402
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.workbook.defined_name import DefinedName  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402
from openpyxl.worksheet.formula import ArrayFormula  # noqa: E402
from openpyxl.worksheet.table import Table, TableStyleInfo  # noqa: E402

TODAY = str(date.today())
PROG = "programmatic (generator assertions only; no human verification yet)"


# --------------------------------------------------------------------------
# helpers

def out(*parts):
    path = os.path.join(FIXTURES, *parts)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    return path


def sidecar(fixture_rel, provenance_app, provenance_notes, features, ground_truth):
    path = os.path.join(FIXTURES, fixture_rel + ".json")
    doc = {
        "fixture": os.path.basename(fixture_rel),
        "provenance": {"app": provenance_app, "version": "", "notes": provenance_notes},
        "features": features,
        "ground_truth": ground_truth,
        "verified_by": PROG,
        "date": TODAY,
    }
    with open(path, "w") as f:
        json.dump(doc, f, indent=2, sort_keys=False)
        f.write("\n")


def rewrite_zip(src, dst, transform):
    """Copy zip src->dst passing (name, payload) through transform; extras via
    transform(None, None) -> [(name, payload)]."""
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            payload = zin.read(item.filename)
            res = transform(item.filename, payload)
            if res is None:
                continue
            name, payload = res
            zout.writestr(name, payload)
        extra = transform(None, None)
        if extra:
            for name, payload in extra:
                zout.writestr(name, payload)


def find_soffice():
    for name in ("soffice", "libreoffice"):
        p = shutil.which(name)
        if p:
            return p
    mac = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
    return mac if os.path.exists(mac) else None


def lo_convert(src, fmt):
    soffice = find_soffice()
    assert soffice, "generator requires LibreOffice"
    work = tempfile.mkdtemp(prefix="paper_gen_lo_")
    try:
        profile = os.path.join(work, "profile")
        outdir = os.path.join(work, "out")
        os.makedirs(outdir)
        tmp_in = os.path.join(work, os.path.basename(src))
        shutil.copyfile(src, tmp_in)
        proc = subprocess.run(
            [soffice, "--headless", "-env:UserInstallation=file://" + profile,
             "--convert-to", fmt, "--outdir", outdir, tmp_in],
            capture_output=True, timeout=180, start_new_session=True,
        )
        stem = os.path.splitext(os.path.basename(src))[0]
        out_path = os.path.join(outdir, stem + "." + fmt.split(":")[0])
        if proc.returncode != 0 or not os.path.exists(out_path):
            raise RuntimeError(
                "LO convert failed rc={0} out={1} stderr={2!r}".format(
                    proc.returncode, os.path.exists(out_path), proc.stderr[-300:]))
        with open(out_path, "rb") as f:
            return f.read()
    finally:
        shutil.rmtree(work, ignore_errors=True)


def verify_loads(path, expect_lo=True):
    load_workbook(path)  # stock openpyxl must load every fixture
    if expect_lo:
        lo_convert(path, "xlsx")  # independent loader must load it too


def sheet_part_for(path, sheet_title_marker):
    """Return (part_name, payload) of the worksheet part containing marker."""
    with zipfile.ZipFile(path) as z:
        for name in z.namelist():
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
                payload = z.read(name)
                if sheet_title_marker in payload:
                    return name, payload
    raise AssertionError("no sheet part contains %r" % sheet_title_marker)


# --------------------------------------------------------------------------
# XML blocks for surgery (namespaces self-contained; mirror real-Excel shapes)

SPARKLINE_EXT = (
    b'<ext uri="{05C60535-1F16-4fd2-B633-F4F36F0B64E0}" '
    b'xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main">'
    b'<x14:sparklineGroups xmlns:xm="http://schemas.microsoft.com/office/excel/2006/main">'
    b'<x14:sparklineGroup displayEmptyCellsAs="gap">'
    b'<x14:colorSeries rgb="FF376092"/><x14:colorNegative rgb="FFD00000"/>'
    b'<x14:colorAxis rgb="FF000000"/><x14:colorMarkers rgb="FFD00000"/>'
    b'<x14:colorFirst rgb="FFD00000"/><x14:colorLast rgb="FFD00000"/>'
    b'<x14:colorHigh rgb="FFD00000"/><x14:colorLow rgb="FFD00000"/>'
    b'<x14:sparklines><x14:sparkline><xm:f>Model!B3:E3</xm:f><xm:sqref>F3</xm:sqref></x14:sparkline>'
    b'<x14:sparkline><xm:f>Model!B4:E4</xm:f><xm:sqref>F4</xm:sqref></x14:sparkline>'
    b'</x14:sparklines></x14:sparklineGroup></x14:sparklineGroups></ext>'
)

X14_CF_GUID = b"{DA7ABA51-AAAA-BBBB-CCCC-123456789012}"

X14_CF_EXT = (
    b'<ext uri="{78C0D931-6437-407d-A8EE-F0AAD7539E65}" '
    b'xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main">'
    b'<x14:conditionalFormattings>'
    b'<x14:conditionalFormatting xmlns:xm="http://schemas.microsoft.com/office/excel/2006/main">'
    b'<x14:cfRule type="dataBar" id="' + X14_CF_GUID + b'">'
    b'<x14:dataBar minLength="0" maxLength="100" negativeBarColorSameAsPositive="0">'
    b'<x14:cfvo type="autoMin"/><x14:cfvo type="autoMax"/>'
    b'<x14:negativeFillColor rgb="FFFF0000"/><x14:axisColor rgb="FF000000"/>'
    b'</x14:dataBar></x14:cfRule><xm:sqref>B6:E6</xm:sqref>'
    b'</x14:conditionalFormatting></x14:conditionalFormattings></ext>'
)

CF_RULE_X14_ID = (
    b'<extLst><ext uri="{B025F937-C7B1-47D3-B67F-A62EFF666E3E}" '
    b'xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main">'
    b'<x14:id>' + X14_CF_GUID + b'</x14:id></ext></extLst>'
)


def inject_sheet_extlst(payload, ext_blocks):
    """Append <extLst> with the given ext blocks before </worksheet>."""
    assert b"</worksheet>" in payload
    # a cfRule-level extLst (the x14 twin pointer) is fine; a SHEET-level one
    # (immediately before </worksheet>) would need merging, not appending
    assert not payload.rstrip().endswith(b"</extLst></worksheet>"), \
        "sheet already has a sheet-level extLst"
    blob = b"<extLst>" + b"".join(ext_blocks) + b"</extLst></worksheet>"
    return payload.replace(b"</worksheet>", blob)


def inject_cfrule_twin_pointer(payload):
    """Attach the x14:id twin pointer to the first dataBar cfRule."""
    marker = b"</dataBar></cfRule>"
    assert marker in payload, "no dataBar cfRule found"
    return payload.replace(marker, b"</dataBar>" + CF_RULE_X14_ID + b"</cfRule>", 1)


def make_shared_formulas(payload, host_ref, si, cells):
    """Convert a run of ordinary formula cells into a shared-formula group.

    ``cells`` = ordered list of (coordinate, formula_text); first is the host.
    Stock openpyxl writes '<c r="C2"><f>B2*2</f><v></v></c>' shapes.
    """
    host_coord, host_formula = cells[0]
    old = ("<f>" + host_formula + "</f>").encode()
    new = ('<f t="shared" ref="' + host_ref + '" si="' + str(si) + '">'
           + host_formula + "</f>").encode()
    assert old in payload, "host formula %r not found" % host_formula
    payload = payload.replace(old, new, 1)
    for coord, formula in cells[1:]:
        old = ("<f>" + formula + "</f>").encode()
        new = ('<f t="shared" si="' + str(si) + '"/>').encode()
        assert old in payload, "follower formula %r not found" % formula
        payload = payload.replace(old, new, 1)
    return payload


# --------------------------------------------------------------------------
# fixture builders

def build_minimal_clean():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Item", "Qty", "Price", "Total"])
    for i, (item, qty, price) in enumerate(
            [("apples", 4, 1.5), ("pears", 2, 2.25), ("plums", 10, 0.4)], start=2):
        ws["A{0}".format(i)] = item
        ws["B{0}".format(i)] = qty
        ws["C{0}".format(i)] = price
        ws["D{0}".format(i)] = "=B{0}*C{0}".format(i)
    ws["D5"] = "=SUM(D2:D4)"
    path = out("minimal", "minimal_clean.xlsx")
    wb.save(path)
    verify_loads(path)
    sidecar("minimal/minimal_clean.xlsx", "openpyxl 3.1.5 (this checkout)",
            "fresh Workbook; 4 formulas, no cached values", ["formulas"],
            {"formula_count": 4, "chart_count": 0, "pivot_count": 0, "vba_present": False})
    return path


def build_chart_image():
    wb = Workbook()
    ws = wb.active
    ws.title = "Model"
    ws.append(["Line", "Q1", "Q2", "Q3", "Q4"])
    ws.append(["Revenue", 100, 110, 121, 133])
    ws.append(["COGS", 40, 44, 48, 53])
    chart = BarChart()
    chart.title = "Revenue by quarter"
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_col=5, max_row=2),
                   titles_from_data=True)
    ws.add_chart(chart, "H2")
    from PIL import Image as PILImage
    from openpyxl.drawing.image import Image as XLImage
    png = os.path.join(tempfile.gettempdir(), "paper_gen_logo.png")
    PILImage.new("RGB", (24, 24), (200, 30, 30)).save(png)
    ws.add_image(XLImage(png), "H20")
    path = out("features", "chart_image.xlsx")
    wb.save(path)
    verify_loads(path)
    sidecar("features/chart_image.xlsx", "openpyxl 3.1.5 (this checkout)",
            "bar chart + embedded PNG; both openpyxl-modeled (stock round-trips them "
            "lossily-but-presently; real-Excel chart fixtures are in FIXTURE-REQUESTS.md)",
            ["chart", "image"],
            {"chart_count": 1, "image_count": 1, "formula_count": 0, "vba_present": False})
    return path


def build_defined_names():
    wb = Workbook()
    ws = wb.active
    ws.title = "Model"
    ws["A1"] = "Growth"
    ws["B1"] = 0.05
    ws["A2"] = "Base"
    ws["B2"] = 1000
    ws["A3"] = "Projected"
    ws["B3"] = "=Base*(1+Growth)"
    ws2 = wb.create_sheet("Data")
    ws2["A1"] = "Regional"
    ws2["B1"] = 250
    ws["A4"] = "Cross-sheet"
    ws["B4"] = "=Data!B1*2"
    wb.defined_names["Growth"] = DefinedName("Growth", attr_text="Model!$B$1")
    wb.defined_names["Base"] = DefinedName("Base", attr_text="Model!$B$2")
    # sheet-scoped defined name (localSheetId hazard, OPEN-QUESTIONS G7)
    ws2.defined_names["LocalTotal"] = DefinedName("LocalTotal", attr_text="Data!$B$1")
    path = out("features", "defined_names.xlsx")
    wb.save(path)
    verify_loads(path)
    sidecar("features/defined_names.xlsx", "openpyxl 3.1.5 (this checkout)",
            "global + sheet-scoped defined names; cross-sheet formula",
            ["defined_names", "cross_sheet_formulas", "sheet_scoped_defined_name"],
            {"formula_count": 2, "defined_names": ["Growth", "Base", "LocalTotal"],
             "vba_present": False})
    return path


def build_tables():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Region", "Amount"])
    for i, region in enumerate(["North", "South", "East", "West"], start=2):
        ws["A{0}".format(i)] = region
        ws["B{0}".format(i)] = i * 10
    tab = Table(displayName="RegionTable", ref="A1:B5")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tab)
    path = out("features", "tables.xlsx")
    wb.save(path)
    verify_loads(path)
    sidecar("features/tables.xlsx", "openpyxl 3.1.5 (this checkout)",
            "one ListObject/table", ["table"],
            {"table_names": ["RegionTable"], "formula_count": 0, "vba_present": False})
    return path


def build_merged():
    wb = Workbook()
    ws = wb.active
    ws.merge_cells("A1:D1")
    ws["A1"] = "Merged banner"
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A3:B4")
    ws["A3"] = "block"
    ws["E1"] = 1
    path = out("features", "merged.xlsx")
    wb.save(path)
    verify_loads(path)
    sidecar("features/merged.xlsx", "openpyxl 3.1.5 (this checkout)",
            "two merged ranges", ["merged_cells"],
            {"merged_ranges": ["A1:D1", "A3:B4"], "formula_count": 0, "vba_present": False})
    return path


def build_datavalidation():
    wb = Workbook()
    ws = wb.active
    dv = DataValidation(type="list", formula1='"Low,Base,High"', allow_blank=True)
    dv.add("B2")
    ws.add_data_validation(dv)
    ws["A2"] = "Scenario"
    ws["B2"] = "Base"
    dv2 = DataValidation(type="whole", operator="between", formula1="1", formula2="10")
    dv2.add("B3")
    ws.add_data_validation(dv2)
    ws["A3"] = "Level"
    ws["B3"] = 5
    path = out("features", "datavalidation.xlsx")
    wb.save(path)
    verify_loads(path)
    sidecar("features/datavalidation.xlsx", "openpyxl 3.1.5 (this checkout)",
            "list dropdown + whole-number DV", ["data_validation"],
            {"dv_count": 2, "formula_count": 0, "vba_present": False})
    return path


def build_hidden():
    wb = Workbook()
    ws = wb.active
    ws.title = "Visible"
    for i in range(1, 6):
        ws["A{0}".format(i)] = i
    ws.row_dimensions[3].hidden = True
    ws.column_dimensions["C"].hidden = True
    ws2 = wb.create_sheet("HiddenNotes")
    ws2["A1"] = "internal"
    ws2.sheet_state = "hidden"
    path = out("features", "hidden.xlsx")
    wb.save(path)
    verify_loads(path)
    sidecar("features/hidden.xlsx", "openpyxl 3.1.5 (this checkout)",
            "hidden row, hidden column, hidden sheet", ["hidden_rows", "hidden_sheet"],
            {"hidden_sheets": ["HiddenNotes"], "formula_count": 0, "vba_present": False})
    return path


def build_schedule():
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"
    ws.append(["Item", "Amount"])
    for i in range(2, 12):  # rows 2-11
        ws["A{0}".format(i)] = "Item {0}".format(i - 1)
        ws["B{0}".format(i)] = i * 100
    ws["A12"] = "TOTAL"
    ws["B12"] = "=SUM(B2:B11)"
    ws["A13"] = "Total x growth"
    ws["B13"] = "=B12*(1+Growth)"
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Grand total"
    ws2["B1"] = "=Schedule!B12"
    wb.defined_names["Growth"] = DefinedName("Growth", attr_text="Schedule!$B$15")
    ws["A15"] = "Growth"
    ws["B15"] = 0.05
    path = out("features", "schedule.xlsx")
    wb.save(path)
    verify_loads(path)
    sidecar("features/schedule.xlsx", "openpyxl 3.1.5 (this checkout)",
            "SUM schedule + defined name + cross-sheet ref; NO cached values "
            "(openpyxl never calculates) — battery jobs 3 and stale-value evidence",
            ["formulas", "defined_names", "cross_sheet_formulas"],
            {"formula_count": 3, "cached_values": {}, "vba_present": False})
    return path


def build_schedule_calc(schedule_path):
    data = lo_convert(schedule_path, "xlsx")
    path = out("features", "schedule_calc.xlsx")
    with open(path, "wb") as f:
        f.write(data)
    verify_loads(path)
    # ground truth from the recalculated cache
    wb = load_workbook(path, data_only=True)
    b12 = wb["Schedule"]["B12"].value
    b13 = wb["Schedule"]["B13"].value
    sb1 = wb["Summary"]["B1"].value
    assert b12 == 6500 and sb1 == 6500 and abs(b13 - 6825) < 1e-9, (b12, b13, sb1)
    sidecar("features/schedule_calc.xlsx", "LibreOffice (headless convert)",
            "schedule.xlsx recalculated by LibreOffice: real cached <v> values, "
            "LO producer XML (declaration, t=\"s\" shared strings, attr order) — "
            "battery job 5 and certification baselines",
            ["formulas", "cached_values", "shared_strings", "lo_producer"],
            {"formula_count": 3,
             "cached_values": {"Schedule!B12": 6500, "Schedule!B13": 6825,
                               "Summary!B1": 6500},
             "vba_present": False})
    return path


def build_macro_stub(schedule_path):
    vba_stub = (b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
                + b"PAPER-XLSX-SYNTHETIC-VBA-STUB" + b"\x00" * 4096)

    def transform(name, payload):
        if name is None:
            return [("xl/vbaProject.bin", vba_stub)]
        if name == "[Content_Types].xml":
            payload = payload.replace(
                b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
                b"application/vnd.ms-excel.sheet.macroEnabled.main+xml")
            payload = payload.replace(
                b"</Types>",
                b'<Override PartName="/xl/vbaProject.bin" '
                b'ContentType="application/vnd.ms-office.vbaProject"/></Types>')
        elif name == "xl/_rels/workbook.xml.rels":
            payload = payload.replace(
                b"</Relationships>",
                b'<Relationship Id="rIdVBA" '
                b'Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" '
                b'Target="vbaProject.bin"/></Relationships>')
        return name, payload

    path = out("features", "macro_stub.xlsm")
    rewrite_zip(schedule_path, path, transform)
    load_workbook(path, keep_vba=True)
    verify_loads(path)
    sidecar("features/macro_stub.xlsm", "openpyxl 3.1.5 + zip surgery",
            "SYNTHETIC vbaProject.bin stub (OLE magic + padding; not a real VBA "
            "project — real .xlsm is in FIXTURE-REQUESTS.md). Battery job 4.",
            ["vba_stub", "formulas"],
            {"formula_count": 3, "vba_present": True, "vba_is_stub": True})
    return path


def build_shared_formulas():
    wb = Workbook()
    ws = wb.active
    ws.title = "Calc"
    ws.append(["x", "double", "note"])
    for i in range(2, 7):  # rows 2-6
        ws["A{0}".format(i)] = i * 10
        ws["B{0}".format(i)] = "=A{0}*2".format(i)
    # a real array formula written by openpyxl's own model
    ws["D2"] = ArrayFormula("D2:D4", "=A2:A4*3")
    path = out("features", "shared_formulas.xlsx")
    wb.save(path)

    # surgery: convert B2:B6 into one shared-formula group (host + followers),
    # the byte shape real Excel writes for filled ranges (G1 in OPEN-QUESTIONS)
    part, payload = sheet_part_for(path, b"double")
    cells = [("B{0}".format(i), "A{0}*2".format(i)) for i in range(2, 7)]
    payload = make_shared_formulas(payload, "B2:B6", 0, cells)

    def transform(name, data):
        if name is None:
            return None
        if name == part:
            return name, payload
        return name, data

    tmp = path + ".tmp"
    rewrite_zip(path, tmp, transform)
    os.replace(tmp, path)

    # verification: group present in bytes; openpyxl expands followers; LO loads
    with zipfile.ZipFile(path) as z:
        sheet = z.read(part)
    assert b'<f t="shared" ref="B2:B6" si="0">' in sheet
    assert sheet.count(b'<f t="shared" si="0"/>') == 4
    assert b't="array" ref="D2:D4"' in sheet
    wb2 = load_workbook(path)
    assert wb2["Calc"]["B4"].value == "=A4*2", wb2["Calc"]["B4"].value
    verify_loads(path)
    sidecar("features/shared_formulas.xlsx", "openpyxl 3.1.5 + zip surgery",
            "shared-formula group B2:B6 (host + 4 si-followers, the real-Excel "
            "filled-range shape, via surgery) + a genuine openpyxl array formula "
            "D2:D4. Gates splice shared-group handling (OPEN-QUESTIONS G1).",
            ["shared_formulas", "array_formula", "formulas"],
            {"shared_groups": [{"si": 0, "ref": "B2:B6", "host": "B2"}],
             "array_formulas": [{"ref": "D2:D4"}],
             "formula_count": 6, "vba_present": False})
    return path


def build_gauntlet():
    wb = Workbook()
    ws = wb.active
    ws.title = "Model"
    ws.merge_cells("A1:F1")
    ws["A1"] = "Quarterly Model"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append(["Line", "Q1", "Q2", "Q3", "Q4", "Trend"])  # row 2
    for row in (("Revenue", 100, 110, 121, 133), ("COGS", 40, 44, 48, 53),
                ("Opex", 30, 31, 32, 33)):
        ws.append(row)  # rows 3-5
    ws["A6"] = "Profit"
    for col in "BCDE":
        ws["{0}6".format(col)] = "={0}3-{0}4-{0}5".format(col)
    ws["A8"] = "Growth rate"
    ws["B8"] = 0.10
    ws["B8"].comment = Comment("Input assumption: quarterly growth", "paper-xlsx")

    named = NamedStyle(name="paper_input")
    named.font = Font(italic=True)
    named.fill = PatternFill("solid", fgColor="FFF2CC")
    wb.add_named_style(named)
    ws["B8"].style = "paper_input"

    ws.conditional_formatting.add(
        "B3:E5", ColorScaleRule(start_type="min", start_color="FFAA0000",
                                end_type="max", end_color="FF00AA00"))
    ws.conditional_formatting.add(
        "B6:E6", DataBarRule(start_type="min", end_type="max", color="FF638EC6"))
    ws.conditional_formatting.add(
        "B8:B8", CellIsRule(operator="greaterThan", formula=["0.5"],
                            fill=PatternFill("solid", fgColor="FFFF0000")))

    dv = DataValidation(type="list", formula1='"Low,Base,High"', allow_blank=True)
    dv.add("B9")
    ws.add_data_validation(dv)
    ws["A9"] = "Scenario"
    ws["B9"] = "Base"

    ws.freeze_panes = "A3"
    ws["A11"] = "Docs"
    ws["A11"].hyperlink = "https://example.com/model-docs"
    ws.row_dimensions[10].hidden = True

    chart = BarChart()
    chart.title = "Revenue by quarter"
    chart.add_data(Reference(ws, min_col=2, min_row=2, max_col=5, max_row=3))
    ws.add_chart(chart, "H2")
    from PIL import Image as PILImage
    from openpyxl.drawing.image import Image as XLImage
    png = os.path.join(tempfile.gettempdir(), "paper_gen_logo2.png")
    PILImage.new("RGB", (24, 24), (30, 30, 200)).save(png)
    ws.add_image(XLImage(png), "H20")

    ws2 = wb.create_sheet("Data")
    ws2.append(["Region", "Amount"])
    for i, region in enumerate(["North", "South", "East", "West"], start=2):
        ws2["A{0}".format(i)] = region
        ws2["B{0}".format(i)] = i * 10
    tab = Table(displayName="RegionTable", ref="A1:B5")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws2.add_table(tab)
    ws["A12"] = "Region total"
    ws["B12"] = "=SUM(Data!B2:B5)"

    ws3 = wb.create_sheet("Calc")
    ws3.append(["x", "double"])
    for i in range(2, 7):
        ws3["A{0}".format(i)] = i * 10
        ws3["B{0}".format(i)] = "=A{0}*2".format(i)

    ws4 = wb.create_sheet("HiddenNotes")
    ws4["A1"] = "internal"
    ws4.sheet_state = "hidden"

    wb.defined_names["GrowthRate"] = DefinedName("GrowthRate", attr_text="Model!$B$8")
    ws2.defined_names["LocalTotal"] = DefinedName("LocalTotal", attr_text="Data!$B$5")

    path = out("gauntlet", "gauntlet.xlsx")
    wb.save(path)

    # surgery 1: sparklines + x14 CF twin block in the Model sheet's extLst
    model_part, payload = sheet_part_for(path, b"Quarterly Model")
    payload = inject_cfrule_twin_pointer(payload)
    payload = inject_sheet_extlst(payload, [SPARKLINE_EXT, X14_CF_EXT])
    # surgery 2: shared-formula group on Calc
    calc_part, calc_payload = sheet_part_for(path, b"double")
    cells = [("B{0}".format(i), "A{0}*2".format(i)) for i in range(2, 7)]
    calc_payload = make_shared_formulas(calc_payload, "B2:B6", 0, cells)

    def transform(name, data):
        if name is None:
            return None
        if name == model_part:
            return name, payload
        if name == calc_part:
            return name, calc_payload
        return name, data

    tmp = path + ".tmp"
    rewrite_zip(path, tmp, transform)
    os.replace(tmp, path)

    with zipfile.ZipFile(path) as z:
        model = z.read(model_part)
        assert b"sparklineGroups" in model
        assert b"x14:conditionalFormattings" in model
        assert b"<x14:id>" in model
        assert b"<drawing" in model  # chart attachment element (rels-driven)
    verify_loads(path)
    sidecar("gauntlet/gauntlet.xlsx", "openpyxl 3.1.5 + zip surgery",
            "real-model-shaped: chart, image, comment, named style, merges, "
            "3 CF rules (incl. dataBar with x14 twin pointer + x14 block), DV, "
            "freeze panes, hyperlink, hidden row + hidden sheet, table, global + "
            "sheet-scoped defined names, cross-sheet formulas, shared-formula "
            "group (surgery), sparklines (surgery). No cached values.",
            ["chart", "image", "comment", "named_style", "merged_cells",
             "conditional_formatting", "x14_cf_twin", "sparklines",
             "data_validation", "freeze_panes", "hyperlink", "hidden_rows",
             "hidden_sheet", "table", "defined_names",
             "sheet_scoped_defined_name", "cross_sheet_formulas",
             "shared_formulas"],
            {"chart_count": 1, "image_count": 1, "formula_count": 12,
             "shared_groups": [{"si": 0, "ref": "B2:B6", "host": "B2"}],
             "cached_values": {}, "vba_present": False})
    return path


def build_lo_authored(chart_image_path):
    data = lo_convert(chart_image_path, "xlsx")
    path = out("features", "lo_authored.xlsx")
    with open(path, "wb") as f:
        f.write(data)
    verify_loads(path)
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        has_sst = "xl/sharedStrings.xml" in names
    sidecar("features/lo_authored.xlsx", "LibreOffice (headless convert)",
            "chart_image.xlsx re-authored by LibreOffice: XML declarations, "
            "t=\"s\" shared strings ({0}), LO attribute order, xdr-prefixed "
            "drawing — producer-variance fixture".format(
                "sst present" if has_sst else "NO sst"),
            ["lo_producer", "chart", "image", "shared_strings"],
            {"vba_present": False})
    return path


def build_corrupt(minimal_path):
    trunc = out("corrupt", "truncated.xlsx")
    with open(minimal_path, "rb") as f:
        head = f.read(1000)
    with open(trunc, "wb") as f:
        f.write(head)
    notzip = out("corrupt", "not_a_zip.xlsx")
    with open(notzip, "wb") as f:
        f.write(b"this is not a zip archive, it just wears the extension\n" * 20)
    for p in (trunc, notzip):
        try:
            load_workbook(p)
            raise AssertionError("corrupt fixture unexpectedly loaded: %s" % p)
        except Exception:
            pass
    sidecar("corrupt/truncated.xlsx", "zip surgery",
            "first 1000 bytes of minimal_clean.xlsx (truncated central directory)",
            ["corrupt"], {})
    sidecar("corrupt/not_a_zip.xlsx", "zip surgery",
            "plain text wearing an .xlsx extension", ["corrupt"], {})
    return trunc, notzip


def build_legacy(minimal_path):
    xls = out("legacy", "legacy.xls")
    data = lo_convert(minimal_path, 'xls:MS Excel 97')
    with open(xls, "wb") as f:
        f.write(data)
    sidecar("legacy/legacy.xls", "LibreOffice (headless convert)",
            "minimal_clean.xlsx converted to real BIFF .xls — refusal tests only",
            ["legacy_xls"], {})

    xlsb = out("legacy", "binary.xlsb")
    made_real = False
    try:
        data = lo_convert(minimal_path, "xlsb")
        with open(xlsb, "wb") as f:
            f.write(data)
        made_real = True
    except Exception:
        # LibreOffice cannot write .xlsb: refusal triggers on extension before
        # content (reader/excel.py), so a labeled dummy suffices for the test.
        with open(xlsb, "wb") as f:
            f.write(b"PK\x03\x04PAPER-XLSX-DUMMY-XLSB-EXTENSION-TRIGGER-ONLY" + b"\x00" * 100)
    sidecar("legacy/binary.xlsb", "LibreOffice convert" if made_real else "dummy bytes",
            ("real LibreOffice-written xlsb" if made_real else
             "DUMMY bytes: openpyxl's refusal triggers on the extension before "
             "reading content; real .xlsb is in FIXTURE-REQUESTS.md"),
            ["legacy_xlsb"], {"real_xlsb": made_real})
    return xls, xlsb


def build_large(rows=10000, cols=15):
    wb = Workbook()
    ws = wb.active
    ws.title = "Big"
    header = ["Col{0}".format(c) for c in range(1, cols)] + ["RowSum"]
    ws.append(header)
    last_data_col = get_column_letter(cols - 1)
    for r in range(2, rows + 2):
        row = []
        for c in range(1, cols):
            if c % 5 == 0:
                row.append("label-{0}".format(r % 997))
            else:
                row.append(r * c % 10007)
        row.append("=SUM(A{0}:{1}{0})".format(r, last_data_col))
        ws.append(row)
    path = out("large", "large150k.xlsx")
    wb.save(path)
    verify_loads(path, expect_lo=False)  # LO verify skipped for speed; loads checked below
    sidecar("large/large150k.xlsx", "openpyxl 3.1.5 (this checkout)",
            "{0}x{1} = {2} cells incl. {3} SUM formulas and repeated strings — "
            "performance smoke (>=100k cells per CONVENTIONS taxonomy)".format(
                rows, cols, rows * cols, rows),
            ["large", "formulas"],
            {"rows": rows, "cols": cols, "cell_count": rows * cols,
             "formula_count": rows, "vba_present": False})
    return path


def write_manifest():
    lines = []
    for root, dirs, files in os.walk(FIXTURES):
        if os.path.basename(root) == "generators":
            dirs[:] = []
            continue
        for name in sorted(files):
            if name.endswith((".json", ".md", ".sha256")):
                continue
            path = os.path.join(root, name)
            rel = os.path.relpath(path, FIXTURES)
            with open(path, "rb") as f:
                digest = hashlib.sha256(f.read()).hexdigest()
            lines.append("{0}  {1}".format(digest, rel.replace(os.sep, "/")))
    manifest = os.path.join(FIXTURES, "MANIFEST.sha256")
    with open(manifest, "w") as f:
        f.write("\n".join(sorted(lines, key=lambda l: l.split("  ", 1)[1])) + "\n")
    print("MANIFEST.sha256: {0} fixtures frozen".format(len(lines)))


def main():
    minimal = build_minimal_clean()
    chart_image = build_chart_image()
    build_defined_names()
    build_tables()
    build_merged()
    build_datavalidation()
    build_hidden()
    schedule = build_schedule()
    build_schedule_calc(schedule)
    build_macro_stub(schedule)
    build_shared_formulas()
    build_gauntlet()
    build_lo_authored(chart_image)
    build_corrupt(minimal)
    build_legacy(minimal)
    build_large()
    write_manifest()
    print("corpus generated OK")


if __name__ == "__main__":
    main()
