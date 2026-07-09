"""Batch 4 (PLAN-v0.1, PR-1 §3): charts and images under preserve —
creation on added sheets, fresh drawings on loaded sheets, appends into
anchor-only originals, and per-property chart edits via chartpatch."""
from __future__ import annotations

import io
import zipfile

import pytest

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.errors import UnsupportedStructureError

from .support.partdiff import part_payloads

def _png_image():
    from PIL import Image as PILImageMod

    from openpyxl.drawing.image import Image

    buf = io.BytesIO()
    PILImageMod.new("RGB", (1, 1), "red").save(buf, format="png")
    buf.seek(0)
    return Image(buf)


def _chart_for(ws, min_col=1, min_row=1, max_row=3):
    chart = BarChart()
    chart.add_data(Reference(ws, min_col=min_col, min_row=min_row,
                             max_row=max_row))
    return chart


class TestAddedSheetDrawings:

    def test_chart_and_image_on_added_sheet(self, fixture_copy, tmp_path):
        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb.create_sheet("Report")
        for i in range(1, 6):
            ws.cell(row=i, column=2, value=i * 10)
        ws.add_chart(_chart_for(ws, min_col=2, max_row=5), "D2")
        ws.add_image(_png_image(), "A8")
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        payloads = part_payloads(out)
        assert any(n.startswith("xl/drawings/drawing") for n in payloads)
        assert any(n.startswith("xl/charts/chart") for n in payloads)
        assert any(n.startswith("xl/media/image") for n in payloads)
        assert b'ContentType="image/png"' in payloads["[Content_Types].xml"] \
            or b'image/png' in payloads["[Content_Types].xml"]
        wb2 = load_workbook(out)
        assert len(wb2["Report"]._charts) == 1
        assert len(wb2["Report"]._images) == 1
        # the loaded sheets are untouched bytes
        wb3 = load_workbook(out)
        assert wb3["Sheet1"]["A1"].value == \
            load_workbook(src)["Sheet1"]["A1"].value

    def test_added_sheet_chart_plus_comment_coexist(self, fixture_copy,
                                                    tmp_path):
        from openpyxl.comments import Comment

        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb.create_sheet("Both")
        ws["A1"] = 1
        ws.add_chart(_chart_for(ws, max_row=1), "C1")
        ws["A1"].comment = Comment("note", "author")
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert len(wb2["Both"]._charts) == 1
        assert wb2["Both"]["A1"].comment is not None


class TestLoadedSheetFreshDrawing:

    def test_chart_on_machinery_free_sheet(self, fixture_copy, tmp_path):
        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Sheet1"]
        ws.add_chart(_chart_for(ws), "F2")
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        payloads = part_payloads(out)
        sheet = next(p for n, p in payloads.items()
                     if n.startswith("xl/worksheets/sheet"))
        assert b"<drawing" in sheet
        assert sheet.count(b"<drawing") == 1
        wb2 = load_workbook(out)
        assert len(wb2["Sheet1"]._charts) == 1
        assert wb2["Sheet1"]["A1"].value is not None   # content intact

    def test_image_on_machinery_free_sheet(self, fixture_copy, tmp_path):
        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Sheet1"].add_image(_png_image(), "E5")
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert len(wb2["Sheet1"]._images) == 1

    def test_add_chart_composes_with_cell_edits(self, fixture_copy,
                                                tmp_path):
        src = fixture_copy("features/schedule.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Schedule"]
        ws["A1"] = "edited"
        ws.add_chart(_chart_for(ws, min_col=2, max_row=4), "H2")
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert wb2["Schedule"]["A1"].value == "edited"
        assert len(wb2["Schedule"]._charts) == 1


class TestExistingDrawingAppend:

    def test_append_chart_preserves_existing_anchors(self, fixture_copy,
                                                     tmp_path):
        src = fixture_copy("features/chart_image.xlsx")
        with zipfile.ZipFile(src) as z:
            original_drawing = next(
                z.read(n) for n in z.namelist()
                if n.startswith("xl/drawings/drawing"))
        wb = load_workbook(src, preserve=True)
        ws = wb["Model"]
        before = len(ws._charts)
        ws.add_chart(_chart_for(ws, min_col=2, max_row=5), "K2")
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        payloads = part_payloads(out)
        new_drawing = next(p for n, p in payloads.items()
                           if n.startswith("xl/drawings/drawing"))
        # every original anchor byte survives verbatim, in order
        head = original_drawing[:original_drawing.rfind(b"</")]
        assert new_drawing.startswith(head)
        wb2 = load_workbook(out)
        assert len(wb2["Model"]._charts) == before + 1
        assert len(wb2["Model"]._images) == 1          # image survived

    def test_append_image_into_existing_drawing(self, fixture_copy,
                                                 tmp_path):
        src = fixture_copy("features/chart_image.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Model"]
        before = len(ws._images)
        ws.add_image(_png_image(), "M2")
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert len(wb2["Model"]._images) == before + 1

    def test_non_anchor_drawing_refuses_at_add_time(self, fixture_copy,
                                                    tmp_path):
        src = fixture_copy("features/chart_image.xlsx")
        # plant a top-level shape into the drawing: no longer anchor-only
        out = str(tmp_path / "shaped.xlsx")
        with zipfile.ZipFile(src) as zin, zipfile.ZipFile(out, "w") as zout:
            for name in zin.namelist():
                payload = zin.read(name)
                if name.startswith("xl/drawings/drawing"):
                    payload = payload.replace(
                        b"</wsDr>", b"<sp><nvSpPr/></sp></wsDr>", 1)
                zout.writestr(name, payload)
        wb = load_workbook(out, preserve=True)
        ws = wb["Model"]
        with pytest.raises(UnsupportedStructureError, match="anchor"):
            ws.add_chart(_chart_for(ws), "K2")
        assert len(ws._charts) == 1                    # atomic: not added


class TestChartPropertyEdits:

    def test_repoint_patches_series_bytes(self, fixture_copy, tmp_path):
        src = fixture_copy("features/chart_image.xlsx")
        wb = load_workbook(src, preserve=True)
        chart = wb["Model"]._charts[0]
        chart.repoint(0, "Model!$B$1:$B$4")
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        chart_xml = next(p for n, p in part_payloads(out).items()
                         if n.startswith("xl/charts/chart"))
        assert b"Model!$B$1:$B$4" in chart_xml \
            or b"'Model'!$B$1:$B$4" in chart_xml

    def test_repoint_validates_range(self, fixture_copy):
        src = fixture_copy("features/chart_image.xlsx")
        wb = load_workbook(src, preserve=True)
        chart = wb["Model"]._charts[0]
        with pytest.raises(ValueError, match="sheet-qualified"):
            chart.repoint(0, "$B$1:$B$4")              # no sheet
        with pytest.raises(ValueError, match="sheet-qualified"):
            chart.repoint(0, "Model!$B$1:$B$4,Model!$D$1")  # multi-area
        with pytest.raises(ValueError, match="series"):
            chart.repoint(5, "Model!$B$1:$B$4")        # no such series

    def test_repoint_to_missing_sheet_refuses_at_save(self, fixture_copy,
                                                      tmp_path):
        src = fixture_copy("features/chart_image.xlsx")
        with open(src, "rb") as f:
            before = f.read()
        wb = load_workbook(src, preserve=True)
        wb["Model"]._charts[0].repoint(0, "Nowhere!$B$1:$B$4")
        with pytest.raises(UnsupportedStructureError, match="Nowhere"):
            wb.save(str(tmp_path / "o.xlsx"))
        with open(src, "rb") as f:
            assert f.read() == before

    def test_title_edit_lands_and_reloads(self, fixture_copy, tmp_path):
        src = fixture_copy("features/chart_image.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Model"]._charts[0].title = "New & Improved <Title>"
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        chart = wb2["Model"]._charts[0]
        runs = chart.title.tx.rich.p[0].r
        assert runs[0].t == "New & Improved <Title>"   # escape round-trip

    def test_inexpressible_property_refuses_named(self, fixture_copy,
                                                  tmp_path):
        src = fixture_copy("features/chart_image.xlsx")
        with open(src, "rb") as f:
            before = f.read()
        wb = load_workbook(src, preserve=True)
        wb["Model"]._charts[0].style = 33
        with pytest.raises(UnsupportedStructureError, match="style"):
            wb.save(str(tmp_path / "o.xlsx"))
        with open(src, "rb") as f:
            assert f.read() == before

    def test_series_add_refuses(self, fixture_copy, tmp_path):
        from openpyxl.chart import Series

        src = fixture_copy("features/chart_image.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Model"]
        chart = ws._charts[0]
        chart.append(Series(Reference(ws, min_col=3, min_row=2, max_row=5)))
        with pytest.raises(UnsupportedStructureError,
                           match="added or removed"):
            wb.save(str(tmp_path / "o.xlsx"))

    def test_shift_plus_property_edit_refuses(self, fixture_copy,
                                              tmp_path):
        # the shift already rewrote the chart's <c:f> texts; a same-session
        # property edit cannot verify against the arm state — refuse, never
        # guess (separate sessions compose fine)
        src = fixture_copy("features/chart_image.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Model"]
        chart = ws._charts[0]
        ws.insert_rows(1)
        chart.repoint(0, "Model!$D$1:$D$4")
        with pytest.raises(UnsupportedStructureError,
                           match="separate sessions"):
            wb.save(str(tmp_path / "o.xlsx"))
