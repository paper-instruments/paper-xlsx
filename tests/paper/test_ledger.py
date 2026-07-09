"""Phase 2b: the dirty ledger — chokepoint marking, refusals, arming
semantics (CONVENTIONS §3.3; PR-0 D5/D7/D8)."""
from __future__ import annotations

import pytest

from openpyxl import load_workbook
from openpyxl.errors import (
    PaperRefusal,
    TargetNotFoundError,
    UnsupportedStructureError,
)
from openpyxl.styles import Font, NamedStyle, PatternFill


@pytest.fixture
def preserved(fixture_copy):
    """A preserve-mode gauntlet workbook plus its ledger."""
    wb = load_workbook(fixture_copy("gauntlet/gauntlet.xlsx"), preserve=True)
    return wb, wb._paper_ledger


def dirty(led, ws):
    return led.dirty_coordinates(ws)


class TestArming:

    def test_ledger_armed_and_empty_after_load(self, preserved):
        wb, led = preserved
        assert led.armed
        assert led.cells == {}                  # loading itself fired no dirt
        assert not led.formulas_changed
        assert led.loaded_sheet_titles == frozenset(wb.sheetnames)

    def test_stock_workbook_has_no_ledger(self, fixture_copy):
        wb = load_workbook(fixture_copy("minimal/minimal_clean.xlsx"))
        assert wb._paper_ledger is None
        wb["Sheet1"]["A1"] = "mutate freely"    # no ledger, no refusals

    def test_reads_and_materialization_do_not_dirty(self, preserved):
        wb, led = preserved
        ws = wb["Model"]
        _ = ws["Z99"]                    # materializes an empty cell
        _ = ws.row_dimensions[55]        # materializes a dimension
        for _row in ws.iter_rows(min_row=1, max_row=30):
            pass
        _ = ws["B8"].font                # style read
        assert led.cells == {}, "reads must never mark dirt"


class TestCellChokepoints:

    def test_value_set_marks_cell(self, preserved):
        wb, led = preserved
        wb["Model"]["B8"] = 0.15
        assert (8, 2) in dirty(led, wb["Model"])
        assert not led.formulas_changed

    def test_new_formula_flags_formulas(self, preserved):
        wb, led = preserved
        wb["Model"]["B99"] = "=SUM(1,2)"        # new formula
        assert led.formulas_changed

    def test_overwriting_a_formula_flags_formulas(self, preserved):
        wb, led = preserved
        assert wb["Model"]["B6"].data_type == "f"
        wb["Model"]["B6"] = 42                   # formula -> literal
        assert led.formulas_changed

    def test_cell_helper_and_setitem_route_through(self, preserved):
        wb, led = preserved
        ws = wb["Data"]
        ws.cell(row=9, column=1, value="via cell()")
        ws["A10"] = "via setitem"
        assert {(9, 1), (10, 1)} <= dirty(led, ws)

    def test_append_marks_cells(self, preserved):
        wb, led = preserved
        ws = wb["Data"]
        before_max = ws.max_row
        ws.append(["x", None, 3])
        marked = dirty(led, ws)
        assert (before_max + 1, 1) in marked
        assert (before_max + 1, 3) in marked
        # None cells carry nothing to splice and are not marked
        assert (before_max + 1, 2) not in marked

    def test_delete_cell_marks_and_flags_formula(self, preserved):
        wb, led = preserved
        ws = wb["Model"]
        assert ws["B6"].data_type == "f"
        del ws["B6"]
        assert (6, 2) in dirty(led, ws)
        assert led.formulas_changed

    def test_style_assignment_marks_cell(self, preserved):
        wb, led = preserved
        ws = wb["Model"]
        ws["A2"].font = Font(bold=True)
        ws["A3"].number_format = "0.00%"
        ws["A4"].style = "paper_input"           # named style, exists in fixture
        assert {(2, 1), (3, 1), (4, 1)} <= dirty(led, ws)

    def test_hyperlink_and_comment_mark_cell(self, preserved):
        from openpyxl.comments import Comment

        wb, led = preserved
        ws = wb["Data"]
        ws["B2"].hyperlink = "https://example.org/x"
        ws["B3"].comment = Comment("note", "author")
        assert {(2, 2), (3, 2)} <= dirty(led, ws)

    def test_data_type_direct_set_marks_and_flags(self, preserved):
        wb, led = preserved
        ws = wb["Model"]
        cell = ws["B6"]                          # a formula cell
        assert cell.data_type == "f"
        cell.data_type = "s"                     # silent demotion chokepoint
        assert (6, 2) in dirty(led, ws)
        assert led.formulas_changed


class TestStructuralRefusals:

    @pytest.mark.parametrize("op,args", [
        ("insert_rows", (3,)),
        ("insert_cols", (2,)),
        ("delete_rows", (3,)),
        ("delete_cols", (2,)),
    ])
    def test_shift_operations_refuse_before_mutating(self, preserved, op, args):
        wb, led = preserved
        ws = wb["Model"]
        b6 = ws["B6"].value
        max_row = ws.max_row
        with pytest.raises(UnsupportedStructureError, match=op):
            getattr(ws, op)(*args)
        assert ws["B6"].value == b6              # model untouched
        assert ws.max_row == max_row
        assert led.cells == {}                   # and no dirt recorded

    def test_move_range_refuses(self, preserved):
        wb, _ = preserved
        with pytest.raises(UnsupportedStructureError, match="move_range"):
            wb["Model"].move_range("B3:C4", rows=2)

    def test_structural_edits_allowed_on_added_sheets(self, preserved):
        wb, _ = preserved
        ws = wb.create_sheet("Scratch")
        ws.append([1, 2, 3])
        ws.insert_rows(1)                        # generated whole at save
        ws.delete_cols(2)

    def test_refusals_are_paper_refusals(self, preserved):
        wb, _ = preserved
        with pytest.raises(PaperRefusal):
            wb["Model"].insert_rows(1)


class TestSheetLifecycle:

    def test_create_sheet_allowed_and_recorded(self, preserved):
        wb, led = preserved
        ws = wb.create_sheet("Appended")
        assert ws in led.added_sheets
        assert not led.is_loaded_sheet(ws)

    def test_remove_loaded_sheet_refuses(self, preserved):
        wb, _ = preserved
        with pytest.raises(UnsupportedStructureError, match="removing sheet"):
            wb.remove(wb["Data"])
        assert "Data" in wb.sheetnames
        with pytest.raises(UnsupportedStructureError):
            del wb["Data"]

    def test_remove_added_sheet_is_a_net_noop(self, preserved):
        wb, led = preserved
        ws = wb.create_sheet("Temp")
        ws["A1"] = 1
        wb.remove(ws)
        assert "Temp" not in wb.sheetnames
        assert ws not in led.added_sheets
        assert ws not in led.cells

    def test_move_sheet_records_reorder(self, preserved):
        # FLIPPED by v0.1 Batch 3 (was a refusal): reorder is expressed at
        # save by rebuilding the sheets element from original entry bytes
        wb, led = preserved
        before = list(wb.sheetnames)
        wb.move_sheet("Data", -1)
        after = list(wb.sheetnames)
        assert set(before) == set(after) and before != after

    def test_copy_worksheet_registers_as_added(self, preserved):
        # FLIPPED by v0.1 Batch 3 (was a refusal): the copy is an ADDED
        # sheet, generated whole at save (battery job 11 covers the file)
        wb, led = preserved
        cp = wb.copy_worksheet(wb["Data"])
        assert cp in led.added_sheets

    def test_rename_cascades_and_added_sheet_still_free(self, preserved):
        # FLIPPED by v0.1 Batch 3 (was a refusal): loaded-sheet renames
        # cascade (full coverage in battery job 8); in-session sheets
        # rename with no ledger involvement at all
        wb, led = preserved
        wb["Data"].title = "Records"
        assert "Records" in wb.sheetnames
        assert "Records" in led.loaded_sheet_titles     # still LOADED
        assert "Data" not in led.loaded_sheet_titles
        ws = wb.create_sheet("New")
        ws.title = "Renamed"                     # in-session sheets may rename
        assert ws.title == "Renamed"
        assert ws not in led.renames             # no cascade recorded


class TestMarkDirty:

    def test_range_form(self, preserved):
        wb, led = preserved
        wb.mark_dirty("Model!B2:C3")
        assert {(2, 2), (2, 3), (3, 2), (3, 3)} == dirty(led, wb["Model"])

    def test_single_cell_and_quoted_title(self, preserved):
        wb, led = preserved
        wb.mark_dirty("'Model'!B7")
        assert (7, 2) in dirty(led, wb["Model"])

    def test_part_form(self, preserved):
        wb, led = preserved
        wb.mark_dirty("xl/media/image1.png")
        assert "xl/media/image1.png" in led.parts

    def test_unknown_sheet_and_part_raise_target_not_found(self, preserved):
        wb, _ = preserved
        with pytest.raises(TargetNotFoundError):
            wb.mark_dirty("Nope!A1")
        with pytest.raises(TargetNotFoundError):
            wb.mark_dirty("xl/media/missing.png")

    def test_stock_workbook_raises_value_error(self, fixture_copy):
        wb = load_workbook(fixture_copy("minimal/minimal_clean.xlsx"))
        with pytest.raises(ValueError, match="preserve=True"):
            wb.mark_dirty("Sheet1!A1")

    def test_non_string_raises_type_error(self, preserved):
        wb, _ = preserved
        with pytest.raises(TypeError):
            wb.mark_dirty(42)


class TestStyleRegistryGuard:

    def test_in_place_mutation_of_shared_style_detected(self, preserved):
        wb, led = preserved
        led.check_style_registry(wb)             # clean workbook passes
        # the StyleProxy nested-object leak: mutates the SHARED interned fill
        ws = wb["Model"]
        ws["B8"].fill.start_color.rgb = "FF12AB34"
        with pytest.raises(UnsupportedStructureError, match="mutated in place"):
            led.check_style_registry(wb)

    def test_appending_new_styles_is_legal(self, preserved):
        wb, led = preserved
        wb["Model"]["A2"].font = Font(name="Menlo", size=9)      # new font
        wb["Model"]["A3"].fill = PatternFill("solid", fgColor="FF00AA00")
        led.check_style_registry(wb)             # appends never trip the guard

    def test_new_named_style_is_legal(self, preserved):
        wb, led = preserved
        ns = NamedStyle(name="fresh_style")
        ns.font = Font(italic=True)
        wb.add_named_style(ns)
        led.check_style_registry(wb)


class TestSheetLifecycleCascade:
    """PLAN-v0.1 3.2: delete and reorder on LOADED sheets."""

    def test_remove_clean_sheet_cascades(self, fixture_copy, tmp_path):
        from openpyxl import load_workbook as _load

        # gauntlet 'Notes' sheet: nothing references it
        src = fixture_copy("gauntlet/gauntlet.xlsx")
        wb = _load(src, preserve=True)
        victim_title = next(t for t in wb.sheetnames
                            if t not in ("Model", "Data", "Summary")
                            and not wb[t]._charts)
        report = wb.remove(wb[victim_title])
        assert report is not None
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = _load(out)
        assert victim_title not in wb2.sheetnames
        assert "Model" in wb2.sheetnames
        # the traps on surviving sheets are intact
        from .support.partdiff import part_payloads

        sheet = next(p for n, p in part_payloads(out).items()
                     if b"Quarterly Model" in p)
        assert b"sparklineGroups" in sheet

    def test_remove_referenced_sheet_refuses_with_enumeration(
            self, fixture_copy, tmp_path):
        from openpyxl import load_workbook as _load

        src = fixture_copy("features/schedule.xlsx")
        with open(src, "rb") as f:
            before = f.read()
        wb = _load(src, preserve=True)
        # Summary!B1 references Schedule -> audit must refuse
        with pytest.raises(UnsupportedStructureError, match="Summary!B1"):
            wb.remove(wb["Schedule"])
        assert "Schedule" in wb.sheetnames          # nothing changed
        with open(src, "rb") as f:
            assert f.read() == before

    def test_reorder_round_trips_with_scoped_names(self, fixture_copy,
                                                   tmp_path):
        from openpyxl import load_workbook as _load

        src = fixture_copy("features/defined_names.xlsx")
        wb = _load(src, preserve=True)
        original_order = list(wb.sheetnames)
        wb.move_sheet(original_order[0], 1)
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = _load(out)
        assert wb2.sheetnames == [original_order[1], original_order[0]] \
            + original_order[2:]
        # sheet-scoped names still resolve on their sheets
        for ws in wb2.worksheets:
            for name, dn in ws.defined_names.items():
                assert dn.value
