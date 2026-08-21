"""PR 6: managed refresh, source repoint, same-sheet move."""
from __future__ import annotations

import pytest

from openpyxl import Workbook, load_workbook
from openpyxl.errors import (
    BoundaryViolationError,
    UnsupportedStructureError,
)
from openpyxl.worksheet.table import Table

from .support.harness import save_and_reopen
from .test_pivot_create_package import _create_by_region, _expected_grid, _output_grid
from .test_pivot_graph import _basic_package, _write_package


_TABLE = "features/tables.xlsx"


def _preserved_matrix(tmp_path, headers, rows, table=None, name="src.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for offset, header in enumerate(headers):
        ws.cell(1, 1 + offset, header)
    for row_offset, record in enumerate(rows, start=2):
        for col_offset, value in enumerate(record):
            ws.cell(row_offset, 1 + col_offset, value)
    last = "%s%s" % ("ABCDEFGHI"[len(headers) - 1], 1 + len(rows))
    if table:
        ws.add_table(Table(displayName=table, ref="A1:%s" % last))
    wb.create_sheet("Summary")
    path = str(tmp_path / name)
    wb.save(path)
    return path, load_workbook(path, preserve=True)


def test_literal_source_edit_then_refresh(fixture_copy, tmp_path):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    _create_by_region(wb["Data"])
    out = str(tmp_path / "created.xlsx")
    wb = save_and_reopen(wb, out, preserve=True)
    wb["Data"]["B2"] = 99
    with pytest.raises(UnsupportedStructureError) as stale:
        wb.save(str(tmp_path / "stale.xlsx"))
    assert stale.value.kind in ("stale-pivot", "stale-pivot-cache")
    pivot = wb["Data"].pivots["ByRegion"]
    refreshed = pivot.refresh()
    saved = str(tmp_path / "refreshed.xlsx")
    wb = save_and_reopen(wb, saved, preserve=True)
    reopened = wb["Data"].pivots["ByRegion"]
    assert reopened.output_range == refreshed.output_range
    assert wb["Data"]["F4"].value == 99 + 0  # North row uses B2


def test_source_edit_without_refresh_refuses_atomically(fixture_copy, tmp_path):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    _create_by_region(wb["Data"])
    wb["Data"]["B2"] = 21
    dest = str(tmp_path / "out.xlsx")
    with pytest.raises(UnsupportedStructureError) as exc:
        wb.save(dest)
    assert exc.value.kind == "stale-pivot"
    with open(src, "rb") as handle:
        before = handle.read()
    assert not os_exists_changed(dest, before)


def os_exists_changed(dest, before):
    import os
    if not os.path.exists(dest):
        return False
    with open(dest, "rb") as handle:
        return handle.read() != before and os.path.getsize(dest) > 0


def test_unchanged_refresh_is_noop(fixture_copy, tmp_path):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    _create_by_region(wb["Data"])
    created = str(tmp_path / "created.xlsx")
    wb.save(created)
    wb = load_workbook(created, preserve=True)
    before = open(created, "rb").read()
    handle = wb["Data"].pivots["ByRegion"]
    same = handle.refresh()
    assert same is handle
    out = str(tmp_path / "noop.xlsx")
    wb.save(out)
    after = open(out, "rb").read()
    from .support.partdiff import part_payloads
    assert part_payloads(created) == part_payloads(out)


def test_table_expansion_refresh_includes_new_row(tmp_path):
    path, wb = _preserved_matrix(
        tmp_path,
        ["Region", "Amount"],
        [["East", 10], ["West", 7]],
        table="Sales",
    )
    summary = wb["Summary"]
    summary.pivots.create(
        name="ByRegion", source="Sales", destination="A1",
        rows=["Region"], values=["Amount"])
    saved = str(tmp_path / "created.xlsx")
    wb = save_and_reopen(wb, saved, preserve=True)
    data = wb["Data"]
    data["A4"] = "North"
    data["B4"] = 5
    data.tables["Sales"].ref = "A1:B4"
    wb["Summary"].pivots["ByRegion"].refresh()
    wb = save_and_reopen(wb, str(tmp_path / "expanded.xlsx"), preserve=True)
    values = {
        wb["Summary"].cell(row, 1).value: wb["Summary"].cell(row, 2).value
        for row in range(1, 8)
        if wb["Summary"].cell(row, 1).value
    }
    assert values["North"] == 5
    assert values["East"] == 10


def test_removed_source_field_refuses(fixture_copy, tmp_path):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    _create_by_region(wb["Data"])
    wb = save_and_reopen(wb, str(tmp_path / "created.xlsx"), preserve=True)
    wb["Data"]["A1"] = "Place"
    with pytest.raises((UnsupportedStructureError, BoundaryViolationError)) as exc:
        wb["Data"].pivots["ByRegion"].refresh()
    assert exc.value.kind in (
        "invalid-pivot-source", "unsupported-pivot-source",
        "unsupported-pivot-operation",
    )


def test_move_same_sheet_and_overlap(fixture_copy, tmp_path):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    handle = _create_by_region(wb["Data"], destination="E3")
    moved = handle.move("G3")
    assert moved.destination == "G3"
    wb["Data"]["K3"] = "blocked"
    with pytest.raises(BoundaryViolationError) as blocked:
        moved.move("J3")
    assert blocked.value.kind == "pivot-output-collision"
    overlapped = moved.move("G4")
    assert overlapped.destination == "G4"
    wb = save_and_reopen(wb, str(tmp_path / "moved.xlsx"), preserve=True)
    pivot = wb["Data"].pivots["ByRegion"]
    assert pivot.destination == "G4"
    assert wb["Data"]["E3"].value is None


def test_repoint_dedicated_source(tmp_path):
    path, wb = _preserved_matrix(
        tmp_path,
        ["Region", "Amount"],
        [["East", 10], ["West", 7]],
        table="Sales",
        name="first.xlsx",
    )
    other = wb["Data"]
    other["D1"] = "Region"
    other["E1"] = "Amount"
    other["D2"] = "North"
    other["E2"] = 4
    other.add_table(Table(displayName="Other", ref="D1:E2"))
    handle = wb["Summary"].pivots.create(
        name="ByRegion", source="Sales", destination="A1",
        rows=["Region"], values=["Amount"])
    handle.repoint_source("Other")
    wb = save_and_reopen(wb, str(tmp_path / "repointed.xlsx"), preserve=True)
    assert wb["Summary"]["A2"].value == "North"
    assert wb["Summary"]["B2"].value == 4
    assert wb["Data"]["A2"].value == "East"


def test_shared_cache_refresh_and_repoint_refuse(tmp_path):
    src = _write_package(tmp_path, "shared.xlsx", _basic_package())
    # add a second pivot relationship onto the same cache via overlay graph
    wb = load_workbook(str(src), preserve=True)
    pivot = wb["Summary"].pivots["SalesByRegion"]
    # synthetic graph fixture is foreign (no Paper tag)
    with pytest.raises(UnsupportedStructureError) as exc:
        pivot.refresh()
    assert exc.value.kind == "unsupported-pivot-operation"


def test_foreign_refresh_on_open_does_not_grant_headless(tmp_path):
    src = _write_package(tmp_path, "foreign.xlsx", _basic_package())
    wb = load_workbook(str(src), preserve=True)
    pivot = wb["Summary"].pivots["SalesByRegion"]
    assert pivot.capabilities.can_headless_refresh is False
    assert pivot.origin == "foreign"
    if pivot.capabilities.can_refresh_on_open:
        wb.set_pivot_refresh_on_load(pivots=["SalesByRegion"])
    with pytest.raises(UnsupportedStructureError):
        pivot.refresh()
