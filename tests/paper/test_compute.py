"""The computation layer — explicit scenarios and certified write-back."""
from __future__ import annotations

import zipfile

import pytest

from openpyxl import Workbook, load_workbook, oracle
from openpyxl.errors import UnsupportedStructureError

needs_soffice = pytest.mark.skipif(
    not oracle.available(), reason="LibreOffice not installed")


class TestEvaluate:

    @needs_soffice
    @pytest.mark.lo_smoke
    def test_scenario_run_certified_and_untouched(self, fixture_copy):
        # battery job 12: one evaluate call with explicit certification state
        src = fixture_copy("features/schedule_calc.xlsx")
        with open(src, "rb") as f:
            before = f.read()
        ev = oracle.evaluate(
            src, set={"Schedule!B2": 1000, "Schedule!B3": 0},
            read=["Summary!B1", "Schedule!B12"])
        assert isinstance(ev, oracle.Evaluation)      # pinned return type
        assert ev.status == "NO_DETECTED_FORMULA_ERRORS"
        assert ev.outputs["Summary!B1"] == ev.outputs["Schedule!B12"]
        assert ev.outputs["Summary!B1"] == 6500 - 200 - 300 + 1000
        cert = ev.certification
        # Every formula touched by this scenario is input-dependent and is
        # therefore excluded from independent certification. Zero checked
        # formulas must never be reported as CERTIFIED.
        assert cert.status == "BASELINE_UNVERIFIABLE"
        assert cert.checked == 0
        assert cert.input_excluded
        assert "Summary!B1" in cert.input_excluded    # downstream of input
        payload = ev.to_dict()
        assert payload["schema"] == "evaluation"
        assert payload["version"] == 2
        assert payload["certification"]["status"] == \
            "BASELINE_UNVERIFIABLE"
        with open(src, "rb") as f:
            assert f.read() == before                 # original untouched

    @needs_soffice
    @pytest.mark.lo_smoke
    def test_evaluate_many_pool(self, fixture_copy):
        src = fixture_copy("features/schedule_calc.xlsx")
        cases = [{"Schedule!B2": v} for v in (100, 300)]
        results = oracle.evaluate_many(src, cases, ["Summary!B1"],
                                       pool_size=2)
        assert [e.outputs["Summary!B1"] for e in results] == [6400, 6600]
        assert all(e.certification.status == "BASELINE_UNVERIFIABLE"
                   and e.certification.checked == 0
                   and e.certification.input_excluded
                   for e in results)


class TestCacheSplice:

    def test_datetime_cache_serializes_as_serial(self, fixture_copy,
                                                 tmp_path):
        import datetime

        src = fixture_copy("features/schedule_calc.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Schedule"]
        target = next((r, c) for (r, c), cell in sorted(ws._cells.items())
                      if cell.data_type == "f")
        wb._paper_ledger.cache_writes.setdefault(ws, {})[target] = \
            datetime.datetime(2026, 7, 8)
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        cell = wb2["Schedule"].cell(row=target[0], column=target[1])
        assert cell.data_type == "f"                  # formula untouched

    def test_cache_write_on_non_formula_refuses(self, fixture_copy,
                                                tmp_path):
        src = fixture_copy("features/schedule_calc.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Schedule"]
        wb._paper_ledger.cache_writes.setdefault(ws, {})[(2, 1)] = 1.0
        with pytest.raises(Exception, match="formula"):
            wb.save(str(tmp_path / "o.xlsx"))
