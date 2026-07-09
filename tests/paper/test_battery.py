"""The five-job acceptance battery (PLAN Phase 1).

Two halves:

- ``TestStockCarnageBaseline`` runs each job against STOCK behavior and
  asserts the exact damage measured in Phase 0 (OPEN-QUESTIONS.md Q11). These
  tests PASS today. They are the fork's justification artifact and its
  permanent regression guard: if upstream behavior ever changes, these fail
  and the damage model gets re-derived, not assumed.

- ``TestBatterySafety`` encodes the forever pass-criterion: each job ends
  correct or loudly refused — never silently wrong. Each is a strict xfail
  naming the phase that flips it; when that phase lands, the xfail must be
  removed (strict=True makes an unexpected pass a failure, so flipping is
  forced, not optional).

Fixture note: the corpus is openpyxl-authored (+ zip surgery), so stock
carnage here UNDERSTATES real-file damage — openpyxl round-trips content it
modeled itself (chart survival is asserted below as fixture-specific
fairness, not as a general claim). Real-Excel fixtures land via
FIXTURE-REQUESTS.md and extend this baseline.
"""
from __future__ import annotations

import io
import zipfile

import pytest

from openpyxl import load_workbook

from .support.partdiff import part_payloads


def _sheet_payload(path_or_bytes, marker):
    parts = part_payloads(path_or_bytes)
    for name, payload in parts.items():
        if name.startswith("xl/worksheets/sheet") and marker in payload:
            return name, payload
    raise AssertionError("no sheet part contains {0!r}".format(marker))


class TestStockCarnageBaseline:
    """Damage-model rows reproduced against stock (PLAN's table, corrected by
    Phase-0 evidence). Every assertion here is a measured stock behavior."""

    def test_job1_assumption_flip_kills_extensions_and_shared_groups(
            self, fixture_copy, tmp_path):
        src = fixture_copy("gauntlet/gauntlet.xlsx")
        _, model_before = _sheet_payload(src, b"Quarterly Model")
        _, calc_before = _sheet_payload(src, b"double")
        assert b"sparklineGroups" in model_before
        assert b"x14:conditionalFormattings" in model_before
        assert b"<x14:id>" in model_before
        assert b'<f t="shared"' in calc_before

        # extension drop is warned at LOAD (Q11: not silent, and load-time)
        with pytest.warns(UserWarning, match="extension is not supported"):
            wb = load_workbook(src)
        wb["Model"]["B8"] = 0.15
        out = str(tmp_path / "job1_out.xlsx")
        wb.save(out)

        _, model_after = _sheet_payload(out, b"Quarterly Model")
        _, calc_after = _sheet_payload(out, b"double")
        # the intended edit landed...
        assert b'r="B8"' in model_after and b"0.15" in model_after
        # ...and everything half-understood died:
        assert b"sparklineGroups" not in model_after            # sparklines gone
        assert b"x14:conditionalFormattings" not in model_after  # x14 CF twin gone
        assert b"<x14:id>" not in model_after                    # twin pointer gone
        assert b'<f t="shared"' not in calc_after                # shared group dissolved
        # fixture-specific fairness: openpyxl-authored chart/image survive
        parts_after = part_payloads(out)
        assert any(n.startswith("xl/charts/") for n in parts_after)
        assert any(n.startswith("xl/media/") for n in parts_after)
        # values are stale by construction: formulas carry no cached value
        assert b"<f>SUM(Data!B2:B5)</f>" in model_after

    def test_job2_pandas_append_kills_extensions(self, fixture_copy, tmp_path):
        pd = pytest.importorskip("pandas")
        src = fixture_copy("gauntlet/gauntlet.xlsx")
        _, model_before = _sheet_payload(src, b"Quarterly Model")
        assert b"sparklineGroups" in model_before
        rels_before = part_payloads(src)["xl/_rels/workbook.xml.rels"]

        df = pd.DataFrame({"a": [1, 2], "b": [3, 4]})
        with pd.ExcelWriter(src, engine="openpyxl", mode="a") as writer:
            df.to_excel(writer, sheet_name="Appended", index=False)

        parts = part_payloads(src)
        # the intended append landed...
        wb = load_workbook(src)
        assert "Appended" in wb.sheetnames
        # ...but the sheet nobody touched lost its extensions,
        _, model_after = _sheet_payload(src, b"Quarterly Model")
        assert b"sparklineGroups" not in model_after
        # and the whole package was regenerated (existing rels renumbered)
        assert parts["xl/_rels/workbook.xml.rels"] != rels_before

    def test_job3_insert_rows_corrupts_references_silently(
            self, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule.xlsx")
        wb = load_workbook(src)
        ws = wb["Schedule"]
        assert ws["B12"].value == "=SUM(B2:B11)"
        ws.insert_rows(5)
        out = str(tmp_path / "job3_out.xlsx")
        wb.save(out)

        wb2 = load_workbook(out)
        ws2 = wb2["Schedule"]
        # data now spans B2:B12 (a row was inserted above the range end)...
        # ...but the moved TOTAL still sums the OLD range: silent corruption
        assert ws2["B13"].value == "=SUM(B2:B11)"
        # the defined name still points at the OLD input cell (now empty)
        assert wb2.defined_names["Growth"].value == "Schedule!$B$15"
        assert ws2["B15"].value is None          # input moved to B16
        assert ws2["B16"].value == 0.05
        # the cross-sheet reference still points at the old TOTAL slot,
        # which now holds a data row — a plausible-looking wrong number
        assert wb2["Summary"]["B1"].value == "=Schedule!B12"
        assert ws2["B12"].value == 1100          # was "Item 10"'s amount, shifted

    def test_job4_xlsm_roundtrip_strips_vba_silently(self, fixture_copy, tmp_path):
        src = fixture_copy("features/macro_stub.xlsm")
        vba_before = part_payloads(src)["xl/vbaProject.bin"]

        wb = load_workbook(src)  # no keep_vba
        out = str(tmp_path / "job4_out.xlsm")
        wb.save(out)
        parts = part_payloads(out)
        assert "xl/vbaProject.bin" not in parts               # VBA gone
        assert b"macroEnabled" not in parts["[Content_Types].xml"]

        # contrast: keep_vba=True preserves it byte-identically (the in-tree
        # retention precedent the spine generalizes)
        wb2 = load_workbook(src, keep_vba=True)
        out2 = str(tmp_path / "job4_keepvba.xlsm")
        wb2.save(out2)
        assert part_payloads(out2)["xl/vbaProject.bin"] == vba_before

    def test_job5_data_only_save_destroys_all_formulas(self, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule_calc.xlsx")
        # marker is the formula text: LO-written files keep strings in the
        # shared-strings part, so cell text like "TOTAL" is not in sheet XML
        _, sheet_before = _sheet_payload(src, b"SUM(B2:B11)")
        assert sheet_before.count(b"<f") >= 2  # formulas present in the file

        wb = load_workbook(src, data_only=True)
        assert wb["Schedule"]["B12"].value == 6500  # cached values read fine
        out = str(tmp_path / "job5_out.xlsx")
        wb.save(out)                                # ...and this is the trap

        _, sheet_after = _sheet_payload(out, b"6500")
        assert b"<f" not in sheet_after             # every formula destroyed
        wb2 = load_workbook(out)
        assert wb2["Schedule"]["B12"].value == 6500  # literal, not =SUM(...)

    def test_stale_cached_values_row_pipelines_read_nothing(self, fixture_copy):
        pd = pytest.importorskip("pandas")
        # openpyxl-written formulas carry no cached values -> pandas sees NaN
        stale = fixture_copy("features/schedule.xlsx")
        df = pd.read_excel(stale, sheet_name="Schedule", header=0)
        total_cell = df[df.iloc[:, 0] == "TOTAL"].iloc[0, 1]
        assert pd.isna(total_cell)
        # the LibreOffice-recalculated twin reads real numbers
        calc = fixture_copy("features/schedule_calc.xlsx")
        df2 = pd.read_excel(calc, sheet_name="Schedule", header=0)
        total_cell2 = df2[df2.iloc[:, 0] == "TOTAL"].iloc[0, 1]
        assert total_cell2 == 6500


class TestBatterySafety:
    """The forever criterion: correct or loudly refused — never silently
    wrong. Strict xfails; each names the phase that must flip it."""

    # GREEN since Phase 2c: the splice preserves sheet-internal extensions
    def test_job1_assumption_flip_preserves_everything(self, fixture_copy, tmp_path):
        src = fixture_copy("gauntlet/gauntlet.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Model"]["B8"] = 0.15
        out = str(tmp_path / "job1_safe.xlsx")
        wb.save(out)
        _, model_after = _sheet_payload(out, b"Quarterly Model")
        assert b"sparklineGroups" in model_after
        assert b"x14:conditionalFormattings" in model_after
        assert b"<x14:id>" in model_after
        wb2 = load_workbook(out)
        assert wb2["Model"]["B8"].value == 0.15

    # GREEN since Phase 2d: sheet addition composes with the preserved package
    def test_job2_pandas_append_preserves_everything(self, fixture_copy):
        pd = pytest.importorskip("pandas")
        src = fixture_copy("gauntlet/gauntlet.xlsx")
        df = pd.DataFrame({"a": [1, 2], "b": [3, 4]})
        with pd.ExcelWriter(src, engine="openpyxl", mode="a",
                            engine_kwargs={"preserve": True}) as writer:
            df.to_excel(writer, sheet_name="Appended", index=False)
        _, model_after = _sheet_payload(src, b"Quarterly Model")
        assert b"sparklineGroups" in model_after
        wb = load_workbook(src)
        assert "Appended" in wb.sheetnames

    # GREEN since Phase 2a via the blanket preserve-save refusal (any refusal
    # satisfies "correct or loudly refused"); Phase 6a narrows the refusal to
    # the specific stranded references and 6b upgrades it to a correct rewrite
    # — this test must STAY green through both.
    def test_job3_insert_rows_refuses_or_rewrites(self, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule.xlsx")
        with open(src, "rb") as f:
            before = f.read()
        wb = load_workbook(src, preserve=True)
        ws = wb["Schedule"]
        try:
            ws.insert_rows(5)
            wb.save(src)
        except Exception as exc:
            # refusal path: must be typed and atomic
            from openpyxl.errors import PaperRefusal
            assert isinstance(exc, PaperRefusal)
            with open(src, "rb") as f:
                assert f.read() == before
        else:
            # rewrite path (Phase 6b): references must have followed the shift
            wb2 = load_workbook(src)
            assert wb2["Schedule"]["B13"].value == "=SUM(B2:B12)"
            assert wb2.defined_names["Growth"].value == "Schedule!$B$16"
            assert wb2["Summary"]["B1"].value == "=Schedule!B13"

    # GREEN since Phase 2c: untouched parts raw-copy byte-identically
    def test_job4_xlsm_roundtrip_preserves_vba(self, fixture_copy, tmp_path):
        src = fixture_copy("features/macro_stub.xlsm")
        vba_before = part_payloads(src)["xl/vbaProject.bin"]
        wb = load_workbook(src, preserve=True)  # note: NO keep_vba flag needed
        out = str(tmp_path / "job4_safe.xlsm")
        wb.save(out)
        parts = part_payloads(out)
        assert parts["xl/vbaProject.bin"] == vba_before
        assert b"macroEnabled" in parts["[Content_Types].xml"]

    # GREEN since Phase 2a via the blanket preserve-save refusal; Phase 3
    # narrows it to the data_only-specific trap (with allow_formula_loss
    # override) — this test must STAY green through that change.
    def test_job5_data_only_save_refuses(self, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule_calc.xlsx")
        with open(src, "rb") as f:
            before = f.read()
        wb = load_workbook(src, preserve=True, data_only=True)
        assert wb["Schedule"]["B12"].value == 6500  # reading stays legal
        out = str(tmp_path / "job5_safe.xlsx")
        from openpyxl.errors import PaperRefusal
        with pytest.raises(PaperRefusal):
            wb.save(out)
        import os
        assert not os.path.exists(out)  # refusal left nothing behind
        with open(src, "rb") as f:
            assert f.read() == before

    # Battery job 6 (PLAN-v0.1): edit inside a shared-formula group.
    # Settled CORRECT by the Batch-0 item-zero probe: dissolve-on-touch via
    # observed si= members (splice.resolve_dirty_cells), refuse orphans.
    def test_job6_shared_group_edit_correct_or_refused(
            self, fixture_copy, tmp_path):
        src = fixture_copy("features/shared_formulas.xlsx")

        # master formula edit: whole group dissolves, meaning preserved
        wb = load_workbook(src, preserve=True)
        wb["Calc"]["B2"] = "=A2*5"
        out = str(tmp_path / "job6_master.xlsx")
        wb.save(out)
        sheet = next(p for n, p in part_payloads(out).items()
                     if n.startswith("xl/worksheets/"))
        assert b't="shared"' not in sheet
        wb2 = load_workbook(out)
        assert wb2["Calc"]["B2"].value == "=A2*5"
        assert wb2["Calc"]["B6"].value == "=A6*2"    # follower kept meaning
        # dissolve is formula-affecting: the recalc flag must be set
        assert b"fullCalcOnLoad" in part_payloads(out)["xl/workbook.xml"]

        # two groups, one touched: the untouched group survives verbatim
        surgical = str(tmp_path / "job6_two_groups.xlsx")
        with zipfile.ZipFile(fixture_copy("features/shared_formulas.xlsx")) \
                as zin, zipfile.ZipFile(surgical, "w") as zout:
            for name in zin.namelist():
                payload = zin.read(name)
                if name.startswith("xl/worksheets/"):
                    payload = payload.replace(
                        b'<c r="D2">',
                        b'<c r="C2"><f t="shared" ref="C2:C3" si="1">A2+1'
                        b'</f><v></v></c><c r="D2">', 1)
                    payload = payload.replace(
                        b'<c r="B3"><f t="shared" si="0"/><v></v></c>',
                        b'<c r="B3"><f t="shared" si="0"/><v></v></c>'
                        b'<c r="C3"><f t="shared" si="1"/><v></v></c>', 1)
                zout.writestr(name, payload)
        wb = load_workbook(surgical, preserve=True)
        wb["Calc"]["B3"] = 999
        out2 = str(tmp_path / "job6_isolation.xlsx")
        wb.save(out2)
        sheet = next(p for n, p in part_payloads(out2).items()
                     if n.startswith("xl/worksheets/"))
        assert b'<f t="shared" ref="C2:C3" si="1">A2+1</f>' in sheet
        assert sheet.count(b'si="0"') == 0
        assert load_workbook(out2)["Calc"]["C3"].value == "=A3+1"

        # orphan follower (host never seen): typed refusal, atomic
        orphan = str(tmp_path / "job6_orphan.xlsx")
        with zipfile.ZipFile(fixture_copy("features/shared_formulas.xlsx")) \
                as zin, zipfile.ZipFile(orphan, "w") as zout:
            for name in zin.namelist():
                payload = zin.read(name)
                if name.startswith("xl/worksheets/"):
                    payload = payload.replace(
                        b'<c r="B4"><f t="shared" si="0"/><v></v></c>',
                        b'<c r="B4"><f t="shared" si="9"/><v></v></c>', 1)
                zout.writestr(name, payload)
        with open(orphan, "rb") as f:
            before = f.read()
        wb = load_workbook(orphan, preserve=True)
        wb["Calc"]["B4"] = 5
        from openpyxl.errors import UnsupportedStructureError
        with pytest.raises(UnsupportedStructureError, match="si=9"):
            wb.save(str(tmp_path / "job6_refused.xlsx"))
        with open(orphan, "rb") as f:
            assert f.read() == before

    # Battery job 14 (PLAN-v0.1): a zero-edit save must be byte-identical on
    # EVERY part — including sheets whose <cols> render trips upstream's
    # impure DimensionHolder.to_tree() (the Batch-0 false-dirty bug).
    def test_job14_noop_save_is_byte_identical_on_cols_sheet(
            self, fixture_copy, tmp_path):
        src = fixture_copy("features/hidden.xlsx")
        wb = load_workbook(src, preserve=True)
        out = str(tmp_path / "job14.xlsx")
        wb.save(out)
        assert part_payloads(src) == part_payloads(out)

    # Battery job 15 (PLAN-v0.1): editing a region whose ORIGINAL element is
    # self-closing (pageMargins/pageSetup/autoFilter/sheetPr...) must splice
    # correctly — the Batch-0 scanner bug emitted malformed XML here.
    def test_job15_self_closing_region_edit_is_correct(
            self, fixture_copy, tmp_path):
        import xml.etree.ElementTree as ET

        # pageMargins is self-closing in the fixture as authored
        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Sheet1"].page_margins.left = 1.25
        out = str(tmp_path / "job15.xlsx")
        wb.save(out)
        sheet = next(p for n, p in part_payloads(out).items()
                     if n.startswith("xl/worksheets/"))
        ET.fromstring(sheet)                       # well-formed
        assert sheet.count(b"<pageMargins") == 1
        assert b'left="1.25"' in sheet
        wb2 = load_workbook(out)
        assert wb2["Sheet1"].page_margins.left == 1.25
        assert wb2["Sheet1"]["B2"].value is not None   # data intact

    # Battery job 2 at Batch-0 exit (PLAN-v0.1 0.5): with the internal
    # default flipped (PAPER_PRESERVE_DEFAULT=1, set in paper harness
    # images), plain pandas mode="a" — no engine_kwargs — preserves.
    def test_job2_pandas_append_under_internal_default_flip(
            self, fixture_copy, monkeypatch):
        pd = pytest.importorskip("pandas")
        monkeypatch.setenv("PAPER_PRESERVE_DEFAULT", "1")
        src = fixture_copy("gauntlet/gauntlet.xlsx")
        df = pd.DataFrame({"a": [1, 2], "b": [3, 4]})
        with pd.ExcelWriter(src, engine="openpyxl", mode="a") as writer:
            df.to_excel(writer, sheet_name="Appended", index=False)
        _, model_after = _sheet_payload(src, b"Quarterly Model")
        assert b"sparklineGroups" in model_after
        wb = load_workbook(src)
        assert "Appended" in wb.sheetnames


class TestBatteryToday:
    """PLAN-v0.1 battery rows 7-13, 16-24 implemented AT THEIR TODAY
    STATES. Several assert dishonest behavior on purpose — they exist to
    be FLIPPED by the batch named in each comment, and only by it
    (weakening an expected state to make a batch pass is prohibited).
    """

    def _surgery(self, fixture_copy, tmp_path, name, old, new):
        src = fixture_copy("minimal/minimal_clean.xlsx")
        out = str(tmp_path / name)
        with zipfile.ZipFile(src) as zin, zipfile.ZipFile(out, "w") as zout:
            for part in zin.namelist():
                payload = zin.read(part)
                if part.startswith("xl/worksheets/sheet"):
                    payload = payload.replace(old, new, 1)
                zout.writestr(part, payload)
        return out

    # job 7 — today: refuse. Batch 1 keeps refuse, adds in_spill context.
    def test_job7_spill_write_refuses(self, fixture_copy, tmp_path):
        from openpyxl.errors import UnsupportedStructureError

        src = self._surgery(fixture_copy, tmp_path, "spill.xlsx",
                            b'<c r="B2"', b'<c r="B2" vm="1"')
        wb = load_workbook(src, preserve=True)
        wb["Sheet1"]["B2"] = 5
        with pytest.raises(UnsupportedStructureError, match="cm/vm"):
            wb.save(str(tmp_path / "o.xlsx"))

    # job 8 — today: refuse at set-time. Batch 3 flips to cascade rewrite.
    def test_job8_sheet_rename_refuses(self, fixture_copy):
        from openpyxl.errors import UnsupportedStructureError

        wb = load_workbook(fixture_copy("features/schedule.xlsx"),
                           preserve=True)
        with pytest.raises(UnsupportedStructureError, match="renam"):
            wb["Schedule"].title = "Renamed"

    # job 9 — Batch-1 state (flipped by 1.6): warn by default, refuse
    # under wb.strict_protection. Protection is reported, never bypassed
    # silently — and never enforced beyond what the caller asked for.
    def test_job9_locked_cell_write_warns_or_refuses(self, tmp_path):
        from openpyxl import Workbook
        from openpyxl.errors import (
            ProtectedWriteWarning,
            UnsupportedStructureError,
        )

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "original"
        ws["A2"] = "original2"
        ws.protection.sheet = True          # cells are locked by default
        src = str(tmp_path / "protected.xlsx")
        wb.save(src)

        wb2 = load_workbook(src, preserve=True)
        with pytest.warns(ProtectedWriteWarning, match="locked"):
            wb2.active["A1"] = "overwritten"
        import warnings as _w
        with _w.catch_warnings():           # once per sheet, not per cell
            _w.simplefilter("error")
            wb2.active["A2"] = "again"
        out = str(tmp_path / "o.xlsx")
        wb2.save(out)                       # the write itself proceeds
        assert load_workbook(out).active["A1"].value == "overwritten"

        wb3 = load_workbook(src, preserve=True)
        wb3.strict_protection = True
        with pytest.raises(UnsupportedStructureError, match="locked"):
            wb3.active["A1"] = "refused"
        assert wb3.active["A1"].value == "original"   # atomic

        # the manifest reports protection so agents can check BEFORE writing
        doc = wb3.manifest().to_dict()
        assert doc["sheets"][0]["protection"] is True
        assert doc["workbook_protection"] is False

    # job 10 — Batch-2 state: CORRECT table append discipline (flipped
    # from the Batch-1 refusal by the lifecycle engine + table verbs).
    def test_job10_table_append_is_correct(self, fixture_copy, tmp_path):
        from openpyxl.preserve.tables import append_row

        src = fixture_copy("features/tables.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb.worksheets[0]
        append_row(ws, "RegionTable", {"Region": "Central", "Amount": 60})
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        parts = part_payloads(out)
        table_part = parts["xl/tables/table1.xml"]
        assert b'ref="A1:B6"' in table_part            # range extended
        wb2 = load_workbook(out)
        ws2 = wb2.worksheets[0]
        assert ws2.tables["RegionTable"].ref == "A1:B6"
        assert ws2["A6"].value == "Central"
        assert ws2["B6"].value == 60
        assert ws2["B5"].value == 50                   # neighbours intact

        # geometry guards refuse atomically: anchor moves and column-count
        # mismatches never reach the file
        src2 = fixture_copy("features/tables.xlsx")
        with open(src2, "rb") as f:
            before = f.read()
        wb3 = load_workbook(src2, preserve=True)
        wb3.worksheets[0].tables["RegionTable"].ref = "A1:C5"  # 3 cols vs 2
        from openpyxl.errors import UnsupportedStructureError
        with pytest.raises(UnsupportedStructureError, match="tableColumns"):
            wb3.save(str(tmp_path / "refused.xlsx"))
        with open(src2, "rb") as f:
            assert f.read() == before                 # atomic

    # job 11 — today: refuse at call time. Batch 3 flips to correct copy.
    def test_job11_copy_sheet_refuses(self, fixture_copy):
        from openpyxl.errors import UnsupportedStructureError

        wb = load_workbook(fixture_copy("gauntlet/gauntlet.xlsx"),
                           preserve=True)
        with pytest.raises(UnsupportedStructureError, match="copy"):
            wb.copy_worksheet(wb["Model"])

    # job 12 — today: no scenario API. Batch 5 ships wb.evaluate().
    def test_job12_no_evaluate_api_yet(self):
        from openpyxl.workbook import Workbook

        assert not hasattr(Workbook, "evaluate")

    # job 13 — Batch-1 state (flipped by 1.5): typed refusal naming the
    # encryption and the decrypt route, on both load arms.
    def test_job13_encrypted_cfb_gets_typed_refusal(
            self, fixture_copy, tmp_path):
        from openpyxl.errors import UnsupportedStructureError

        src = str(tmp_path / "encrypted.xlsx")
        with open(src, "wb") as f:
            f.write(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 4096)
        with pytest.raises(UnsupportedStructureError, match="ENCRYPT"):
            load_workbook(src, preserve=True)
        with pytest.raises(UnsupportedStructureError, match="password"):
            load_workbook(src)                        # stock arm too
        # file-like arm rewinds after sniffing on ordinary files
        import io as _io
        with open(fixture_copy("minimal/minimal_clean.xlsx"), "rb") as f:
            buf = _io.BytesIO(f.read())
        assert load_workbook(buf).active is not None

    # 1.5 input honesty: duplicate entry names are a parser differential
    # (reader takes the LAST copy, raw copy keeps BOTH) — preserve refuses
    def test_duplicate_zip_entries_refuse_under_preserve(
            self, fixture_copy, tmp_path):
        from openpyxl.errors import UnsupportedStructureError

        src = fixture_copy("minimal/minimal_clean.xlsx")
        dup = str(tmp_path / "dup.xlsx")
        with zipfile.ZipFile(src) as zin, zipfile.ZipFile(dup, "w") as zout:
            for name in zin.namelist():
                zout.writestr(name, zin.read(name))
                if name.startswith("xl/worksheets/sheet"):
                    zout.writestr(name, zin.read(name))   # duplicate
        with pytest.raises(UnsupportedStructureError, match="duplicate"):
            load_workbook(dup, preserve=True)
        load_workbook(dup)                       # stock keeps upstream arm

    # job 16 — Batch-1 state: REFUSE (flipped from silent staleness by
    # the 1.1 object guards). Batch 4 flips chart editing to correct.
    def test_job16_chart_title_edit_refuses(self, fixture_copy, tmp_path):
        from openpyxl.errors import UnsupportedStructureError

        src = fixture_copy("features/chart_image.xlsx")
        with open(src, "rb") as f:
            before = f.read()
        wb = load_workbook(src, preserve=True)
        ws = next(w for w in wb.worksheets if w._charts)
        ws._charts[0].title = "TAMPERED"
        with pytest.raises(UnsupportedStructureError, match="chart"):
            wb.save(str(tmp_path / "o.xlsx"))
        with open(src, "rb") as f:
            assert f.read() == before                 # atomic

        # a MUTATION-ONLY session refuses too (no other dirt on the sheet)
        wb2 = load_workbook(src, preserve=True)
        ws2 = next(w for w in wb2.worksheets if w._images)
        ws2._images[0].anchor._from.col = 9
        with pytest.raises(UnsupportedStructureError, match="image"):
            wb2.save(str(tmp_path / "o2.xlsx"))

    # job 17 — Batch-1 state (flipped by 1.2): a value edit feeding
    # formulas forces fullCalcOnLoad so stale caches can never masquerade
    # as current to a human opener.
    def test_job17_value_edit_feeding_formulas_sets_recalc_flag(
            self, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule_calc.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Schedule"]["B2"] = 1200            # feeds =SUM(B2:B11)
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        assert b"fullCalcOnLoad" in part_payloads(out)["xl/workbook.xml"]

        # a value edit feeding NO formula keeps workbook.xml untouched —
        # the flag is targeted, not a blanket stamp (reads never dirty,
        # and neither do isolated writes)
        src2 = fixture_copy("minimal/minimal_clean.xlsx")
        wb2 = load_workbook(src2, preserve=True)
        wb2["Sheet1"]["A1"] = "note"
        out2 = str(tmp_path / "o2.xlsx")
        wb2.save(out2)
        assert part_payloads(out2)["xl/workbook.xml"] == \
            part_payloads(src2)["xl/workbook.xml"]

    # job 18 — Batch-2 state: CORRECT ("make this range a table" via the
    # part-lifecycle engine; was a save-time refusal).
    def test_job18_make_range_a_table_is_correct(self, fixture_copy,
                                                 tmp_path):
        from openpyxl.worksheet.table import Table

        wb = load_workbook(fixture_copy("minimal/minimal_clean.xlsx"),
                           preserve=True)
        wb["Sheet1"].add_table(Table(displayName="New", ref="A1:B3"))
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        parts = part_payloads(out)
        table_part = next(p for n, p in parts.items()
                          if n.startswith("xl/tables/"))
        assert b'displayName="New"' in table_part
        assert b'ref="A1:B3"' in table_part
        sheet = next(p for n, p in parts.items()
                     if n.startswith("xl/worksheets/"))
        assert b"<tableParts count=\"1\">" in sheet
        assert b"table+xml" in parts["[Content_Types].xml"]
        wb2 = load_workbook(out)
        assert "New" in wb2["Sheet1"].tables
        assert wb2["Sheet1"].tables["New"].ref == "A1:B3"
        assert wb2["Sheet1"]["B2"].value is not None   # data intact

    # job 19 — today: refuse at save. Batch 2 flips to correct.
    def test_job19_comment_on_comment_free_sheet_refuses(
            self, fixture_copy, tmp_path):
        from openpyxl.comments import Comment
        from openpyxl.errors import UnsupportedStructureError

        wb = load_workbook(fixture_copy("minimal/minimal_clean.xlsx"),
                           preserve=True)
        wb["Sheet1"]["B2"].comment = Comment("note", "paper")
        with pytest.raises(UnsupportedStructureError, match="comment"):
            wb.save(str(tmp_path / "o.xlsx"))

    # job 20 — today: refuse (x14 twin desync gate). Batch 3 flips to
    # correct twin-sync.
    def test_job20_x14_cf_edit_refuses(self, fixture_copy, tmp_path):
        from openpyxl.errors import UnsupportedStructureError
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import PatternFill

        wb = load_workbook(fixture_copy("gauntlet/gauntlet.xlsx"),
                           preserve=True)
        wb["Model"].conditional_formatting.add(
            "B3:B5", CellIsRule(
                operator="greaterThan", formula=["1000"],
                fill=PatternFill(start_color="FFC7CE", fill_type="solid")))
        with pytest.raises(UnsupportedStructureError, match="x14"):
            wb.save(str(tmp_path / "o.xlsx"))

    # job 21 — today: refuse (cm metadata). Batch 3 flips to correct
    # cm/vm bookkeeping.
    def test_job21_value_write_on_excel365_cell_refuses(
            self, fixture_copy, tmp_path):
        from openpyxl.errors import UnsupportedStructureError

        src = self._surgery(fixture_copy, tmp_path, "e365.xlsx",
                            b'<c r="B2"', b'<c r="B2" cm="1"')
        wb = load_workbook(src, preserve=True)
        wb["Sheet1"]["B2"] = 5
        with pytest.raises(UnsupportedStructureError, match="cm/vm"):
            wb.save(str(tmp_path / "o.xlsx"))

    # job 22 — today: refuse at call time. Batch 4 flips to correct.
    def test_job22_add_chart_refuses(self, fixture_copy):
        from openpyxl.chart import BarChart
        from openpyxl.errors import UnsupportedStructureError

        wb = load_workbook(fixture_copy("gauntlet/gauntlet.xlsx"),
                           preserve=True)
        with pytest.raises(UnsupportedStructureError, match="chart"):
            wb["Model"].add_chart(BarChart(), "H1")

    # job 23 — today: no localization API. Batch 6 ships it (and finally
    # raises the pinned AmbiguousTargetError).
    def test_job23_no_label_localization_api_yet(self, fixture_copy):
        wb = load_workbook(fixture_copy("minimal/minimal_clean.xlsx"))
        assert not hasattr(wb, "locate")
        assert not hasattr(wb.active, "find_by_label")

    # job 24 — today: no oracle write-back. Batch 5 ships it,
    # certification-gated.
    def test_job24_no_oracle_writeback_yet(self):
        from openpyxl import oracle

        assert not hasattr(oracle, "write_back")
