"""PR 9: cache-backed raw foreign output ownership."""
from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.pivot.adopt_qualify import analyze_adoption
from .test_pivot_adoption_qualification import (
    _codes,
    _load_payload,
    _paper_then_foreign,
    _rewrite_zip,
)
from .test_pivot_graph import _basic_package


def test_stale_source_does_not_invalidate_old_ownership(
        fixture_copy, tmp_path, monkeypatch):
    wb, path = _paper_then_foreign(fixture_copy, tmp_path)
    ws = wb["Data"]
    # Source values change; persisted cache/output stay internally consistent.
    ws["B2"] = 999
    result = analyze_adoption(ws.pivots["ByRegion"])
    assert result.ownership is not None
    assert result.public.strategy == "dedicated-replacement"
    monkeypatch.setattr(
        "openpyxl.pivot.adopt_qualify.excel_equivalence_proved",
        lambda: True,
    )
    monkeypatch.setattr(
        "openpyxl.pivot.adopt_inventory.excel_equivalence_proved",
        lambda: True,
    )
    qualified = ws.pivots["ByRegion"].qualify_adoption()
    assert qualified.eligible is True


def test_unexplained_blank_cell_node_disables_adoption(
        fixture_copy, tmp_path):
    wb, path = _paper_then_foreign(fixture_copy, tmp_path)
    pivot = wb["Data"].pivots["ByRegion"]
    analysis = analyze_adoption(pivot)
    assert analysis.ownership is not None
    # Inject an unexplained empty <c> inside location.ref.
    def add_empty(_name, body):
        if not _name.startswith("xl/worksheets/sheet") or not _name.endswith(".xml"):
            return body
        if b'r="F3"' in body or b"ref=\"E3:F9\"" in body:
            pass
        marker = b"</sheetData>"
        extra = b'<row r="3"><c r="F3"/></row>'
        if extra in body or b'r="F3"' in body:
            return body
        if b"<sheetData" not in body:
            return body
        return body.replace(marker, extra + marker, 1)

    _rewrite_zip(path, add_empty)
    reopened = load_workbook(path, preserve=True)
    result = reopened["Data"].pivots["ByRegion"].qualify_adoption()
    assert result.eligible is False
    assert "foreign-output-unproved" in _codes(result)


def test_cache_output_disagreement_disables_adoption(tmp_path):
    payload = _basic_package()
    wb = _load_payload(tmp_path, payload)
    result = wb["Summary"].pivots["SalesByRegion"].qualify_adoption()
    assert result.eligible is False
    assert "foreign-output-unproved" in _codes(result)


def test_report_filter_cells_are_outside_location_ref():
    from openpyxl.pivot.api_types import (
        PivotAxisField, PivotItemFilter, PivotMeasure, PivotSource, PivotSpec,
    )
    from openpyxl.pivot.layout import ROLE_FILTER
    from openpyxl.pivot.plan import plan_pivot
    from openpyxl.pivot.source import snapshot_from_matrix
    from openpyxl.utils.cell import range_boundaries

    spec = PivotSpec(
        name="Filtered",
        source=PivotSource.table("SalesData"),
        destination="E5",
        rows=[PivotAxisField("Region")],
        filters=[PivotItemFilter("Status")],
        values=[PivotMeasure("Amount")],
    )
    snapshot = snapshot_from_matrix(
        ["Region", "Status", "Amount"],
        [["East", "Open", 10], ["West", "Open", 7]],
        source=spec.source,
    )
    plan = plan_pivot(spec, snapshot)
    filter_coords = [
        (cell.row, cell.column)
        for cell in plan.output.cells
        if cell.role == ROLE_FILTER
    ]
    min_col, min_row, max_col, max_row = range_boundaries(plan.output.ref)
    assert filter_coords
    assert all(row < min_row for row, _column in filter_coords)
    body = {
        (cell.row, cell.column)
        for cell in plan.output.cells
        if cell.role != ROLE_FILTER
    }
    assert not body.intersection(filter_coords)
