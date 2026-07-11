from __future__ import annotations

import zipfile
from copy import copy

import pytest

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell, MergedCell
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.errors import (
    BoundaryViolationError,
    PaperRefusal,
    ProtectedWriteWarning,
)
from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula


def _preserved(tmp_path, wb, name="source.xlsx"):
    path = tmp_path / name
    wb.save(path)
    return load_workbook(path, preserve=True)


def _ledger_state(wb):
    led = wb._paper_ledger
    return {
        "cells": {ws: set(coords) for ws, coords in led.cells.items()},
        "formulas_changed": led.formulas_changed,
        "loaded_sheet_titles": led.loaded_sheet_titles,
        "protection_warned": set(led.protection_warned),
        "renames": dict(led.renames),
        "sheet_states": dict(led.sheet_states),
        "shifts": {ws: list(shifts) for ws, shifts in led.shifts.items()},
        "value_overwrites": {
            ws: set(coords) for ws, coords in led.value_overwrites.items()
        },
    }


def test_shift_rewrites_cf_and_dv_formulas_cross_sheet(tmp_path):
    wb = Workbook()
    data = wb.active
    data.title = "Data"
    rules = wb.create_sheet("Rules")
    data["A2"] = 1
    rules.conditional_formatting.add(
        "B2", FormulaRule(formula=["Data!$A$2>0"]))
    dv = DataValidation(type="custom", formula1="=Data!$A$2>0",
                        formula2="Data!$A$2")
    dv.add("C2")
    rules.add_data_validation(dv)

    wb = _preserved(tmp_path, wb)
    wb["Data"].insert_rows(2)
    wb["Data"].insert_cols(1)

    cf = next(iter(wb["Rules"].conditional_formatting))
    assert cf.rules[0].formula == ["Data!$B$3>0"]
    shifted_dv = wb["Rules"].data_validations.dataValidation[0]
    assert shifted_dv.formula1 == "=Data!$B$3>0"
    assert shifted_dv.formula2 == "Data!$B$3"

    out = tmp_path / "shifted.xlsx"
    wb.save(out)
    reloaded = load_workbook(out)
    cf = next(iter(reloaded["Rules"].conditional_formatting))
    assert cf.rules[0].formula == ["Data!$B$3>0"]
    assert reloaded["Rules"].data_validations.dataValidation[0].formula1 \
        == "=Data!$B$3>0"


def test_shift_rekeys_hashed_ranges_and_clears_deleted_filter(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws["A2"] = 1
    ws.merge_cells("B2:C2")
    ws.conditional_formatting.add("A2", FormulaRule(formula=["A2>0"]))
    dv = DataValidation(type="custom", formula1="A2>0")
    dv.add("D2")
    ws.add_data_validation(dv)
    ws.auto_filter.ref = "A2:D2"

    wb = _preserved(tmp_path, wb)
    ws = wb.active
    ws.insert_rows(2)
    ws.unmerge_cells("B3:C3")
    assert ws.conditional_formatting["A3"][0].formula == ["A3>0"]
    assert str(ws.data_validations.dataValidation[0].sqref) == "D3"

    ws.delete_rows(3)
    assert ws.auto_filter.ref is None


def test_column_dimensions_shift_keys_and_group_bounds(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions.group("B", "D", outline_level=2, hidden=True)
    ws.column_dimensions["F"].width = 22
    ws["A1"] = "seed"

    wb = _preserved(tmp_path, wb)
    ws = wb.active
    ws.insert_cols(3)
    grouped = ws.column_dimensions["B"]
    assert (grouped.index, grouped.min, grouped.max) == ("B", 2, 5)
    assert grouped.hidden and grouped.outlineLevel == 2
    assert "F" not in ws.column_dimensions
    assert ws.column_dimensions["G"].width == 22

    ws.delete_cols(2)
    grouped = ws.column_dimensions["B"]
    assert (grouped.index, grouped.min, grouped.max) == ("B", 2, 4)
    assert ws.column_dimensions["F"].width == 22


def test_column_dimension_delete_resolves_destination_key_collision(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws["B1"] = "seed"

    wb = _preserved(tmp_path, wb, "column-collision.xlsx")
    ws = wb.active
    ws.delete_cols(1)

    assert set(ws.column_dimensions) == {"A"}
    assert ws.column_dimensions["A"].index == "A"
    assert ws.column_dimensions["A"].width == 20


def test_rename_quotes_three_d_spans_and_rewrites_cf_dv(tmp_path):
    wb = Workbook()
    first = wb.active
    first.title = "First"
    wb.create_sheet("Last")
    rules = wb.create_sheet("Rules")
    rules["A1"] = "=SUM(First:Last!A1)"
    rules.conditional_formatting.add(
        "B1", FormulaRule(formula=["SUM(First:Last!A1)>0"]))
    dv = DataValidation(type="custom", formula1="SUM(First:Last!A1)>0")
    dv.add("C1")
    rules.add_data_validation(dv)

    wb = _preserved(tmp_path, wb)
    wb["First"].title = "First Sheet"

    assert wb["Rules"]["A1"].value == "=SUM('First Sheet:Last'!A1)"
    cf = next(iter(wb["Rules"].conditional_formatting))
    assert cf.rules[0].formula == ["SUM('First Sheet:Last'!A1)>0"]
    assert wb["Rules"].data_validations.dataValidation[0].formula1 == \
        "SUM('First Sheet:Last'!A1)>0"

    out = tmp_path / "renamed.xlsx"
    wb.save(out)
    reloaded = load_workbook(out)
    assert reloaded["Rules"]["A1"].value == \
        "=SUM('First Sheet:Last'!A1)"


@pytest.mark.parametrize("kind", [
    "formula",
    "defined-name",
    "conditional-formatting",
    "data-validation",
    "filter",
])
def test_overflow_refuses_before_mutation(tmp_path, kind):
    wb = Workbook()
    target = wb.active
    target.title = "Target"
    target["A1"] = "sentinel"
    other = wb.create_sheet("Other")
    if kind == "formula":
        other["A1"] = "=Target!XFD1"
    elif kind == "defined-name":
        wb.defined_names.add(DefinedName(
            "Edge", attr_text="'Target'!$XFD$1"))
    elif kind == "conditional-formatting":
        target.conditional_formatting.add(
            "XFD1", FormulaRule(formula=["XFD1>0"]))
    elif kind == "data-validation":
        dv = DataValidation(type="custom", formula1="XFD1>0")
        dv.add("XFD1")
        target.add_data_validation(dv)
    else:
        target.auto_filter.ref = "XFD1:XFD2"

    wb = _preserved(tmp_path, wb, kind + ".xlsx")
    target = wb["Target"]
    with pytest.raises(BoundaryViolationError):
        target.insert_cols(1)
    assert target["A1"].value == "sentinel"
    if kind == "formula":
        assert wb["Other"]["A1"].value == "=Target!XFD1"


def test_row_reference_overflow_refuses_before_mutation(tmp_path):
    wb = Workbook()
    target = wb.active
    target.title = "Target"
    target["A1"] = "sentinel"
    wb.create_sheet("Other")["A1"] = "=Target!A1048576"

    wb = _preserved(tmp_path, wb, "row-overflow.xlsx")
    with pytest.raises(BoundaryViolationError):
        wb["Target"].insert_rows(1)
    assert wb["Target"]["A1"].value == "sentinel"
    assert wb["Other"]["A1"].value == "=Target!A1048576"


def test_append_totals_moves_hyperlink_without_residue(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Region", "Amount"])
    ws.append(["West", 10])
    ws.append(["Total", "=SUBTOTAL(109,T[Amount])"])
    table = Table(displayName="T", ref="A1:B3")
    table.totalsRowCount = 1
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2")
    ws.add_table(table)

    wb = _preserved(tmp_path, wb)
    ws = wb.active
    ws["A3"].hyperlink = "https://example.com/total"
    from openpyxl.preserve.tables import append_row

    append_row(ws, "T", ["East", 20])
    assert ws["A4"].hyperlink.target == "https://example.com/total"
    assert ws["A4"].hyperlink.ref == "A4"
    assert ws["A3"].hyperlink is None

    output = tmp_path / "totals-hyperlink.xlsx"
    wb.save(output)
    reloaded = load_workbook(output)
    assert reloaded.active["A4"].hyperlink.target == \
        "https://example.com/total"
    assert reloaded.active["A3"].hyperlink is None


def test_append_totals_with_preserved_comment_refuses_before_mutation(
        tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Region", "Amount"])
    ws.append(["West", 10])
    ws.append(["Total", "=SUBTOTAL(109,T[Amount])"])
    table = Table(displayName="T", ref="A1:B3")
    table.totalsRowCount = 1
    ws.add_table(table)
    ws["A3"].comment = Comment("totals note", "paper")

    wb = _preserved(tmp_path, wb, "totals-comment.xlsx")
    ws = wb.active
    from openpyxl.preserve.tables import append_row

    with pytest.raises(PaperRefusal, match="comment/VML anchor"):
        append_row(ws, "T", ["East", 20])
    assert ws.tables["T"].ref == "A1:B3"
    assert ws["A3"].value == "Total"
    assert ws["A3"].comment.text == "totals note"
    assert ws["A4"].value is None


def test_append_totals_with_preserved_hyperlink_refuses_before_mutation(
        tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Region", "Amount"])
    ws.append(["West", 10])
    ws.append(["Total", "=SUBTOTAL(109,T[Amount])"])
    table = Table(displayName="T", ref="A1:B3")
    table.totalsRowCount = 1
    ws.add_table(table)
    ws["A3"].hyperlink = "https://example.com/total"

    wb = _preserved(tmp_path, wb, "totals-link.xlsx")
    ws = wb.active
    from openpyxl.preserve.tables import append_row

    with pytest.raises(PaperRefusal, match="hyperlink relationship"):
        append_row(ws, "T", ["East", 20])
    assert ws.tables["T"].ref == "A1:B3"
    assert ws["A3"].value == "Total"
    assert ws["A3"].hyperlink.target == "https://example.com/total"
    assert ws["A4"].value is None


def test_numeric_entity_chart_reference_refuses_before_mutation(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"
    for row in range(1, 4):
        ws.cell(row=row, column=1, value=row)
    chart = BarChart()
    chart.add_data(Reference(ws, min_col=1, min_row=1, max_row=3))
    ws.add_chart(chart, "C1")
    source = tmp_path / "chart.xlsx"
    wb.save(source)

    patched = tmp_path / "numeric-chart.xlsx"
    with zipfile.ZipFile(source) as zin, zipfile.ZipFile(patched, "w") as zout:
        for item in zin.infolist():
            payload = zin.read(item.filename)
            if item.filename.startswith("xl/charts/chart"):
                payload = payload.replace(b"'Sheet'!", b"'&#83;heet'!")
            zout.writestr(copy(item), payload)

    wb = load_workbook(patched, preserve=True)
    with pytest.raises(PaperRefusal, match="numeric character"):
        wb["Sheet"].insert_rows(1)
    assert wb["Sheet"]["A1"].value == 1


def test_numeric_entity_chart_reference_blocks_rename_atomically(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"
    for row in range(1, 4):
        ws.cell(row=row, column=1, value=row)
    chart = BarChart()
    chart.add_data(Reference(ws, min_col=1, min_row=1, max_row=3))
    ws.add_chart(chart, "C1")
    source = tmp_path / "rename-chart.xlsx"
    wb.save(source)

    patched = tmp_path / "rename-numeric-chart.xlsx"
    with zipfile.ZipFile(source) as zin, zipfile.ZipFile(patched, "w") as zout:
        for item in zin.infolist():
            payload = zin.read(item.filename)
            if item.filename.startswith("xl/charts/chart"):
                payload = payload.replace(b"'Sheet'!", b"'&#83;heet'!")
            zout.writestr(copy(item), payload)

    wb = load_workbook(patched, preserve=True)
    before = _ledger_state(wb)
    with pytest.raises(PaperRefusal, match="numeric character"):
        wb["Sheet"].title = "Inputs"
    assert wb.sheetnames == ["Sheet"]
    assert _ledger_state(wb) == before


def test_numeric_entity_in_sibling_chart_series_blocks_rename(tmp_path):
    wb = Workbook()
    target = wb.active
    target.title = "Sheet"
    other = wb.create_sheet("Other")
    for row in range(1, 4):
        target.cell(row=row, column=1, value=row)
        other.cell(row=row, column=1, value=row * 10)
    chart = BarChart()
    chart.add_data(Reference(target, min_col=1, min_row=1, max_row=3))
    chart.add_data(Reference(other, min_col=1, min_row=1, max_row=3))
    target.add_chart(chart, "C1")
    source = tmp_path / "mixed-chart.xlsx"
    wb.save(source)

    patched = tmp_path / "mixed-numeric-chart.xlsx"
    with zipfile.ZipFile(source) as zin, zipfile.ZipFile(patched, "w") as zout:
        for item in zin.infolist():
            payload = zin.read(item.filename)
            if item.filename.startswith("xl/charts/chart"):
                payload = payload.replace(b"'Other'!", b"'&#79;ther'!")
            zout.writestr(copy(item), payload)

    wb = load_workbook(patched, preserve=True)
    before = _ledger_state(wb)
    with pytest.raises(PaperRefusal, match="numeric character"):
        wb["Sheet"].title = "Inputs"
    assert wb.sheetnames == ["Sheet", "Other"]
    assert _ledger_state(wb) == before


@pytest.mark.parametrize("operation", ["insert_rows", "delete_rows"])
def test_shift_preflights_protected_dependent_formula_atomically(
        tmp_path, operation):
    wb = Workbook()
    data = wb.active
    data.title = "Data"
    data["A2"] = "keep"
    data["A3"] = "move"
    calc = wb.create_sheet("Calc")
    calc["A1"] = "=Data!A3"
    calc.protection.sheet = True

    wb = _preserved(tmp_path, wb, operation + "-protected.xlsx")
    wb.strict_protection = True
    before = _ledger_state(wb)
    before_cells = {
        key: (cell.value, cell.data_type)
        for key, cell in wb["Data"]._cells.items()
    }

    with pytest.raises(PaperRefusal, match="strict_protection"):
        getattr(wb["Data"], operation)(2)

    assert {
        key: (cell.value, cell.data_type)
        for key, cell in wb["Data"]._cells.items()
    } == before_cells
    assert wb["Calc"]["A1"].value == "=Data!A3"
    assert _ledger_state(wb) == before


def test_rename_preflights_every_protected_formula_atomically(tmp_path):
    wb = Workbook()
    data = wb.active
    data.title = "Data"
    data["A1"] = 1
    early = wb.create_sheet("Early")
    early["A1"] = "=Data!A1"
    protected = wb.create_sheet("Protected")
    protected["A1"] = "=Data!A1"
    protected.protection.sheet = True

    wb = _preserved(tmp_path, wb, "rename-protected.xlsx")
    wb.strict_protection = True
    before = _ledger_state(wb)

    with pytest.raises(PaperRefusal, match="strict_protection"):
        wb["Data"].title = "Inputs"

    assert wb.sheetnames == ["Data", "Early", "Protected"]
    assert wb["Early"]["A1"].value == "=Data!A1"
    assert wb["Protected"]["A1"].value == "=Data!A1"
    assert _ledger_state(wb) == before


@pytest.mark.parametrize("attribute", ["r1", "r2"])
def test_rename_refuses_explicit_data_table_reference_atomically(
        tmp_path, attribute):
    wb = Workbook()
    data = wb.active
    data.title = "Data"
    data["A1"] = 1
    calc = wb.create_sheet("Calc")
    kwargs = {attribute: "Data!A1"}
    calc["A1"] = DataTableFormula(ref="A1:B2", **kwargs)

    wb = _preserved(tmp_path, wb, "rename-data-table-" + attribute + ".xlsx")
    before = _ledger_state(wb)
    with pytest.raises(PaperRefusal, match="data table"):
        wb["Data"].title = "Inputs"

    formula = wb["Calc"]["A1"].value
    assert wb.sheetnames == ["Data", "Calc"]
    assert getattr(formula, attribute) == "Data!A1"
    assert _ledger_state(wb) == before


def test_merged_placeholders_rebuild_after_insert_and_round_trip(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "heading"
    ws.merge_cells("A1:A3")

    wb = _preserved(tmp_path, wb, "merge-insert.xlsx")
    ws = wb.active
    ws.insert_rows(2)

    assert {str(rng) for rng in ws.merged_cells.ranges} == {"A1:A4"}
    assert isinstance(ws["A1"], Cell)
    assert all(isinstance(ws.cell(row, 1), MergedCell)
               for row in range(2, 5))
    with pytest.raises(AttributeError, match="read-only"):
        ws["A2"].value = "hidden"

    output = tmp_path / "merge-insert-output.xlsx"
    wb.save(output)
    reloaded = load_workbook(output)
    assert {str(rng) for rng in reloaded.active.merged_cells.ranges} == {
        "A1:A4"
    }
    assert isinstance(reloaded.active["A1"], Cell)
    assert all(isinstance(reloaded.active.cell(row, 1), MergedCell)
               for row in range(2, 5))


def test_merged_anchor_rebuilds_after_delete_and_round_trip(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "deleted heading"
    ws.merge_cells("A1:A3")

    wb = _preserved(tmp_path, wb, "merge-delete.xlsx")
    ws = wb.active
    ws.delete_rows(1)

    assert {str(rng) for rng in ws.merged_cells.ranges} == {"A1:A2"}
    assert isinstance(ws["A1"], Cell)
    assert ws["A1"].value is None
    assert isinstance(ws["A2"], MergedCell)
    ws["A1"] = "replacement heading"

    output = tmp_path / "merge-delete-output.xlsx"
    wb.save(output)
    reloaded = load_workbook(output)
    assert {str(rng) for rng in reloaded.active.merged_cells.ranges} == {
        "A1:A2"
    }
    assert isinstance(reloaded.active["A1"], Cell)
    assert reloaded.active["A1"].value == "replacement heading"
    assert isinstance(reloaded.active["A2"], MergedCell)


def test_cross_sheet_array_formula_blocks_shift_before_mutation(tmp_path):
    wb = Workbook()
    target = wb.active
    target.title = "Data"
    target["A1"] = 10
    calc = wb.create_sheet("Calc")
    calc["A1"] = ArrayFormula(ref="A1:A2", text="=Data!A1+{1;2}")

    wb = _preserved(tmp_path, wb, "array-cross-sheet.xlsx")
    with pytest.raises(PaperRefusal, match="array formula"):
        wb["Data"].insert_rows(1)
    assert wb["Data"]["A1"].value == 10
    assert wb["Calc"]["A1"].value.text == "=Data!A1+{1;2}"


def test_cross_sheet_array_formula_blocks_rename_before_mutation(tmp_path):
    wb = Workbook()
    target = wb.active
    target.title = "Data"
    target["A1"] = 10
    calc = wb.create_sheet("Calc")
    calc["A1"] = ArrayFormula(ref="A1:A2", text="=Data!A1+{1;2}")

    wb = _preserved(tmp_path, wb, "array-rename.xlsx")
    with pytest.raises(PaperRefusal, match="loaded array"):
        wb["Data"].title = "Inputs"
    assert wb.sheetnames == ["Data", "Calc"]
    assert wb["Calc"]["A1"].value.text == "=Data!A1+{1;2}"


def test_reversed_formula_range_keeps_orientation_through_shift(tmp_path):
    wb = Workbook()
    data = wb.active
    data.title = "Data"
    data["A1"] = 1
    data["B2"] = 2
    calc = wb.create_sheet("Calc")
    calc["A1"] = "=SUM(Data!B2:A1)"

    wb = _preserved(tmp_path, wb, "reversed-range.xlsx")
    wb["Data"].insert_rows(1)
    assert wb["Calc"]["A1"].value == "=SUM(Data!B3:A2)"
    output = tmp_path / "reversed-range-output.xlsx"
    wb.save(output)
    assert load_workbook(output)["Calc"]["A1"].value == \
        "=SUM(Data!B3:A2)"


def test_three_d_formula_blocks_single_sheet_shift(tmp_path):
    wb = Workbook()
    first = wb.active
    first.title = "First"
    middle = wb.create_sheet("Middle")
    wb.create_sheet("Last")
    summary = wb.create_sheet("Summary")
    middle["A2"] = 10
    summary["A1"] = "=SUM(First:Last!A2)"

    wb = _preserved(tmp_path, wb, "three-d.xlsx")
    with pytest.raises(PaperRefusal, match="3-D formula"):
        wb["Middle"].insert_rows(1)
    assert wb["Middle"]["A2"].value == 10
    assert wb["Summary"]["A1"].value == "=SUM(First:Last!A2)"


def test_append_preflights_inherited_formula_grid_boundary(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Value", "Formula"])
    ws.append([1, "=A1048576"])
    ws.add_table(Table(displayName="T", ref="A1:B2"))

    wb = _preserved(tmp_path, wb, "append-boundary.xlsx")
    ws = wb.active
    from openpyxl.preserve.tables import append_row

    with pytest.raises(BoundaryViolationError, match="outside Excel"):
        append_row(ws, "T", [2])
    assert ws.tables["T"].ref == "A1:B2"
    assert (3, 1) not in ws._cells
    assert (3, 2) not in ws._cells


def test_append_preflights_all_values_before_writing_any_cell(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["First", "Second"])
    ws.append([1, 2])
    ws.add_table(Table(displayName="T", ref="A1:B2"))

    wb = _preserved(tmp_path, wb, "append-invalid.xlsx")
    ws = wb.active
    from openpyxl.preserve.tables import append_row

    with pytest.raises(ValueError, match="Cannot convert"):
        append_row(ws, "T", [3, object()])
    assert ws.tables["T"].ref == "A1:B2"
    assert (3, 1) not in ws._cells
    assert (3, 2) not in ws._cells


def test_shift_rewrites_unquoted_unicode_sheet_reference(tmp_path):
    wb = Workbook()
    data = wb.active
    data.title = "数据"
    data["A1"] = 10
    wb.create_sheet("Summary")["A1"] = "=数据!A1"

    wb = _preserved(tmp_path, wb, "unicode-sheet.xlsx")
    wb["数据"].insert_rows(1)

    assert wb["数据"]["A2"].value == 10
    assert wb["Summary"]["A1"].value == "=数据!A2"


def test_unparseable_formula_refuses_shift_before_mutation(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 10
    ws["B1"] = "=[broken"

    wb = _preserved(tmp_path, wb, "unparseable-formula.xlsx")
    ws = wb.active
    before_ledger = _ledger_state(wb)

    with pytest.raises(PaperRefusal, match="tokenizer rejected"):
        ws.insert_rows(1)

    assert ws["A1"].value == 10
    assert ws["B1"].value == "=[broken"
    assert _ledger_state(wb) == before_ledger


def test_late_structural_refusal_rolls_back_same_model_objects(
        tmp_path, monkeypatch):
    from openpyxl.preserve import structural

    wb = Workbook()
    data = wb.active
    data.title = "Data"
    data["A2"] = 10
    summary = wb.create_sheet("Summary")
    summary["A1"] = "=Data!A2"
    wb.defined_names.add(DefinedName("Input", attr_text="Data!$A$2"))

    wb = _preserved(tmp_path, wb, "late-refusal.xlsx")
    data = wb["Data"]
    moved_cell = data["A2"]
    formula_cell = wb["Summary"]["A1"]
    defined_name = wb.defined_names["Input"]
    before_ledger = _ledger_state(wb)

    def refuse(*_args, **_kwargs):
        raise PaperRefusal("forced late refusal")

    monkeypatch.setattr(structural, "_rebuild_merged_cells", refuse)
    with pytest.raises(PaperRefusal, match="forced late"):
        data.insert_rows(2)

    assert data["A2"] is moved_cell
    assert moved_cell.value == 10
    assert wb["Summary"]["A1"] is formula_cell
    assert formula_cell.value == "=Data!A2"
    assert wb.defined_names["Input"] is defined_name
    assert defined_name.value == "Data!$A$2"
    assert _ledger_state(wb) == before_ledger


def test_structural_warning_promoted_to_error_rolls_back(tmp_path):
    import warnings

    wb = Workbook()
    wb.active["A2"] = "keep"
    wb = _preserved(tmp_path, wb, "warning-error.xlsx")
    ws = wb.active
    ws.protection.sheet = True
    original_cell = ws["A2"]
    before_ledger = _ledger_state(wb)

    with warnings.catch_warnings():
        warnings.simplefilter("error", ProtectedWriteWarning)
        with pytest.raises(ProtectedWriteWarning):
            ws.insert_rows(2)

    assert ws["A2"] is original_cell
    assert original_cell.value == "keep"
    assert _ledger_state(wb) == before_ledger


def test_late_structural_refusal_restores_style_registries(
        tmp_path, monkeypatch):
    import openpyxl.formatting.formatting as formatting
    from openpyxl.styles import Border, Side

    wb = Workbook()
    ws = wb.active
    ws["A2"] = "anchor"
    ws["A2"].border = Border(bottom=Side(style="thin"))
    ws.merge_cells("A2:B2")
    ws.conditional_formatting.add(
        "A2", FormulaRule(formula=["A2<>\"\""]))
    wb = _preserved(tmp_path, wb, "style-rollback.xlsx")
    ws = wb.active
    before_lengths = tuple(len(getattr(wb, name)) for name in (
        "_fonts", "_fills", "_borders", "_alignments", "_protections",
        "_number_formats", "_cell_styles"))

    def refuse(*_args, **_kwargs):
        raise PaperRefusal("forced after merge rebuild")

    monkeypatch.setattr(formatting, "ConditionalFormatting", refuse)
    with pytest.raises(PaperRefusal, match="after merge rebuild"):
        ws.insert_rows(2)

    assert ws["A2"].value == "anchor"
    assert str(ws.merged_cells) == "A2:B2"
    assert tuple(len(getattr(wb, name)) for name in (
        "_fonts", "_fills", "_borders", "_alignments", "_protections",
        "_number_formats", "_cell_styles")) == before_lengths
