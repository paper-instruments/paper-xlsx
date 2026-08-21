"""Supported delivery helpers and package hardening."""
from __future__ import annotations

import io
import json
from pathlib import Path
import warnings
import zipfile

import pytest

from openpyxl import load_workbook
from openpyxl.errors import (
    AmbiguousTargetError,
    ProtectedWriteWarning,
    TargetNotFoundError,
    UnsupportedStructureError,
)


def _with_pivot_graph(workbook, pivots, *, refresh_on_load=False,
                      cache_parts=None, prefixed_cache=False,
                      prefixed_pivot=False,
                      orphan_parts=(), worksheet_sources=None,
                      cache_source_types=None, cache_attributes=None):
    """Attach a small relationship-complete pivot graph to source bytes.

    The preservation API indexes package relationships, not openpyxl's pivot
    model. Keeping the fixture at the package layer makes each resolution
    contract explicit without relying on the upstream pivot deserializer.
    ``pivots`` is an iterable of ``(name, cache_id)`` pairs.
    """
    workbook_xml = (
        b'<pivotCaches xmlns:r="http://schemas.openxmlformats.org/office'
        b'Document/2006/relationships">'
        + b"".join(
            b'<pivotCache cacheId="%s" r:id="rIdCache%s"/>'
            % (cache_id.encode("ascii"), cache_id.encode("ascii"))
            for cache_id in sorted({cache_id for _name, cache_id in pivots})
        )
        + b"</pivotCaches>"
    )
    cache_ids = sorted({cache_id for _name, cache_id in pivots})
    if cache_parts is None:
        cache_parts = {
            cache_id: "xl/pivotCache/pivotCacheDefinition{0}.xml".format(
                cache_id)
            for cache_id in cache_ids
        }
    workbook_rels = b"".join(
        b'<Relationship Id="rIdCache%s" Type="http://schemas.openxml'
        b'formats.org/officeDocument/2006/relationships/'
        b'pivotCacheDefinition" Target="/%s"/>'
        % (cache_id.encode("ascii"),
           cache_parts[cache_id].encode("utf-8"))
        for cache_id in cache_ids
    )
    sheet_rels = b"".join(
        b'<Relationship Id="rIdPivot%s" Type="http://schemas.openxml'
        b'formats.org/officeDocument/2006/relationships/pivotTable" '
        b'Target="../pivotTables/pivotTable%s.xml"/>'
        % (str(index).encode("ascii"), str(index).encode("ascii"))
        for index, _pivot in enumerate(pivots, 1)
    )
    content_type_overrides = b"".join(
        b'<Override PartName="/xl/pivotTables/pivotTable%s.xml" '
        b'ContentType="application/vnd.openxmlformats-officedocument.'
        b'spreadsheetml.pivotTable+xml"/>' % str(index).encode("ascii")
        for index, _pivot in enumerate(pivots, 1)
    ) + b"".join(
        b'<Override PartName="/%s" ContentType="application/vnd.'
        b'openxmlformats-officedocument.spreadsheetml.'
        b'pivotCacheDefinition+xml"/>' % part.encode("utf-8")
        for part in tuple(cache_parts.values()) + tuple(orphan_parts)
    )
    sheet_rels_part = "xl/worksheets/_rels/sheet1.xml.rels"
    sheet_rels_seen = False
    source = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(workbook._paper_source)) as zin, \
            zipfile.ZipFile(source, "w") as zout:
        for info in zin.infolist():
            payload = zin.read(info.filename)
            if info.filename == "xl/workbook.xml":
                payload = payload.replace(b"</workbook>",
                                          workbook_xml + b"</workbook>")
            elif info.filename == "xl/_rels/workbook.xml.rels":
                payload = payload.replace(b"</Relationships>",
                                          workbook_rels
                                          + b"</Relationships>")
            elif info.filename == "[Content_Types].xml":
                payload = payload.replace(
                    b"</Types>", content_type_overrides + b"</Types>")
            elif info.filename == sheet_rels_part:
                sheet_rels_seen = True
                payload = payload.replace(
                    b"</Relationships>", sheet_rels + b"</Relationships>")
            zout.writestr(info, payload)
        if not sheet_rels_seen:
            zout.writestr(
                sheet_rels_part,
                b'<Relationships xmlns="http://schemas.openxmlformats.org/'
                b'package/2006/relationships">' + sheet_rels
                + b"</Relationships>")
        for index, (name, cache_id) in enumerate(pivots, 1):
            root = b"x:pivotTableDefinition" if prefixed_pivot \
                else b"pivotTableDefinition"
            namespace = (b' xmlns:x="http://schemas.openxmlformats.org/'
                         b'spreadsheetml/2006/main"') if prefixed_pivot \
                else (b' xmlns="http://schemas.openxmlformats.org/'
                      b'spreadsheetml/2006/main"')
            zout.writestr(
                "xl/pivotTables/pivotTable%s.xml" % index,
                b'<' + root + namespace + b' name="'
                + name.encode("utf-8") + b'" cacheId="'
                + cache_id.encode("ascii") + b'" dataCaption="Values">'
                + b'<location ref="F1:G2" firstHeaderRow="1" '
                + b'firstDataRow="1" firstDataCol="1"/></' + root + b'>')
        refresh = b' refreshOnLoad="1"' if refresh_on_load else b""
        worksheet_sources = worksheet_sources or {}
        cache_source_types = cache_source_types or {}
        cache_attributes = cache_attributes or {}
        for cache_id in cache_ids:
            root = b"x:pivotCacheDefinition" if prefixed_cache \
                else b"pivotCacheDefinition"
            namespace = (b' xmlns:x="http://schemas.openxmlformats.org/'
                         b'spreadsheetml/2006/main"') if prefixed_cache \
                else (b' xmlns="http://schemas.openxmlformats.org/'
                      b'spreadsheetml/2006/main"')
            source_tag = b"x:cacheSource" if prefixed_cache \
                else b"cacheSource"
            source_type = cache_source_types.get(
                cache_id, "worksheet").encode("ascii")
            worksheet_source = worksheet_sources.get(cache_id)
            if worksheet_source is None:
                source_payload = (b'<' + source_tag
                                  + b' type="' + source_type + b'"/>')
            else:
                source_payload = (
                    b'<' + source_tag + b' type="' + source_type + b'">'
                    + worksheet_source + b'</' + source_tag + b'>')
            zout.writestr(
                cache_parts[cache_id],
                b'<' + root + namespace + b' recordCount="0"'
                + refresh + cache_attributes.get(cache_id, b"")
                + b'>' + source_payload
                + b'</' + root + b'>')
        for part in orphan_parts:
            zout.writestr(
                part,
                b'<pivotCacheDefinition xmlns="http://schemas.openxml'
                b'formats.org/spreadsheetml/2006/main" recordCount="0"/>'
            )
    workbook._paper_source = source.getvalue()


def test_copy_format_through_the_splice(fixture_copy, tmp_path):
    from openpyxl.preserve import copy_format
    from openpyxl.styles import Font

    wb = load_workbook(
        fixture_copy("features/schedule.xlsx"), preserve=True)
    ws = wb["Schedule"]
    ws["B2"].font = Font(bold=True, italic=True)
    assert copy_format(ws, "B2", "B3:B4") == 2
    out = tmp_path / "o.xlsx"
    wb.save(out)
    wb2 = load_workbook(out)
    assert wb2["Schedule"]["B3"].font.bold is True
    assert wb2["Schedule"]["B4"].font.italic is True


def test_copy_format_rolls_back_model_and_ledger(
        fixture_copy, monkeypatch):
    from copy import copy

    from openpyxl.preserve import copy_format
    from openpyxl.styles import Font
    import openpyxl.preserve.styleverbs as styleverbs

    wb = load_workbook(
        fixture_copy("features/schedule.xlsx"), preserve=True)
    ws = wb["Schedule"]
    ws["B2"].font = Font(bold=True, italic=True)
    before_styles = {
        coordinate: copy(ws[coordinate]._style)
        for coordinate in ("B3", "B4")
    }
    before_dirty = set(wb._paper_ledger.dirty_coordinates(ws))
    before_row = ws._current_row
    calls = []

    def fail_after_first(coordinate):
        calls.append(coordinate)
        raise RuntimeError("injected format-copy failure")

    monkeypatch.setattr(
        styleverbs, "_copy_format_commit_point", fail_after_first)
    with pytest.raises(RuntimeError, match="injected"):
        copy_format(ws, "B2", "B3:B4")
    assert calls == [(3, 2)]
    assert {coordinate: ws[coordinate]._style
            for coordinate in ("B3", "B4")} == before_styles
    assert set(wb._paper_ledger.dirty_coordinates(ws)) == before_dirty
    assert ws._current_row == before_row


def test_copy_format_rolls_back_new_cells_on_interrupt(
        fixture_copy, monkeypatch):
    from openpyxl.preserve import copy_format
    import openpyxl.preserve.styleverbs as styleverbs

    wb = load_workbook(
        fixture_copy("features/schedule.xlsx"), preserve=True)
    ws = wb["Schedule"]
    before_cells = set(ws._cells)
    before_dirty = set(wb._paper_ledger.dirty_coordinates(ws))
    before_row = ws._current_row
    calls = []

    def interrupt_after_second(coordinate):
        calls.append(coordinate)
        if len(calls) == 2:
            raise KeyboardInterrupt("injected format-copy interruption")

    monkeypatch.setattr(
        styleverbs, "_copy_format_commit_point", interrupt_after_second)
    with pytest.raises(KeyboardInterrupt, match="injected"):
        copy_format(ws, "B2", "Z100:Z101")

    assert calls == [(100, 26), (101, 26)]
    assert set(ws._cells) == before_cells
    assert set(wb._paper_ledger.dirty_coordinates(ws)) == before_dirty
    assert ws._current_row == before_row


def test_copy_format_preflights_merges_and_protection(
        fixture_copy, tmp_path):
    from openpyxl.preserve import copy_format

    src = fixture_copy("minimal/minimal_clean.xlsx")
    wb = load_workbook(src, preserve=True)
    ws = wb.active
    ws["A1"].number_format = "$0.00"
    ws.merge_cells("B1:C1")
    before_anchor = ws["B1"]._style
    with pytest.raises(UnsupportedStructureError,
                       match="non-anchor merged cell"):
        copy_format(ws, "A1", "B1:C1")
    assert ws["B1"]._style == before_anchor

    protected = tmp_path / "protected.xlsx"
    wb = load_workbook(src, preserve=True)
    ws = wb.active
    ws["A1"].number_format = "$0.00"
    ws["B1"].number_format = "0.00"
    ws.protection.sheet = True
    wb.save(protected)
    wb = load_workbook(protected, preserve=True)
    ws = wb.active
    wb.strict_protection = True
    before = ws["B1"]._style
    before_cells = set(ws._cells)
    with pytest.raises(UnsupportedStructureError,
                       match="strict_protection"):
        copy_format(ws, "A1", "B1:C1")
    assert ws["B1"]._style == before
    assert set(ws._cells) == before_cells

    wb.strict_protection = False
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        assert copy_format(ws, "A1", "B1:C1") == 2
    assert len([item for item in caught
                if issubclass(item.category, ProtectedWriteWarning)]) == 1


def test_pivot_refresh_requires_explicit_scope(fixture_copy):
    wb = load_workbook(
        fixture_copy("minimal/minimal_clean.xlsx"), preserve=True)
    with pytest.raises(ValueError, match="exactly one"):
        wb.set_pivot_refresh_on_load()


def test_all_pivot_refresh_patches_only_root_attribute(
        fixture_copy, tmp_path):
    wb = load_workbook(
        fixture_copy("minimal/minimal_clean.xlsx"), preserve=True)
    custom_part = "xl/custom/pivot-cache-a.xml"
    orphan_part = "xl/pivotCache/pivotCacheDefinition99.xml"
    _with_pivot_graph(
        wb, [("SalesPivot", "1")], cache_parts={"1": custom_part},
        orphan_parts=(orphan_part,))
    original = wb._paper_source

    assert wb.set_pivot_refresh_on_load(all=True) == [
        custom_part]
    out = tmp_path / "o.xlsx"
    wb.save(out)
    with zipfile.ZipFile(io.BytesIO(original)) as before, \
            zipfile.ZipFile(out) as after:
        assert b'refreshOnLoad="1"' in after.read(custom_part)
        assert after.read(orphan_part) == before.read(orphan_part)


def test_targeted_pivot_uses_current_sheet_name_after_rename(
        fixture_copy, tmp_path):
    wb = load_workbook(
        fixture_copy("minimal/minimal_clean.xlsx"), preserve=True)
    _with_pivot_graph(
        wb, [("SalesPivot", "1")],
        cache_source_types={"1": "external"})
    wb["Sheet1"].title = "Renamed's"

    assert wb.set_pivot_refresh_on_load(
        pivots=["'Renamed''s'!SalesPivot"]
    ) == ["xl/pivotCache/pivotCacheDefinition1.xml"]
    out = tmp_path / "renamed-pivot.xlsx"
    wb.save(out)
    with zipfile.ZipFile(out) as archive:
        assert b'refreshOnLoad="1"' in archive.read(
            "xl/pivotCache/pivotCacheDefinition1.xml")


def test_prefixed_pivot_parts_are_indexed_patched_and_reported(
        fixture_copy, tmp_path):
    wb = load_workbook(
        fixture_copy("minimal/minimal_clean.xlsx"), preserve=True)
    _with_pivot_graph(
        wb, [("SalesPivot", "1")], prefixed_cache=True,
        prefixed_pivot=True)

    wb.set_pivot_refresh_on_load(all=True)
    receipt = wb.save(tmp_path / "prefixed-pivot.xlsx", receipt=True)

    assert [effect for effect in receipt.derived_effects
            if effect["kind"] == "pivot_refresh_on_load_enabled"] == [{
                "kind": "pivot_refresh_on_load_enabled",
                "part": "xl/pivotCache/pivotCacheDefinition1.xml",
                "cause": "explicit_request",
            }]


def test_targeted_pivots_follow_relationships_and_deduplicate_shared_cache(
        fixture_copy, tmp_path):
    wb = load_workbook(
        fixture_copy("minimal/minimal_clean.xlsx"), preserve=True)
    _with_pivot_graph(wb, [("SalesPivot", "1"), ("MarginPivot", "1")])

    assert wb.set_pivot_refresh_on_load(
        pivots=["Sheet1!SalesPivot", "MarginPivot"]
    ) == ["xl/pivotCache/pivotCacheDefinition1.xml"]
    receipt = wb.save(tmp_path / "targeted-pivots.xlsx", receipt=True)

    assert [effect for effect in receipt.derived_effects
            if effect["kind"] == "pivot_refresh_on_load_enabled"] == [{
                "kind": "pivot_refresh_on_load_enabled",
                "part": "xl/pivotCache/pivotCacheDefinition1.xml",
                "cause": "explicit_request",
            }]


def test_targeted_pivot_reports_ambiguity_and_missing_names(fixture_copy):
    wb = load_workbook(
        fixture_copy("minimal/minimal_clean.xlsx"), preserve=True)
    _with_pivot_graph(wb, [("Duplicate", "1"), ("Duplicate", "2")])

    with pytest.raises(AmbiguousTargetError, match="sheet-qualified"):
        wb.set_pivot_refresh_on_load(pivots=["Duplicate"])
    with pytest.raises(TargetNotFoundError, match="Missing"):
        wb.set_pivot_refresh_on_load(pivots=["Missing"])


def test_targeted_pivot_is_idempotent_when_refresh_is_already_enabled(
        fixture_copy, tmp_path):
    wb = load_workbook(
        fixture_copy("minimal/minimal_clean.xlsx"), preserve=True)
    _with_pivot_graph(wb, [("ReadyPivot", "1")], refresh_on_load=True)
    original = wb._paper_source

    assert wb.set_pivot_refresh_on_load(pivots=["ReadyPivot"]) == [
        "xl/pivotCache/pivotCacheDefinition1.xml"]
    receipt = wb.save(tmp_path / "already-enabled.xlsx", receipt=True)

    with zipfile.ZipFile(io.BytesIO(original)) as before, \
            zipfile.ZipFile(tmp_path / "already-enabled.xlsx") as after:
        part = "xl/pivotCache/pivotCacheDefinition1.xml"
        assert after.read(part) == before.read(part)
    assert not any(effect["kind"] == "pivot_refresh_on_load_enabled"
                   for effect in receipt.derived_effects)


def test_pivot_refresh_requires_preserve(fixture_copy):
    wb = load_workbook(
        fixture_copy("minimal/minimal_clean.xlsx"), preserve=False)
    with pytest.raises(ValueError, match="preserve"):
        wb.set_pivot_refresh_on_load(all=True)


def _excel_pivot_fixture():
    return (Path(__file__).parents[2]
            / "openpyxl/reader/tests/data/pivot.xlsx")


def test_defined_name_pivot_source_edit_refuses_before_delivery(tmp_path):
    source = _excel_pivot_fixture()
    wb = load_workbook(source, preserve=True)
    wb["raw"]["A2"] = 999

    with pytest.raises(UnsupportedStructureError) as caught:
        wb.validate()
    assert caught.value.kind == "stale-pivot-cache"
    assert caught.value.anchor == "ptsheet!PivotTable1"
    assert "mydata" in str(caught.value)
    assert "headless readers" in str(caught.value)

    output = tmp_path / "refused.xlsx"
    with pytest.raises(UnsupportedStructureError,
                       match="stale pivot results"):
        wb.save(output)
    assert not output.exists()


def test_pivot_source_retarget_requires_refresh_and_reports_cause(tmp_path):
    source = _excel_pivot_fixture()
    wb = load_workbook(source, preserve=True)
    wb.defined_names["mydata"].attr_text = "raw!$A$1:$F$17"

    with pytest.raises(UnsupportedStructureError) as caught:
        wb.validate()
    assert caught.value.kind == "stale-pivot-cache"
    assert caught.value.anchor == "ptsheet!PivotTable1"
    wb.set_pivot_refresh_on_load(pivots=["ptsheet!PivotTable1"])
    output = tmp_path / "retargeted-source.xlsx"

    receipt = wb.save(output, receipt=True)

    effects = [
        effect for effect in receipt.derived_effects
        if effect["kind"] == "pivot_source_changed_requires_refresh"]
    assert effects == [{
        "kind": "pivot_source_changed_requires_refresh",
        "part": "xl/pivotCache/pivotCacheDefinition1.xml",
        "cause": "source_changed",
        "pivots": ["ptsheet!PivotTable1"],
        "source": "mydata",
        "requirement": "excel_refresh_on_open",
    }]
    reopened = load_workbook(output, preserve=True)
    assert reopened.defined_names["mydata"].attr_text == \
        "raw!$A$1:$F$17"
    with zipfile.ZipFile(output) as archive:
        assert b'refreshOnLoad="1"' in archive.read(
            "xl/pivotCache/pivotCacheDefinition1.xml")


@pytest.mark.parametrize("formulas", [
    {"A2": "=C2*2", "C2": "=D2+1"},
    {"A2": '=INDIRECT("D2")'},
])
def test_formula_fed_pivot_source_requires_refresh(tmp_path, formulas):
    from openpyxl import Workbook

    source = tmp_path / "formula-source.xlsx"
    created = Workbook()
    ws = created.active
    ws.title = "Data"
    ws["A1"] = "Pivot input"
    ws["D2"] = 1
    for coordinate, formula in formulas.items():
        ws[coordinate] = formula
    created.save(source)

    wb = load_workbook(source, preserve=True)
    _with_pivot_graph(
        wb, [("FormulaPivot", "1")],
        worksheet_sources={
            "1": b'<worksheetSource ref="A1:A10" sheet="Data"/>'})
    wb["Data"]["D2"] = 5

    with pytest.raises(UnsupportedStructureError) as caught:
        wb.validate()
    assert caught.value.kind == "stale-pivot-cache"
    assert caught.value.anchor == "Data!FormulaPivot"


def test_data_only_pivot_dependency_uses_retained_formula(tmp_path):
    from openpyxl import Workbook

    source = tmp_path / "data-only-formula-source.xlsx"
    created = Workbook()
    ws = created.active
    ws.title = "Data"
    ws["A2"] = "=D2"
    ws["D2"] = 1
    created.save(source)

    wb = load_workbook(source, preserve=True, data_only=True)
    _with_pivot_graph(
        wb, [("FormulaPivot", "1")],
        worksheet_sources={
            "1": b'<worksheetSource ref="A2" sheet="Data"/>'})
    wb["Data"]["D2"] = 5

    with pytest.raises(UnsupportedStructureError) as caught:
        wb.save(tmp_path / "refused.xlsx", allow_formula_loss=True)
    assert caught.value.kind == "stale-pivot-cache"


def test_defined_name_dependency_change_in_pivot_source_requires_refresh(
        tmp_path):
    from openpyxl import Workbook
    from openpyxl.workbook.defined_name import DefinedName

    source = tmp_path / "named-formula-source.xlsx"
    created = Workbook()
    data = created.active
    data.title = "Data"
    data["A1"] = "=Input"
    inputs = created.create_sheet("Inputs")
    inputs["A1"] = 1
    inputs["A2"] = 2
    created.defined_names.add(
        DefinedName("Input", attr_text="Inputs!$A$1"))
    created.save(source)

    wb = load_workbook(source, preserve=True)
    _with_pivot_graph(
        wb, [("FormulaPivot", "1")],
        worksheet_sources={
            "1": b'<worksheetSource ref="A1" sheet="Data"/>'})
    wb.defined_names["Input"].attr_text = "Inputs!$A$2"

    with pytest.raises(UnsupportedStructureError) as caught:
        wb.validate()
    assert caught.value.kind == "stale-pivot-cache"

    wb.set_pivot_refresh_on_load(pivots=["Data!FormulaPivot"])
    output = tmp_path / "named-formula-output.xlsx"
    wb.save(output)
    with zipfile.ZipFile(output) as archive:
        workbook_xml = archive.read("xl/workbook.xml")
    assert b'calcMode="auto"' in workbook_xml
    assert b'fullCalcOnLoad="1"' in workbook_xml


def test_dynamic_name_dependency_change_in_pivot_source_requires_refresh(
        tmp_path):
    from openpyxl import Workbook
    from openpyxl.workbook.defined_name import DefinedName

    source = tmp_path / "dynamic-name-formula-source.xlsx"
    created = Workbook()
    data = created.active
    data.title = "Data"
    data["A1"] = "=Input"
    data["D1"] = 1
    data["D2"] = 2
    created.defined_names.add(DefinedName(
        "Input", attr_text="OFFSET(Data!$D$1,0,0)"))
    created.save(source)

    wb = load_workbook(source, preserve=True)
    _with_pivot_graph(
        wb, [("FormulaPivot", "1")],
        worksheet_sources={
            "1": b'<worksheetSource ref="A1" sheet="Data"/>'})
    wb.defined_names["Input"].attr_text = "OFFSET(Data!$D$2,0,0)"

    with pytest.raises(UnsupportedStructureError) as caught:
        wb.validate()
    assert caught.value.kind == "stale-pivot-cache"


def test_table_dependency_change_in_pivot_source_requires_refresh(tmp_path):
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table

    source = tmp_path / "table-formula-source.xlsx"
    created = Workbook()
    data = created.active
    data.title = "Data"
    data.append(["Amount"])
    data.append([1])
    data.append([2])
    data.add_table(Table(displayName="Sales", ref="A1:A3"))
    data["C1"] = "=SUM(Sales[Amount])"
    created.save(source)

    wb = load_workbook(source, preserve=True)
    _with_pivot_graph(
        wb, [("FormulaPivot", "1")],
        worksheet_sources={
            "1": b'<worksheetSource ref="C1" sheet="Data"/>'})
    wb["Data"].tables["Sales"].ref = "A1:A4"

    with pytest.raises(UnsupportedStructureError) as caught:
        wb.validate()
    assert caught.value.kind == "stale-pivot-cache"


def test_array_formula_followers_in_pivot_source_require_refresh(tmp_path):
    from openpyxl import Workbook
    from openpyxl.worksheet.formula import ArrayFormula

    source = tmp_path / "array-formula-source.xlsx"
    created = Workbook()
    ws = created.active
    ws.title = "Data"
    ws["A1"] = ArrayFormula(ref="A1:A3", text="=D1*{1;2;3}")
    ws["D1"] = 1
    created.save(source)

    wb = load_workbook(source, preserve=True)
    _with_pivot_graph(
        wb, [("FormulaPivot", "1")],
        worksheet_sources={
            "1": b'<worksheetSource ref="A2:A3" sheet="Data"/>'})
    wb["Data"]["D1"] = 5

    with pytest.raises(UnsupportedStructureError) as caught:
        wb.validate()
    assert caught.value.kind == "stale-pivot-cache"


def test_data_table_followers_in_pivot_source_require_refresh(tmp_path):
    from openpyxl import Workbook
    from openpyxl.worksheet.formula import DataTableFormula

    source = tmp_path / "data-table-source.xlsx"
    created = Workbook()
    ws = created.active
    ws.title = "Data"
    ws["A1"] = DataTableFormula(ref="A1:A3", r1="D1")
    ws["D1"] = 1
    created.save(source)

    wb = load_workbook(source, preserve=True)
    _with_pivot_graph(
        wb, [("FormulaPivot", "1")],
        worksheet_sources={
            "1": b'<worksheetSource ref="A2:A3" sheet="Data"/>'})
    wb["Data"]["D1"] = 5

    with pytest.raises(UnsupportedStructureError) as caught:
        wb.validate()
    assert caught.value.kind == "stale-pivot-cache"


def test_formula_cache_write_in_pivot_source_requires_refresh(tmp_path):
    from openpyxl import Workbook

    source = tmp_path / "formula-cache-source.xlsx"
    created = Workbook()
    ws = created.active
    ws.title = "Data"
    ws["A1"] = "=1+1"
    created.save(source)

    wb = load_workbook(source, preserve=True)
    _with_pivot_graph(
        wb, [("FormulaPivot", "1")],
        worksheet_sources={
            "1": b'<worksheetSource ref="A1" sheet="Data"/>'})
    wb._paper_ledger.cache_writes[wb["Data"]] = {(1, 1): 2}

    with pytest.raises(UnsupportedStructureError) as caught:
        wb.validate()
    assert caught.value.kind == "stale-pivot-cache"


@pytest.mark.parametrize("formula", ["=NOW()", "=TODAY()", "=RAND()"])
def test_volatile_formula_in_pivot_source_requires_refresh(
        tmp_path, formula):
    from openpyxl import Workbook

    source = tmp_path / "volatile-pivot-source.xlsx"
    created = Workbook()
    ws = created.active
    ws.title = "Data"
    ws["A1"] = formula
    ws["D1"] = 1
    created.save(source)

    wb = load_workbook(source, preserve=True)
    _with_pivot_graph(
        wb, [("VolatilePivot", "1")],
        worksheet_sources={
            "1": b'<worksheetSource ref="A1" sheet="Data"/>'})
    wb["Data"]["D1"] = 2

    with pytest.raises(UnsupportedStructureError) as caught:
        wb.validate()
    assert caught.value.kind == "stale-pivot-cache"


@pytest.mark.parametrize(
    "mutation", ["cell-format", "hidden-row", "cross-column-filter"])
def test_non_value_formula_precedent_requires_pivot_refresh(
        tmp_path, mutation):
    from openpyxl import Workbook

    source = tmp_path / (mutation + "-pivot-source.xlsx")
    created = Workbook()
    ws = created.active
    ws.title = "Data"
    ws["A1"] = ('=CELL("color",D1)' if mutation == "cell-format"
                else "=SUBTOTAL(109,D1:D2)")
    ws["D1"] = -1
    ws["D2"] = 20
    if mutation == "cross-column-filter":
        ws["C1"] = "Filter"
        ws["C2"] = "drop"
    created.save(source)

    wb = load_workbook(source, preserve=True)
    _with_pivot_graph(
        wb, [("ContextPivot", "1")],
        worksheet_sources={
            "1": b'<worksheetSource ref="A1" sheet="Data"/>'})
    if mutation == "cell-format":
        wb["Data"]["D1"].number_format = "[Red]0"
    elif mutation == "hidden-row":
        wb["Data"].row_dimensions[2].hidden = True
    else:
        wb["Data"].auto_filter.ref = "C1:C2"
        wb["Data"].auto_filter.add_filter_column(0, ["keep"])

    assert not any(wb._paper_ledger.value_overwrites.values())
    with pytest.raises(UnsupportedStructureError) as caught:
        wb.validate()
    assert caught.value.kind == "stale-pivot-cache"


def test_non_contextual_formula_allows_row_display_change(tmp_path):
    from openpyxl import Workbook

    source = tmp_path / "sum-hidden-row-pivot-source.xlsx"
    created = Workbook()
    ws = created.active
    ws.title = "Data"
    ws["A1"] = "=SUM(D1:D2)"
    ws["D1"] = 10
    ws["D2"] = 20
    created.save(source)

    wb = load_workbook(source, preserve=True)
    _with_pivot_graph(
        wb, [("SumPivot", "1")],
        worksheet_sources={
            "1": b'<worksheetSource ref="A1" sheet="Data"/>'})
    wb["Data"].row_dimensions[2].hidden = True

    wb.validate()


def test_dirty_closure_processes_high_fanout_once(monkeypatch):
    from openpyxl import Workbook
    import openpyxl.preserve.pivots as pivots

    wb = Workbook()
    ws = wb.active
    ws["A1"] = 1
    formula_count = 5000
    for row in range(1, formula_count + 1):
        ws.cell(row, 2, "=$A$1")

    calls = 0
    original = pivots._DependencyIndex.pop_cell

    def counted_pop_cell(self, cell):
        nonlocal calls
        calls += 1
        return original(self, cell)

    monkeypatch.setattr(
        pivots._DependencyIndex, "pop_cell", counted_pop_cell)
    tainted, tainted_ranges = pivots._dirty_closure(
        wb, {ws: {(1, 1)}})

    assert len(tainted) == formula_count + 1
    assert tainted_ranges == set()
    assert calls == formula_count + 1


def test_dirty_closure_follows_reverse_order_chain():
    from openpyxl import Workbook
    from openpyxl.preserve.pivots import _dirty_closure

    wb = Workbook()
    ws = wb.active
    formula_count = 2000
    for row in range(1, formula_count + 1):
        ws.cell(row, 1, "=A{0}".format(row + 1))
    ws.cell(formula_count + 1, 1, 1)

    tainted, tainted_ranges = _dirty_closure(
        wb, {ws: {(formula_count + 1, 1)}})

    assert len(tainted) == formula_count + 1
    assert tainted_ranges == set()


def test_oracle_recalc_requests_formula_backed_pivot_refresh(
        tmp_path, monkeypatch):
    from openpyxl import Workbook, oracle

    raw = tmp_path / "formula-pivot-raw.xlsx"
    created = Workbook()
    ws = created.active
    ws.title = "Data"
    ws["A1"] = "=D1*2"
    ws["D1"] = 3
    created.save(raw)

    source = tmp_path / "formula-pivot.xlsx"
    wb = load_workbook(raw, preserve=True)
    _with_pivot_graph(
        wb, [("FormulaPivot", "1")],
        worksheet_sources={
            "1": b'<worksheetSource ref="A1" sheet="Data"/>'})
    wb.save(source)
    original = source.read_bytes()

    def recalculated(data, _timeout):
        output = io.BytesIO()
        with zipfile.ZipFile(io.BytesIO(data)) as zin, \
                zipfile.ZipFile(output, "w") as zout:
            for info in zin.infolist():
                payload = zin.read(info.filename)
                if info.filename == "xl/worksheets/sheet1.xml":
                    old = next((candidate for candidate in (
                        b"<f>D1*2</f><v></v>",
                        b"<f>D1*2</f><v/>",
                        b"<f>D1*2</f><v />",
                    ) if candidate in payload), None)
                    assert old is not None
                    payload = payload.replace(
                        old, b"<f>D1*2</f><v>6</v>", 1)
                zout.writestr(info, payload)
        return output.getvalue()

    monkeypatch.setattr(oracle, "_recalculate_bytes", recalculated)
    candidate = tmp_path / "candidate.xlsx"
    result = oracle.recalc(source, output_path=candidate)

    assert source.read_bytes() == original
    assert result.pivot_refreshes == [{
        "part": "xl/pivotCache/pivotCacheDefinition1.xml",
        "pivots": ["Data!FormulaPivot"],
        "source": "Data!A1",
        "requirement": "excel_refresh_on_open",
    }]
    assert result.to_dict()["pivot_refreshes"] == result.pivot_refreshes
    with zipfile.ZipFile(candidate) as archive:
        assert b'refreshOnLoad="1"' in archive.read(
            "xl/pivotCache/pivotCacheDefinition1.xml")
    assert load_workbook(candidate, data_only=True)["Data"]["A1"].value == 6


def test_formula_change_outside_pivot_source_remains_allowed(tmp_path):
    from openpyxl import Workbook

    source = tmp_path / "unrelated-formula.xlsx"
    created = Workbook()
    ws = created.active
    ws.title = "Data"
    ws["A1"] = "Pivot input"
    ws["A2"] = 10
    ws["B2"] = "=D2*2"
    ws["D2"] = 1
    created.save(source)

    wb = load_workbook(source, preserve=True)
    _with_pivot_graph(
        wb, [("FormulaPivot", "1")],
        worksheet_sources={
            "1": b'<worksheetSource ref="A1:A10" sheet="Data"/>'})
    wb["Data"]["D2"] = 5

    wb.validate()
    wb.save(tmp_path / "unrelated-formula-output.xlsx")


def test_unrelated_edit_preserves_pivot_cache_byte_identically(tmp_path):
    source = _excel_pivot_fixture()
    wb = load_workbook(source, preserve=True)
    wb["ptsheet"]["G20"] = "unrelated"
    output = tmp_path / "unrelated.xlsx"

    wb.save(output)

    cache_part = "xl/pivotCache/pivotCacheDefinition1.xml"
    with zipfile.ZipFile(source) as before, zipfile.ZipFile(output) as after:
        assert after.read(cache_part) == before.read(cache_part)
        changed = sorted(
            name for name in set(before.namelist()) & set(after.namelist())
            if before.read(name) != after.read(name))
    assert changed == ["xl/worksheets/sheet1.xml"]


def test_explicit_refresh_allows_source_edit_and_reports_requirement(
        tmp_path):
    source = _excel_pivot_fixture()
    wb = load_workbook(source, preserve=True)
    wb["raw"]["A2"] = 999
    wb.set_pivot_refresh_on_load(pivots=["ptsheet!PivotTable1"])
    output = tmp_path / "refresh.xlsx"

    wb.validate()
    receipt = wb.save(output, receipt=True)

    assert receipt.parts_changed == [
        "xl/pivotCache/pivotCacheDefinition1.xml",
        "xl/worksheets/sheet2.xml",
    ]
    assert [effect for effect in receipt.derived_effects
            if effect["kind"] ==
            "pivot_source_changed_requires_refresh"] == [{
                "kind": "pivot_source_changed_requires_refresh",
                "part": "xl/pivotCache/pivotCacheDefinition1.xml",
                "cause": "input_changed",
                "pivots": ["ptsheet!PivotTable1"],
                "source": "mydata",
                "requirement": "excel_refresh_on_open",
            }]
    with zipfile.ZipFile(output) as archive:
        assert b'refreshOnLoad="1"' in archive.read(
            "xl/pivotCache/pivotCacheDefinition1.xml")


def test_direct_range_pivot_source_edit_refuses(fixture_copy):
    wb = load_workbook(
        fixture_copy("minimal/minimal_clean.xlsx"), preserve=True)
    _with_pivot_graph(
        wb, [("SalesPivot", "1")],
        worksheet_sources={
            "1": b'<worksheetSource ref="A1:B10" sheet="Sheet1"/>'})
    wb["Sheet1"]["B2"] = "changed"

    with pytest.raises(UnsupportedStructureError,
                       match=r"Sheet1!A1:B10") as caught:
        wb.validate()
    assert caught.value.kind == "stale-pivot-cache"


@pytest.mark.parametrize(("source_ref", "outside_cell"), [
    ("A:D", "E1"),
    ("1:4", "A5"),
])
def test_open_pivot_source_allows_unrelated_edit(
        fixture_copy, source_ref, outside_cell):
    wb = load_workbook(
        fixture_copy("minimal/minimal_clean.xlsx"), preserve=True)
    _with_pivot_graph(
        wb, [("SalesPivot", "1")],
        worksheet_sources={
            "1": ('<worksheetSource ref="{0}" sheet="Sheet1"/>'
                  .format(source_ref).encode("ascii"))})
    wb["Sheet1"][outside_cell] = "unrelated"

    wb.validate()


def test_structural_shift_intersecting_encoded_pivot_source_refuses_atomically(
        fixture_copy):
    wb = load_workbook(
        fixture_copy("minimal/minimal_clean.xlsx"), preserve=True)
    _with_pivot_graph(
        wb, [("SalesPivot", "1")],
        worksheet_sources={
            "1": b'<worksheetSource ref="A1:A3" sheet="Sh&#101;et1"/>'})
    before = [wb["Sheet1"].cell(row=row, column=1).value
              for row in range(1, 5)]

    with pytest.raises(UnsupportedStructureError) as caught:
        wb["Sheet1"].insert_rows(2)
    assert "pivot" in str(caught.value)
    assert [wb["Sheet1"].cell(row=row, column=1).value
            for row in range(1, 5)] == before


def test_unresolved_local_pivot_source_refuses_conservatively(fixture_copy):
    wb = load_workbook(
        fixture_copy("minimal/minimal_clean.xlsx"), preserve=True)
    _with_pivot_graph(wb, [("SalesPivot", "1")])
    wb["Sheet1"]["A1"] = "changed"

    with pytest.raises(UnsupportedStructureError,
                       match="unresolved worksheet pivot source") as caught:
        wb.validate()
    assert caught.value.kind == "stale-pivot-cache"


def test_external_pivot_source_does_not_block_local_edits(
        fixture_copy, tmp_path):
    wb = load_workbook(
        fixture_copy("minimal/minimal_clean.xlsx"), preserve=True)
    _with_pivot_graph(
        wb, [("ExternalPivot", "1")],
        cache_source_types={"1": "external"})
    wb["Sheet1"]["A1"] = "local change"

    wb.validate()
    wb.save(tmp_path / "external-pivot.xlsx")


def test_named_table_expansion_requires_explicit_pivot_refresh(tmp_path):
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table

    source = tmp_path / "table-source.xlsx"
    created = Workbook()
    ws = created.active
    ws.title = "Data"
    ws.append(["Region", "Amount"])
    ws.append(["West", 10])
    ws.add_table(Table(displayName="SalesData", ref="A1:B2"))
    created.save(source)

    wb = load_workbook(source, preserve=True)
    _with_pivot_graph(
        wb, [("SalesPivot", "1")],
        worksheet_sources={
            "1": b'<worksheetSource name="SalesData"/>'})
    wb["Data"].append_table_row("SalesData", ["East", 20])

    with pytest.raises(UnsupportedStructureError,
                       match="SalesData"):
        wb.validate()

    wb.set_pivot_refresh_on_load(pivots=["Data!SalesPivot"])
    wb.validate()


def test_already_enabled_refresh_still_requires_explicit_consent(
        fixture_copy, tmp_path):
    wb = load_workbook(
        fixture_copy("minimal/minimal_clean.xlsx"), preserve=True)
    _with_pivot_graph(
        wb, [("SalesPivot", "1")], refresh_on_load=True,
        worksheet_sources={
            "1": b'<worksheetSource ref="A1:B10" sheet="Sheet1"/>'})
    wb["Sheet1"]["A1"] = "changed"

    with pytest.raises(UnsupportedStructureError,
                       match="stale pivot results"):
        wb.validate()

    wb.set_pivot_refresh_on_load(pivots=["Sheet1!SalesPivot"])
    receipt = wb.save(tmp_path / "already-enabled-source-edit.xlsx",
                      receipt=True)
    effects = [effect for effect in receipt.derived_effects
               if effect["kind"] ==
               "pivot_source_changed_requires_refresh"]
    assert len(effects) == 1
    assert not any(effect["kind"] == "pivot_refresh_on_load_enabled"
                   for effect in receipt.derived_effects)


def test_explicit_refresh_reenables_disabled_pivot_cache(
        fixture_copy, tmp_path):
    wb = load_workbook(
        fixture_copy("minimal/minimal_clean.xlsx"), preserve=True)
    _with_pivot_graph(
        wb, [("SalesPivot", "1")], refresh_on_load=True,
        worksheet_sources={
            "1": b'<worksheetSource ref="A1:B10" sheet="Sheet1"/>'},
        cache_attributes={"1": b' enableRefresh="0"'})
    wb["Sheet1"]["A1"] = "changed"
    wb.set_pivot_refresh_on_load(pivots=["Sheet1!SalesPivot"])

    output = tmp_path / "refresh-reenabled.xlsx"
    receipt = wb.save(output, receipt=True)

    with zipfile.ZipFile(output) as archive:
        payload = archive.read(
            "xl/pivotCache/pivotCacheDefinition1.xml")
    assert b'enableRefresh="1"' in payload
    assert any(effect["kind"] == "pivot_refresh_on_load_enabled"
               for effect in receipt.derived_effects)


def test_highly_compressed_valid_extra_part_is_not_an_eligibility_error(
        fixture_copy, tmp_path):
    src = fixture_copy("minimal/minimal_clean.xlsx")
    crafted = tmp_path / "compressed.xlsx"
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(
            crafted, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            zout.writestr(info.filename, zin.read(info.filename))
        zout.writestr("xl/media/blob.bin", b"\x00" * (2 << 20))
    wb = load_workbook(crafted, preserve=True)
    assert wb.active is not None


def test_zip_confusion_is_normalized(fixture_copy, tmp_path):
    src = fixture_copy("features/schedule.xlsx")
    with open(src, "rb") as handle:
        data = bytearray(handle.read())
    with zipfile.ZipFile(io.BytesIO(bytes(data))) as archive:
        info = archive.getinfo("docProps/core.xml")
    data[info.header_offset + 14] ^= 0xFF
    crafted = tmp_path / "confused.xlsx"
    with open(crafted, "wb") as handle:
        handle.write(data)
    wb = load_workbook(crafted, preserve=True)
    wb["Schedule"]["A2"] = "edit"
    out = tmp_path / "o.xlsx"
    wb.save(out)
    with zipfile.ZipFile(out) as archive:
        assert b"cp:coreProperties" in archive.read("docProps/core.xml")


def test_spooled_save_is_correct_zip(fixture_copy, tmp_path):
    src = fixture_copy("features/schedule.xlsx")
    wb = load_workbook(src, preserve=True)
    wb["Schedule"]["A2"] = "spooled"
    out = tmp_path / "o.xlsx"
    wb.save(out)
    with zipfile.ZipFile(out) as archive:
        assert archive.testzip() is None
    assert load_workbook(out)["Schedule"]["A2"].value == "spooled"


def test_validate_runs_the_save_plan_without_building_an_archive(
        fixture_copy, monkeypatch):
    from openpyxl.preserve import zipio

    workbook = load_workbook(
        fixture_copy("features/schedule.xlsx"), preserve=True)
    workbook["Schedule"]["A2"] = "edited"

    def unexpected(*args, **kwargs):
        raise AssertionError("validation assembled an archive")

    monkeypatch.setattr(zipio, "build_archive_bytes", unexpected)
    monkeypatch.setattr(zipio, "build_and_deliver", unexpected)
    assert workbook.validate() is None


def test_receipt_reports_deterministic_derived_effects(
        fixture_copy, tmp_path):
    workbook = load_workbook(
        fixture_copy("features/schedule_calc.xlsx"), preserve=True)
    workbook["Schedule"]["B2"] = 123
    receipt = workbook.save(tmp_path / "receipt.xlsx", receipt=True)
    payload = receipt.to_dict()

    assert payload["derived_effects_version"] == 1
    kinds = [effect["kind"] for effect in payload["derived_effects"]]
    assert "formula_cache_removed" in kinds
    assert "recalculation_metadata_changed" in kinds
    assert json.loads(json.dumps(payload))["derived_effects"] == \
        payload["derived_effects"]


def test_receipt_does_not_report_unchanged_calc_properties(
        fixture_copy, tmp_path):
    workbook = load_workbook(
        fixture_copy("minimal/minimal_clean.xlsx"), preserve=True)
    workbook.create_sheet("Added")

    receipt = workbook.save(tmp_path / "added-sheet.xlsx", receipt=True)

    assert not any(
        effect["kind"] == "recalculation_metadata_changed"
        for effect in receipt.derived_effects
    )
