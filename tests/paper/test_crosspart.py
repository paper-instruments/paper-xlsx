"""Phase 2d: cross-part edits — new sheets, styles append, workbook.xml
splice, calcChain cascade, hyperlink relationships (PR-0 D2/D9/D11-D13)."""
from __future__ import annotations

import zipfile

import pytest

from openpyxl import load_workbook
from openpyxl.errors import PaperRefusal, UnsupportedStructureError
from openpyxl.package import diff_package
from openpyxl.styles import Font, PatternFill

from .support.harness import assert_part_budget
from .support.partdiff import part_payloads

GAUNTLET = "gauntlet/gauntlet.xlsx"


def _model_sheet_name(path):
    for name, payload in part_payloads(path).items():
        if name.startswith("xl/worksheets/") and b"Quarterly Model" in payload:
            return name
    raise AssertionError("no Model sheet")


class TestStylesAppend:

    def test_new_font_appends_to_styles(self, fixture_copy, tmp_path):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb["Model"]["A2"].font = Font(name="Menlo", size=7)
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        assert_part_budget(src, out, expect_changed={
            _model_sheet_name(src), "xl/styles.xml"})
        wb2 = load_workbook(out)
        font = wb2["Model"]["A2"].font
        assert (font.name, font.size) == ("Menlo", 7)
        # append-only: every original font element still present, in order,
        # before the appended one (only the fonts count attribute changed)
        before = part_payloads(src)["xl/styles.xml"]
        after = part_payloads(out)["xl/styles.xml"]
        fonts_open_end = before.index(b">", before.index(b"<fonts")) + 1
        fonts_before = before[fonts_open_end:before.index(b"</fonts>")]
        assert fonts_before in after
        assert after.index(fonts_before) < after.index(b"Menlo")

    def test_count_attributes_bumped(self, fixture_copy, tmp_path):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb["Model"]["A2"].font = Font(name="Menlo", size=7)
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        import re
        before = part_payloads(src)["xl/styles.xml"]
        after = part_payloads(out)["xl/styles.xml"]

        def count_of(payload, tag):
            m = re.search(br'<%s count="(\d+)"' % tag, payload)
            return int(m.group(1)) if m else None

        assert count_of(after, b"fonts") == count_of(before, b"fonts") + 1
        assert count_of(after, b"cellXfs") == count_of(before, b"cellXfs") + 1

    def test_custom_number_format_appends(self, fixture_copy, tmp_path):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb["Model"]["B3"].number_format = '#,##0.000 "kg"'
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert wb2["Model"]["B3"].number_format == '#,##0.000 "kg"'

    def test_new_named_style_refuses(self, fixture_copy, tmp_path):
        from openpyxl.styles import NamedStyle

        src = fixture_copy(GAUNTLET)
        with open(src, "rb") as f:
            before = f.read()
        wb = load_workbook(src, preserve=True)
        ns = NamedStyle(name="brand_new")
        ns.font = Font(bold=True)
        wb.add_named_style(ns)
        wb["Model"]["A2"].style = "brand_new"
        with pytest.raises(UnsupportedStructureError, match="named styles"):
            wb.save(str(tmp_path / "o.xlsx"))
        with open(src, "rb") as f:
            assert f.read() == before

    @pytest.mark.lo_smoke
    def test_styles_append_loads_in_libreoffice(self, fixture_copy, tmp_path, lo):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb["Model"]["A2"].font = Font(name="Menlo", size=7)
        wb["Model"]["A3"].fill = PatternFill("solid", fgColor="FF00AA00")
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        assert lo.lo_loads(out)


class TestAddedSheets:

    def test_added_sheet_composes_with_preserved_package(
            self, fixture_copy, tmp_path):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        ws = wb.create_sheet("Appended")
        ws["A1"] = "header"
        ws["B2"] = 42
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        d = diff_package(src, out)
        assert set(c.part for c in d.changed) == {
            "[Content_Types].xml", "xl/_rels/workbook.xml.rels",
            "xl/workbook.xml"}
        assert len(d.added) == 1 and d.added[0].startswith("xl/worksheets/")
        assert not d.removed
        wb2 = load_workbook(out)
        assert wb2["Appended"]["B2"].value == 42
        assert wb2.sheetnames[-1] == "Appended"
        # traps on the untouched sheet survive
        parts = part_payloads(out)
        assert b"sparklineGroups" in parts[_model_sheet_name(src)]

    def test_added_sheet_with_hyperlink_gets_a_rels_part(
            self, fixture_copy, tmp_path):
        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb.create_sheet("Links")
        ws["A1"].hyperlink = "https://example.org/doc"
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        d = diff_package(src, out)
        rels_parts = [n for n in d.added if "_rels" in n]
        assert len(rels_parts) == 1
        wb2 = load_workbook(out)
        assert wb2["Links"]["A1"].hyperlink.target == "https://example.org/doc"

    def test_added_sheet_with_chart_refuses(self, fixture_copy, tmp_path):
        from openpyxl.chart import BarChart, Reference

        src = fixture_copy(GAUNTLET)
        with open(src, "rb") as f:
            before = f.read()
        wb = load_workbook(src, preserve=True)
        ws = wb.create_sheet("Charted")
        ws.append([1, 2, 3])
        chart = BarChart()
        chart.add_data(Reference(ws, min_col=1, min_row=1, max_col=3, max_row=1))
        ws.add_chart(chart, "E1")
        with pytest.raises(UnsupportedStructureError, match="charts or images"):
            wb.save(str(tmp_path / "o.xlsx"))
        with open(src, "rb") as f:
            assert f.read() == before

    def test_pandas_append_full_flow(self, fixture_copy):
        pd = pytest.importorskip("pandas")
        src = fixture_copy(GAUNTLET)
        df = pd.DataFrame({"a": [1, 2], "b": [3.5, 4.5]})
        with pd.ExcelWriter(src, engine="openpyxl", mode="a",
                            engine_kwargs={"preserve": True}) as xw:
            df.to_excel(xw, sheet_name="Data2", index=False)
        wb = load_workbook(src)
        assert wb["Data2"]["B3"].value == 4.5
        parts = part_payloads(src)
        sheet = next(p for n, p in parts.items()
                     if n.startswith("xl/worksheets/") and b"Quarterly Model" in p)
        assert b"sparklineGroups" in sheet
        assert b"x14:conditionalFormattings" in sheet

    @pytest.mark.lo_smoke
    def test_added_sheet_loads_in_libreoffice(self, fixture_copy, tmp_path, lo):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        ws = wb.create_sheet("Fresh")
        ws["A1"] = "check"
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        assert lo.lo_loads(out)


class TestWorkbookXmlSplice:

    def test_defined_name_add(self, fixture_copy, tmp_path):
        from openpyxl.workbook.defined_name import DefinedName

        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb.defined_names["Fresh"] = DefinedName("Fresh", attr_text="Model!$A$1")
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        assert_part_budget(src, out, expect_changed={"xl/workbook.xml"})
        wb2 = load_workbook(out)
        assert wb2.defined_names["Fresh"].value == "Model!$A$1"
        assert wb2.defined_names["GrowthRate"].value == "Model!$B$8"

    def test_print_area_serializes_into_defined_names(self, fixture_copy, tmp_path):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb["Model"].print_area = "A1:F12"
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        # quote_sheetname always quotes and print areas are absolute
        assert wb2["Model"].print_area == "'Model'!$A$1:$F$12"

    def test_calc_properties_change(self, fixture_copy, tmp_path):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb.calculation.fullCalcOnLoad = True
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        parts = part_payloads(out)
        assert b'fullCalcOnLoad="1"' in parts["xl/workbook.xml"]

    def test_sheet_state_toggle(self, fixture_copy, tmp_path):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb["HiddenNotes"].sheet_state = "visible"
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        assert_part_budget(src, out, expect_changed={"xl/workbook.xml"})
        wb2 = load_workbook(out)
        assert wb2["HiddenNotes"].sheet_state == "visible"

    def test_active_sheet_change(self, fixture_copy, tmp_path):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb.active = wb["Data"]
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert wb2.active.title == "Data"


class TestCalcChainCascade:

    @pytest.fixture
    def calcchain_file(self, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule_calc.xlsx")
        out = str(tmp_path / "with_calcchain.xlsx")
        with zipfile.ZipFile(src) as zin, zipfile.ZipFile(out, "w") as zout:
            for name in zin.namelist():
                payload = zin.read(name)
                if name == "[Content_Types].xml":
                    payload = payload.replace(
                        b"</Types>",
                        b'<Override PartName="/xl/calcChain.xml" ContentType='
                        b'"application/vnd.openxmlformats-officedocument.'
                        b'spreadsheetml.calcChain+xml"/></Types>')
                if name == "xl/_rels/workbook.xml.rels":
                    payload = payload.replace(
                        b"</Relationships>",
                        b'<Relationship Id="rId99" Type="http://schemas.'
                        b'openxmlformats.org/officeDocument/2006/relationships/'
                        b'calcChain" Target="calcChain.xml"/></Relationships>')
                zout.writestr(name, payload)
            zout.writestr(
                "xl/calcChain.xml",
                b'<?xml version="1.0"?><calcChain xmlns="http://schemas.'
                b'openxmlformats.org/spreadsheetml/2006/main">'
                b'<c r="B12" i="1"/></calcChain>')
        return out

    def test_formula_edit_cascades_calcchain_removal(self, calcchain_file, tmp_path):
        wb = load_workbook(calcchain_file, preserve=True)
        wb["Schedule"]["B20"] = "=SUM(B2:B3)"
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        parts = part_payloads(out)
        assert "xl/calcChain.xml" not in parts                       # part gone
        assert b"calcChain" not in parts["[Content_Types].xml"]      # override gone
        assert b"calcChain" not in parts["xl/_rels/workbook.xml.rels"]  # rel gone
        wb2 = load_workbook(out)
        assert wb2["Schedule"]["B20"].value == "=SUM(B2:B3)"

    def test_value_only_edit_keeps_calcchain(self, calcchain_file, tmp_path):
        wb = load_workbook(calcchain_file, preserve=True)
        wb["Schedule"]["A1"] = "renamed header"      # no formula involved
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        assert "xl/calcChain.xml" in part_payloads(out)


class TestHyperlinks:

    def test_hyperlink_add_touches_sheet_and_its_rels(self, fixture_copy, tmp_path):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb["Data"]["B2"].hyperlink = "https://example.org/x"
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        d = diff_package(src, out)
        changed = set(c.part for c in d.changed)
        assert any(n.endswith(".rels") for n in changed)
        assert any(n.startswith("xl/worksheets/sheet") for n in changed)
        wb2 = load_workbook(out)
        assert wb2["Data"]["B2"].hyperlink.target == "https://example.org/x"

    def test_existing_hyperlink_and_rels_survive_additions(
            self, fixture_copy, tmp_path):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb["Model"]["A13"].hyperlink = "https://example.org/new"
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        # the fixture's original hyperlink still resolves
        assert wb2["Model"]["A11"].hyperlink.target == \
            "https://example.com/model-docs"
        assert wb2["Model"]["A13"].hyperlink.target == "https://example.org/new"

    def test_hyperlink_removal_refuses(self, fixture_copy, tmp_path):
        src = fixture_copy(GAUNTLET)
        with open(src, "rb") as f:
            before = f.read()
        wb = load_workbook(src, preserve=True)
        wb["Model"]["A11"].hyperlink = None
        with pytest.raises(UnsupportedStructureError, match="ADDITION"):
            wb.save(str(tmp_path / "o.xlsx"))
        with open(src, "rb") as f:
            assert f.read() == before


class TestConditionalFormattingLift:

    def test_cf_add_on_sheet_without_x14(self, fixture_copy, tmp_path):
        from openpyxl.formatting.rule import CellIsRule

        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Sheet1"].conditional_formatting.add(
            "D2:D5", CellIsRule(operator="greaterThan", formula=["5"],
                                fill=PatternFill("solid", fgColor="FFFFAA00")))
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        cfs = list(wb2["Sheet1"].conditional_formatting)
        assert len(cfs) == 1
        rule = cfs[0].rules[0]
        assert rule.operator == "greaterThan"
        # the dxf allocated for the fill landed in styles.xml
        assert rule.dxfId is not None
        assert b"<dxfs" in part_payloads(out)["xl/styles.xml"] or \
            b"<dxf>" in part_payloads(out)["xl/styles.xml"]

    @pytest.mark.lo_smoke
    def test_cf_add_loads_in_libreoffice(self, fixture_copy, tmp_path, lo):
        from openpyxl.formatting.rule import CellIsRule

        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Sheet1"].conditional_formatting.add(
            "D2:D5", CellIsRule(operator="greaterThan", formula=["5"],
                                fill=PatternFill("solid", fgColor="FFFFAA00")))
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        assert lo.lo_loads(out)


class TestV0CrosspartRefusals:

    def test_table_add_refuses(self, fixture_copy, tmp_path):
        from openpyxl.worksheet.table import Table

        src = fixture_copy("features/tables.xlsx")
        with open(src, "rb") as f:
            before = f.read()
        wb = load_workbook(src, preserve=True)
        ws = wb["Data"]
        ws.add_table(Table(displayName="T2", ref="D1:E3"))
        with pytest.raises(PaperRefusal, match="table"):
            wb.save(str(tmp_path / "o.xlsx"))
        with open(src, "rb") as f:
            assert f.read() == before

    def test_mark_dirty_part_refuses(self, fixture_copy, tmp_path):
        src = fixture_copy(GAUNTLET)
        wb = load_workbook(src, preserve=True)
        wb.mark_dirty("xl/media/image1.png")
        with pytest.raises(UnsupportedStructureError, match="mark_dirty"):
            wb.save(str(tmp_path / "o.xlsx"))
