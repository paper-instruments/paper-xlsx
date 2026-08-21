"""PR 7: managed update, rename, delete, and session composition."""
from __future__ import annotations

import pytest

from openpyxl import load_workbook
from openpyxl.errors import (
    AmbiguousTargetError,
    TargetNotFoundError,
    UnsupportedStructureError,
)

from .support.harness import save_and_reopen
from .support.partdiff import part_payloads
from .test_pivot_create_package import _create_by_region
from .test_pivot_graph import _basic_package, _write_package


_TABLE = "features/tables.xlsx"


def test_update_values_and_layout(fixture_copy, tmp_path):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    handle = _create_by_region(wb["Data"])
    updated = handle.update(layout="outline", values=["Amount"])
    assert updated.spec.layout == "outline"
    assert len(updated.spec.values) == 1
    wb = save_and_reopen(wb, str(tmp_path / "updated.xlsx"), preserve=True)
    reopened = wb["Data"].pivots["ByRegion"]
    assert reopened.spec.layout == "outline"
    assert [item.field for item in reopened.spec.values] == ["Amount"]


def test_rename_and_reopen(fixture_copy, tmp_path):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    handle = _create_by_region(wb["Data"])
    renamed = handle.rename("RegionalSales")
    assert renamed.name == "RegionalSales"
    names = [pivot.name for pivot in wb["Data"].pivots]
    assert names == ["RegionalSales"]
    with pytest.raises(TargetNotFoundError) as stale:
        handle.name
    assert stale.value.kind == "stale-pivot-handle"
    wb = save_and_reopen(wb, str(tmp_path / "renamed.xlsx"), preserve=True)
    names = [pivot.name for pivot in wb["Data"].pivots]
    assert names == ["RegionalSales"]
    assert wb["Data"].pivots["RegionalSales"].origin == "paper"


def test_rename_uniqueness_refuses(fixture_copy):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    _create_by_region(wb["Data"])
    _create_by_region(wb["Data"], name="Other", destination="H3")
    with pytest.raises(AmbiguousTargetError) as exc:
        wb["Data"].pivots["ByRegion"].rename("other")
    assert exc.value.kind == "ambiguous-pivot"
    assert wb["Data"].pivots["ByRegion"].name == "ByRegion"
    assert wb["Data"].pivots["Other"].name == "Other"


def test_create_then_delete_is_noop(fixture_copy, tmp_path):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    handle = _create_by_region(wb["Data"])
    handle.delete()
    out = str(tmp_path / "noop.xlsx")
    wb.save(out)
    assert part_payloads(src) == part_payloads(out)


def test_receipt_consolidates_lifecycle(fixture_copy, tmp_path):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    handle = _create_by_region(wb["Data"])
    handle.update(layout="compact")
    created = wb.save(str(tmp_path / "created.xlsx"), receipt=True)
    created_kinds = [
        item["kind"] for item in created.derived_effects
        if str(item.get("kind", "")).startswith("pivot_")
    ]
    assert created_kinds == ["pivot_created"]

    wb = load_workbook(src, preserve=True)
    handle = _create_by_region(wb["Data"])
    handle.delete()
    cancelled = wb.save(str(tmp_path / "cancelled.xlsx"), receipt=True)
    cancelled_kinds = [
        item["kind"] for item in cancelled.derived_effects
        if str(item.get("kind", "")).startswith("pivot_")
    ]
    assert cancelled_kinds == []

    wb = load_workbook(src, preserve=True)
    _create_by_region(wb["Data"])
    wb = save_and_reopen(wb, str(tmp_path / "reopened.xlsx"), preserve=True)
    wb["Data"].pivots["ByRegion"].update(layout="outline")
    wb["Data"].pivots["ByRegion"].delete()
    deleted = wb.save(str(tmp_path / "deleted.xlsx"), receipt=True)
    deleted_kinds = [
        item["kind"] for item in deleted.derived_effects
        if str(item.get("kind", "")).startswith("pivot_")
    ]
    assert deleted_kinds == ["pivot_deleted"]


def test_delete_reopened_dedicated_pivot(fixture_copy, tmp_path):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    _create_by_region(wb["Data"])
    created = str(tmp_path / "created.xlsx")
    wb = save_and_reopen(wb, created, preserve=True)
    handle = wb["Data"].pivots["ByRegion"]
    handle.delete()
    assert list(wb["Data"].pivots) == []
    with pytest.raises(TargetNotFoundError) as stale:
        handle.refresh()
    assert stale.value.kind == "stale-pivot-handle"
    wb = save_and_reopen(wb, str(tmp_path / "deleted.xlsx"), preserve=True)
    assert list(wb["Data"].pivots) == []
    assert wb["Data"]["E3"].value is None


def test_stale_handle_after_delete(fixture_copy):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    handle = _create_by_region(wb["Data"])
    handle.delete()
    with pytest.raises(TargetNotFoundError) as exc:
        handle.refresh()
    assert exc.value.kind == "stale-pivot-handle"


def test_foreign_mutators_refuse(tmp_path):
    src = _write_package(tmp_path, "foreign.xlsx", _basic_package())
    wb = load_workbook(src, preserve=True)
    pivot = wb["Summary"].pivots["SalesByRegion"]
    with pytest.raises(UnsupportedStructureError) as exc:
        pivot.update(layout="compact")
    assert exc.value.kind == "unsupported-pivot-operation"
    with pytest.raises(UnsupportedStructureError):
        pivot.rename("Other")
    with pytest.raises(UnsupportedStructureError):
        pivot.delete()


def test_create_update_save_one_graph(fixture_copy, tmp_path):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    handle = _create_by_region(wb["Data"])
    handle.update(layout="compact")
    wb = save_and_reopen(wb, str(tmp_path / "composed.xlsx"), preserve=True)
    assert wb["Data"].pivots["ByRegion"].spec.layout == "compact"
    assert len(list(wb["Data"].pivots)) == 1
