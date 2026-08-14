from __future__ import annotations

import zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.formula import ArrayFormula


def test_array_formula_text_participates_in_dependencies():
    wb = Workbook()
    ws = wb.active
    ws["B1"] = 5
    ws["A1"] = ArrayFormula(ref="A1:A2", text="=B1*{1;2}")

    from openpyxl.preserve.perception import dependency_sketch

    sketch = dependency_sketch(wb)
    assert sketch.references["'Sheet'!A1"][0][1] == (2, 1, 2, 1)


def test_preserve_save_retains_macro_package_content_type(fixture_copy,
                                                          tmp_path):
    source = fixture_copy("features/macro_stub.xlsm")
    wb = load_workbook(source, preserve=True)
    from openpyxl.xml.constants import XLSM

    assert wb.mime_type == XLSM
    wb.active["A1"] = "edited"
    output = tmp_path / "retained.xlsm"
    wb.save(output)
    with zipfile.ZipFile(output) as archive:
        content_types = archive.read("[Content_Types].xml")
        assert b"macroEnabled.main+xml" in content_types
        assert "xl/vbaProject.bin" in archive.namelist()
