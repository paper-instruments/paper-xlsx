"""Preserve-mode load — byte retention and flag semantics."""
from __future__ import annotations

import io
import os
import warnings

import pytest

from openpyxl import Workbook, load_workbook
from openpyxl.errors import PaperRefusal, UnsupportedStructureError
from openpyxl.reader.excel import _preserve_by_default
from openpyxl.utils.exceptions import InvalidFileException


def test_file_like_default_sniffs_ooxml_and_restores_position(
        fixture_copy):
    with open(fixture_copy("minimal/minimal_clean.xlsx"), "rb") as handle:
        stream = io.BytesIO(handle.read())
    stream.name = "misleading.xlsb"
    stream.seek(37)
    assert _preserve_by_default(stream, False) is True
    assert stream.tell() == 37
    assert not stream.closed

    non_ooxml = io.BytesIO(b"not a zip")
    non_ooxml.name = "misleading.xlsx"
    non_ooxml.seek(3)
    assert _preserve_by_default(non_ooxml, False) is False
    assert non_ooxml.tell() == 3


class TestPreserveLoad:

    def test_blob_retained_equals_source_bytes(self, fixture_copy):
        src = fixture_copy("gauntlet/gauntlet.xlsx")
        wb = load_workbook(src, preserve=True)
        assert wb.preserve is True
        with open(src, "rb") as f:
            assert wb._paper_source == f.read()

    def test_default_load_retains_source(self, fixture_copy):
        wb = load_workbook(fixture_copy("minimal/minimal_clean.xlsx"))
        assert wb.preserve is True
        assert wb._paper_source is not None

    @pytest.mark.parametrize("suffix", [b".xls", b".xlsb", b".csv"])
    def test_unsupported_bytes_path_defaults_to_stock(self, suffix):
        assert _preserve_by_default(b"/tmp/workbook" + suffix, False) is False

    def test_supported_bytes_path_defaults_to_preserve(self):
        assert _preserve_by_default(b"/tmp/workbook.xlsx", False) is True

    def test_extensionless_file_like_defaults_to_preserve(
            self, fixture_copy, tmp_path):
        src = fixture_copy("minimal/minimal_clean.xlsx")
        upload = tmp_path / "upload"
        with open(src, "rb") as source:
            data = source.read()
        upload.write_bytes(data)

        with upload.open("rb") as stream:
            wb = load_workbook(stream)

        assert wb.preserve is True
        assert wb._paper_source == data

    def test_explicit_stock_load_retains_nothing(self, fixture_copy):
        wb = load_workbook(
            fixture_copy("minimal/minimal_clean.xlsx"), preserve=False)
        assert wb.preserve is False
        assert wb._paper_source is None

    def test_fresh_workbook_is_not_preserve_mode(self):
        wb = Workbook()
        assert wb.preserve is False

    def test_file_like_source_is_read_eagerly(self, fixture_copy):
        src = fixture_copy("features/schedule.xlsx")
        with open(src, "rb") as f:
            data = f.read()
        fh = io.BytesIO(data)
        fh.seek(100)  # arbitrary position — retention must rewind
        wb = load_workbook(fh, preserve=True)
        assert wb._paper_source == data
        # the source handle being overwritten afterwards must not matter
        fh.seek(0)
        fh.write(b"\x00" * len(data))
        assert wb._paper_source == data

    def test_preserve_plus_read_only_refuses_with_value_error(self, fixture_copy):
        src = fixture_copy("minimal/minimal_clean.xlsx")
        with pytest.raises(ValueError, match="preserve=True cannot be combined"):
            load_workbook(src, read_only=True, preserve=True)

    def test_default_read_only_load_uses_stock_mode(self, fixture_copy):
        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, read_only=True)
        assert wb.preserve is False
        assert wb._paper_source is None
        wb.close()

    def test_preserve_keeps_extension_refusals(self, fixture_copy):
        # the .xls/.xlsb check is not bypassed by retention; under preserve
        # it is the typed refusal, on the stock path unchanged
        src = fixture_copy("legacy/legacy.xls")
        with pytest.raises(UnsupportedStructureError, match="xls"):
            load_workbook(src, preserve=True)
        with pytest.raises(InvalidFileException, match="xls"):
            load_workbook(src)

    def test_preserve_with_data_only_loads(self, fixture_copy):
        wb = load_workbook(fixture_copy("features/schedule_calc.xlsx"),
                           preserve=True, data_only=True)
        assert wb["Schedule"]["B12"].value == 6500

    def test_extension_warning_suppressed_under_preserve(self, fixture_copy):
        src = fixture_copy("gauntlet/gauntlet.xlsx")
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            load_workbook(src, preserve=True)
        assert not [w for w in caught
                    if "extension is not supported" in str(w.message)]
        # ...and still fires on the stock path (stock save WILL remove them)
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            load_workbook(src, preserve=False)
        assert [w for w in caught
                if "extension is not supported" in str(w.message)]


class TestStockCompatibility:

    def test_rich_file_stock_save_emits_no_paper_warning(
            self, fixture_copy, tmp_path):
        src = fixture_copy("gauntlet/gauntlet.xlsx")
        wb = load_workbook(src, preserve=False)
        out = str(tmp_path / "out.xlsx")
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            wb.save(out)
        assert not [w for w in caught
                    if w.message.__class__.__module__ == "openpyxl.errors"]

    def test_vba_stock_save_emits_no_paper_warning(
            self, fixture_copy, tmp_path):
        src = fixture_copy("features/macro_stub.xlsm")
        wb = load_workbook(src, preserve=False)
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            wb.save(str(tmp_path / "o1.xlsm"))
        assert not [w for w in caught
                    if w.message.__class__.__module__ == "openpyxl.errors"]

        wb2 = load_workbook(src, keep_vba=True, preserve=False)
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            wb2.save(str(tmp_path / "o2.xlsm"))
        assert not [w for w in caught
                    if w.message.__class__.__module__ == "openpyxl.errors"]

    def test_clean_file_stock_save_stays_silent(self, fixture_copy, tmp_path):
        wb = load_workbook(
            fixture_copy("minimal/minimal_clean.xlsx"), preserve=False)
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            wb.save(str(tmp_path / "out.xlsx"))
        assert not [w for w in caught
                    if w.message.__class__.__module__ == "openpyxl.errors"]

    def test_fresh_workbook_save_stays_silent(self, tmp_path):
        wb = Workbook()
        wb.active["A1"] = 1
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            wb.save(str(tmp_path / "out.xlsx"))
        assert not [w for w in caught
                    if w.message.__class__.__module__ == "openpyxl.errors"]
