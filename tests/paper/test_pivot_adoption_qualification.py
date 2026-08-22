"""PR 9: read-only foreign PivotTable adoption qualification."""
from __future__ import annotations

import io
import zipfile

import pytest

from openpyxl import load_workbook
from openpyxl.pivot import PivotAdoptionQualification, PivotTable
from openpyxl.pivot.api_types import ADOPTION_TO_DICT_SCHEMA, ADOPTION_TO_DICT_VERSION
from openpyxl.pivot.qualify import PAPER_TAG

from .support.harness import save_and_reopen
from .test_pivot_create_package import _create_by_region
from .test_pivot_graph import _basic_package, _write_package


_TABLE = "features/tables.xlsx"


def _codes(qualification, capability=None):
    return [
        item.code for item in qualification.reasons
        if capability is None or item.capability == capability
    ]


def _load_payload(tmp_path, payload, name="foreign.xlsx"):
    return load_workbook(_write_package(tmp_path, name, payload), preserve=True)


def _rewrite_zip(path, transform):
    with open(path, "rb") as handle:
        payload = handle.read()
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(payload)) as before, \
            zipfile.ZipFile(output, "w") as after:
        for info in before.infolist():
            body = transform(info.filename, before.read(info.filename))
            after.writestr(info, body)
    with open(path, "wb") as handle:
        handle.write(output.getvalue())
    return path


def _strip_paper_tag(path):
    marker = (' tag="%s"' % PAPER_TAG).encode("ascii")

    def transform(_name, body):
        return body.replace(marker, b"")

    return _rewrite_zip(path, transform)


def _paper_then_foreign(fixture_copy, tmp_path, name="ByRegion"):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    _create_by_region(wb["Data"], name=name)
    dest = src + ".adopt-qual.xlsx"
    save_and_reopen(wb, dest, preserve=True)
    _strip_paper_tag(dest)
    return load_workbook(dest, preserve=True), dest


def test_adoption_qualification_export_and_schema(tmp_path):
    wb = _load_payload(tmp_path, _basic_package())
    result = wb["Summary"].pivots["SalesByRegion"].qualify_adoption()
    assert isinstance(result, PivotAdoptionQualification)
    payload = result.to_dict()
    assert list(payload)[:2] == ["schema", "version"]
    assert payload["schema"] == ADOPTION_TO_DICT_SCHEMA
    assert payload["version"] == ADOPTION_TO_DICT_VERSION
    assert payload["eligible"] is False
    assert "adopt" not in PivotTable.__dict__ or not callable(
        getattr(PivotTable, "adopt", None))


def test_managed_pivot_is_already_managed(fixture_copy):
    src = fixture_copy(_TABLE)
    wb = load_workbook(src, preserve=True)
    handle = _create_by_region(wb["Data"])
    dest = src + ".managed.xlsx"
    reopened = save_and_reopen(wb, dest, preserve=True)
    pivot = reopened["Data"].pivots["ByRegion"]
    assert pivot.origin == "paper"
    result = pivot.qualify_adoption()
    assert result.eligible is False
    assert result.strategy is None
    assert "already-managed" in _codes(result)
    assert handle.origin == "paper"


def test_complete_dedicated_is_gated_until_excel_evidence(
        fixture_copy, tmp_path, monkeypatch):
    wb, path = _paper_then_foreign(fixture_copy, tmp_path)
    before = open(path, "rb").read()
    pivot = wb["Data"].pivots["ByRegion"]
    assert pivot.origin == "foreign"
    result = pivot.qualify_adoption()
    assert result.strategy == "dedicated-replacement"
    assert result.eligible is False
    assert "foreign-managed-equivalence-unproved" in _codes(result)
    after = open(path, "rb").read()
    assert after == before

    monkeypatch.setattr(
        "openpyxl.pivot.adopt_qualify.excel_equivalence_proved",
        lambda: True,
    )
    monkeypatch.setattr(
        "openpyxl.pivot.adopt_inventory.excel_equivalence_proved",
        lambda: True,
    )
    qualified = wb["Data"].pivots["ByRegion"].qualify_adoption()
    assert qualified.strategy == "dedicated-replacement"
    assert "foreign-managed-equivalence-unproved" not in _codes(qualified)
    assert qualified.eligible is True
    assert qualified.requires_calculation is False
    assert qualified.calculation_engine is None
    assert open(path, "rb").read() == before


def test_shared_cache_selects_isolation_strategy(tmp_path):
    payload = _basic_package(
        extra_pivots=(
            ("MarginByRegion", "xl/pivotTables/pivotTable2.xml",
             "rIdPivot2", "1"),
        ),
    )
    wb = _load_payload(tmp_path, payload)
    result = wb["Summary"].pivots["SalesByRegion"].qualify_adoption()
    assert result.strategy == "shared-isolation"
    assert result.eligible is False


def test_defined_name_source_is_not_adoptable(tmp_path):
    payload = _basic_package(
        source={"kind": "defined-name", "name": "SalesRange"},
        defined_name="SalesRange",
    )
    wb = _load_payload(tmp_path, payload)
    pivot = wb["Summary"].pivots["SalesByRegion"]
    descriptor = pivot.source
    result = pivot.qualify_adoption()
    assert result.eligible is False
    if descriptor is not None and descriptor.kind == "defined-name":
        assert "unsupported-pivot-source" in _codes(result)
    else:
        # Synthetic named sources may project as a range; they still refuse.
        assert result.eligible is False
        assert "foreign-managed-equivalence-unproved" in _codes(result)


def test_missing_records_and_grouping_refuse(tmp_path):
    missing = _load_payload(
        tmp_path, _basic_package(include_records=False), name="norec.xlsx")
    assert "foreign-cache-records-unavailable" in _codes(
        missing["Summary"].pivots["SalesByRegion"].qualify_adoption())
    grouped = _load_payload(
        tmp_path, _basic_package(grouping=True), name="group.xlsx")
    assert "unsupported-grouping" in _codes(
        grouped["Summary"].pivots["SalesByRegion"].qualify_adoption())
    calculated = _load_payload(
        tmp_path, _basic_package(calculated=True), name="calc.xlsx")
    assert "unsupported-calculated" in _codes(
        calculated["Summary"].pivots["SalesByRegion"].qualify_adoption())


def test_unknown_extension_and_unproved_known_extension(tmp_path, monkeypatch):
    unknown = _load_payload(
        tmp_path, _basic_package(ext_uri="{DEADBEEF-0000}"), name="ext.xlsx")
    assert "unsupported-extension" in _codes(
        unknown["Summary"].pivots["SalesByRegion"].qualify_adoption())

    known = _load_payload(
        tmp_path,
        _basic_package(ext_uri="{747A6164-185A-40DC-8AA5-F01512510D54}"),
        name="known-ext.xlsx",
    )
    result = known["Summary"].pivots["SalesByRegion"].qualify_adoption()
    assert result.eligible is False
    assert any(code in _codes(result) for code in (
        "foreign-extension-unproved", "unsupported-extension",
        "foreign-managed-equivalence-unproved",
    ))


def test_xlsm_vba_and_nondefault_core_schema(tmp_path):
    payload = _basic_package()
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(payload)) as before, \
            zipfile.ZipFile(output, "w") as after:
        for info in before.infolist():
            after.writestr(info, before.read(info.filename))
        after.writestr("xl/vbaProject.bin", b"VBA")
    wb = _load_payload(tmp_path, output.getvalue(), name="macro.xlsx")
    result = wb["Summary"].pivots["SalesByRegion"].qualify_adoption()
    assert result.eligible is False
    assert "foreign-vba-dependency-unproved" in _codes(result)

    mutated = _basic_package()
    mutated = _rewrite_bytes(
        mutated, "xl/pivotTables/pivotTable1.xml",
        lambda body: body.replace(
            b"<pivotTableDefinition",
            b'<pivotTableDefinition multipleFieldFilters="1"',
            1,
        ))
    core = _load_payload(tmp_path, mutated, name="core.xlsx")
    codes = _codes(core["Summary"].pivots["SalesByRegion"].qualify_adoption())
    assert "foreign-core-semantics-unclassified" in codes \
        or "unsupported-pivot-feature" in codes


def test_namespace_lookalike_cannot_qualify(tmp_path):
    payload = _basic_package()

    def evil(body):
        return body.replace(
            b'<Relationship Id="rIdPivot1"',
            b'<evil:Relationship xmlns:evil="urn:paper-test:evil" '
            b'Id="rIdPivot1"',
            1,
        )

    payload = _rewrite_bytes(
        payload, "xl/worksheets/_rels/sheet1.xml.rels", evil)
    wb = _load_payload(tmp_path, payload, name="lookalike.xlsx")
    assert list(wb["Summary"].pivots) == []


def test_qualification_does_not_invoke_libreoffice(tmp_path, monkeypatch):
    called = []

    def boom(*_args, **_kwargs):
        called.append(True)
        raise AssertionError("LibreOffice must not run during qualification")

    monkeypatch.setattr("openpyxl.oracle.find_soffice", boom)
    monkeypatch.setattr("openpyxl.pivot.calculate.find_soffice", boom, raising=False)
    wb = _load_payload(tmp_path, _basic_package())
    wb["Summary"].pivots["SalesByRegion"].qualify_adoption()
    assert called == []


def test_omitted_item_t_is_schema_default_data(tmp_path):
    payload = _basic_package()
    wb = _load_payload(tmp_path, payload)
    result = wb["Summary"].pivots["SalesByRegion"].qualify_adoption()
    assert "foreign-core-semantics-unclassified" not in _codes(result)
    mutated = _rewrite_bytes(
        _basic_package(),
        "xl/pivotTables/pivotTable1.xml",
        lambda body: body.replace(b"<item ", b'<item t="default" ', 1),
    )
    other = _load_payload(tmp_path, mutated, name="default-t.xlsx")
    defaulted = other["Summary"].pivots["SalesByRegion"].qualify_adoption()
    assert defaulted.eligible is False
    # t="default" is classified, not treated as omitted data.
    assert "foreign-core-semantics-unclassified" not in _codes(defaulted)


def test_ordinary_enumeration_does_not_qualify_adoption(tmp_path, monkeypatch):
    called = []

    def wrap(original):
        def _wrapped(*args, **kwargs):
            called.append(True)
            return original(*args, **kwargs)
        return _wrapped

    from openpyxl.pivot import adopt_qualify
    monkeypatch.setattr(
        adopt_qualify, "analyze_adoption",
        wrap(adopt_qualify.analyze_adoption),
    )
    wb = _load_payload(tmp_path, _basic_package())
    collection = wb["Summary"].pivots
    assert len(collection) == 1
    pivot = collection["SalesByRegion"]
    pivot.to_dict()
    assert called == []
    pivot.qualify_adoption()
    assert called == [True]


def _rewrite_bytes(payload, part, transform):
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(payload)) as before, \
            zipfile.ZipFile(output, "w") as after:
        for info in before.infolist():
            body = before.read(info.filename)
            if info.filename == part:
                body = transform(body)
            after.writestr(info, body)
    return output.getvalue()
