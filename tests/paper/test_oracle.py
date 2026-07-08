"""Phase 5: the LibreOffice oracle — driver rules, recalc, certification
(PLAN Phase 5; PR-0 §7/D16/D17)."""
from __future__ import annotations

import os
import zipfile

import pytest

from openpyxl import Workbook, load_workbook
from openpyxl import oracle
from openpyxl.errors import OracleTimeoutError, OracleUnavailableError


class TestDriverRules:

    def test_unavailable_raises_typed_error(self, fixture_copy, monkeypatch):
        monkeypatch.setattr(oracle, "find_soffice", lambda: None)
        with pytest.raises(OracleUnavailableError, match="LibreOffice"):
            oracle.recalc(fixture_copy("features/schedule.xlsx"))
        with pytest.raises(OracleUnavailableError):
            oracle.certify(fixture_copy("features/schedule_calc.xlsx"))

    def test_output_path_and_in_place_are_exclusive(self, fixture_copy):
        with pytest.raises(ValueError, match="not both"):
            oracle.recalc(fixture_copy("features/schedule.xlsx"),
                          output_path="/tmp/x.xlsx", in_place=True)

    def test_in_place_requires_a_path(self, fixture_copy):
        with open(fixture_copy("features/schedule.xlsx"), "rb") as f:
            data = f.read()
        with pytest.raises(ValueError, match="path"):
            oracle.recalc(data, in_place=True)

    @pytest.mark.lo_smoke
    def test_original_path_never_reaches_libreoffice(
            self, lo, fixture_copy, monkeypatch):
        # the tested invariant (CONVENTIONS §4): temp copies only
        import subprocess as sp

        src = fixture_copy("features/schedule.xlsx")
        seen = []
        real_run = sp.run

        def spy(cmd, **kw):
            seen.append(list(cmd))
            return real_run(cmd, **kw)

        monkeypatch.setattr(oracle.subprocess, "run", spy)
        with open(src, "rb") as f:
            before = f.read()
        oracle.recalc(src)
        assert seen, "LibreOffice was never invoked"
        for cmd in seen:
            assert src not in cmd, "the original path was handed to soffice"
        with open(src, "rb") as f:
            assert f.read() == before

    @pytest.mark.lo_smoke
    def test_timeout_raises_typed_error(self, lo, fixture_copy):
        with pytest.raises(OracleTimeoutError, match="within"):
            oracle.recalc(fixture_copy("features/schedule.xlsx"),
                          timeout=0.001)


@pytest.mark.lo_smoke
class TestRecalc:

    def test_recalc_computes_cached_values(self, lo, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule.xlsx")   # empty <v></v> caches
        out = str(tmp_path / "recalced.xlsx")
        result = oracle.recalc(src, output_path=out)
        assert result.status == "ok"
        assert result.formula_cells == 3
        assert result.to_dict()["error_cells"] == 0
        wb = load_workbook(out, data_only=True)
        assert wb["Schedule"]["B12"].value == 6500
        assert wb["Summary"]["B1"].value == 6500

    def test_in_place_replaces_atomically(self, lo, fixture_copy):
        src = fixture_copy("features/schedule.xlsx")
        oracle.recalc(src, in_place=True)
        wb = load_workbook(src, data_only=True)
        assert wb["Schedule"]["B12"].value == 6500

    def test_error_scan_finds_tokens(self, lo, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = 10
        ws["B1"] = "=A1/0"
        ws["B2"] = "=A1*2"
        src = str(tmp_path / "err.xlsx")
        wb.save(src)
        result = oracle.recalc(src)
        assert result.status == "errors"
        doc = result.to_dict()
        assert doc["error_cells"] == 1
        assert doc["errors"][0]["value"] == "#DIV/0!"
        assert doc["errors"][0]["cell"] == "B1"
        assert doc["schema"] == "oracle_recalc"

    def test_scan_only_writes_nothing(self, lo, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule.xlsx")   # lives in tmp_path
        with open(src, "rb") as f:
            before = f.read()
        oracle.recalc(src)
        with open(src, "rb") as f:
            assert f.read() == before
        assert os.listdir(str(tmp_path)) == [os.path.basename(src)]


class TestCertify:

    def test_baseline_unverifiable_without_cached_values(
            self, fixture_copy, monkeypatch):
        # no LibreOffice needed: the answer key is missing, so the oracle
        # never runs (openpyxl-written files carry empty <v></v>)
        monkeypatch.setattr(oracle, "find_soffice", lambda: None)
        result = oracle.certify(fixture_copy("features/schedule.xlsx"))
        assert result.status == "BASELINE_UNVERIFIABLE"
        assert result.checked == 0
        assert result.unverifiable          # the formula cells are listed

    @pytest.mark.lo_smoke
    def test_lo_recalced_file_certifies(self, lo, fixture_copy):
        result = oracle.certify(fixture_copy("features/schedule_calc.xlsx"))
        assert result.status == "CERTIFIED"
        assert result.checked == 3
        assert result.divergences == []

    @pytest.mark.lo_smoke
    def test_certification_is_deterministic(self, lo, fixture_copy):
        src = fixture_copy("features/schedule_calc.xlsx")
        first = oracle.certify(src).to_dict()
        second = oracle.certify(src).to_dict()
        assert first == second

    @pytest.mark.lo_smoke
    def test_tampered_cache_diverges_with_both_values(
            self, lo, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule_calc.xlsx")
        tampered = str(tmp_path / "tampered.xlsx")
        with zipfile.ZipFile(src) as zin, zipfile.ZipFile(tampered, "w") as zout:
            for name in zin.namelist():
                payload = zin.read(name)
                if name.startswith("xl/worksheets/") and b"6500" in payload:
                    payload = payload.replace(b"<v>6500</v>", b"<v>9999</v>", 1)
                zout.writestr(name, payload)
        result = oracle.certify(tampered)
        assert result.status == "DIVERGED"
        div = result.divergences[0]
        assert div["cached"] == 9999
        assert div["computed"] == 6500
        assert "B12" in div["address"]

    @pytest.mark.lo_smoke
    def test_volatile_cells_and_downstream_excluded(self, lo, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "=RAND()"          # nondeterministic volatile
        ws["A2"] = "=A1*2"            # downstream of it
        ws["B1"] = 21
        ws["B2"] = "=B1*2"            # independent, verifiable
        raw = str(tmp_path / "vol.xlsx")
        wb.save(raw)
        recalced = str(tmp_path / "vol_calc.xlsx")
        oracle.recalc(raw, output_path=recalced)   # produce an answer key
        result = oracle.certify(recalced)
        excluded = set(result.volatile_excluded)
        assert any("A1" in a for a in excluded)
        assert any("A2" in a for a in excluded)    # downstream taint
        assert result.status == "CERTIFIED"        # B2 checked and matching
        assert result.checked >= 1
