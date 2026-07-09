"""The part-lifecycle engine and its dividends (PLAN-v0.1 Batch 2;
PR-1 §1). Every part a save creates or deletes routes through
PartPlan.add_part/remove_part — no bespoke cascades."""
from __future__ import annotations

import io
import zipfile

import pytest

from openpyxl import load_workbook
from openpyxl.errors import (
    RelationshipPolicyError,
    TargetNotFoundError,
)

from .support.partdiff import part_payloads


class TestCustomPropsLifecycle:

    def test_first_custom_prop_creates_the_part(self, fixture_copy,
                                                 tmp_path):
        # v0 refused; the engine creates part + CT override + package rel
        src = fixture_copy("minimal/minimal_clean.xlsx")
        assert "docProps/custom.xml" not in part_payloads(src)
        wb = load_workbook(src, preserve=True)
        from openpyxl.packaging.custom import StringProperty

        wb.custom_doc_props.append(StringProperty(name="Reviewed",
                                                  value="yes"))
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        parts = part_payloads(out)
        assert b"Reviewed" in parts["docProps/custom.xml"]
        assert b"custom.xml" in parts["[Content_Types].xml"]
        assert b"custom-properties" in parts["_rels/.rels"]
        wb2 = load_workbook(out)
        assert wb2.custom_doc_props["Reviewed"].value == "yes"

    def test_removing_last_custom_prop_drops_the_part(self, fixture_copy,
                                                      tmp_path):
        # build a file WITH one custom prop first, then remove it
        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        from openpyxl.packaging.custom import StringProperty

        wb.custom_doc_props.append(StringProperty(name="Tmp", value="x"))
        staged = str(tmp_path / "staged.xlsx")
        wb.save(staged)

        wb2 = load_workbook(staged, preserve=True)
        del wb2.custom_doc_props["Tmp"]
        out = str(tmp_path / "o.xlsx")
        wb2.save(out)
        parts = part_payloads(out)
        assert "docProps/custom.xml" not in parts
        assert b"custom.xml" not in parts["[Content_Types].xml"]
        assert b"custom-properties" not in parts["_rels/.rels"]
        load_workbook(out)                        # reloadable


class TestStylesPartCreation:

    def _styleless(self, fixture_copy, tmp_path):
        import re

        src = fixture_copy("minimal/minimal_clean.xlsx")
        out = str(tmp_path / "styleless.xlsx")
        with zipfile.ZipFile(src) as zin, zipfile.ZipFile(out, "w") as zout:
            for name in zin.namelist():
                if name == "xl/styles.xml":
                    continue
                payload = zin.read(name)
                if name == "[Content_Types].xml":
                    payload = payload.replace(
                        b'<Override PartName="/xl/styles.xml" ContentType='
                        b'"application/vnd.openxmlformats-officedocument.'
                        b'spreadsheetml.styles+xml"/>', b"")
                if name == "xl/_rels/workbook.xml.rels":
                    payload = re.sub(
                        br'<Relationship [^>]*Target="styles.xml"[^>]*/>',
                        b"", payload)
                if name.startswith("xl/worksheets/sheet"):
                    payload = re.sub(br' s="\d+"', b"", payload)
                zout.writestr(name, payload)
        return out

    def test_styled_write_creates_styles_part(self, fixture_copy, tmp_path):
        src = self._styleless(fixture_copy, tmp_path)
        wb = load_workbook(src, preserve=True)
        from openpyxl.styles import Font

        wb["Sheet1"]["B2"] = 42
        wb["Sheet1"]["B2"].font = Font(bold=True)
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        parts = part_payloads(out)
        assert "xl/styles.xml" in parts
        assert b"/xl/styles.xml" in parts["[Content_Types].xml"]
        wb2 = load_workbook(out)
        assert wb2["Sheet1"]["B2"].font.bold is True
        assert wb2["Sheet1"]["B2"].value == 42


class TestReplacePart:

    def test_media_swap_lands_and_guards_hold(self, fixture_copy, tmp_path):
        src = fixture_copy("features/chart_image.xlsx")
        wb = load_workbook(src, preserve=True)
        new_png = part_payloads(src)["xl/media/image1.png"] + b"\x00"
        wb.replace_part("xl/media/image1.png", new_png)
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        assert part_payloads(out)["xl/media/image1.png"] == new_png

        wb2 = load_workbook(src, preserve=True)
        with pytest.raises(TargetNotFoundError):
            wb2.replace_part("xl/media/nope.png", b"x")
        with pytest.raises(RelationshipPolicyError, match="managed"):
            wb2.replace_part("xl/workbook.xml", b"x")
        with pytest.raises(RelationshipPolicyError, match="managed"):
            wb2.replace_part("xl/worksheets/sheet1.xml", b"x")

    def test_replace_part_meaningless_on_stock(self, fixture_copy):
        wb = load_workbook(fixture_copy("features/chart_image.xlsx"))
        with pytest.raises(ValueError, match="preserve"):
            wb.replace_part("xl/media/image1.png", b"x")


class TestEnginePrimitive:

    def test_add_part_refuses_existing_names(self):
        from openpyxl.preserve.lifecycle import PartPlan

        plan = PartPlan({"xl/workbook.xml"})
        with pytest.raises(RelationshipPolicyError, match="never overwrites"):
            plan.add_part("xl/workbook.xml", b"x")

    def test_remove_part_requires_existence(self):
        from openpyxl.preserve.lifecycle import PartPlan

        plan = PartPlan({"xl/workbook.xml"})
        with pytest.raises(TargetNotFoundError):
            plan.remove_part("xl/ghost.xml")

    def test_relative_targets(self):
        from openpyxl.preserve.lifecycle import _relative_target

        assert _relative_target("xl/workbook.xml",
                                "xl/styles.xml") == "styles.xml"
        assert _relative_target("", "docProps/custom.xml") == \
            "docProps/custom.xml"
        assert _relative_target("xl/worksheets/sheet1.xml",
                                "xl/comments1.xml") == "../comments1.xml"


class TestTableLifecycle:

    def test_table_removal_drops_part_and_element(self, fixture_copy,
                                                  tmp_path):
        src = fixture_copy("features/tables.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb.worksheets[0]
        del ws.tables["RegionTable"]
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        parts = part_payloads(out)
        assert not any(n.startswith("xl/tables/") for n in parts)
        sheet = next(p for n, p in parts.items()
                     if n.startswith("xl/worksheets/"))
        assert b"<tableParts" not in sheet
        assert b"table+xml" not in parts["[Content_Types].xml"]
        wb2 = load_workbook(out)
        assert not wb2.worksheets[0].tables
        assert wb2.worksheets[0]["B2"].value == 20     # data survives

    def test_second_table_coexists_with_original(self, fixture_copy,
                                                 tmp_path):
        from openpyxl.worksheet.table import Table

        src = fixture_copy("features/tables.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb.worksheets[0]
        ws["D1"] = "K"
        ws["D2"] = 1
        ws.add_table(Table(displayName="Second", ref="D1:D2"))
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        parts = part_payloads(out)
        # original part untouched byte-for-byte
        assert parts["xl/tables/table1.xml"] == \
            part_payloads(src)["xl/tables/table1.xml"]
        sheet = next(p for n, p in parts.items()
                     if n.startswith("xl/worksheets/"))
        assert b'<tableParts count="2">' in sheet
        wb2 = load_workbook(out)
        assert set(wb2.worksheets[0].tables) == {"RegionTable", "Second"}

    def test_totals_row_moves_with_append(self, fixture_copy, tmp_path):
        # craft a totals-row table: extend RegionTable with a totals row
        src = fixture_copy("features/tables.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb.worksheets[0]
        tbl = ws.tables["RegionTable"]
        ws["A6"] = "Total"
        ws["B6"] = "=SUBTOTAL(109,RegionTable[Amount])"
        tbl.totalsRowCount = 1
        tbl.ref = "A1:B6"
        staged = str(tmp_path / "staged.xlsx")
        wb.save(staged)

        wb2 = load_workbook(staged, preserve=True)
        ws2 = wb2.worksheets[0]
        from openpyxl.preserve.tables import append_row

        append_row(ws2, "RegionTable", ["Central", 60])
        out = str(tmp_path / "o.xlsx")
        wb2.save(out)
        wb3 = load_workbook(out)
        ws3 = wb3.worksheets[0]
        assert ws3["A6"].value == "Central"
        assert ws3["B6"].value == 60
        assert ws3["A7"].value == "Total"              # totals stayed last
        assert ws3.tables["RegionTable"].ref == "A1:B7"

    def test_append_refuses_content_below(self, fixture_copy, tmp_path):
        from openpyxl.errors import UnsupportedStructureError
        from openpyxl.preserve.tables import append_row

        src = fixture_copy("features/tables.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb.worksheets[0]
        ws["A8"] = "in the way"
        with pytest.raises(UnsupportedStructureError, match="below"):
            append_row(ws, "RegionTable", ["X", 1])


class TestCommentCreation:

    def test_comment_edit_and_remove_before_save(self, fixture_copy,
                                                  tmp_path):
        from openpyxl.comments import Comment

        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Sheet1"]
        ws["A1"].comment = Comment("draft", "paper")
        ws["A1"].comment = Comment("final", "paper")   # replaced pre-save
        ws["B2"].comment = Comment("gone", "paper")
        ws["B2"].comment = None                        # removed pre-save
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert "final" in wb2["Sheet1"]["A1"].comment.text
        assert wb2["Sheet1"]["B2"].comment is None

    def test_comment_on_sheet_with_existing_machinery_refuses(
            self, fixture_copy, tmp_path):
        from openpyxl.comments import Comment
        from openpyxl.errors import UnsupportedStructureError

        src = fixture_copy("gauntlet/gauntlet.xlsx")   # Model has comments
        with open(src, "rb") as f:
            before = f.read()
        wb = load_workbook(src, preserve=True)
        wb["Model"]["G1"].comment = Comment("new", "paper")
        with pytest.raises(UnsupportedStructureError, match="already"):
            wb.save(str(tmp_path / "o.xlsx"))
        with open(src, "rb") as f:
            assert f.read() == before

    def test_comments_and_new_table_coexist_via_reserved_rids(
            self, fixture_copy, tmp_path):
        from openpyxl.comments import Comment
        from openpyxl.worksheet.table import Table

        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Sheet1"]
        ws["A1"].comment = Comment("note", "paper")
        ws.add_table(Table(displayName="T", ref="A1:B3"))
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        wb2 = load_workbook(out)
        assert wb2["Sheet1"]["A1"].comment is not None
        assert "T" in wb2["Sheet1"].tables
