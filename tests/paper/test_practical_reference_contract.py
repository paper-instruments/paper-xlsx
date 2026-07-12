import io
import warnings

import pytest

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.data_source import AxDataSource, MultiLevelStrRef, StrRef
from openpyxl.chart.text import Text
from openpyxl.chart.title import Title
from openpyxl.errors import PaperRefusal
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.preserve.references import chart_source_ref_objects
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.filters import FilterColumn, SortCondition, SortState
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.scenario import InputCells, Scenario, ScenarioList
from openpyxl.worksheet.table import Table, TableFormula


def _preserved(tmp_path, workbook, name="source.xlsx"):
    source = tmp_path / name
    workbook.save(source)
    return load_workbook(source, preserve=True)


def _table_with_formulas(sheet):
    sheet.append(["Amount"])
    sheet.append([1])
    sheet.append([2])
    table = Table(displayName="Calculations", ref="A1:A3")
    table._initialise_columns()
    column = table.tableColumns[0]
    column.name = "Amount"
    column.calculatedColumnFormula = TableFormula(attr_text="Data!A2*2")
    column.totalsRowFormula = TableFormula(attr_text="SUM(Data!A2)")
    sheet.add_table(table)


def _reference_workbook():
    workbook = Workbook()
    data = workbook.active
    data.title = "Data"
    data["A1"] = "Input"
    data["A2"] = 10
    data.print_title_rows = "1:2"
    data.auto_filter.ref = "A1:A3"
    data.auto_filter.sortState = SortState(
        ref="A2:A3", sortCondition=[SortCondition(ref="A2:A3")])
    data.scenarios = ScenarioList(
        sqref="A2",
        scenario=[Scenario(
            name="Base", inputCells=[InputCells(r="A2", val="10")])],
    )
    rules = workbook.create_sheet("Rules")
    _table_with_formulas(rules)
    rules.conditional_formatting.add(
        "B1",
        ColorScaleRule(
            start_type="formula", start_value="Data!A2",
            start_color="FF0000", end_type="max", end_color="00FF00"))
    return workbook


def _table_formulas(sheet):
    column = sheet.tables["Calculations"].tableColumns[0]
    return (
        column.calculatedColumnFormula.attr_text,
        column.totalsRowFormula.attr_text,
    )


def test_shift_rewrites_modeled_professional_workbook_surfaces(tmp_path):
    workbook = _preserved(tmp_path, _reference_workbook())
    workbook["Rules"]["C1"].hyperlink = Hyperlink(
        ref="C1", location="Data!A2")

    workbook["Data"].insert_rows(2)

    data = workbook["Data"]
    rules = workbook["Rules"]
    assert data.auto_filter.ref == "A1:A4"
    assert data.auto_filter.sortState.ref == "A3:A4"
    assert data.auto_filter.sortState.sortCondition[0].ref == "A3:A4"
    assert data.print_title_rows == "$1:$3"
    assert str(data.scenarios.sqref) == "A3"
    assert data.scenarios.scenario[0].inputCells[0].r == "A3"
    conditional = next(iter(rules.conditional_formatting))
    assert conditional.rules[0].colorScale.cfvo[0].val == "Data!A3"
    assert rules["C1"].hyperlink.location == "Data!A3"
    assert _table_formulas(rules) == ("Data!A3*2", "SUM(Data!A3)")

    output = tmp_path / "shifted.xlsx"
    workbook.save(output)
    reopened = load_workbook(output)
    assert reopened["Data"].auto_filter.sortState.ref == "A3:A4"
    assert _table_formulas(reopened["Rules"]) == (
        "Data!A3*2", "SUM(Data!A3)")


def test_added_sheet_shift_updates_existing_sheet_formula(tmp_path):
    workbook = Workbook()
    workbook.active.title = "Existing"
    workbook = _preserved(tmp_path, workbook, "added.xlsx")
    added = workbook.create_sheet("Added")
    added["A2"] = 10
    workbook["Existing"]["A1"] = "=Added!A2"

    added.insert_rows(2)

    assert workbook["Existing"]["A1"].value == "=Added!A3"


def test_sheet_local_name_follows_shift(tmp_path):
    workbook = Workbook()
    workbook.active.title = "Data"
    workbook.active["A2"] = 10
    workbook.active.defined_names.add(
        DefinedName("LocalRange", attr_text="$A$2"))
    workbook = _preserved(tmp_path, workbook, "local-name.xlsx")

    workbook["Data"].insert_rows(2)

    assert workbook["Data"].defined_names["LocalRange"].attr_text == "$A$3"


def test_rename_rewrites_modeled_formula_surfaces(tmp_path):
    workbook = _preserved(tmp_path, _reference_workbook(), "rename.xlsx")
    workbook["Rules"]["C1"].hyperlink = Hyperlink(
        ref="C1", location="Data!A2")

    workbook["Data"].title = "Inputs"

    rules = workbook["Rules"]
    conditional = next(iter(rules.conditional_formatting))
    assert conditional.rules[0].colorScale.cfvo[0].val == "'Inputs'!A2"
    assert rules["C1"].hyperlink.location == "'Inputs'!A2"
    assert _table_formulas(rules) == (
        "'Inputs'!A2*2", "SUM('Inputs'!A2)")


@pytest.mark.parametrize(
    "formula",
    ['=INDIRECT("Da"&"ta!A2")', '=EVALUATE("Da"&"ta!A2")',
     "=OFFSET(Data!A1,1,0)", "=INDEX(Data!A:A,2)"],
)
def test_dynamic_structural_reference_refuses_before_mutation(
        tmp_path, formula):
    workbook = Workbook()
    workbook.active.title = "Data"
    workbook.active["A2"] = 10
    workbook.create_sheet("Other")["A1"] = formula
    workbook = _preserved(tmp_path, workbook, "dynamic.xlsx")

    with pytest.raises(PaperRefusal) as refusal:
        workbook["Data"].insert_rows(2)

    assert refusal.value.kind == "dynamic-structural-reference"
    assert workbook["Data"]["A2"].value == 10


def test_dynamic_function_name_inside_text_does_not_block_shift(tmp_path):
    workbook = Workbook()
    workbook.active.title = "Data"
    workbook.active["A2"] = 10
    workbook.create_sheet("Other")["A1"] = (
        '=IF(Data!A2="INDEX(",Data!A2,Data!A3)')
    workbook = _preserved(tmp_path, workbook, "quoted-function-shift.xlsx")

    workbook["Data"].insert_rows(2)

    assert workbook["Other"]["A1"].value == (
        '=IF(Data!A3="INDEX(",Data!A3,Data!A4)')


def test_dynamic_function_name_inside_text_does_not_block_rename(tmp_path):
    workbook = Workbook()
    workbook.active.title = "Data"
    workbook.create_sheet("Other")["A1"] = (
        '=IF(Data!A1="INDIRECT(",Data!A2,Data!A3)')
    workbook = _preserved(tmp_path, workbook, "quoted-function-rename.xlsx")

    workbook["Data"].title = "Inputs"

    assert workbook["Other"]["A1"].value == (
        '=IF(\'Inputs\'!A1="INDIRECT(",\'Inputs\'!A2,\'Inputs\'!A3)')


def test_chart_reference_walker_includes_titles_axes_and_multilevel():
    workbook = Workbook()
    sheet = workbook.active
    for row in range(1, 4):
        sheet.cell(row, 1, row)
    chart = BarChart()
    chart.add_data(Reference(sheet, min_col=1, min_row=1, max_row=3))
    chart.title = Title(tx=Text(strRef=StrRef(f="Sheet!$A$1")))
    chart.x_axis.title = Title(tx=Text(strRef=StrRef(f="Sheet!$A$2")))
    chart.series[0].cat = AxDataSource(
        multiLvlStrRef=MultiLevelStrRef(f="Sheet!$A$1:$A$3"))
    references = list(chart_source_ref_objects(chart))

    assert {reference.f for reference in references} >= {
        "Sheet!$A$1", "Sheet!$A$2", "Sheet!$A$1:$A$3"}


def test_repeated_supported_shifts_keep_references_and_remaps_coherent(
        tmp_path):
    workbook = Workbook()
    data = workbook.active
    data.title = "Data"
    data["A2"] = 10
    data.defined_names.add(DefinedName("Input", attr_text="$A$2"))
    other = workbook.create_sheet("Other")
    other["A1"] = "=Data!A2"
    workbook = _preserved(tmp_path, workbook, "repeated.xlsx")

    chart = BarChart()
    chart.add_data(Reference(
        workbook["Data"], min_col=1, min_row=2, max_row=2))
    workbook["Other"].add_chart(chart, "D1")

    first = workbook["Data"].insert_rows(2)
    second = workbook["Data"].insert_rows(3)

    assert first.map("Data!A2") == "Data!A3"
    assert second.map("Data!A3") == "Data!A4"
    assert workbook["Other"]["A1"].value == "=Data!A4"
    assert workbook["Data"].defined_names["Input"].attr_text == "$A$4"
    assert next(chart_source_ref_objects(chart)).f == "'Data'!$A$4"

    output = tmp_path / "repeated-output.xlsx"
    workbook.save(output)
    reopened = load_workbook(output)
    assert reopened["Other"]["A1"].value == "=Data!A4"
    assert reopened["Data"].defined_names["Input"].attr_text == "$A$4"


def test_implicit_intersection_reference_follows_shift_and_rename(tmp_path):
    workbook = Workbook()
    workbook.active.title = "Data"
    workbook.active["A2"] = 10
    workbook.create_sheet("Other")["A1"] = "=@Data!A2"
    workbook = _preserved(tmp_path, workbook, "implicit.xlsx")

    workbook["Data"].insert_rows(2)
    workbook["Data"].title = "Inputs"

    assert workbook["Other"]["A1"].value == "=@'Inputs'!A3"


def test_second_shift_checks_chart_ranges_after_first_shift(tmp_path):
    workbook = Workbook()
    data = workbook.active
    data.title = "Data"
    data["A2"] = 1
    data["A3"] = 2
    other = workbook.create_sheet("Other")
    chart = BarChart()
    chart.add_data(Reference(data, min_col=1, min_row=2, max_row=3))
    other.add_chart(chart, "D1")
    workbook = _preserved(tmp_path, workbook, "chart-shifts.xlsx")

    workbook["Data"].insert_rows(2)
    before = dict(workbook["Data"]._cells)
    with pytest.raises(PaperRefusal, match="delete data charted"):
        workbook["Data"].delete_rows(3, 2)

    assert workbook["Data"]._cells == before


def test_renamed_sheet_chart_delete_refuses_before_mutation(tmp_path):
    workbook = Workbook()
    data = workbook.active
    data.title = "Data"
    data["A2"] = 1
    data["A3"] = 2
    chart = BarChart()
    chart.add_data(Reference(data, min_col=1, min_row=2, max_row=3))
    workbook.create_sheet("Other").add_chart(chart, "D1")
    workbook = _preserved(tmp_path, workbook, "renamed-chart.xlsx")
    workbook["Data"].title = "Inputs"
    before = dict(workbook["Inputs"]._cells)

    with pytest.raises(PaperRefusal, match="delete data charted"):
        workbook["Inputs"].delete_rows(2, 2)

    assert workbook["Inputs"]._cells == before


def test_table_extent_rewrites_and_destructive_delete_refuses(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    _table_with_formulas(sheet)
    workbook = _preserved(tmp_path, workbook, "table-range.xlsx")

    workbook["Data"].insert_rows(2)
    assert workbook["Data"].tables["Calculations"].ref == "A1:A4"
    before = dict(workbook["Data"]._cells)
    with pytest.raises(PaperRefusal):
        workbook["Data"].delete_rows(1, 4)
    assert workbook["Data"]._cells == before


@pytest.mark.parametrize(
    "operation",
    [lambda sheet: sheet.insert_cols(1 + 1),
     lambda sheet: sheet.delete_cols(1),
     lambda sheet: sheet.delete_rows(1)],
)
def test_table_metadata_changing_shifts_refuse_before_mutation(
        tmp_path, operation):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Amount", "Rate"])
    sheet.append([1, 2])
    table = Table(displayName="Inputs", ref="A1:B2")
    sheet.add_table(table)
    workbook = _preserved(tmp_path, workbook, "table-metadata.xlsx")
    before = dict(workbook["Data"]._cells)

    with pytest.raises(PaperRefusal) as refusal:
        operation(workbook["Data"])

    assert refusal.value.kind == "table-structure-edit-unsupported"
    assert workbook["Data"]._cells == before


def test_deleting_table_only_data_row_refuses_before_mutation(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Amount"])
    sheet.append([1])
    sheet.add_table(Table(displayName="Inputs", ref="A1:A2"))
    workbook = _preserved(tmp_path, workbook, "short-table.xlsx")

    with pytest.raises(PaperRefusal) as refusal:
        workbook.active.delete_rows(2)

    assert refusal.value.kind == "table-structure-edit-unsupported"
    assert workbook.active.tables["Inputs"].ref == "A1:A2"


def test_three_dimensional_defined_name_refuses(tmp_path):
    workbook = Workbook()
    workbook.active.title = "First"
    workbook.create_sheet("Middle")["A2"] = 2
    workbook.create_sheet("Last")
    workbook.defined_names.add(DefinedName(
        "Across", attr_text="First:Last!$A$2"))
    workbook = _preserved(tmp_path, workbook, "three-d-name.xlsx")

    with pytest.raises(PaperRefusal) as refusal:
        workbook["Middle"].insert_rows(2)

    assert refusal.value.kind == "three-dimensional-structural-reference"


def test_filter_columns_drop_deleted_field_and_rebase_remaining(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    for row in (["A", "B", "C", "D"], [1, 2, 3, 4]):
        sheet.append(row)
    sheet.auto_filter.ref = "A1:D2"
    sheet.auto_filter.filterColumn = [
        FilterColumn(colId=1), FilterColumn(colId=2)]
    workbook = _preserved(tmp_path, workbook, "filters.xlsx")

    workbook["Data"].delete_cols(2)

    assert workbook["Data"].auto_filter.ref == "A1:C2"
    assert [column.colId for column in
            workbook["Data"].auto_filter.filterColumn] == [1]


def test_late_structural_refusal_restores_serialized_workbook(tmp_path):
    workbook = _preserved(tmp_path, _reference_workbook(), "rollback.xlsx")
    before = io.BytesIO()
    workbook.save(before)

    with pytest.raises(PaperRefusal):
        workbook["Data"].delete_rows(2)

    after = io.BytesIO()
    workbook.save(after)
    assert after.getvalue() == before.getvalue()


def test_print_area_follows_structural_shift(tmp_path):
    workbook = Workbook()
    workbook.active["A3"] = 3
    workbook.active.print_area = "A1:A3"
    workbook = _preserved(tmp_path, workbook, "print-area.xlsx")

    workbook.active.insert_rows(2)

    assert workbook.active.print_area.endswith("$A$1:$A$4")


def test_workbook_name_without_sheet_context_refuses(tmp_path):
    workbook = Workbook()
    workbook.active["A2"] = 2
    workbook.defined_names.add(DefinedName("Localish", attr_text="$A$2"))
    workbook = _preserved(tmp_path, workbook, "ambiguous-name.xlsx")

    with pytest.raises(PaperRefusal) as refusal:
        workbook.active.insert_rows(2)

    assert refusal.value.kind == "ambiguous-structural-reference"


def test_move_range_translation_failure_rolls_back(tmp_path):
    workbook = Workbook()
    workbook.active["A1"] = 1
    workbook.active["B2"] = "=A1"
    workbook = _preserved(tmp_path, workbook, "move-rollback.xlsx")
    before = io.BytesIO()
    workbook.save(before)

    with pytest.raises(Exception):
        workbook.active.move_range("B2", rows=-1, cols=-1, translate=True)

    after = io.BytesIO()
    workbook.save(after)
    assert after.getvalue() == before.getvalue()


def test_rename_warning_failure_rolls_back_prior_rewrites(tmp_path):
    from openpyxl.errors import ProtectedWriteWarning

    workbook = Workbook()
    data = workbook.active
    data.title = "Data"
    first = workbook.create_sheet("First")
    first["A1"] = "=Data!A1"
    protected = workbook.create_sheet("Protected")
    protected["A1"] = "=Data!A1"
    protected.protection.sheet = True
    workbook = _preserved(tmp_path, workbook, "rename-rollback.xlsx")

    with warnings.catch_warnings():
        warnings.simplefilter("error", ProtectedWriteWarning)
        with pytest.raises(ProtectedWriteWarning):
            workbook["Data"].title = "Inputs"

    assert workbook["Data"].title == "Data"
    assert workbook["First"]["A1"].value == "=Data!A1"
    assert workbook["Protected"]["A1"].value == "=Data!A1"
