import io
import datetime
import re
from types import SimpleNamespace
import zipfile

import pytest

from openpyxl import Workbook, load_workbook
from openpyxl import oracle
from openpyxl.errors import UnsupportedStructureError
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils.datetime import MAC_EPOCH, WINDOWS_EPOCH, to_excel


def _package_with_caches(formulas, caches):
    workbook = Workbook()
    sheet = workbook.active
    for coordinate, formula in formulas.items():
        sheet[coordinate] = formula
    target = io.BytesIO()
    workbook.save(target)

    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(target.getvalue())) as source, \
            zipfile.ZipFile(output, "w") as destination:
        for info in source.infolist():
            payload = source.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                for coordinate, (data_type, value) in caches.items():
                    pattern = re.compile(
                        rb'(<c r="' + coordinate.encode("ascii")
                        + rb'"[^>]*>\s*<f[^>]*>.*?</f>)\s*'
                        + rb'(?:<v\s*/>|<v></v>)', re.S)
                    type_attribute = (b' t="' + data_type.encode("ascii")
                                      + b'"') if data_type else b""
                    replacement = (
                        b'<c r="' + coordinate.encode("ascii")
                        + b'"' + type_attribute + b'><f>'
                        + formulas[coordinate][1:].encode("utf-8")
                        + b'</f><v>' + str(value).encode("utf-8") + b'</v>')
                    payload, count = pattern.subn(replacement, payload)
                    assert count == 1
            destination.writestr(info, payload)
    return output.getvalue()


def _replace_formula_cache(package, coordinate, value):
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(package)) as source, \
            zipfile.ZipFile(output, "w") as destination:
        for info in source.infolist():
            payload = source.read(info.filename)
            if info.filename == "xl/worksheets/sheet1.xml":
                pattern = re.compile(
                    rb'(<c r="' + coordinate.encode("ascii")
                    + rb'"[^>]*>\s*<f[^>]*>.*?</f>)\s*'
                    + rb'(?:<v\s*/>|<v></v>)', re.S)
                replacement = (rb'\1<v>' + repr(value).encode("ascii")
                               + b'</v>')
                payload, count = pattern.subn(replacement, payload)
                assert count == 1
            destination.writestr(info, payload)
    return output.getvalue()


def test_formula_result_comparison_requires_matching_excel_types():
    assert not oracle._formula_results_match("#N/A", "e", "#N/A", "s")
    assert not oracle._formula_results_match(True, "b", 1, "n")
    assert not oracle._formula_results_match("42", "s", 42, "n")


def test_numeric_comparison_accepts_rounding_noise_not_material_difference():
    adjacent = 1.0 + 2 ** -52
    assert oracle._values_match(1.0, adjacent)
    assert not oracle._values_match(1_000_000_000_000.0,
                                    1_000_000_000_100.0)


def test_error_scan_distinguishes_text_from_excel_error_type():
    workbook = Workbook()
    workbook.active["A1"] = "#N/A"
    workbook.active["A1"].data_type = "s"
    workbook.active["A2"] = "#N/A"
    workbook.active["A2"].data_type = "e"
    target = io.BytesIO()
    workbook.save(target)

    assert oracle._scan_errors(target.getvalue()) == [
        {"sheet": "Sheet", "cell": "A2", "value": "#N/A"}]


def test_matching_formula_error_never_certifies():
    package = _package_with_caches(
        {"A1": "=1/0"}, {"A1": ("e", "#DIV/0!")})

    result, _recalculated = oracle._certify_impl(
        package, 1, recalculated=package)

    assert result.status == "DIVERGED"
    assert result.divergences == [{
        "address": "Sheet!A1",
        "cached": "#DIV/0!",
        "computed": "#DIV/0!",
        "reason": "formula-error",
    }]


def test_excluded_formula_prevents_complete_certification():
    package = _package_with_caches(
        {"A1": "=RAND()", "B1": "=1+1"},
        {"A1": (None, 0.5), "B1": (None, 2)})

    result, _recalculated = oracle._certify_impl(
        package, 1, recalculated=package)

    assert result.status == "BASELINE_UNVERIFIABLE"
    assert result.checked == 1
    assert result.volatile_excluded == ["Sheet!A1"]


def test_cache_write_type_gate_rejects_mismatched_serializer_type():
    assert oracle._cache_write_preserves_type(42, "n", None)
    assert not oracle._cache_write_preserves_type("42", "n", None)
    assert not oracle._cache_write_preserves_type("#N/A", "s", None)
    assert not oracle._cache_write_preserves_type(
        datetime.date(2025, 1, 1), "d", None)


@pytest.mark.parametrize(
    ("value", "number_format"),
    [
        (datetime.datetime(2025, 1, 2, 3, 4, 5),
         "yyyy-mm-dd h:mm:ss"),
        (datetime.time(3, 4, 5), "h:mm:ss"),
        (datetime.timedelta(hours=27, minutes=4), "[h]:mm:ss"),
    ],
)
def test_cache_write_type_gate_accepts_proven_temporal_round_trip(
        value, number_format):
    assert oracle._cache_write_preserves_type(
        value, "d", WINDOWS_EPOCH, number_format=number_format)
    assert not oracle._cache_write_preserves_type(
        value, "d", WINDOWS_EPOCH, number_format="General")


@pytest.mark.parametrize("epoch", [WINDOWS_EPOCH, MAC_EPOCH])
def test_preserved_recalc_candidate_writes_temporal_cache_with_source_style(
        monkeypatch, epoch):
    workbook = Workbook()
    workbook.epoch = epoch
    cell = workbook.active["A1"]
    cell.value = "=DATE(2025,1,2)+TIME(3,4,5)"
    cell.number_format = "yyyy-mm-dd h:mm:ss"
    source = io.BytesIO()
    workbook.save(source)
    computed_value = datetime.datetime(2025, 1, 2, 3, 4, 5)
    recalculated = _replace_formula_cache(
        source.getvalue(), "A1", to_excel(computed_value, epoch))
    computed_formulas = load_workbook(
        io.BytesIO(recalculated), data_only=False)
    computed_values = load_workbook(io.BytesIO(recalculated), data_only=True)
    certification = SimpleNamespace(
        volatile_excluded=[], external_excluded=[],
        unsupported_excluded=[], input_excluded=[])
    monkeypatch.setattr(
        oracle, "_certify_impl",
        lambda *_args, **_kwargs: (certification, recalculated))

    candidate, written, _unchanged, excluded, _parts = \
        oracle._preserved_recalc_candidate(
            source.getvalue(), recalculated, computed_formulas,
            computed_values, 1)

    assert written == ["Sheet!A1"]
    assert excluded == {}
    reloaded = load_workbook(io.BytesIO(candidate), data_only=True)
    assert reloaded.active["A1"].value == computed_value


def test_preserved_recalc_candidate_refuses_temporal_cache_without_date_style(
        monkeypatch):
    workbook = Workbook()
    cell = workbook.active["A1"]
    cell.value = "=DATE(2025,1,2)"
    source = io.BytesIO()
    workbook.save(source)

    styled = load_workbook(io.BytesIO(source.getvalue()))
    styled.active["A1"].number_format = "yyyy-mm-dd"
    styled_source = io.BytesIO()
    styled.save(styled_source)
    computed_value = datetime.datetime(2025, 1, 2)
    recalculated = _replace_formula_cache(
        styled_source.getvalue(), "A1",
        to_excel(computed_value, WINDOWS_EPOCH))
    computed_formulas = load_workbook(
        io.BytesIO(recalculated), data_only=False)
    computed_values = load_workbook(io.BytesIO(recalculated), data_only=True)
    certification = SimpleNamespace(
        volatile_excluded=[], external_excluded=[],
        unsupported_excluded=[], input_excluded=[])
    monkeypatch.setattr(
        oracle, "_certify_impl",
        lambda *_args, **_kwargs: (certification, recalculated))

    _candidate, written, _unchanged, excluded, _parts = \
        oracle._preserved_recalc_candidate(
            source.getvalue(), recalculated, computed_formulas,
            computed_values, 1)

    assert written == []
    assert excluded == {"Sheet!A1": "computed-cache-type-not-writable"}


def test_defined_name_formula_exclusions_prevent_certification():
    workbook = Workbook()
    workbook.defined_names.add(DefinedName("Clock", attr_text="TODAY()"))
    workbook.active["A1"] = "=Clock"
    target = io.BytesIO()
    workbook.save(target)
    package = target.getvalue()

    result, _recalculated = oracle._certify_impl(
        package, 1, recalculated=package)

    assert result.status == "BASELINE_UNVERIFIABLE"
    assert result.volatile_excluded == ["Sheet!A1"]


def test_defined_name_formula_propagates_input_taint():
    workbook = Workbook()
    workbook.active["A1"] = 1
    workbook.defined_names.add(DefinedName(
        "Calc", attr_text="SUM(Sheet!$A$1)"))
    workbook.active["B1"] = "=Calc"
    target = io.BytesIO()
    workbook.save(target)
    package = target.getvalue()

    result, _ = oracle._certify_impl(
        package, 1, recalculated=package,
        input_seeds=[("Sheet", 1, 1)])

    assert result.status == "BASELINE_UNVERIFIABLE"
    assert result.input_excluded == ["Sheet!B1"]


def test_separate_path_recalc_never_replaces_source(tmp_path, monkeypatch):
    source = tmp_path / "source.xlsx"
    package = _package_with_caches(
        {"A1": "=1+1"}, {"A1": (None, 2)})
    source.write_bytes(package)

    def return_candidate(data, _timeout):
        return data

    monkeypatch.setattr(oracle, "_recalculate_bytes", return_candidate)
    output = tmp_path / "candidate.xlsx"
    result = oracle.recalc(source, output_path=output)
    assert source.read_bytes() == package
    assert result.written == []
    assert result.verified_unchanged == ["Sheet!A1"]
    assert result.package_diff == ["xl/workbook.xml"]
    with zipfile.ZipFile(output) as archive:
        calc_properties = archive.read("xl/workbook.xml")
        assert b'fullCalcOnLoad="1"' in calc_properties
        assert b'forceFullCalc="1"' in calc_properties
