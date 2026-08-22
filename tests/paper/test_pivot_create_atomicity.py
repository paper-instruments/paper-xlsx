"""Atomic refusals for PR 4 pivot creation."""
from __future__ import annotations

import os

import pytest

from openpyxl import load_workbook
from openpyxl.errors import (
    AmbiguousTargetError,
    BoundaryViolationError,
    TargetNotFoundError,
    UnsupportedStructureError,
)
from openpyxl.pivot import create as create_mod
from openpyxl.worksheet.formula import ArrayFormula

from .support.harness import assert_refusal_atomic


_TABLE = "features/tables.xlsx"


def _fingerprint(wb):
    ledger = wb._paper_ledger
    cells = {}
    for ws in wb.worksheets:
        cells[ws.title] = {
            coord: (cell.value, cell._style, cell.data_type)
            for coord, cell in ws._cells.items()
        }
    return {
        "cells": cells,
        "dimensions": {ws.title: ws.dimensions for ws in wb.worksheets},
        "styles": len(wb._cell_styles),
        "dirty": {
            ws.title: frozenset(ledger.dirty_coordinates(ws))
            for ws in wb.worksheets
        },
        "overwrites": {
            ws.title: frozenset(ledger.value_overwrites.get(ws, ()))
            for ws in wb.worksheets
        },
        "operations": tuple(sorted(ledger.pivot_operations)),
    }


def _assert_untouched(wb, before, src, dest=None):
    expected = {key: value for key, value in before.items() if key != "source"}
    assert _fingerprint(wb) == expected
    with open(src, "rb") as handle:
        source_bytes = handle.read()
    assert source_bytes == before["source"]
    if dest is not None and os.path.exists(dest):
        with open(dest, "rb") as handle:
            assert handle.read() == before["source"]


def _load(path):
    wb = load_workbook(path, preserve=True)
    with open(path, "rb") as handle:
        source = handle.read()
    snap = _fingerprint(wb)
    snap["source"] = source
    return wb, snap


def test_duplicate_name_is_atomic(fixture_copy, tmp_path):
    src = fixture_copy(_TABLE)
    wb, before = _load(src)
    dest = os.path.join(str(tmp_path), "dup.xlsx")
    wb["Data"].pivots.create(
        name="ByRegion", source="RegionTable", destination="E3",
        rows=["Region"], values=["Amount"])
    with pytest.raises(AmbiguousTargetError) as exc:
        wb["Data"].pivots.create(
            name="byregion", source="RegionTable", destination="G3",
            rows=["Region"], values=["Amount"])
    assert exc.value.kind == "ambiguous-pivot"
    assert not os.path.exists(dest)
    with open(src, "rb") as handle:
        assert handle.read() == before["source"]


def test_missing_table_is_atomic(fixture_copy, tmp_path):
    src = fixture_copy(_TABLE)
    dest = os.path.join(str(tmp_path), "missing.xlsx")

    def mutate(wb, path):
        wb["Data"].pivots.create(
            name="ByRegion", source="NoSuchTable", destination="E3",
            rows=["Region"], values=["Amount"])
        wb.save(path)

    error = assert_refusal_atomic(
        src, tmp_path, mutate, TargetNotFoundError, load_kw={"preserve": True})
    assert error.kind in ("pivot-not-found", "invalid-pivot-source")
    assert not os.path.exists(dest)


def test_missing_field_is_atomic(fixture_copy):
    src = fixture_copy(_TABLE)
    wb, before = _load(src)
    with pytest.raises(BoundaryViolationError) as exc:
        wb["Data"].pivots.create(
            name="ByRegion", source="RegionTable", destination="E3",
            rows=["Missing"], values=["Amount"])
    assert exc.value.kind in ("invalid-pivot-source", "unsupported-pivot-feature")
    _assert_untouched(wb, before, src)


def test_source_output_overlap_is_atomic(fixture_copy):
    src = fixture_copy(_TABLE)
    wb, before = _load(src)
    with pytest.raises(BoundaryViolationError) as exc:
        wb["Data"].pivots.create(
            name="ByRegion", source="RegionTable", destination="A1",
            rows=["Region"], values=["Amount"])
    assert exc.value.kind == "pivot-source-output-overlap"
    _assert_untouched(wb, before, src)


def test_nonblank_destination_is_atomic(fixture_copy):
    src = fixture_copy(_TABLE)
    wb, before = _load(src)
    wb["Data"]["E4"] = "blocked"
    before = _fingerprint(wb)
    with open(src, "rb") as handle:
        before["source"] = handle.read()
    with pytest.raises(BoundaryViolationError) as exc:
        wb["Data"].pivots.create(
            name="ByRegion", source="RegionTable", destination="E3",
            rows=["Region"], values=["Amount"])
    assert exc.value.kind == "pivot-output-collision"
    _assert_untouched(wb, before, src)


def test_table_destination_is_atomic(fixture_copy):
    src = fixture_copy(_TABLE)
    wb, before = _load(src)
    with pytest.raises(BoundaryViolationError) as exc:
        wb["Data"].pivots.create(
            name="ByRegion", source="'Data'!A1:B5", destination="A3",
            rows=["Region"], values=["Amount"])
    assert exc.value.kind in (
        "pivot-output-collision", "pivot-source-output-overlap")
    _assert_untouched(wb, before, src)


def test_merge_destination_is_atomic(fixture_copy):
    src = fixture_copy("features/merged.xlsx")
    wb, before = _load(src)
    wb["Sheet"].cell(10, 1, "Region")
    wb["Sheet"].cell(10, 2, "Amount")
    wb["Sheet"].cell(11, 1, "East")
    wb["Sheet"].cell(11, 2, 4)
    before = _fingerprint(wb)
    with open(src, "rb") as handle:
        before["source"] = handle.read()
    with pytest.raises(BoundaryViolationError) as exc:
        wb["Sheet"].pivots.create(
            name="ByRegion",
            source="'Sheet'!A10:B11",
            destination="A1",
            rows=["Region"],
            values=["Amount"],
        )
    assert exc.value.kind == "pivot-output-collision"
    _assert_untouched(wb, before, src)


def test_array_formula_destination_is_atomic(fixture_copy):
    src = fixture_copy(_TABLE)
    wb, before = _load(src)
    wb["Data"]["E4"] = ArrayFormula(ref="E4", text="=1")
    before = _fingerprint(wb)
    with open(src, "rb") as handle:
        before["source"] = handle.read()
    with pytest.raises(BoundaryViolationError) as exc:
        wb["Data"].pivots.create(
            name="ByRegion", source="RegionTable", destination="E3",
            rows=["Region"], values=["Amount"])
    assert exc.value.kind == "pivot-output-collision"
    _assert_untouched(wb, before, src)


def test_formula_source_without_oracle_is_atomic(fixture_copy, monkeypatch):
    monkeypatch.setattr("openpyxl.oracle.find_soffice", lambda: None)
    src = fixture_copy(_TABLE)
    wb, before = _load(src)
    wb["Data"]["B2"] = "=1+19"
    before = _fingerprint(wb)
    with open(src, "rb") as handle:
        before["source"] = handle.read()
    with pytest.raises(UnsupportedStructureError) as exc:
        wb["Data"].pivots.create(
            name="ByRegion", source="RegionTable", destination="E3",
            rows=["Region"], values=["Amount"])
    assert exc.value.kind == "unsupported-pivot-source"
    _assert_untouched(wb, before, src)


def test_defined_name_source_is_atomic(fixture_copy):
    from openpyxl.pivot.api_types import PivotSource

    src = fixture_copy(_TABLE)
    wb, before = _load(src)
    with pytest.raises(UnsupportedStructureError) as exc:
        wb["Data"].pivots.create(
            name="ByRegion",
            source=PivotSource(kind="defined-name", name="SalesRange"),
            destination="E3",
            rows=["Region"],
            values=["Amount"],
        )
    assert exc.value.kind == "unsupported-pivot-source"
    _assert_untouched(wb, before, src)


def test_stale_source_refuses_at_save(fixture_copy, tmp_path):
    src = fixture_copy(_TABLE)
    dest = os.path.join(str(tmp_path), "stale.xlsx")
    wb = load_workbook(src, preserve=True)
    wb["Data"].pivots.create(
        name="ByRegion", source="RegionTable", destination="E3",
        rows=["Region"], values=["Amount"])
    wb["Data"]["B5"] = 99
    with pytest.raises(UnsupportedStructureError) as exc:
        wb.save(dest)
    assert exc.value.kind == "stale-pivot"
    assert not os.path.exists(dest)
    with open(src, "rb") as handle:
        original = handle.read()
    assert open(src, "rb").read() == original


def test_interrupt_before_each_commit_point(fixture_copy):
    src = fixture_copy(_TABLE)
    for point in ("start", "validated", "planned", "built", "cells", "ledger"):
        wb, before = _load(src)

        def hook(name, workbook, expected=point):
            if name == expected:
                raise RuntimeError("injected %s" % expected)

        create_mod.CREATE_CHECKPOINT = hook
        try:
            with pytest.raises(RuntimeError):
                wb["Data"].pivots.create(
                    name="ByRegion", source="RegionTable", destination="E3",
                    rows=["Region"], values=["Amount"])
        finally:
            create_mod.CREATE_CHECKPOINT = None
        _assert_untouched(wb, before, src)


def test_existing_part_collision_refuses_at_save(fixture_copy, tmp_path):
    from openpyxl.preserve.lifecycle import PartPlan
    from openpyxl.errors import RelationshipPolicyError

    plan = PartPlan({"xl/pivotTables/pivotTable1.xml"})
    with pytest.raises(RelationshipPolicyError):
        plan.add_part("xl/pivotTables/pivotTable1.xml", b"<x/>")
    src = fixture_copy(_TABLE)
    wb, before = _load(src)
    wb["Data"].pivots.create(
        name="ByRegion", source="RegionTable", destination="E3",
        rows=["Region"], values=["Amount"])
    operation = next(iter(wb._paper_ledger.pivot_operations.values()))
    # Force the staged allocation onto a name that already exists in the
    # source package so save-time add_part refuses.
    object.__setattr__(
        operation.allocation, "pivot_part", "xl/tables/table1.xml")
    dest = os.path.join(str(tmp_path), "collision.xlsx")
    with pytest.raises(RelationshipPolicyError):
        wb.save(dest)
    assert not os.path.exists(dest)
    with open(src, "rb") as handle:
        assert handle.read() == before["source"]
