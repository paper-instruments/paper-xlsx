"""Supported perception, validation, receipt, and diff APIs."""
from __future__ import annotations

import io
import zipfile

import pytest

from openpyxl import Workbook, load_workbook
from openpyxl.errors import (
    AmbiguousTargetError,
    TargetNotFoundError,
    UnsupportedStructureError,
)


class TestSearchAndScan:

    def test_search_values_and_formulas(self, fixture_copy):
        wb = load_workbook(fixture_copy("features/schedule.xlsx"))
        hits = wb.search("Grand")
        assert {"address": "Summary!A1", "match": "Grand",
                "kind": "value"} in hits
        formula_hits = wb.search("Schedule!", values=False)
        assert all(h["kind"] == "formula" for h in formula_hits)
        assert formula_hits

    def test_search_regex(self, fixture_copy):
        wb = load_workbook(fixture_copy("features/schedule.xlsx"))
        hits = wb.search(r"Item \d", regex=True, formulas=False)
        assert len(hits) >= 2

    def test_scan_errors_sees_cached_and_formula_refs(self, fixture_copy,
                                                      tmp_path):
        from openpyxl.preserve import scan_errors

        src = fixture_copy("features/schedule.xlsx")
        crafted = str(tmp_path / "err.xlsx")
        with zipfile.ZipFile(src) as zin, \
                zipfile.ZipFile(crafted, "w") as zout:
            for name in zin.namelist():
                payload = zin.read(name)
                if name == "xl/worksheets/sheet1.xml":
                    payload = payload.replace(
                        b"</sheetData>",
                        b'<row r="30"><c r="A30" t="e"><f>1/0</f>'
                        b"<v>#DIV/0!</v></c></row></sheetData>", 1)
                zout.writestr(name, payload)
        wb = load_workbook(crafted, preserve=True)
        wb["Summary"]["C1"] = "=#REF!+1"
        results = scan_errors(wb)
        sources = {r["source"] for r in results}
        assert {"cache", "formula"} <= sources

    def test_scan_errors_uses_formula_operands_not_string_substrings(self):
        from openpyxl.preserve import scan_errors

        wb = Workbook()
        ws = wb.active
        ws.title = "Model"
        ws["A1"] = '=IF(B1="#REF!", 1, 0)'
        ws["A2"] = "=#REF!+1"
        ws["A3"] = "=Other!#REF!"
        ws["A4"] = "=#SPILL!"
        results = scan_errors(wb)
        assert results == [
            {"address": "Model!A2", "value": "#REF!",
             "source": "formula"},
            {"address": "Model!A3", "value": "#REF!",
             "source": "formula"},
            {"address": "Model!A4", "value": "#SPILL!",
             "source": "formula"},
        ]

    def test_scan_errors_refuses_unscannable_formula_without_partial_report(
            self):
        from openpyxl.preserve import scan_errors

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "=#REF!"
        ws["A2"] = '="unterminated'

        with pytest.raises(UnsupportedStructureError,
                           match="No partial formula-error report"):
            scan_errors(wb)

    def test_scan_errors_normalizes_tokenizer_index_error(self):
        from openpyxl.preserve import scan_errors

        wb = Workbook()
        wb.active["A1"] = "=)"

        with pytest.raises(
                UnsupportedStructureError,
                match="No partial formula-error report") as caught:
            scan_errors(wb)
        assert caught.value.kind == "unscannable-formula"
        assert caught.value.anchor == "Sheet!A1"

    def test_allowed_values_literal_and_range(self, fixture_copy):
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = load_workbook(fixture_copy("features/schedule.xlsx"))
        ws = wb["Summary"]
        dv = DataValidation(type="list", formula1='"Yes,No,Maybe"')
        dv.add("D1")
        ws.add_data_validation(dv)
        assert ws.allowed_values("D1") == ["Yes", "No", "Maybe"]
        dv2 = DataValidation(type="list",
                             formula1="=Schedule!$A$2:$A$4")
        dv2.add("D2")
        ws.add_data_validation(dv2)
        assert ws.allowed_values(ws["D2"]) == ["Item 1", "Item 2",
                                               "Item 3"]
        assert ws.allowed_values("E9") is None

    def test_allowed_values_preserves_literal_text_and_blanks(self):
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Yes"
        ws["A3"] = " No "
        literal = DataValidation(
            type="list", formula1='" Yes,No ,A""B,"')
        literal.add("D1")
        ws.add_data_validation(literal)
        ranged = DataValidation(type="list", formula1="=$A$1:$A$3")
        ranged.add("D2")
        ws.add_data_validation(ranged)
        assert ws.allowed_values("D1") == [" Yes", "No ", 'A"B', ""]
        assert ws.allowed_values("D2") == ["Yes", None, " No "]

    def test_allowed_values_handles_quoted_cross_sheet_and_reversed_ranges(
            self):
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = Workbook()
        ws = wb.active
        source = wb.create_sheet("Owner's Inputs")
        source["A1"] = "Low"
        source["A2"] = "High"
        validation = DataValidation(
            type="list", formula1="='Owner''s Inputs'!$A$2:$A$1")
        validation.add("B1")
        ws.add_data_validation(validation)

        assert ws.allowed_values("B1") == ["Low", "High"]

    def test_allowed_values_refuses_unsupported_or_ambiguous_sources(self):
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "=1+1"
        cases = {
            "D1": "=ModelInputs",
            "D2": "=Missing!$A$1:$A$2",
            "D3": "=$A$1:$A$2",
            "D4": "=$A$1:$B$2",
        }
        for target, formula in cases.items():
            validation = DataValidation(type="list", formula1=formula)
            validation.add(target)
            ws.add_data_validation(validation)
        with pytest.raises(UnsupportedStructureError, match="defined name"):
            ws.allowed_values("D1")
        with pytest.raises(TargetNotFoundError, match="missing worksheet"):
            ws.allowed_values("D2")
        with pytest.raises(UnsupportedStructureError, match="formula cell"):
            ws.allowed_values("D3")
        with pytest.raises(UnsupportedStructureError,
                           match="two-dimensional"):
            ws.allowed_values("D4")

        first = DataValidation(type="list", formula1='"A,B"')
        second = DataValidation(type="list", formula1='"C,D"')
        first.add("E1")
        second.add("E1")
        ws.add_data_validation(first)
        ws.add_data_validation(second)
        with pytest.raises(AmbiguousTargetError,
                           match="covered by 2 list validations"):
            ws.allowed_values("E1")

        ws.merge_cells("F1:G1")
        merged = DataValidation(type="list", formula1='"A,B"')
        merged.add("G1")
        ws.add_data_validation(merged)
        with pytest.raises(UnsupportedStructureError,
                           match="non-anchor merged cell"):
            ws.allowed_values("G1")


class TestValidateAndReceipt:

    def test_validate_raises_what_save_would(self, fixture_copy):
        wb = load_workbook(fixture_copy("features/chart_image.xlsx"),
                           preserve=True)
        wb["Model"]._charts[0].style = 31     # inexpressible mutation
        with pytest.raises(UnsupportedStructureError, match="style"):
            wb.validate()

    def test_validate_clean_session_returns_none(self, fixture_copy):
        wb = load_workbook(fixture_copy("features/schedule.xlsx"),
                           preserve=True)
        wb["Schedule"]["A2"] = "edited"
        assert wb.validate() is None

    def test_save_receipt(self, fixture_copy, tmp_path):
        from openpyxl.preserve.receipts import EditReceipt

        wb = load_workbook(fixture_copy("features/schedule.xlsx"),
                           preserve=True)
        wb["Schedule"]["A2"] = "renamed item"
        out = str(tmp_path / "o.xlsx")
        result = wb.save(out, receipt=True)
        assert isinstance(result, EditReceipt)
        payload = result.to_dict()
        assert payload["schema"] == "edit_receipt"
        assert payload["version"] == 2
        sheet_part = next(iter(result.cells_changed))
        assert result.cells_changed[sheet_part] == {"A2": "changed"}
        assert sheet_part in result.parts_changed

    def test_receipt_requires_preserve(self, fixture_copy, tmp_path):
        wb = load_workbook(
            fixture_copy("features/schedule.xlsx"), preserve=False)
        with pytest.raises(ValueError, match="preserve"):
            wb.save(str(tmp_path / "o.xlsx"), receipt=True)


class TestDiffWorkbooks:

    def test_content_vs_shifted(self, fixture_copy, tmp_path):
        from openpyxl.preserve import diff_workbooks

        src = fixture_copy("features/schedule.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb["Schedule"]
        remap = ws.insert_rows(1)             # AddressRemap
        ws["B3"] = 999                        # content change (was B2)
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        report = diff_workbooks(src, out, remaps=[remap])
        payload = report.to_dict()
        assert payload["schema"] == "workbook_diff"
        changed_addrs = {e["address"] for e in report.changed}
        assert "Schedule!B2" in changed_addrs  # the value edit
        # everything else moved intact: classified shifted, not changed
        shifted_from = {e["from"] for e in report.shifted}
        assert "Schedule!B4" in shifted_from or "Schedule!A4" in \
            shifted_from
        assert not any(e["address"].startswith("Schedule!A")
                       and e["address"] != "Schedule!A1"
                       for e in report.changed
                       if e["before"] is not None
                       and e["after"] is not None)

    def test_sheet_membership(self, fixture_copy, tmp_path):
        from openpyxl.preserve import diff_workbooks

        src = fixture_copy("features/schedule.xlsx")
        wb = load_workbook(src, preserve=True)
        ws = wb.create_sheet("New")
        ws["A1"] = 1
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        report = diff_workbooks(src, out)
        assert report.added_sheets == ["New"]
        assert report.removed_sheets == []
