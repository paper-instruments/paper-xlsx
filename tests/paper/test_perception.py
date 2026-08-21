"""Perception helpers for cell diffs and dependency sketches."""
from __future__ import annotations

from openpyxl import Workbook, load_workbook
from openpyxl.package import diff_cells
from openpyxl.preserve.perception import dependency_sketch
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.table import Table, TableColumn

class TestDiffCells:

    def test_value_and_formula_changes(self, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Schedule"]["B2"] = 999
        wb["Schedule"]["B20"] = "=SUM(B2:B4)"
        out = str(tmp_path / "b.xlsx")
        wb.save(out)
        d = diff_cells(src, out)
        by_addr = {c["address"]: c for c in d.changes}
        assert by_addr["'Schedule'!B2"]["old_value"] == 200
        assert by_addr["'Schedule'!B2"]["new_value"] == 999
        assert by_addr["'Schedule'!B20"]["new_formula"] == "=SUM(B2:B4)"
        assert not d.sheets_added and not d.sheets_removed

    def test_identical_files_are_clean(self, fixture_copy):
        a = fixture_copy("features/schedule.xlsx", "a.xlsx")
        b = fixture_copy("features/schedule.xlsx", "b.xlsx")
        assert diff_cells(a, b).clean

    def test_added_sheet_reported(self, fixture_copy, tmp_path):
        src = fixture_copy("minimal/minimal_clean.xlsx")
        wb = load_workbook(src, preserve=True)
        wb.create_sheet("Extra")["A1"] = 1
        out = str(tmp_path / "b.xlsx")
        wb.save(out)
        d = diff_cells(src, out)
        assert d.sheets_added == ["Extra"]

    def test_to_dict_schema(self, fixture_copy):
        a = fixture_copy("minimal/minimal_clean.xlsx")
        doc = diff_cells(a, a).to_dict()
        assert doc["schema"] == "cells_diff" and doc["version"] == 1


class TestDependencySketch:

    def test_known_edges(self, fixture_copy):
        wb = load_workbook(fixture_copy("gauntlet/gauntlet.xlsx"),
                           preserve=True)
        sk = dependency_sketch(wb)
        doc = sk.to_dict()
        assert doc["references"]["'Model'!B6"] == ["B3", "B4", "B5"]
        assert doc["references"]["'Model'!B12"] == ["Data!B2:B5"]

    def test_intersection_query_cross_sheet(self, fixture_copy):
        wb = load_workbook(fixture_copy("gauntlet/gauntlet.xlsx"),
                           preserve=True)
        sk = dependency_sketch(wb)
        # who references Data!B2:B5?
        assert sk.cells_referencing("Data", (2, 2, 2, 5)) == ["'Model'!B12"]
        # nobody references Data column D
        assert sk.cells_referencing("Data", (4, 1, 4, 100)) == []

    def test_defined_name_expansion(self, fixture_copy):
        wb = load_workbook(fixture_copy("features/schedule.xlsx"),
                           preserve=True)
        sk = dependency_sketch(wb)
        # B13 = B12*(1+Growth); Growth -> Schedule!$B$15
        hits = sk.cells_referencing("Schedule", (2, 15, 2, 15))
        assert "'Schedule'!B13" in hits

    def test_common_structured_refs_resolve_to_table_ranges(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["Name", "Amount", "Tax"])
        ws.append(["A", 10, 1])
        ws.append(["B", 20, 2])
        ws.append(["C", 30, 3])
        ws.append(["Total", 60, 6])
        table = Table(displayName="Table1", ref="A1:C5")
        table.totalsRowCount = 1
        ws.add_table(table)
        ws["E1"] = "=SUM(Table1[Amount])"
        ws["E2"] = "=SUM(Table1[[#Data],[Amount]])"
        ws["E3"] = "=SUM(Table1[[Amount]:[Tax]])"
        ws["E4"] = "=SUM(Table1[#Headers])"
        ws["E5"] = "=SUM(Table1[#Totals])"
        ws["E6"] = "=SUM(Table1[#All])"
        ws["E7"] = "=SUM(Table1[[#Headers],[Amount]:[Tax]])"
        ws["E8"] = "=SUM(tAbLe1[amount])"
        path = tmp_path / "structured.xlsx"
        wb.save(path)
        wb = load_workbook(path, preserve=True)

        sk = dependency_sketch(wb)
        refs = {
            address: bounds
            for address, [(_sheet, bounds, _raw)] in sk.references.items()
        }
        assert refs["'Data'!E1"] == (2, 2, 2, 4)
        assert refs["'Data'!E2"] == (2, 2, 2, 4)
        assert refs["'Data'!E3"] == (2, 2, 3, 4)
        assert refs["'Data'!E4"] == (1, 1, 3, 1)
        assert refs["'Data'!E5"] == (1, 5, 3, 5)
        assert refs["'Data'!E6"] == (1, 1, 3, 5)
        assert refs["'Data'!E7"] == (2, 1, 3, 1)
        assert refs["'Data'!E8"] == (2, 2, 2, 4)
        assert sk.unresolved == {}

    def test_current_row_structured_refs_resolve_exact_row(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Amount", "Calc"])
        ws.append([10, "=[@Amount]"])
        ws.append([20, "=Table1[@Amount]"])
        ws.add_table(Table(displayName="Table1", ref="A1:B3"))
        path = tmp_path / "current-row.xlsx"
        wb.save(path)
        wb = load_workbook(path, preserve=True)

        sk = dependency_sketch(wb)
        refs = {
            address: bounds
            for address, [(_sheet, bounds, _raw)] in sk.references.items()
        }
        assert refs["'Sheet'!B2"] == (1, 2, 1, 2)
        assert refs["'Sheet'!B3"] == (1, 3, 1, 3)
        assert sk.cells_referencing("Sheet", (1, 3, 1, 3)) == ["'Sheet'!B3"]

    def test_external_workbook_reference_stays_unresolved(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "=[Book.xlsx]Sheet!A1"

        sk = dependency_sketch(wb)
        assert sk.to_dict()["unresolved"] == {
            "'Sheet'!A1": ["[Book.xlsx]Sheet!A1"]
        }
        assert "'Sheet'!A1" in sk.cells_referencing("Anywhere", (1, 1, 1, 1))

    def test_current_row_in_multi_cell_formula_stays_unresolved(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Amount", "Calc"])
        ws.append([10, None])
        ws.append([20, None])
        ws.add_table(Table(displayName="Table1", ref="A1:B3"))
        ws["B2"] = ArrayFormula(ref="B2:B3", text="=[@Amount]")

        sk = dependency_sketch(wb)

        assert sk.to_dict()["unresolved"] == {
            "'Sheet'!B2": ["[@Amount]"]
        }

    def test_unproven_structured_refs_stay_unresolved(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Amount"])
        ws.append([10])
        ws.add_table(Table(displayName="Only", ref="A1:A2"))
        ws["C1"] = "=Missing[Amount]"
        ws["C2"] = "=Only[Missing]"

        duplicate = Table(displayName="DupCols", ref="A1:B2",
                          tableColumns=[
                              TableColumn(id=1, name="Amount"),
                              TableColumn(id=2, name="amount"),
                          ])
        ws["B1"] = "amount"
        ws._tables.add(duplicate)
        ws["C3"] = "=DupCols[Amount]"
        ws["C5"] = "=Only[[#Headers],[#Data],[Amount]]"

        empty = wb.create_sheet("Empty")
        empty["A1"] = "Amount"
        empty.add_table(Table(displayName="EmptyTable", ref="A1:A1"))
        ws["C6"] = "=EmptyTable[Missing]"

        other = wb.create_sheet("Other")
        other.append(["Amount"])
        other.append([20])
        ws["D1"] = "Amount"
        ws["D2"] = 10
        ws._tables.add(Table(displayName="Table1", ref="D1:D2",
                             tableColumns=[
                                 TableColumn(id=1, name="Amount"),
                             ]))
        other._tables.add(Table(displayName="table1", ref="A1:A2",
                                tableColumns=[
                                    TableColumn(id=1, name="Amount"),
                                ]))
        ws["C4"] = "=Table1[Amount]"

        sk = dependency_sketch(wb)
        assert sk.to_dict()["unresolved"] == {
            "'Sheet'!C1": ["Missing[Amount]"],
            "'Sheet'!C2": ["Only[Missing]"],
            "'Sheet'!C3": ["DupCols[Amount]"],
            "'Sheet'!C4": ["Table1[Amount]"],
            "'Sheet'!C5": ["Only[[#Headers],[#Data],[Amount]]"],
            "'Sheet'!C6": ["EmptyTable[Missing]"],
        }
