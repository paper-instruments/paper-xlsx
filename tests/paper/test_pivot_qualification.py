"""Pivot validity and per-operation capability qualification."""
from __future__ import annotations

import json
import io
import os
import zipfile

import pytest

from openpyxl import load_workbook
from openpyxl.errors import UnsupportedStructureError
from openpyxl.pivot.api_types import PivotCapabilities
from openpyxl.pivot.graph import load_pivot_graph
from openpyxl.pivot.inspect import project_pivot
from openpyxl.pivot.qualify import PAPER_TAG, qualify_pivot

from .conftest import FIXTURES_DIR
from .test_pivot_graph import _basic_package, _sidecar_binaries, _write_package


_PIVOTS = os.path.join(FIXTURES_DIR, "pivots")


def _qualify(payload, workbook=None, ownership_proved=False, name=None):
    graph = load_pivot_graph(payload, workbook=workbook)
    nodes = list(graph.pivots)
    if name is not None:
        nodes = [node for node in nodes if node.identity.name == name]
    node = nodes[0]
    cache = graph.caches_by_part.get(node.cache_definition_part)
    projection = project_pivot(node, cache, source=payload, workbook=workbook)
    return qualify_pivot(
        node, cache, projection, graph, workbook=workbook,
        ownership_proved=ownership_proved), projection, node, cache


def _codes(qualification, capability=None):
    return [
        item.code for item in qualification.reasons
        if capability is None or item.capability == capability
    ]


def _rewrite_payload(payload, part, transform):
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(payload)) as before, \
            zipfile.ZipFile(output, "w") as after:
        for info in before.infolist():
            body = before.read(info.filename)
            if info.filename == part:
                body = transform(body)
            after.writestr(info, body)
    return output.getvalue()


def test_foreign_valid_pivot_receives_only_refresh_on_open():
    qualification, projection, node, cache = _qualify(_basic_package())
    assert qualification.origin == "foreign"
    assert qualification.valid is True
    assert projection.complete is True
    assert qualification.capabilities == PivotCapabilities(
        can_refresh_on_open=True)
    assert qualification.refresh_on_open_scope == ("Summary!SalesByRegion",)
    assert "foreign-operation-deferred" in _codes(
        qualification, "can_edit_layout")
    assert "foreign-operation-deferred" in _codes(
        qualification, "can_delete")


def test_paper_marker_does_not_grant_ownership_dependent_caps():
    qualification, projection, node, _cache = _qualify(
        _basic_package(tag=PAPER_TAG))
    assert qualification.origin == "paper"
    assert qualification.valid is True
    assert projection.complete is True
    assert qualification.capabilities.can_refresh_on_open is True
    assert qualification.capabilities.can_edit_layout is False
    assert qualification.capabilities.can_rename is True
    assert qualification.capabilities.can_headless_refresh is False
    assert qualification.capabilities.can_delete is False
    assert "output-ownership-unproved" in _codes(
        qualification, "can_delete")
    assert "output-ownership-unproved" in _codes(
        qualification, "can_edit_layout")
    assert node.tag == PAPER_TAG


def test_duplicate_external_worksheet_rid_invalidates_affected_pivot():
    payload = _basic_package(tag=PAPER_TAG)

    def add_duplicate(body):
        closing = b"</Relationships>"
        duplicate = (
            b'<Relationship Id="rIdPivot1" Type="urn:external" '
            b'Target="https://example.invalid/pivot" '
            b'TargetMode="External"/>'
        )
        assert closing in body
        return body.replace(closing, duplicate + closing, 1)

    payload = _rewrite_payload(
        payload, "xl/worksheets/_rels/sheet1.xml.rels", add_duplicate)
    qualification = _qualify(payload, ownership_proved=True)[0]
    assert qualification.valid is False
    assert qualification.capabilities.can_edit_layout is False
    assert "duplicate-relationship-id" in _codes(qualification)


def test_foreign_namespace_cannot_define_workbook_cache_registry():
    payload = _basic_package(tag=PAPER_TAG)

    def replace_registry(body):
        assert b"<pivotCaches>" in body
        return body.replace(
            b"<pivotCaches>",
            b'<evil:pivotCaches xmlns:evil="urn:paper-test:evil">',
            1,
        ).replace(b"</pivotCaches>", b"</evil:pivotCaches>", 1)

    payload = _rewrite_payload(payload, "xl/workbook.xml", replace_registry)
    qualification = _qualify(payload, ownership_proved=True)[0]
    assert qualification.valid is False
    assert qualification.capabilities.can_edit_layout is False


def test_foreign_namespace_cannot_define_relationship_nodes():
    payload = _basic_package(tag=PAPER_TAG)

    def replace_relationship(body):
        marker = b'<Relationship Id="rIdPivot1"'
        replacement = (
            b'<evil:Relationship xmlns:evil="urn:paper-test:evil" '
            b'Id="rIdPivot1"'
        )
        assert marker in body
        return body.replace(marker, replacement, 1)

    payload = _rewrite_payload(
        payload, "xl/worksheets/_rels/sheet1.xml.rels",
        replace_relationship)
    graph = load_pivot_graph(payload)
    assert not graph.pivots
    assert any(
        reason.code == "unexpected-namespace"
        for reason in graph.reasons
    )


def test_missing_field_unknown_aggregate_and_bad_item_index():
    missing = _qualify(_basic_package(tag=PAPER_TAG, rows=(5,)))[0]
    assert missing.capabilities.can_edit_layout is False
    assert missing.capabilities.can_rename is False
    assert "missing-field" in _codes(missing, "can_edit_layout")
    assert "missing-field" in _codes(missing, "can_rename")

    unknown = _qualify(_basic_package(
        tag=PAPER_TAG,
        values=(("Amount", 1, "product"),)))[0]
    assert "unknown-aggregate" in _codes(unknown, "can_edit_layout")
    assert unknown.capabilities.can_headless_refresh is False

    bad_item = _qualify(_basic_package(
        tag=PAPER_TAG, pages=((0, 99),)))[0]
    assert "invalid-item-index" in _codes(bad_item, "can_edit_layout")


def test_grouping_extension_and_data_model_disable_mutation():
    grouping = _qualify(_basic_package(tag=PAPER_TAG, grouping=True))[0]
    assert grouping.origin == "paper"
    assert "unsupported-grouping" in _codes(grouping, "can_edit_layout")
    assert grouping.capabilities.can_rename is False
    assert "unsupported-grouping" in _codes(grouping, "can_rename")
    assert grouping.capabilities.can_refresh_on_open is True

    extension = _qualify(_basic_package(
        tag=PAPER_TAG,
        ext_uri="{725AE2AE-9491-48be-B2B4-4EB974FC3084}",
    ))[0]
    assert "unsupported-extension" in _codes(extension, "can_edit_layout")
    assert extension.capabilities.can_refresh_on_open is True

    data_model = _qualify(_basic_package(
        tag=PAPER_TAG,
        source={"kind": "external"},
    ))[0]
    assert data_model.source_supported is False
    assert any(code in ("unsupported-data-model", "unsupported-source")
               for code in _codes(data_model))
    assert data_model.capabilities.can_edit_layout is False


def test_shared_cache_disables_isolation_sensitive_caps():
    qualification, _projection, _node, cache = _qualify(_basic_package(
        tag=PAPER_TAG,
        extra_pivots=(
            ("MarginByRegion", "xl/pivotTables/pivotTable2.xml",
             "rIdPivot2", "1"),
        ),
    ), name="SalesByRegion")
    assert cache.referenced_by == (
        ("Summary", "MarginByRegion"),
        ("Summary", "SalesByRegion"),
    )
    assert qualification.cache_shared is True
    assert qualification.capabilities.can_refresh_on_open is True
    assert qualification.capabilities.can_edit_layout is False
    assert qualification.capabilities.can_headless_refresh is False
    assert qualification.capabilities.can_move is False
    assert qualification.refresh_on_open_scope == (
        "Summary!MarginByRegion",
        "Summary!SalesByRegion",
    )
    assert "pivot-cache-shared" in _codes(qualification, "can_delete")
    assert "pivot-cache-shared" in _codes(
        qualification, "can_repoint_source")
    assert "pivot-cache-shared" in _codes(
        qualification, "can_edit_layout")


def test_broken_graph_disables_every_capability():
    qualification, _projection, node, _cache = _qualify(
        _basic_package(dangling_pivot=True))
    assert node.valid is False
    assert qualification.valid is False
    assert qualification.capabilities == PivotCapabilities()
    assert qualification.capabilities.can_refresh_on_open is False
    assert any(item.capability is None for item in qualification.reasons)


def test_duplicate_header_is_an_unsupported_source(tmp_path):
    path = _write_package(tmp_path, "headers.xlsx", _basic_package())
    wb = load_workbook(path, preserve=True)
    data = wb["Data"]
    data["A1"] = "Region"
    data["B1"] = "region"
    qualification, projection, _node, _cache = _qualify(
        wb._paper_source, workbook=wb)
    assert projection.source.kind == "range"
    assert qualification.source_supported is False
    assert "duplicate-header" in _codes(qualification)


def test_invalid_lookup_returns_bounded_invalid_handle(tmp_path):
    path = _write_package(tmp_path, "broken.xlsx", _basic_package())
    wb = load_workbook(path, preserve=True)
    wb._paper_source = _basic_package(
        extra_pivots=(
            ("Broken", "xl/pivotTables/pivotTable2.xml",
             "rIdPivot2", "99"),
        ),
    )
    from openpyxl.pivot.api import invalidate_pivot_overlay
    invalidate_pivot_overlay(wb)
    broken = wb["Summary"].pivots["Broken"]
    payload = broken.to_dict()
    assert payload["valid"] is False
    assert payload["name"] == "Broken"
    assert "source" not in payload or payload.get("values") in (None, [])
    assert payload["capabilities"]["can_refresh_on_open"] is False
    assert payload["schema"] == "pivot_table"
    assert "records" not in json.dumps(payload)


@pytest.mark.parametrize("filename", _sidecar_binaries())
def test_real_fixture_capabilities_match_sidecar(filename):
    path = os.path.join(_PIVOTS, filename)
    with open(path + ".json") as handle:
        sidecar = json.load(handle)
    expected = sidecar.get("expected_pivot_qualification")
    if expected is None:
        pytest.skip("sidecar has no expected_pivot_qualification")
    wb = load_workbook(path, preserve=True)
    observed = {}
    for sheet in wb.worksheets:
        try:
            collection = sheet.pivots
        except UnsupportedStructureError:
            continue
        for pivot in collection:
            observed["%s!%s" % (sheet.title, pivot.name)] = {
                "valid": pivot.valid,
                "origin": pivot.origin,
                "capabilities": pivot.capabilities.to_dict(),
                "reasons": [item.to_dict() for item in pivot.qualification_reasons],
            }
    assert observed == expected
