"""Honesty organs — the data_only trap, recalc-on-load, format
refusals."""
from __future__ import annotations

import re
import warnings
import zipfile

import pytest

from openpyxl import Workbook, load_workbook
from openpyxl.errors import (
    LossySaveWarning,
    PaperRefusal,
    UnsupportedStructureError,
)
from openpyxl.utils.exceptions import InvalidFileException

from .support.partdiff import part_payloads


class TestDataOnlyTrap:

    def test_preserve_data_only_save_refuses_naming_the_override(
            self, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule_calc.xlsx")
        with open(src, "rb") as f:
            before = f.read()
        wb = load_workbook(src, preserve=True, data_only=True)
        with pytest.raises(UnsupportedStructureError,
                           match="allow_formula_loss"):
            wb.save(str(tmp_path / "o.xlsx"))
        with open(src, "rb") as f:
            assert f.read() == before

    def test_override_cannot_bypass_formula_protection(
            self, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule_calc.xlsx")
        wb = load_workbook(src, preserve=True, data_only=True)
        with pytest.raises(PaperRefusal, match="retained source"):
            wb["Schedule"]["B12"] = 9999

        assert wb["Schedule"]["B12"].value == 6500

    def test_noop_data_only_save_with_override_is_byte_identical(
            self, fixture_copy, tmp_path):
        from openpyxl.package import diff_package

        src = fixture_copy("features/schedule_calc.xlsx")
        wb = load_workbook(src, preserve=True, data_only=True)
        out = str(tmp_path / "o.xlsx")
        wb.save(out, allow_formula_loss=True)
        assert diff_package(src, out).clean

    def test_stock_data_only_save_warns(self, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule_calc.xlsx")
        wb = load_workbook(src, data_only=True, preserve=False)
        with pytest.warns(LossySaveWarning, match="PERMANENTLY replaces"):
            wb.save(str(tmp_path / "o.xlsx"))

    def test_stock_data_only_warning_silenced_by_override(
            self, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule_calc.xlsx")
        wb = load_workbook(src, data_only=True, preserve=False)
        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            wb.save(str(tmp_path / "o.xlsx"), allow_formula_loss=True)
        assert not [w for w in caught
                    if isinstance(w.message, LossySaveWarning)
                    and "PERMANENTLY" in str(w.message)]


class TestRecalcOnLoad:

    def _cached_formula_workbook(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Model"
        ws["A1"] = 1
        ws["A2"] = 2
        ws["B1"] = "=A1+A2"
        ws["B2"] = "=B1*2"
        summary = wb.create_sheet("Summary")
        summary["A1"] = "=Model!B1*3"
        raw = str(tmp_path / "raw.xlsx")
        wb.save(raw)
        out = str(tmp_path / "cached.xlsx")

        def set_cache(payload, coord, formula, value):
            pattern = (
                br'(<c r="' + coord + br'"[^>]*><f>' + formula
                + br'</f>)(?:<v(?:\s[^>]*)?/>|<v(?:\s[^>]*)?>.*?</v>)'
                + br'(</c>)')
            replacement = br'\1<v>' + value + br'</v>\2'
            payload, count = re.subn(
                pattern, replacement, payload, count=1, flags=re.S)
            assert count == 1
            return payload

        with zipfile.ZipFile(raw) as zin, zipfile.ZipFile(out, "w") as zout:
            for info in zin.infolist():
                payload = zin.read(info.filename)
                if info.filename == "xl/workbook.xml":
                    payload = payload.replace(b' fullCalcOnLoad="1"', b"")
                    payload = payload.replace(b' forceFullCalc="1"', b"")
                if info.filename == "xl/worksheets/sheet1.xml":
                    payload = set_cache(payload, b"B1", b"A1\\+A2", b"3")
                    payload = set_cache(payload, b"B2", b"B1\\*2", b"6")
                if info.filename == "xl/worksheets/sheet2.xml":
                    payload = set_cache(
                        payload, b"A1", b"Model!B1\\*3", b"9")
                zout.writestr(info, payload)
        return out

    def test_formula_edit_sets_full_calc_on_load(self, fixture_copy, tmp_path):
        # schedule_calc is LibreOffice-written: its calcPr lacks the flag
        src = fixture_copy("features/schedule_calc.xlsx")
        assert b"fullCalcOnLoad" not in part_payloads(src)["xl/workbook.xml"]
        wb = load_workbook(src, preserve=True)
        wb["Schedule"]["B20"] = "=SUM(B2:B3)"
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        assert b'fullCalcOnLoad="1"' in part_payloads(out)["xl/workbook.xml"]

    def test_formula_edit_invalidates_preserved_formula_caches(
            self, tmp_path):
        src = self._cached_formula_workbook(tmp_path)
        assert b"fullCalcOnLoad" not in part_payloads(src)["xl/workbook.xml"]
        wb = load_workbook(src, preserve=True)
        wb["Model"]["B1"] = "=A1+A2+10"
        out = str(tmp_path / "o.xlsx")
        wb.save(out)

        parts = part_payloads(out)
        workbook_xml = parts["xl/workbook.xml"]
        assert b'calcMode="auto"' in workbook_xml
        assert b'fullCalcOnLoad="1"' in workbook_xml
        assert b'forceFullCalc="1"' in workbook_xml
        sheet = parts["xl/worksheets/sheet1.xml"]
        edited = re.search(br'<c r="B1".*?</c>', sheet, re.S).group(0)
        downstream = re.search(br'<c r="B2".*?</c>', sheet, re.S).group(0)
        assert b"<f>A1+A2+10</f>" in edited
        assert b"<v" not in edited
        assert b"<f>B1*2</f>" in downstream
        assert b"<v" not in downstream
        summary = parts["xl/worksheets/sheet2.xml"]
        assert b"<f>Model!B1*3</f>" in summary
        assert b"<v" not in summary

    def test_value_edit_feeding_formula_invalidates_preserved_caches(
            self, tmp_path):
        src = self._cached_formula_workbook(tmp_path)
        wb = load_workbook(src, preserve=True)
        wb["Model"]["A1"] = 10
        out = str(tmp_path / "o.xlsx")
        wb.save(out)

        parts = part_payloads(out)
        assert b'calcMode="auto"' in parts["xl/workbook.xml"]
        sheet = parts["xl/worksheets/sheet1.xml"]
        assert b"<f>A1+A2</f>" in sheet
        assert b"<f>B1*2</f>" in sheet
        assert b"<v>3</v>" not in sheet
        assert b"<v>6</v>" not in sheet
        assert b"<v>9</v>" not in parts["xl/worksheets/sheet2.xml"]

    def test_value_edit_invalidates_array_follower_caches(self, tmp_path):
        src = self._cached_formula_workbook(tmp_path)
        array_src = str(tmp_path / "array-cached.xlsx")
        with zipfile.ZipFile(src) as zin, \
                zipfile.ZipFile(array_src, "w") as zout:
            for info in zin.infolist():
                payload = zin.read(info.filename)
                if info.filename == "xl/worksheets/sheet1.xml":
                    payload = payload.replace(
                        b"<f>A1+A2</f><v>3</v>",
                        b'<f t="array" ref="B1:B3">A1+A2</f><v>3</v>')
                    payload = payload.replace(
                        b"<f>B1*2</f><v>6</v>", b"<v>6</v>")
                    payload = payload.replace(
                        b"</sheetData>",
                        b'<row r="3"><c r="B3"><v>9</v></c></row>'
                        b"</sheetData>")
                zout.writestr(info, payload)

        wb = load_workbook(array_src, preserve=True)
        wb["Model"]["A1"] = 10
        out = str(tmp_path / "array-output.xlsx")
        wb.save(out)

        sheet = part_payloads(out)["xl/worksheets/sheet1.xml"]
        for coordinate in (b"B1", b"B2", b"B3"):
            cell = re.search(
                br'<c r="' + coordinate + br'".*?</c>', sheet, re.S).group(0)
            assert b"<v" not in cell
        assert b'<f t="array" ref="B1:B3">A1+A2</f>' in sheet

    def test_style_only_precedent_edit_preserves_formula_caches(
            self, tmp_path):
        src = self._cached_formula_workbook(tmp_path)
        before = part_payloads(src)
        wb = load_workbook(src, preserve=True)
        cell = wb["Model"]["A1"]
        cell.font = cell.font.copy(bold=True)
        out = str(tmp_path / "style-only.xlsx")
        wb.save(out)

        after = part_payloads(out)
        assert after["xl/workbook.xml"] == before["xl/workbook.xml"]
        assert b"<f>A1+A2</f><v>3</v>" in \
            after["xl/worksheets/sheet1.xml"]
        assert b"<f>B1*2</f><v>6</v>" in \
            after["xl/worksheets/sheet1.xml"]
        assert b"<f>Model!B1*3</f><v>9</v>" in \
            after["xl/worksheets/sheet2.xml"]

    def test_calcpr_only_save_raw_copies_unspliceable_untouched_sheet(
            self, tmp_path):
        src = self._cached_formula_workbook(tmp_path)
        prefixed_src = str(tmp_path / "prefixed-root.xlsx")
        main = b"http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        with zipfile.ZipFile(src) as zin, \
                zipfile.ZipFile(prefixed_src, "w") as zout:
            for info in zin.infolist():
                payload = zin.read(info.filename)
                if info.filename == "xl/worksheets/sheet2.xml":
                    payload = payload.replace(
                        b"<worksheet ",
                        b'<m:worksheet xmlns:m="' + main + b'" ', 1)
                    payload = payload.replace(
                        b"</worksheet>", b"</m:worksheet>", 1)
                zout.writestr(info, payload)

        before = part_payloads(prefixed_src)
        assert b"<m:worksheet" in before["xl/worksheets/sheet2.xml"]
        wb = load_workbook(prefixed_src, preserve=True)
        wb["Model"]["B1"] = "=A1+A2+1"
        out = str(tmp_path / "prefixed-root-output.xlsx")
        wb.save(out)

        after = part_payloads(out)
        assert after["xl/worksheets/sheet2.xml"] == \
            before["xl/worksheets/sheet2.xml"]
        assert b'fullCalcOnLoad="1"' in after["xl/workbook.xml"]
        assert b'forceFullCalc="1"' in after["xl/workbook.xml"]

    def test_unrelated_value_edit_preserves_formula_caches(self, tmp_path):
        src = self._cached_formula_workbook(tmp_path)
        before = part_payloads(src)
        wb = load_workbook(src, preserve=True)
        wb["Model"]["C1"] = "note"
        out = str(tmp_path / "o.xlsx")
        wb.save(out)

        after = part_payloads(out)
        assert after["xl/workbook.xml"] == before["xl/workbook.xml"]
        sheet = after["xl/worksheets/sheet1.xml"]
        assert b"<f>A1+A2</f><v>3</v>" in sheet
        assert b"<f>B1*2</f><v>6</v>" in sheet
        assert b"<f>Model!B1*3</f><v>9</v>" in \
            after["xl/worksheets/sheet2.xml"]

    def test_value_only_edit_does_not_touch_workbook_xml(
            self, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule_calc.xlsx")
        wb = load_workbook(src, preserve=True)
        wb["Schedule"]["A1"] = "renamed"            # no formula involved
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        before = part_payloads(src)["xl/workbook.xml"]
        after = part_payloads(out)["xl/workbook.xml"]
        assert before == after

    def test_formula_deletion_also_sets_the_flag(self, fixture_copy, tmp_path):
        src = fixture_copy("features/schedule_calc.xlsx")
        wb = load_workbook(src, preserve=True)
        del wb["Schedule"]["B13"]                   # deletes a formula cell
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        assert b'fullCalcOnLoad="1"' in part_payloads(out)["xl/workbook.xml"]


class TestFormatRefusals:

    def test_preserve_xls_refuses_with_conversion_hint(self, fixture_copy):
        src = fixture_copy("legacy/legacy.xls")
        with pytest.raises(UnsupportedStructureError, match="LibreOffice"):
            load_workbook(src, preserve=True)

    def test_preserve_xlsb_refuses_with_conversion_hint(self, fixture_copy):
        src = fixture_copy("legacy/binary.xlsb")
        with pytest.raises(UnsupportedStructureError, match="convert"):
            load_workbook(src, preserve=True)

    def test_refusals_are_paper_refusals(self, fixture_copy):
        with pytest.raises(PaperRefusal):
            load_workbook(fixture_copy("legacy/legacy.xls"), preserve=True)

    def test_stock_path_keeps_upstream_exception(self, fixture_copy):
        # strict superset: stock behavior unchanged
        with pytest.raises(InvalidFileException):
            load_workbook(fixture_copy("legacy/legacy.xls"))
        with pytest.raises(InvalidFileException):
            load_workbook(fixture_copy("legacy/binary.xlsb"))


class TestLossInventoryCompleteness:
    """The damage classes the v0 scan provably missed —
    each was verified silently damaged by a stock save with zero warnings
    in the post-v0 review."""

    def _surgical(self, fixture_copy, tmp_path):
        import zipfile

        src = fixture_copy("minimal/minimal_clean.xlsx")
        out = str(tmp_path / "missed_classes.xlsx")
        with zipfile.ZipFile(src) as zin, zipfile.ZipFile(out, "w") as zout:
            for name in zin.namelist():
                payload = zin.read(name)
                if name == "xl/workbook.xml":
                    payload = payload.replace(
                        b"<sheets>",
                        b'<fileSharing readOnlyRecommended="1"/><sheets>', 1)
                if name.startswith("xl/worksheets/sheet"):
                    payload = payload.replace(
                        b'</sheetData>',
                        b'<row r="9"><c r="A9" t="inlineStr"><is><r><rPr>'
                        b'<b/></rPr><t>bold run</t></r></is></c></row>'
                        b'</sheetData>', 1)
                    # protectedRanges is a WORKSHEET element, not a
                    # workbook.xml one
                    payload = payload.replace(
                        b'</sheetData>',
                        b'</sheetData><protectedRanges><protectedRange '
                        b'name="pr1" sqref="A1:B2"/></protectedRanges>', 1)
                zout.writestr(name, payload)
            zout.writestr("xl/threadedComments/threadedComment1.xml",
                          b'<?xml version="1.0"?><ThreadedComments/>')
        return out

    def test_previously_missed_classes_are_inventoried(
            self, fixture_copy, tmp_path):
        src = self._surgical(fixture_copy, tmp_path)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = load_workbook(src)
        kinds = wb._paper_loss_inventory.kinds()
        assert "rich-text" in kinds
        assert "workbook-content" in kinds       # fileSharing
        assert "worksheet-content" in kinds      # protectedRanges
        assert "threaded-comments" in kinds
        details = " ".join(l["detail"]
                           for l in wb._paper_loss_inventory.losses)
        assert "fileSharing" in details
        assert "protected ranges" in details

    def test_chart_auxiliary_parts_are_inventoried(self, fixture_copy):
        # colors1.xml/style1.xml: the v0 endswith("colors.xml") never
        # matched real producers' numbered names (dead code, verified —
        # the stock save drops both parts)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = load_workbook(
                fixture_copy("features/lo_authored.xlsx"))
        aux = [l["location"] for l in wb._paper_loss_inventory.losses
               if l["kind"] == "chart-auxiliary"]
        assert "xl/charts/colors1.xml" in aux
        assert "xl/charts/style1.xml" in aux

    def test_preserve_keeps_every_missed_class_verbatim(
            self, fixture_copy, tmp_path):
        # the counterpart honesty claim: under preserve these classes
        # survive byte-identical even alongside a cell edit
        src = self._surgical(fixture_copy, tmp_path)
        wb = load_workbook(src, preserve=True)
        wb["Sheet1"]["A2"] = "green apples"
        out = str(tmp_path / "kept.xlsx")
        wb.save(out)
        before = part_payloads(src)
        after = part_payloads(out)
        assert after["xl/workbook.xml"] == before["xl/workbook.xml"]
        assert after["xl/threadedComments/threadedComment1.xml"] == \
            before["xl/threadedComments/threadedComment1.xml"]
        sheet = next(p for n, p in after.items()
                     if n.startswith("xl/worksheets/"))
        assert b"<r><rPr><b/></rPr><t>bold run</t></r>" in sheet


class TestProducerFingerprint:
    """What we stamp into app.xml is PINNED — the field
    incident on record is Excel rendering charts differently on the
    producer string alone, a class LibreOffice smoke checks are blind to.
    Changing this string is a reviewed decision, never drive-by."""

    PINNED_APP_XML = (
        b'<Properties xmlns="http://schemas.openxmlformats.org/'
        b'officeDocument/2006/extended-properties">'
        b'<Application>Microsoft Excel Compatible / Openpyxl 3.1.5'
        b'</Application><AppVersion>3.1</AppVersion></Properties>'
    )

    def test_fresh_workbook_app_xml_is_pinned(self, tmp_path):
        import io
        import zipfile

        from openpyxl import Workbook

        buf = io.BytesIO()
        Workbook().save(buf)
        payload = zipfile.ZipFile(buf).read("docProps/app.xml")
        assert payload.split(b"?>")[-1].lstrip() == self.PINNED_APP_XML

    def test_preserved_file_keeps_original_app_xml_through_edits(
            self, fixture_copy, tmp_path):
        # the no-op property covers zero-edit saves; this pins the
        # EDITED-save case (app.xml is never sanctioned collateral)
        src = fixture_copy("features/lo_authored.xlsx")
        before = part_payloads(src)["docProps/app.xml"]
        wb = load_workbook(src, preserve=True)
        ws = wb.worksheets[0]
        ws.cell(row=1, column=1,
                value=ws.cell(row=1, column=1).value)
        out = str(tmp_path / "o.xlsx")
        wb.save(out)
        assert part_payloads(out)["docProps/app.xml"] == before
